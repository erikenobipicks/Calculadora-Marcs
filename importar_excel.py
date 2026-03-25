"""
Importa el catàleg de l'Excel a la base de dades (SQLite local o PostgreSQL).
Executa: python importar_excel.py [fitxer.xlsx]
"""
import sys, os
import openpyxl

DATABASE_URL = os.environ.get('DATABASE_URL', '')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

USE_PG = bool(DATABASE_URL)

if USE_PG:
    import psycopg2
    import psycopg2.extras
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = False
    c = conn.cursor()
    def upsert(table, cols, vals):
        placeholders = ','.join(['%s']*len(vals))
        col_str = ','.join(cols)
        c.execute(f"INSERT INTO {table} ({col_str}) VALUES ({placeholders}) ON CONFLICT ({cols[0]}) DO UPDATE SET " +
                  ','.join([f"{col}=EXCLUDED.{col}" for col in cols[1:]]), vals)
else:
    import sqlite3
    DB = os.path.join(os.path.dirname(__file__), 'objectiu.db')
    conn = sqlite3.connect(DB)
    c = conn.cursor()
    def upsert(table, cols, vals):
        placeholders = ','.join(['?']*len(vals))
        c.execute(f"INSERT OR REPLACE INTO {table} VALUES ({placeholders})", vals)

excel_path = sys.argv[1] if len(sys.argv) > 1 else 'Marcs_Objectiu_2026.xlsx'
if not os.path.exists(excel_path):
    print(f"ERROR: No trobo {excel_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_path, data_only=True)
print(f"Excel obert: {excel_path}")
print(f"Mode: {'PostgreSQL' if USE_PG else 'SQLite'}")

# Moldures
ws = wb['Lista Molduras']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ref = row[0]
    if not ref or str(ref).strip() in ('', 'Referencia', '-'): continue
    ref   = str(ref).strip()
    preu  = float(row[1]) if row[1] else 0
    gruix = float(row[2]) if row[2] else 0
    cost  = float(row[3]) if row[3] else 0
    prov  = str(row[4]).strip() if row[4] else ''
    ref2  = str(row[5]).strip() if row[5] else ''
    ubi   = str(row[6]).strip() if row[6] else ''
    desc  = str(row[7]).strip() if row[7] else ''
    upsert('moldures', ['referencia','preu_taller','gruix','cost','proveidor','ref2','ubicacio','descripcio','foto'],
           [ref, preu, gruix, cost, prov, ref2, ubi, desc, None])
    count += 1
print(f"Moldures: {count}")

# Vidres
ws = wb['Vidres']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ref, preu = row[0], row[1]
    if not ref or not preu: continue
    upsert('vidres', ['referencia','preu'], [str(ref).strip(), float(preu)])
    count += 1
print(f"Vidres: {count}")

# Passpartout
ws = wb['Passpertout']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ref, preu = row[0], row[1]
    if not ref or not preu: continue
    upsert('passpartout', ['referencia','preu'], [str(ref).strip(), float(preu)])
    count += 1
print(f"Passpartout: {count}")

# Encolat
ws = wb['Encolado_Pro']
count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    ref, preu = row[0], row[1]
    if not ref or not preu: continue
    upsert('encolat_pro', ['referencia','preu'], [str(ref).strip(), float(preu)])
    count += 1
print(f"Encolat: {count}")

conn.commit()
conn.close()
print("Import completat!")
