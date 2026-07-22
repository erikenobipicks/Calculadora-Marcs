import base64, hashlib, hmac, secrets, os, json, re, time, unicodedata, math, ipaddress
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, g, has_request_context)
from datetime import datetime, timedelta
from functools import wraps
from urllib.parse import urlencode, quote as urllib_quote
from urllib import error as urllib_error
from urllib import request as urllib_request
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os as _os
_FONT_DIR = _os.path.join(_os.path.dirname(__file__), 'static', 'fonts')
pdfmetrics.registerFont(TTFont('DejaVu', _os.path.join(_FONT_DIR, 'DejaVuSans.ttf')))
pdfmetrics.registerFont(TTFont('DejaVu-Bold', _os.path.join(_FONT_DIR, 'DejaVuSans-Bold.ttf')))
import io

app = Flask(__name__)
# SECRET_KEY és obligatori. Abans hi havia un fallback a
# secrets.token_hex(32), cosa que feia que cada restart generés un secret
# nou i descarregués totes les sessions existents. Ara exigim que estigui
# a l'env i fallem a l'arrencada si falta. És preferible un crash
# explícit a la fase de deploy que un funcionament inestable.
_SECRET_KEY = (os.environ.get('SECRET_KEY') or '').strip()
if not _SECRET_KEY:
    raise RuntimeError(
        "SECRET_KEY env var is required. Set it on the deploy "
        "(Railway → Variables → SECRET_KEY) with a long random string, "
        "e.g. `python3 -c 'import secrets; print(secrets.token_urlsafe(64))'`."
    )
app.secret_key = _SECRET_KEY

# Railway (i altres proxies) envien HTTPS al client però HTTP cap al Flask.
# ProxyFix fa que Flask respecti X-Forwarded-Proto/Host per detectar HTTPS.
# Imprescindible si SESSION_COOKIE_SECURE=True, sinó les cookies no es creen.
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# Flags de cookie bàsics, sempre. Abans només s'activaven quan hi havia
# SESSION_COOKIE_DOMAIN, deixant configuracions sense subdomini amb
# cookies sense Secure. ProxyFix (vegeu dalt) ja garanteix que Flask
# detecta HTTPS correctament darrere del proxy.
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# Cookie compartida amb reusrevela.cat per SSO entre subdominis
# (.reusrevela.cat permet compartir entre reusrevela.cat i calculadora.reusrevela.cat)
_COOKIE_DOMAIN = os.environ.get('SESSION_COOKIE_DOMAIN', '').strip() or None
if _COOKIE_DOMAIN:
    app.config['SESSION_COOKIE_DOMAIN'] = _COOKIE_DOMAIN


# Capcaleres de seguretat + noindex global. Aquesta eina es interna (login),
# no s'ha d'indexar mai; i afegim defensa-en-profunditat contra clickjacking,
# MIME-sniffing i downgrade a HTTP. La CSP permet 'unsafe-inline' perque les
# plantilles tenen JS/CSS inline i handlers onclick (no es pot endurir sense
# refactoritzar-les); tot i aixi restringeix l'origen dels recursos.
_CSP = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline'; "
    "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
    "font-src 'self' https://fonts.gstatic.com data:; "
    "img-src 'self' data: https:; "
    "connect-src 'self'; "
    "form-action 'self' https://reusrevela.cat https://reusrevela.es; "
    "base-uri 'self'; "
    "frame-ancestors 'none'"
)


@app.after_request
def _security_headers(resp):
    resp.headers.setdefault('X-Frame-Options', 'DENY')
    resp.headers.setdefault('X-Content-Type-Options', 'nosniff')
    resp.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    resp.headers.setdefault('Strict-Transport-Security', 'max-age=31536000; includeSubDomains')
    resp.headers.setdefault('Content-Security-Policy', _CSP)
    # Eina interna amb login: mai indexable per cap cercador.
    resp.headers.setdefault('X-Robots-Tag', 'noindex, nofollow')
    return resp


@app.route('/robots.txt')
def robots_txt():
    from flask import Response
    return Response("User-agent: *\nDisallow: /\n", mimetype='text/plain')


MOLDURA_IMAGE_EXTS = ('jpg', 'jpeg', 'png', 'webp', 'gif')
MOLDURA_COLOR_FILTERS = [
    ('daurat', 'Daurades'),
    ('plata', 'Plata'),
    ('negre', 'Negres'),
    ('blanc', 'Blanques'),
    ('marro', 'Marrons'),
    ('fusta', 'Fusta / natural'),
    ('gris', 'Grises'),
    ('blau', 'Blaves'),
    ('verd', 'Verdes'),
    ('vermell', 'Vermelles'),
]
MOLDURA_COLOR_KEYWORDS = {
    'daurat': ('daurat', 'daurada', 'dorat', 'dorada', 'or ', ' oro', 'oro ', ' orro', 'or vell'),
    'plata': ('plata', 'argent', 'silver'),
    'negre': ('negre', 'negra', 'negro', 'mate'),
    'blanc': ('blanc', 'blanca', 'blanco', 'blanca'),
    'marro': ('marron', 'marró', 'caoba', 'wengé', 'wenge', 'noguera'),
    'fusta': ('fusta', 'natural', 'pi', 'pino', 'roure', 'roble', 'bambu', 'canya', 'fresno'),
    'gris': ('gris', 'antracita'),
    'blau': ('blau', 'azul', 'marino'),
    'verd': ('verd', 'verde', 'oliva'),
    'vermell': ('vermell', 'rojo', 'granate', 'burdeos'),
}
MOLDURA_COLOR_KEYWORDS['marro'] = ('marro', 'marron', 'caoba', 'wenge', 'noguera')
MOLDURA_GRUIX_FILTERS = [
    ('fina', 'Fines fins a 2 cm'),
    ('mitjana', 'Mitjanes de 2 a 4 cm'),
    ('gruixuda', 'Gruixudes de 4 a 6 cm'),
    ('extra', 'Extra de mes de 6 cm'),
]

COMMERCIAL_MARGIN_DEFAULTS = {
    'general': 60.0,
    'frames': 60.0,
    'canvas': 60.0,
    'prints': 60.0,
    'foam': 60.0,
    'laminate_foam': 60.0,
    'fine_art': 60.0,
    'albums': 60.0,
}

DEFAULT_BRAND_COLOR = '#1A6B45'
DEFAULT_BRAND_SECONDARY_COLOR = '#C8873A'
DEFAULT_BRAND_MENU_COLOR = DEFAULT_BRAND_COLOR


def _normalize_hex_color(value, default=DEFAULT_BRAND_COLOR):
    text = str(value or '').strip()
    if not text:
        return default
    if not text.startswith('#'):
        text = '#' + text
    if len(text) != 7:
        return default
    try:
        int(text[1:], 16)
    except ValueError:
        return default
    return text.upper()


def _mix_with_white(hex_color, ratio=0.88):
    base = _normalize_hex_color(hex_color)
    ratio = min(max(float(ratio), 0.0), 1.0)
    r = int(base[1:3], 16)
    g = int(base[3:5], 16)
    b = int(base[5:7], 16)
    nr = int(r + (255 - r) * ratio)
    ng = int(g + (255 - g) * ratio)
    nb = int(b + (255 - b) * ratio)
    return f'#{nr:02X}{ng:02X}{nb:02X}'


def _mix_hex(hex_color, target_hex, ratio=0.5):
    base = _normalize_hex_color(hex_color)
    target = _normalize_hex_color(target_hex, '#FFFFFF')
    ratio = min(max(float(ratio), 0.0), 1.0)
    r = int(base[1:3], 16)
    g = int(base[3:5], 16)
    b = int(base[5:7], 16)
    tr = int(target[1:3], 16)
    tg = int(target[3:5], 16)
    tb = int(target[5:7], 16)
    nr = int(r + (tr - r) * ratio)
    ng = int(g + (tg - g) * ratio)
    nb = int(b + (tb - b) * ratio)
    return f'#{nr:02X}{ng:02X}{nb:02X}'


def _hex_luminance(hex_color):
    base = _normalize_hex_color(hex_color)
    r = int(base[1:3], 16) / 255.0
    g = int(base[3:5], 16) / 255.0
    b = int(base[5:7], 16) / 255.0
    return (0.2126 * r) + (0.7152 * g) + (0.0722 * b)


def _contrast_text_color(hex_color):
    return '#1C1B18' if _hex_luminance(hex_color) >= 0.62 else '#FFFFFF'


def _current_brand_palette():
    brand_color = DEFAULT_BRAND_COLOR
    secondary_color = DEFAULT_BRAND_SECONDARY_COLOR
    menu_color = DEFAULT_BRAND_MENU_COLOR
    if has_request_context():
        brand_color = _normalize_hex_color(session.get('brand_color', DEFAULT_BRAND_COLOR))
        secondary_color = _normalize_hex_color(
            session.get('brand_color_secondary', DEFAULT_BRAND_SECONDARY_COLOR),
            DEFAULT_BRAND_SECONDARY_COLOR,
        )
        menu_color = _normalize_hex_color(
            session.get('brand_color_menu', brand_color or DEFAULT_BRAND_MENU_COLOR),
            brand_color or DEFAULT_BRAND_MENU_COLOR,
        )
    nav_text_color = _contrast_text_color(menu_color)
    nav_muted_color = (
        _mix_hex(menu_color, '#FFFFFF', 0.62)
        if nav_text_color == '#FFFFFF'
        else _mix_hex(menu_color, '#1C1B18', 0.58)
    )
    nav_pill_color = (
        'rgba(255,255,255,.14)'
        if nav_text_color == '#FFFFFF'
        else 'rgba(28,27,24,.10)'
    )
    return {
        'brand_color': brand_color,
        'brand_color_light': _mix_with_white(brand_color, 0.88),
        'brand_secondary_color': secondary_color,
        'brand_secondary_color_light': _mix_with_white(secondary_color, 0.86),
        'nav_color': menu_color,
        'nav_text_color': nav_text_color,
        'nav_muted_color': nav_muted_color,
        'nav_pill_color': nav_pill_color,
    }


@app.context_processor
def inject_brand_theme():
    return _current_brand_palette()


@app.context_processor
def inject_pendents_albara():
    if session.get('is_admin'):
        row = query('''SELECT COUNT(*) as n FROM comandes
                       WHERE observacions LIKE '%[ACCEPTAT]%'
                         AND (fd_albara IS NULL OR fd_albara='')''', one=True)
        return {'nav_pendents_albara': row['n'] if row else 0}
    return {'nav_pendents_albara': 0}

LAMINATE_ONLY_PRICES = {
    '20x30': 7.35,
    '24x30': 7.35,
    '24x36': 8.15,
    '28x35': 9.55,
    '30x30': 9.55,
    '30x40': 9.95,
    '30x45': 10.35,
    '30x50': 11.35,
    '35x50': 12.20,
    '30x60': 12.60,
    '40x50': 12.50,
    '40x60': 13.90,
    '50x50': 16.40,
    '50x60': 21.30,
    '50x70': 22.00,
    '50x75': 22.60,
    '50x80': 24.20,
    '60x60': 22.00,
    '60x70': 20.15,
    '60x80': 23.50,
    '60x90': 25.95,
    '60x100': 31.00,
    '70x100': 33.00,
    '80x100': 35.00,
    '80x120': 31.00,
    '80x150': 36.00,
    '80x180': 56.75,
    '80x200': 60.00,
    '90x100': 24.80,
    '90x120': 27.55,
    '90x150': 36.25,
    '90x180': 59.00,
    '90x200': 74.65,
    '100x100': 32.00,
    '100x150': 36.75,
    '100x200': 85.00,
}


def _safe_moldura_ref(ref):
    ref = (ref or '').strip().lower()
    safe = ''.join(ch if ch.isalnum() else '-' for ch in ref)
    return safe.strip('-')


def _to_public_photo_url(value):
    value = str(value or '').strip().replace('\\', '/')
    if not value:
        return ''
    if value.startswith(('http://', 'https://', 'data:', '/')):
        return value
    if value.startswith('static/'):
        return '/' + value
    full = os.path.join(app.root_path, value.lstrip('/').replace('/', os.sep))
    if os.path.isfile(full):
        return '/' + value.lstrip('/')
    return ''


def _find_local_moldura_photo(ref):
    fotos_dir = os.path.join(app.root_path, 'static', 'fotos')
    if not ref or not os.path.isdir(fotos_dir):
        return ''

    ref_strip = ref.strip()
    # Supplier codes use separators: "1717-206" or "4017/530" → photo is "1717206.jpg"
    ref_clean = ref_strip.replace('-', '').replace('/', '')

    candidates = []
    for stem in [ref_strip, ref_strip.lower(), ref_strip.upper(), _safe_moldura_ref(ref_strip),
                 ref_clean, ref_clean.lower()]:
        if stem and stem not in candidates:
            candidates.append(stem)

    # Handle leading-zero suffix: "1815-0171" (clean "18150171") → also try "1815171"
    # "7184/0086" (clean "71840086") → also try "7184086" (remove exactly ONE leading zero)
    if len(ref_clean) == 8 and ref_clean[:4].isdigit() and ref_clean[4] == '0':
        alt = ref_clean[:4] + ref_clean[5:]  # remove the single leading zero at position 4
        if alt and alt not in candidates:
            candidates.append(alt)

    for stem in candidates:
        for ext in MOLDURA_IMAGE_EXTS:
            path = os.path.join(fotos_dir, f'{stem}.{ext}')
            if os.path.isfile(path):
                return f'/static/fotos/{stem}.{ext}'
    return ''


PASSPARTOU_IMAGE_EXTS = ('jpg', 'jpeg', 'png', 'webp', 'gif')

def _passpartou_photo_url(ref):
    """Retorna la URL pública de la foto d'un passpartú, o '' si no n'hi ha.
    Cerca a /static/passpartous/ qualsevol fitxer el nom del qual comenci
    per la referència (case-insensitive). Permet noms tipus
    'P001 - Blanc Cru.jpeg' o 'P001.jpg' indistintament."""
    if not ref:
        return ''
    ref_up = ref.strip().upper()
    folder = os.path.join(app.root_path, 'static', 'passpartous')
    if not os.path.isdir(folder):
        return ''
    try:
        entries = os.listdir(folder)
    except OSError:
        return ''
    # Mira primer matches exactes "REF.ext", després "REF " o "REF-"
    for fname in entries:
        stem, dot, ext = fname.rpartition('.')
        if not dot or ext.lower() not in PASSPARTOU_IMAGE_EXTS:
            continue
        if stem.upper() == ref_up:
            return f'/static/passpartous/{fname}'
    for fname in entries:
        stem, dot, ext = fname.rpartition('.')
        if not dot or ext.lower() not in PASSPARTOU_IMAGE_EXTS:
            continue
        s_up = stem.upper().replace('  ', ' ').strip()
        if (s_up.startswith(ref_up + ' ') or
            s_up.startswith(ref_up + '-') or
            s_up.startswith(ref_up + ' -')):
            return f'/static/passpartous/{fname}'
    return ''


def _resolve_moldura_photo(ref, foto, ref2=''):
    """Resolve photo for a moldura. Tries: stored foto URL, then local file by
    referencia (stripping separators), then local file by ref2 (supplier code)."""
    public_url = _to_public_photo_url(foto)
    if public_url:
        return public_url
    local = _find_local_moldura_photo(ref)
    if local:
        return local
    if ref2:
        local2 = _find_local_moldura_photo(ref2)
        if local2:
            return local2
    return ''


def _serialize_moldura(row):
    if not row:
        return None
    data = dict(row)
    data['foto'] = _resolve_moldura_photo(
        data.get('referencia', ''), data.get('foto', ''), ref2=data.get('ref2', ''))
    # Stock de marcs: valors derivats en metres per a la UI.
    sc = data.get('stock_cm')
    smin = data.get('stock_min_cm')
    data['stock_controlat'] = sc is not None
    data['stock_m'] = _cm_a_m(sc)
    data['stock_min_m'] = _cm_a_m(smin)
    data['stock_baix'] = (sc is not None and smin is not None and float(sc) < float(smin))
    return data


def _serialize_moldures(rows):
    return [_serialize_moldura(row) for row in (rows or [])]


def _normalize_text(value):
    text = str(value or '').strip().lower()
    return ''.join(
        ch for ch in unicodedata.normalize('NFKD', text)
        if not unicodedata.combining(ch)
    )


def _matches_moldura_color(desc, color):
    if not color:
        return True
    keywords = MOLDURA_COLOR_KEYWORDS.get(color, ())
    text = ' ' + _normalize_text(desc) + ' '
    return any(keyword in text for keyword in keywords)


def _matches_moldura_gruix(gruix, bucket):
    value = float(gruix or 0)
    if not bucket:
        return True
    if bucket == 'fina':
        return value <= 2
    if bucket == 'mitjana':
        return 2 < value <= 4
    if bucket == 'gruixuda':
        return 4 < value <= 6
    if bucket == 'extra':
        return value > 6
    return True


def _query_moldures(q='', proveidor='', cols='*'):
    sql = f"SELECT {cols} FROM moldures"
    args = []
    where = []

    if q:
        like = f'%{q.lower()}%'
        where.append("""(
            LOWER(referencia) LIKE ?
            OR LOWER(COALESCE(descripcio, '')) LIKE ?
            OR LOWER(COALESCE(proveidor, '')) LIKE ?
        )""")
        args.extend([like, like, like])

    if proveidor:
        where.append("proveidor=?")
        args.append(proveidor)

    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += " ORDER BY referencia"
    return query(sql, args)


def _moldura_photo_path(ref, foto):
    public_url = _resolve_moldura_photo(ref, foto)
    if not public_url.startswith('/static/'):
        return ''
    rel = public_url[len('/static/'):]
    full = os.path.join(app.root_path, 'static', rel.replace('/', os.sep))
    return full if os.path.isfile(full) else ''


def _save_moldura_photo(upload, referencia):
    if not upload or not getattr(upload, 'filename', ''):
        return ''

    filename = upload.filename.rsplit('/', 1)[-1].rsplit('\\', 1)[-1]
    if '.' not in filename:
        raise ValueError('La imatge ha de tenir extensio (jpg, png, webp o gif).')

    ext = filename.rsplit('.', 1)[-1].lower()
    if ext not in MOLDURA_IMAGE_EXTS:
        raise ValueError('Format d\'imatge no valid. Usa JPG, PNG, WEBP o GIF.')

    stem = _safe_moldura_ref(referencia) or 'moldura'
    fotos_dir = os.path.join(app.root_path, 'static', 'fotos')
    os.makedirs(fotos_dir, exist_ok=True)

    for old_ext in MOLDURA_IMAGE_EXTS:
        old_path = os.path.join(fotos_dir, f'{stem}.{old_ext}')
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass

    upload.save(os.path.join(fotos_dir, f'{stem}.{ext}'))
    return f'/static/fotos/{stem}.{ext}'

def _iter_conflict_check_files(base):
    """Yield text/source files where merge markers should never appear."""
    exts = {'.py', '.html', '.css', '.js', '.md', '.txt', '.toml', '.yml', '.yaml'}
    skip_dirs = {'.git', '__pycache__', '.pytest_cache', '.mypy_cache', '.venv', 'venv'}

    for root, dirs, files in os.walk(base):
        dirs[:] = [d for d in dirs if d not in skip_dirs]
        for name in files:
            if os.path.splitext(name)[1].lower() in exts:
                yield os.path.join(root, name)


def _assert_no_conflict_markers():
    """Fail fast if unresolved merge conflict markers are present in source files."""
    base = os.path.dirname(__file__)
    found = []

    for path in _iter_conflict_check_files(base):
        rel = os.path.relpath(path, base)
        try:
            with open(path, 'r', encoding='utf-8') as f:
                for n, line in enumerate(f, start=1):
                    stripped = line.strip()
                    if line.startswith('<<<<<<< ') or line.startswith('>>>>>>> '):
                        found.append(f"{rel}:{n}: {stripped}")
                    elif line.startswith('=======') and stripped == '=======':
                        found.append(f"{rel}:{n}: =======")
        except (FileNotFoundError, UnicodeDecodeError):
            continue

    if found:
        details = "\n".join(found[:20])
        more = f"\n... and {len(found)-20} more" if len(found) > 20 else ""
        raise RuntimeError(f"Merge conflict markers detected:\n{details}{more}")

_assert_no_conflict_markers()

# â"€â"€ DB layer: PostgreSQL (production) or SQLite (local) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
DATABASE_URL = os.environ.get('DATABASE_URL')

if DATABASE_URL:
    import psycopg2
    import psycopg2.extras
    # Railway uses postgres:// but psycopg2 needs postgresql://
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    USE_PG = True
else:
    import sqlite3
    USE_PG = False
    DB = os.path.join(os.path.dirname(__file__), 'objectiu.db')

def get_db():
    if 'db' not in g:
        if USE_PG:
            g.db = psycopg2.connect(
                DATABASE_URL,
                cursor_factory=psycopg2.extras.RealDictCursor,
                connect_timeout=10,
                options='-c statement_timeout=30000',
            )
        else:
            g.db = sqlite3.connect(DB)
            g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db:
        try: db.close()
        except: pass

def _fix_sql(sql):
    """Convert SQLite ? placeholders to PostgreSQL %s, and escape % in LIKE"""
    if USE_PG:
        import re as _re
        # First escape existing % that are part of LIKE patterns (not already %s or %%)
        # Replace LIKE '...%...' patterns to use %%
        sql = _re.sub(r"LIKE '([^']*)'", lambda m: "LIKE '" + m.group(1).replace('%','%%') + "'", sql)
        sql = sql.replace('?', '%s')
    return sql

def query(sql, args=(), one=False):
    db = get_db()
    sql = _fix_sql(sql)
    if USE_PG:
        # Rollback any failed transaction before executing
        if db.status != 1:  # 1 = STATUS_READY
            try: db.rollback()
            except: pass
        cur = db.cursor()
        cur.execute(sql, list(args))
        r = cur.fetchall()
        # Convert to list of dicts
        r = [dict(row) for row in r]
        return (r[0] if r else None) if one else r
    else:
        cur = db.execute(sql, args)
        r = cur.fetchall()
        return (r[0] if r else None) if one else r

def _pg_fix_insert(sql):
    """Convert SQLite INSERT OR IGNORE/REPLACE to PostgreSQL syntax"""
    sql = sql.replace('INSERT OR IGNORE INTO', 'INSERT INTO').replace(
          'INSERT OR REPLACE INTO', 'INSERT INTO')
    # Add ON CONFLICT DO NOTHING for IGNORE (will be added below per-call)
    return sql

def execute(sql, args=()):
    db = get_db()
    if USE_PG:
        sql2 = _fix_sql(sql)
        is_ignore  = 'INSERT OR IGNORE'  in sql.upper()
        is_replace = 'INSERT OR REPLACE' in sql.upper()
        sql2 = _pg_fix_insert(sql2)
        if 'INSERT' in sql2.upper():
            sql2 = sql2.rstrip().rstrip(';')
            if is_ignore:
                sql2 += ' ON CONFLICT DO NOTHING'
            elif is_replace:
                # SQLite "INSERT OR REPLACE" → Postgres UPSERT. Assumim que la
                # primera columna del llistat és la PK (cas de config.clau, etc.).
                m = re.search(r'INSERT\s+INTO\s+\w+\s*\(([^)]+)\)\s*VALUES', sql2, re.IGNORECASE)
                if m:
                    cols = [c.strip().strip('"').strip("'") for c in m.group(1).split(',')]
                    if len(cols) >= 2:
                        pk = cols[0]
                        update_set = ', '.join('{c} = EXCLUDED.{c}'.format(c=c) for c in cols[1:])
                        sql2 += ' ON CONFLICT (' + pk + ') DO UPDATE SET ' + update_set
            # Only add RETURNING id for tables that actually have a SERIAL id
            # column. Tables with TEXT primary key (referencia) — moldures,
            # vidres, passpartout, encolat_pro, impressio — would fail on
            # 'column id does not exist'.
            has_id = any(t in sql2.upper() for t in [
                'INTO USUARIS', 'INTO COMANDES',
                'INTO FEEDBACK', 'INTO HISTORIAL_PREUS_COST',
                'INTO LAB_SENDS',
            ])
            if has_id and 'RETURNING' not in sql2.upper():
                sql2 += ' RETURNING id'
            cur = db.cursor()
            cur.execute(sql2, list(args))
            db.commit()
            if has_id:
                row = cur.fetchone()
                return row['id'] if row else None
            return None
        else:
            cur = db.cursor()
            cur.execute(sql2, list(args))
            db.commit()
            return None
    else:
        cur = db.execute(sql, args)
        db.commit()
        return cur.lastrowid

from werkzeug.security import generate_password_hash, check_password_hash


def hash_pw(pw):
    """Hash d'una contrasenya amb pbkdf2-sha256 + sal aleatòria (werkzeug).
    Els hashes resultants comencen per 'pbkdf2:sha256:...' i contenen '$'
    separant els camps, cosa que els distingeix dels hashes legacy (64
    chars hex).
    """
    return generate_password_hash(pw or '', method='pbkdf2:sha256', salt_length=16)


def _hash_pw_legacy(pw):
    """SHA-256 sense sal. Només per verificar hashes antics que encara no
    han migrat. Cap ruta hauria d'escriure res nou amb això."""
    return hashlib.sha256((pw or '').encode()).hexdigest()


def _is_legacy_hash(stored):
    """Un hash legacy és un string de 64 caràcters hex (SHA-256 sense sal
    emmagatzemat com a hexdigest). Els hashes werkzeug sempre contenen
    el separador '$'."""
    if not isinstance(stored, str) or not stored:
        return False
    return '$' not in stored


def verify_pw(stored, provided):
    """True si la contrasenya 'provided' és correcta per al hash 'stored',
    independentment de si és un hash legacy (SHA-256) o un de werkzeug."""
    if not stored or provided is None:
        return False
    if _is_legacy_hash(stored):
        return hmac.compare_digest(stored, _hash_pw_legacy(provided))
    try:
        return check_password_hash(stored, provided)
    except Exception:
        return False


def maybe_upgrade_hash(user_id, stored, provided):
    """Si l'usuari acaba de passar un login amb un hash legacy, aprofitem
    per rehashejar amb el format nou i desar-ho a la BD. Migració
    progressiva: no cal cap operació massiva, simplement a cada login
    correcte es moderniza un usuari més.

    No fa fallar el login si l'UPDATE falla — només registra warning.
    """
    if not _is_legacy_hash(stored) or not user_id:
        return
    try:
        execute('UPDATE usuaris SET password=? WHERE id=?', [hash_pw(provided), user_id])
    except Exception as exc:
        app.logger.warning('password_rehash_failed user_id=%s detail=%s', user_id, exc)


def _is_admin_session():
    return bool(session.get('is_admin'))


# ── F2 · Cicle d'estats de comanda ─────────────────────────────────────────
# Ordre del flux. La clau coincideix amb la classe CSS .estat-<clau> de
# brand.css (etiquetes de color ja existents). 'pagat' NO és un estat: és
# dada de cobrament a part.
COMANDA_ESTATS = [
    ('nou', 'Nou'),
    ('revisar', 'Pendent de revisar'),
    ('produccio', 'En producció'),
    ('material', 'Pendent de material'),
    ('preparat', 'Preparat'),
    ('avisat', 'Avisat client'),
    ('entregat', 'Entregat'),
    ('cancelat', 'Cancel·lat'),
]
COMANDA_ESTAT_KEYS = {k for k, _ in COMANDA_ESTATS}
COMANDA_ESTAT_LABELS = dict(COMANDA_ESTATS)
# Estats considerats "tancats" (no pendents de feina al taller).
COMANDA_ESTATS_TANCATS = {'entregat', 'cancelat'}


def _derive_estat(row):
    """Estat efectiu d'una comanda: el camp `estat` si és vàlid; si no, derivat
    dels flags existents (compat amb files antigues encara sense backfill).
    Mateix mapeig que la migració: entregat→entregat, [ACCEPTAT]→produccio,
    altrament nou."""
    e = str(_row_get(row, 'estat', '') or '').strip().lower()
    if e in COMANDA_ESTAT_KEYS:
        return e
    if _row_get(row, 'entregat', 0):
        return 'entregat'
    if '[ACCEPTAT]' in str(_row_get(row, 'observacions', '') or ''):
        return 'produccio'
    return 'nou'


def _comanda_es_urgent(row, dies=21):
    """Urgent = comanda oberta (no tancada) amb més de `dies` d'antiguitat.
    Sense camp de venciment, és una heurística per a la safata d'accions."""
    if _derive_estat(row) in COMANDA_ESTATS_TANCATS:
        return False
    d = _parse_comanda_date(_row_get(row, 'data'))
    if not d:
        return False
    try:
        return (datetime.now().date() - d).days >= dies
    except Exception:
        return False


def _audit_log(action, target_user_id=None, target_username=None, details=''):
    """Registra una acció administrativa a la taula audit_log. Idempotent
    davant errors (no peta la request si la taula no existeix encara — el
    cas freqüent quan encara no s'ha passat /admin/run-migrations)."""
    actor_id = session.get('user_id')
    actor_un = session.get('username') or ''
    if not actor_id:
        return
    try:
        execute(
            "INSERT INTO audit_log (actor_user_id, actor_username, target_user_id, target_username, action, details) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [actor_id, actor_un, target_user_id, target_username or '', action, details or ''],
        )
    except Exception as e:
        # No volem que un fall a l'audit log faci caure l'acció administrativa.
        try:
            print(f'audit_log skip: {e}')
        except Exception:
            pass


def _row_get(row, key, default=None):
    if row is None:
        return default
    try:
        return row[key]
    except Exception:
        pass
    try:
        return row.get(key, default)
    except Exception:
        return default


def _user_access_status(user):
    status = str(_row_get(user, 'access_status', '') or '').strip().lower()
    return status or 'active'


def _get_marge_value(u):
    """Marge professional efectiu. Prioritat: marge_pro_pct > marge > 60.
    Fa servir 'is not None' en lloc de 'or' perquè un valor explícit de 0
    és legítim (admin, col·laborador intern) i no hauria de caure al
    fallback de 60."""
    val = _row_get(u, 'marge_pro_pct')
    if val is not None:
        return float(val)
    val = _row_get(u, 'marge')
    if val is not None:
        return float(val)
    return 60.0


def _get_marge_impressio_value(u):
    """Marge d'impressió. Prioritat: marge_impressio_pro_pct > marge_impressio > 0.
    Mateix raonament que _get_marge_value — 0 és legítim."""
    val = _row_get(u, 'marge_impressio_pro_pct')
    if val is not None:
        return float(val)
    val = _row_get(u, 'marge_impressio')
    if val is not None:
        return float(val)
    return 0.0


def get_marge_impressio_tram(area_cm2, usuari):
    """Marge d'impressions PVD→PVP per àrea de la foto (cm²).
    6 trams configurables; cada usuari té els seus propis valors a
    usuaris.imp_tram1..imp_tram6. Si en falta algun, cau al default
    global de config (imp_tram{N}_marge_default).

    Retorna {'marge': float, 'tram': int (1-6), 'limit': float|None}."""
    try:
        a = float(area_cm2 or 0)
    except (TypeError, ValueError):
        a = 0.0

    limits = [
        float(get_config_value('imp_tram1_area', '900')),
        float(get_config_value('imp_tram2_area', '2000')),
        float(get_config_value('imp_tram3_area', '4200')),
        float(get_config_value('imp_tram4_area', '6000')),
        float(get_config_value('imp_tram5_area', '14400')),
        float('inf'),
    ]
    user_vals = [_row_get(usuari, f'imp_tram{i}') for i in range(1, 7)]
    defaults = [
        float(get_config_value('imp_tram1_marge_default', '80')),
        float(get_config_value('imp_tram2_marge_default', '75')),
        float(get_config_value('imp_tram3_marge_default', '70')),
        float(get_config_value('imp_tram4_marge_default', '60')),
        float(get_config_value('imp_tram5_marge_default', '50')),
        float(get_config_value('imp_tram6_marge_default', '45')),
    ]
    for i, lim in enumerate(limits):
        if a <= lim:
            v = user_vals[i]
            marge = float(v) if v is not None else defaults[i]
            return {'marge': marge, 'tram': i + 1, 'limit': None if lim == float('inf') else lim}
    # Fallback teòricament inabastable
    return {'marge': defaults[-1], 'tram': 6, 'limit': None}


def _user_is_allowed(user):
    return bool(user) and (bool(_row_get(user, 'is_admin', 0)) or _user_access_status(user) == 'active')


def _clean_profile_type(value):
    allowed = {'professional', 'studio', 'gallery', 'association'}
    value = (value or '').strip().lower()
    return value if value in allowed else 'professional'


def _bridge_api_token():
    return os.environ.get('PUBLIC_BRIDGE_TOKEN', '').strip()


def _bridge_signing_secret():
    return os.environ.get('BRIDGE_LOGIN_SECRET', '').strip() or _bridge_api_token()


def _bridge_tokens_valids():
    """Tokens de bridge acceptats per als endpoints /api/public/*. Suporta
    rotació sense downtime: durant una rotació es defineix
    PUBLIC_BRIDGE_TOKEN_NEXT amb el valor nou mentre el vell
    (PUBLIC_BRIDGE_TOKEN) encara és vàlid; quan la web ja envia el nou, es
    promou NEXT → principal i s'esborra NEXT. Si NEXT no està definit, el
    comportament és idèntic al d'abans (un sol token vàlid)."""
    toks = [
        os.environ.get('PUBLIC_BRIDGE_TOKEN', '').strip(),
        os.environ.get('PUBLIC_BRIDGE_TOKEN_NEXT', '').strip(),
    ]
    return [t for t in toks if t]


def _bridge_token_ok(provided):
    """Compara el token rebut amb els vàlids en temps constant
    (hmac.compare_digest) per evitar el canal lateral de temporització que
    tenia la comparació '!=' anterior. Retorna False si no hi ha cap token
    configurat o si el rebut és buit."""
    provided = (provided or '').strip()
    if not provided:
        return False
    return any(hmac.compare_digest(provided, t) for t in _bridge_tokens_valids())


def _bridge_ip_allowed():
    """Allowlist d'IPs opcional per als endpoints /api/public/*. Si
    BRIDGE_ALLOWED_IPS no està definida (cas per defecte), no restringeix res
    i és totalment retrocompatible. Quan es defineix (CSV d'IPs o xarxes
    CIDR), només es permeten peticions des d'aquestes adreces. request.remote_addr
    ja és la IP real del client gràcies a ProxyFix (x_for=1)."""
    allow = os.environ.get('BRIDGE_ALLOWED_IPS', '').strip()
    if not allow:
        return True
    try:
        client = ipaddress.ip_address((request.remote_addr or '').strip())
    except ValueError:
        return False
    for net in (n.strip() for n in allow.split(',')):
        if not net:
            continue
        try:
            if client in ipaddress.ip_network(net, strict=False):
                return True
        except ValueError:
            continue
    return False


def _main_site_url():
    return os.environ.get('MAIN_SITE_URL', 'https://reusrevela.cat').strip().rstrip('/')


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def get_config_value(clau, default=None):
    """Read a single value from the config key-value table."""
    r = query("SELECT valor FROM config WHERE clau=?", [clau], one=True)
    return r['valor'] if r else default


# ── Extras (càrrecs configurables) ────────────────────────────────────────
EXTRAS_DEFAULTS = [
    {
        "key": "samarreta_surcharge",
        "name": "Càrrec samarreta",
        "description": "Feina addicional per emmarcar samarreta.",
        "price_pvd": 20.0,
        "margin_pct": None,
        "mode": "auto",
        "piece_types": ["samarreta"],
        "actiu": True,
        "ordre": 1,
    },
    {
        "key": "desmuntar_marc",
        "name": "Desmuntar i netejar un marc existent",
        "description": "Quan el client porta un marc i vols aprofitar el vidre (o canviar-lo).",
        "price_pvd": 10.0,
        "margin_pct": None,
        "mode": "manual",
        "piece_types": [],
        "actiu": True,
        "ordre": 2,
    },
]


def _normalize_extra(e):
    """Coerce types so the JS/template gets predictable values."""
    pt = e.get('piece_types') or []
    if isinstance(pt, str):
        pt = [x.strip() for x in pt.split(',') if x.strip()]
    mp = e.get('margin_pct')
    try:
        mp = float(mp) if mp not in (None, '', 'null') else None
    except (TypeError, ValueError):
        mp = None
    try:
        price = float(e.get('price_pvd') or 0)
    except (TypeError, ValueError):
        price = 0.0
    try:
        ordre = int(e.get('ordre') or 0)
    except (TypeError, ValueError):
        ordre = 0
    mode = (e.get('mode') or 'manual').strip().lower()
    if mode not in ('manual', 'auto'):
        mode = 'manual'
    return {
        'key': (e.get('key') or '').strip(),
        'name': (e.get('name') or '').strip(),
        'description': (e.get('description') or '').strip(),
        'price_pvd': price,
        'margin_pct': mp,
        'mode': mode,
        'piece_types': pt,
        'actiu': bool(e.get('actiu', True)),
        'ordre': ordre,
    }


def get_extras_list():
    """Return the list of configured extras, seeding defaults on first read."""
    raw = get_config_value('extras_json')
    if not raw:
        save_extras_list(EXTRAS_DEFAULTS)
        return [dict(e) for e in EXTRAS_DEFAULTS]
    try:
        data = json.loads(raw)
        if not isinstance(data, list):
            return [dict(e) for e in EXTRAS_DEFAULTS]
        return [_normalize_extra(e) for e in data]
    except Exception:
        return [dict(e) for e in EXTRAS_DEFAULTS]


def save_extras_list(extras):
    """Persist extras as JSON in the config table."""
    payload = json.dumps([_normalize_extra(e) for e in extras], ensure_ascii=False)
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('extras_json', ?)", [payload])


def calcular_pvd(preu_cost, categoria):
    """cost × (1 + marge_admin_pct/100). categoria: 'moldures'|'vidres'|'passpartu'|'encolat'"""
    if preu_cost is None:
        return None
    marge = float(get_config_value(f'marge_admin_{categoria}_pct', '60'))
    return round(preu_cost * (1 + marge / 100), 4)


def calcular_preu_marc(amplada, alcada, gruix, preu_cost, merma_pct=10.0, minim_cm=100.0):
    """Calcula cost i PVD d'un marc a partir de preu_cost per cm lineal.
    Retorna dict {'cost': float, 'pvd': float} o None si dades insuficients."""
    if not all([amplada, alcada, preu_cost]):
        return None
    perimetre = 2 * (amplada + alcada)
    long_bruta = (perimetre + (gruix or 0) * 8) * (1 + (merma_pct or 10.0) / 100)
    long_final = max(long_bruta, minim_cm or 100.0)
    cost = round(long_final * preu_cost / 100, 4)
    pvd = calcular_pvd(cost, 'moldures')
    return {'cost': cost, 'pvd': pvd}


# ── Control d'stock de marcs (metres lineals) ───────────────────────────────
# L'stock es desa en CM lineals a moldures.stock_cm. NULL = aquesta referència
# NO es controla per stock (opt-in): no es descompta ni es mostra alerta.
# stock_min_cm és el llindar d'avís (NULL/0 = sense avís). Tot el moviment queda
# registrat a la taula moviments_stock_marc (entrada/sortida/ajust).

def _m_a_cm(valor):
    """Converteix una entrada en metres (string del formulari) a cm. Buit → None."""
    if valor is None:
        return None
    s = str(valor).strip().replace(',', '.')
    if s == '':
        return None
    try:
        return round(float(s) * 100, 2)
    except (TypeError, ValueError):
        return None


def _cm_a_m(cm):
    """cm → metres per mostrar (2 decimals). None → None."""
    if cm is None:
        return None
    try:
        return round(float(cm) / 100, 2)
    except (TypeError, ValueError):
        return None


def _consum_marc_cm(final_w, final_h, gruix, merma_pct=10.0, minim_cm=100.0, gruix_extra=0.0):
    """Llargada de motllura consumida (cm lineals) per a un marc, replicant
    EXACTAMENT la fórmula de la calculadora (calcular() a calculadora.html):
        per       = 2·(final_w + final_h)        # mides EXTERIORS del marc
        long_bruta = (per + (gruix+gruix_extra)·8) · (1 + merma/100)
        long_final = max(long_bruta, minim_cm)
    Per al marc principal gruix_extra=0. Per al premarc, gruix=gruix_pre i
    gruix_extra=gruix_marc (suma dels dos gruixos, com fa la calc)."""
    fw = float(final_w or 0)
    fh = float(final_h or 0)
    if fw <= 0 or fh <= 0:
        return 0.0
    per = 2 * (fw + fh)
    g = float(gruix or 0) + float(gruix_extra or 0)
    long_bruta = (per + g * 8) * (1 + (float(merma_pct) if merma_pct is not None else 10.0) / 100)
    return round(max(long_bruta, float(minim_cm) if minim_cm else 100.0), 2)


def _aplica_moviment_stock(ref, cm, tipus, motiu='', usuari_id=None, albara_num=None, activar=False):
    """Aplica un moviment d'stock a una referència de marc i el registra.

    cm: magnitud en cm lineals (sempre positiu). El signe l'aplica `tipus`:
        'entrada' (+cm, compra), 'sortida' (-cm, consum), 'ajust' (=cm, fixa el valor).
    activar: si True i la referència encara no es controla per stock (stock_cm IS
        NULL), la inicialitza partint de 0 (opt-in). El descompte automàtic NO
        l'activa (activar=False), de manera que només es descompten els marcs que
        l'admin ha donat d'alta explícitament.
    Retorna dict {referencia, tipus, cm, stock_resultant, sota_minim} o None si la
    referència no es controla per stock i no s'activa, o no existeix.
    No llança si hi ha problemes: el descompte automàtic és best-effort."""
    ref = (ref or '').strip()
    if not ref or ref == '-':
        return None
    try:
        m = query('SELECT stock_cm, stock_min_cm FROM moldures WHERE LOWER(referencia)=LOWER(?)',
                  [ref], one=True)
    except Exception:
        return None
    if not m:
        return None
    actual = _row_get(m, 'stock_cm', None)
    if actual is None:
        if not activar:
            # Opt-in: referència sense stock controlat → no fem res.
            return None
        actual = 0.0
    actual = float(actual)
    cm = abs(float(cm or 0))
    if tipus == 'ajust':
        nou = cm
        delta = nou - actual
    elif tipus == 'entrada':
        delta = cm
        nou = actual + cm
    else:  # sortida (consum)
        tipus = 'sortida'
        delta = -cm
        nou = actual - cm  # pot quedar negatiu: avisem però no bloquegem
    nou = round(nou, 2)
    min_cm = _row_get(m, 'stock_min_cm', None)
    sota_minim = (min_cm is not None) and (nou < float(min_cm))
    try:
        execute('UPDATE moldures SET stock_cm=? WHERE LOWER(referencia)=LOWER(?)', [nou, ref])
        execute(
            '''INSERT INTO moviments_stock_marc
               (referencia, data, tipus, cm, motiu, albara_num, usuari_id, stock_resultant)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            [ref, datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
             tipus, round(delta, 2), (motiu or '')[:300], albara_num, usuari_id, nou])
    except Exception as e:
        print('moviment stock err:', e)
        return None
    return {'referencia': ref, 'tipus': tipus, 'cm': round(delta, 2),
            'stock_resultant': nou, 'sota_minim': sota_minim}


def _stock_dades_marc(ref):
    """Retorna (gruix, merma_pct, minim_cm, stock_cm) d'una referència o None."""
    ref = (ref or '').strip()
    if not ref or ref == '-':
        return None
    try:
        m = query('SELECT gruix, merma_pct, minim_cm, stock_cm FROM moldures WHERE LOWER(referencia)=LOWER(?)',
                  [ref], one=True)
    except Exception:
        return None
    if not m:
        return None
    return (float(_row_get(m, 'gruix', 0) or 0),
            float(_row_get(m, 'merma_pct', 10.0) or 10.0),
            float(_row_get(m, 'minim_cm', 100.0) or 100.0),
            _row_get(m, 'stock_cm', None))


def _descompta_stock_albara(marc, pre_marc, final_w, final_h, quantitat, num_albara, usuari_id):
    """Descompta l'stock consumit pel marc principal i el premarc en generar un
    albarà. Best-effort: si una referència no es controla per stock o falla, se
    salta. Retorna llista d'avisos (refs que queden sota mínim o negatives) per
    informar l'admin a la calculadora."""
    avisos = []
    qty = max(1.0, float(quantitat or 1))
    motiu = f'Albarà {num_albara}' if num_albara else 'Albarà'
    # Marc principal
    dm = _stock_dades_marc(marc)
    gruix_marc = dm[0] if dm else 0.0
    if dm and dm[3] is not None:
        cm = _consum_marc_cm(final_w, final_h, dm[0], dm[1], dm[2]) * qty
        res = _aplica_moviment_stock(marc, cm, 'sortida', motiu, usuari_id, num_albara)
        if res and (res['sota_minim'] or res['stock_resultant'] < 0):
            avisos.append({'ref': marc, 'stock_resultant': res['stock_resultant'],
                           'negatiu': res['stock_resultant'] < 0})
    # Premarc (consumeix amb gruix_pre + gruix_marc, com la calc)
    dp = _stock_dades_marc(pre_marc)
    if dp and dp[3] is not None:
        cm = _consum_marc_cm(final_w, final_h, dp[0], dp[1], dp[2], gruix_extra=gruix_marc) * qty
        res = _aplica_moviment_stock(pre_marc, cm, 'sortida', motiu, usuari_id, num_albara)
        if res and (res['sota_minim'] or res['stock_resultant'] < 0):
            avisos.append({'ref': pre_marc, 'stock_resultant': res['stock_resultant'],
                           'negatiu': res['stock_resultant'] < 0})
    return avisos


def _normalize_commercial_margins(raw=None, frame_margin=None, print_margin=None):
    data = raw if isinstance(raw, dict) else {}
    general_margin = max(_safe_float(data.get('general'), frame_margin if frame_margin is not None else COMMERCIAL_MARGIN_DEFAULTS['general']), 0.0)
    frame_value = max(_safe_float(data.get('frames'), frame_margin if frame_margin is not None else general_margin), 0.0)
    print_seed = print_margin if print_margin is not None else frame_value
    print_value = max(_safe_float(data.get('prints'), print_seed), 0.0)
    normalized = {
        'general': general_margin,
        'frames': frame_value,
        'canvas': max(_safe_float(data.get('canvas'), frame_value), 0.0),
        'prints': print_value,
        'foam': general_margin,
        'laminate_foam': general_margin,
        'fine_art': max(_safe_float(data.get('fine_art'), frame_value), 0.0),
        'albums': max(_safe_float(data.get('albums'), frame_value), 0.0),
        # Productes de la calculadora amb marge propi configurable (per defecte 40%).
        'digital': max(_safe_float(data.get('digital'), 40.0), 0.0),
        'orles': max(_safe_float(data.get('orles'), 40.0), 0.0),
        'regals': max(_safe_float(data.get('regals'), 40.0), 0.0),
        'offset': max(_safe_float(data.get('offset'), 40.0), 0.0),
    }
    return normalized


def _load_user_commercial_margins(user_row):
    payload = {}
    raw = user_row['margins_json'] if user_row and 'margins_json' in user_row.keys() else ''
    if raw:
        try:
            payload = json.loads(raw)
        except (TypeError, ValueError):
            payload = {}
    frame_margin = user_row['marge'] if user_row and user_row['marge'] is not None else COMMERCIAL_MARGIN_DEFAULTS['frames']
    print_margin = None
    if user_row and 'marge_impressio' in user_row.keys() and user_row['marge_impressio'] is not None:
        print_margin = user_row['marge_impressio']
    return _normalize_commercial_margins(payload, frame_margin=frame_margin, print_margin=print_margin)


def _format_margin_for_view(value):
    value = _safe_float(value, 0.0)
    return int(value) if float(value).is_integer() else round(value, 2)


def _sync_private_commercial_settings(frame_margin, print_margin, margins=None):
    api_token = _bridge_api_token()
    base = _main_site_url()
    if not api_token or not base:
        return {'attempted': False, 'reason': 'missing_config'}

    normalized = _normalize_commercial_margins(margins or {}, frame_margin=frame_margin, print_margin=print_margin)
    payload = dict(normalized)
    req = urllib_request.Request(
        f'{base}/api/private/commercial-settings-sync',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'Content-Type': 'application/json',
            'X-Bridge-Token': api_token,
            # Cloudflare davant de reusrevela.cat pot bloquejar el UA per
            # defecte de Python-urllib. Enviar un UA propi permet whitelist
            # i fa visibles els hits de sortida.
            'User-Agent': 'calculadora-marcs-bridge/1.0 (+https://calculadora.reusrevela.cat)',
        },
        method='POST',
    )
    try:
        with urllib_request.urlopen(req, timeout=12) as resp:
            body = resp.read().decode('utf-8')
            return {'attempted': True, 'ok': True, 'response': json.loads(body or '{}')}
    except urllib_error.HTTPError as exc:
        detail = exc.read().decode('utf-8', errors='ignore')
        return {'attempted': True, 'ok': False, 'status': exc.code, 'detail': detail}
    except (urllib_error.URLError, TimeoutError, ValueError) as exc:
        return {'attempted': True, 'ok': False, 'detail': str(exc)}


def _urlsafe_b64encode(raw):
    return base64.urlsafe_b64encode(raw).decode().rstrip('=')


def _urlsafe_b64decode(text):
    text = str(text or '').strip()
    if not text:
        return b''
    padding = '=' * (-len(text) % 4)
    return base64.urlsafe_b64decode((text + padding).encode())


def _safe_next_path(value, default='/'):
    text = str(value or '').strip()
    if not text.startswith('/') or text.startswith('//') or '://' in text:
        return default
    return text


def _build_bridge_token(data):
    secret = _bridge_signing_secret()
    if not secret:
        raise RuntimeError('Missing bridge signing secret.')
    payload = _urlsafe_b64encode(json.dumps(data, separators=(',', ':'), ensure_ascii=True).encode())
    signature = _urlsafe_b64encode(hmac.new(secret.encode(), payload.encode(), hashlib.sha256).digest())
    return f'{payload}.{signature}'


def _read_bridge_token(token, max_age=45):
    secret = _bridge_signing_secret()
    if not secret or '.' not in str(token or ''):
        return None

    payload, signature = token.split('.', 1)
    expected = _urlsafe_b64encode(hmac.new(secret.encode(), payload.encode(), hashlib.sha256).digest())
    if not hmac.compare_digest(signature, expected):
        return None

    try:
        data = json.loads(_urlsafe_b64decode(payload).decode())
    except Exception:
        return None

    issued_at = int(data.get('iat') or 0)
    if not issued_at or (time.time() - issued_at) > max_age:
        return None
    return data


def _load_empresa_nom_for_session(user):
    try:
        nom_emp = user['nom_empresa'] if user['nom_empresa'] else ''
        if not nom_emp:
            r_cfg = query("SELECT valor FROM config WHERE clau='empresa_nom'", one=True)
            nom_emp = r_cfg['valor'] if r_cfg and r_cfg['valor'] else ''
        return nom_emp or 'Calculadora'
    except Exception:
        return 'Calculadora'


def _build_web_return_url(source=None, lang=None):
    base = _main_site_url()
    if not base:
        return ''

    source = str(source or '').strip().lower()
    lang = (lang or 'ca').strip().lower() or 'ca'
    path = '/area-privada' if source in {'private_area', 'web_private', 'area_privada', 'web'} else '/'
    if path == '/':
        return f'{base}/?lang={lang}'
    return f'{base}{path}?lang={lang}'


def _current_web_return_url():
    lang = session.get('bridge_lang') or request.args.get('lang') or 'ca'
    base = _main_site_url()
    return f'{base}/area-privada?lang={lang}'


def _current_web_order_url():
    base = _main_site_url()
    lang = (session.get('bridge_lang') or request.args.get('lang') or 'ca').strip().lower() or 'ca'
    return f'{base}/area-privada/comanda?lang={lang}'


def _current_web_module_url(path):
    base = _main_site_url()
    lang = (session.get('bridge_lang') or request.args.get('lang') or 'ca').strip().lower() or 'ca'
    clean_path = '/' + str(path or '').strip().lstrip('/')
    return f'{base}{clean_path}?lang={lang}'


def _needs_setup(user_id):
    try:
        u2 = query('SELECT setup_done FROM usuaris WHERE id=?', [user_id], one=True)
        return bool(u2) and not bool(_row_get(u2, 'setup_done', 0))
    except Exception as exc:
        print(f'setup check error: {exc}')
        return False


def _start_user_session(user):
    # Usem _row_get per a TOTS els camps: si el schema no té una columna (p.ex.
    # brand_color en bases antigues) no volem un 500 durant el login.
    session['user_id'] = _row_get(user, 'id')
    session['username'] = _row_get(user, 'username', '')
    session['is_admin'] = bool(_row_get(user, 'is_admin', 0))
    session['nom'] = _row_get(user, 'nom', '') or ''
    session['access_status'] = _user_access_status(user)
    session['profile_type'] = _clean_profile_type(_row_get(user, 'profile_type', 'professional'))
    session['empresa_nom'] = _load_empresa_nom_for_session(user)
    session['brand_color'] = _normalize_hex_color(_row_get(user, 'brand_color', DEFAULT_BRAND_COLOR))
    session['brand_color_secondary'] = _normalize_hex_color(
        _row_get(user, 'brand_color_secondary', DEFAULT_BRAND_SECONDARY_COLOR),
        DEFAULT_BRAND_SECONDARY_COLOR,
    )
    session['brand_color_menu'] = _normalize_hex_color(
        _row_get(user, 'brand_color_menu', session.get('brand_color', DEFAULT_BRAND_MENU_COLOR)),
        session.get('brand_color', DEFAULT_BRAND_MENU_COLOR),
    )


def _can_access_comanda(row):
    return bool(row) and (_is_admin_session() or row['user_id'] == session.get('user_id'))


def _get_comanda_for_session(cid, *, fields='*'):
    row = query(f'SELECT {fields} FROM comandes WHERE id=?', [cid], one=True)
    if not _can_access_comanda(row):
        return None
    return row


def _get_comanda_by_sessio_for_session(sessio_id, *, fields='id, user_id'):
    row = query(f'SELECT {fields} FROM comandes WHERE sessio_id=? LIMIT 1', [sessio_id], one=True)
    if not _can_access_comanda(row):
        return None
    return row


_INTERMOL_REFS = [
    ('1021271','1021'),('1021298','1021'),('1520136','1520'),('1520170','1520'),('1520180','1520'),
    ('1528003','1528'),('17172011','1717'),('17172012','1717'),('17172013','1717'),('17172014','1717'),
    ('17172015','1717'),('17172016','1717'),('1717206','1717'),('1717212','1717'),('1717216','1717'),
    ('1717261','1717'),('1717263','1717'),('1717268','1717'),('1717270','1717'),('1717272','1717'),
    ('1717281','1717'),('1717295','1717'),('1717330','1717'),('1717336','1717'),('1717374','1717'),
    ('1717380','1717'),('1717540','1717'),('1717541','1717'),('1717543','1717'),('1717549','1717'),
    ('1717653','1717'),('1717673','1717'),('1717975','1717'),('1717985','1717'),('18141008','1814'),
    ('18141010','1814'),('1814640','1814'),('1815171','1815'),('1815181','1815'),('18183012','1818'),
    ('18183016','1818'),('1840003','1840'),('1840007','1840'),('18413011','1841'),('18413012','1841'),
    ('18413013','1841'),('1930536','1930'),('1930540','1930'),('1930975','1930'),('1930985','1930'),
    ('2020003','2020'),('2020007','2020'),('2127172','2127'),('2127485','2127'),('2322481','2322'),
    ('2322536','2322'),('2558640','2558'),('2816530','2816'),('2816536','2816'),('28253012','2825'),
    ('28253016','2825'),('3031003','3031'),('3031007','3031'),('3131530','3131'),('3131536','3131'),
    ('3131540','3131'),('3131653','3131'),('3131673','3131'),('3237171','3237'),('3237181','3237'),
    ('3237675','3237'),('4017380','4017'),('4017530','4017'),('4017540','4017'),('4017546','4017'),
    ('4017549','4017'),('4017638','4017'),('4017985','4017'),('4144171','4144'),('4411629','4411'),
    ('4413629','4413'),('5064614','5064'),('5067714','5067'),('5285714','5285'),('5301183','5301'),
    ('5301806','5301'),('5302667','5302'),('5302806','5302'),('5303624','5303'),('5303720','5303'),
    ('5515201','5515'),('5515203','5515'),('5515238','5515'),('5670236','5670'),('5753616','5753'),
    ('5753617','5753'),('5753718','5753'),('5757718','5757'),('61831021','6183'),('61861022','6186'),
    ('6201571','6201'),('6201581','6201'),('6202670','6202'),('64121019','6214'),('64124010','6214'),
    ('64125004','6214'),('6344082','6344'),('6412626','6412'),('6486171','6486'),('6486181','6486'),
    ('6486629','6486'),('6486727','6486'),('66121024','6612'),('66121025','6612'),('66123008','6612'),
    ('66124011','6612'),('66125006','6612'),('69921027','6992'),('69931027','6993'),('69931028','6993'),
    ('69951026','6995'),('69951028','6995'),('7017638','7017'),('7027627','7027'),('7184076','7184'),
    ('7184086','7184'),('7184602','7184'),('7184702','7184'),('7532581','7532'),('7560481','7560'),
    ('7801471','7801'),('7901480','7901'),('7901870','7901'),('7901979','7901'),('7902976','7902'),
    ('7902979','7902'),('8517374','8517'),('8517651','8517'),('8537171','8537'),('8537181','8537'),
    ('8537655','8537'),('8547603','8547'),('8547675','8547'),('8547703','8547'),('85574001','8557'),
]

_PROECO_PREUS = [
    ('PROECO1520', 7.81),  ('PROECO1824', 9.24),  ('PROECO2020', 7.81),
    ('PROECO2025', 8.80),  ('PROECO2030', 10.67), ('PROECO2323', 8.80),
    ('PROECO2430', 11.55), ('PROECO2436', 12.65), ('PROECO2835', 14.52),
    ('PROECO3030', 14.52), ('PROECO3040', 15.51), ('PROECO3045', 16.50),
    ('PROECO3050', 17.93), ('PROECO3060', 20.35), ('PROECO3550', 19.36),
    ('PROECO4050', 20.35), ('PROECO4060', 21.34), ('PROECO5050', 23.21),
    ('PROECO5060', 25.19), ('PROECO5070', 29.04), ('PROECO5075', 31.90),
    ('PROECO60100', 43.56),('PROECO6060', 29.04), ('PROECO6070', 33.88),
    ('PROECO6080', 36.74), ('PROECO6090', 40.70), ('PROECO70100', 48.40),
    ('PROECO80100', 50.38),('PROECO80120', 55.22),('PROECO80150', 61.05),
    ('PROECO80180', 87.12),('PROECO80200', 116.16),('PROECO90100', 53.24),
    ('PROECO90120', 62.92),('PROECO90150', 72.60),('PROECO90180', 92.95),
    ('PROECO90200', 125.84),
]

def _seed_proeco_preus(db, use_pg=False):
    """Inserta els preus de ProEco si no existeixen."""
    ok = 0
    if use_pg:
        cur = db.cursor()
        for ref, preu in _PROECO_PREUS:
            try:
                cur.execute(
                    "INSERT INTO proeco (referencia, preu) VALUES (%s, %s) ON CONFLICT (referencia) DO NOTHING",
                    [ref, preu]
                )
                ok += 1
            except Exception as e:
                print(f'ProEco seed PG skip {ref}: {e}')
    else:
        for ref, preu in _PROECO_PREUS:
            try:
                db.execute(
                    "INSERT OR IGNORE INTO proeco (referencia, preu) VALUES (?, ?)",
                    [ref, preu]
                )
                ok += 1
            except Exception as e:
                print(f'ProEco seed SQLite skip {ref}: {e}')
    print(f'ProEco seed: {ok}/{len(_PROECO_PREUS)} preus inserits')


def _seed_intermol_moldures(db, use_pg=False):
    """Neteja registres Intermol inserits per error (codi proveïdor com a referència primària).
    Les fotos es resolen dinàmicament via _find_local_moldura_photo: el codi intern
    "1717-206" troba automàticament "1717206.jpg" eliminant el guió/barra.
    """
    refs = [r for r, _ in _INTERMOL_REFS]
    deleted = 0

    if use_pg:
        cur = db.cursor()
        for ref in refs:
            try:
                cur.execute(
                    "DELETE FROM moldures WHERE referencia=%s "
                    "AND preu_taller=0 AND cost=0 AND COALESCE(descripcio,'')=''",
                    [ref]
                )
                deleted += cur.rowcount
            except Exception as e:
                print(f'Intermol cleanup PG {ref}: {e}')
        try:
            db.commit()
        except Exception:
            pass
    else:
        for ref in refs:
            try:
                db.execute(
                    "DELETE FROM moldures WHERE referencia=? "
                    "AND preu_taller=0 AND cost=0 AND COALESCE(descripcio,'')=''",
                    [ref]
                )
            except Exception as e:
                print(f'Intermol cleanup SQLite {ref}: {e}')
        try:
            db.commit()
        except Exception:
            pass
    print(f'Intermol cleanup: {deleted} registres erronis eliminats (fotos resoltes per referencia)')


def _fix_ref2_errors(db, use_pg=False):
    """Corregeix errors de transcripció coneguts al camp ref2 de moldures."""
    fixes = [
        # (referencia, ref2_erroni, ref2_correcte)
        ('M2314', '6183/1022', '6183/1021'),  # foto 61831021.jpg disponible
        ('M2315', '6214/1019', '6412/1019'),  # 6214 era 6412 invertit
        ('M2316', '6214/4010', '6412/4010'),
        ('M2317', '6214/5004', '6412/5004'),
    ]
    fixed = 0
    if use_pg:
        cur = db.cursor()
        for ref, wrong, correct in fixes:
            try:
                cur.execute(
                    "UPDATE moldures SET ref2=%s WHERE referencia=%s AND ref2=%s",
                    [correct, ref, wrong])
                fixed += cur.rowcount
            except Exception as e:
                print(f'ref2 fix PG {ref}: {e}')
        try:
            db.commit()
        except Exception:
            pass
    else:
        for ref, wrong, correct in fixes:
            try:
                db.execute(
                    "UPDATE moldures SET ref2=? WHERE referencia=? AND ref2=?",
                    [correct, ref, wrong])
                fixed += 1
            except Exception as e:
                print(f'ref2 fix SQLite {ref}: {e}')
        try:
            db.commit()
        except Exception:
            pass
    if fixed:
        print(f'ref2 fixes: {fixed} correccions aplicades')


def _seed_admin_if_configured(db):
    admin_user = os.environ.get('ADMIN_USERNAME', '').strip()
    admin_pass = os.environ.get('ADMIN_PASSWORD', '').strip()
    admin_name = os.environ.get('ADMIN_NAME', 'Administrador').strip() or 'Administrador'

    if not admin_user or not admin_pass:
        print('Admin bootstrap skipped: set ADMIN_USERNAME and ADMIN_PASSWORD to create the first admin.')
        return

    if USE_PG:
        cur = db.cursor()
        cur.execute('SELECT id FROM usuaris WHERE username=%s', [admin_user])
        if not cur.fetchone():
            cur.execute(
                'INSERT INTO usuaris (username,password,nom,is_admin) VALUES (%s,%s,%s,%s)',
                [admin_user, hash_pw(admin_pass), admin_name, 1]
            )
            db.commit()
            print(f"Admin creat des de variables d'entorn: usuari={admin_user}")
    else:
        cur = db.execute('SELECT id FROM usuaris WHERE username=?', [admin_user])
        if not cur.fetchone():
            db.execute(
                'INSERT INTO usuaris (username,password,nom,is_admin) VALUES (?,?,?,1)',
                [admin_user, hash_pw(admin_pass), admin_name]
            )
            db.commit()
            print(f"Admin creat des de variables d'entorn: usuari={admin_user}")

# â"€â"€ Auth decorators â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
_CANONICAL_HOST = os.environ.get('CANONICAL_HOST', '').strip().lower()


@app.before_request
def _redirect_legacy_host():
    """Redirigeix 301 qualsevol host no canònic (p.ex. calculadora.objectiufotografs.com)
    cap a CANONICAL_HOST (p.ex. calculadora.reusrevela.cat). Opt-in via env var.
    Preserva path i query string. Només actua amb GET/HEAD."""
    if not _CANONICAL_HOST:
        return
    if request.method not in ('GET', 'HEAD'):
        return
    current_host = (request.host or '').lower()
    # No redirigim si ja estem al host canònic
    if current_host == _CANONICAL_HOST:
        return
    # No redirigim localhost/127.0.0.1/railway.app internal domains
    if current_host.startswith(('localhost', '127.', '0.0.0.0')) or 'railway' in current_host:
        return
    target = f'https://{_CANONICAL_HOST}{request.full_path}'
    # full_path acaba amb '?' si no hi ha query → netegem
    if target.endswith('?'):
        target = target[:-1]
    # 302 (temporal) en lloc de 301 perquè els navegadors no el cachegin
    # permanentment i puguin reintentar si la configuració canvia.
    return redirect(target, code=302)


@app.before_request
def _sso_from_shared_cookie():
    """SSO fallback: si ve una sessió del Repo B (reusrevela.cat) via cookie
    compartida amb 'private_professional' però sense 'user_id' (sessió de A),
    carregar l'usuari automàticament perquè no hagi de tornar a loguejar-se."""
    if session.get('user_id'):
        return
    pp = session.get('private_professional') or {}
    username = (pp.get('username') or '').strip().lower()
    if not username:
        return
    try:
        user = query('SELECT * FROM usuaris WHERE username=?', [username], one=True)
    except Exception:
        return
    if not user:
        return
    try:
        if not _user_is_allowed(user):
            return
    except Exception:
        pass
    # Activar la sessió del Repo A reusant les dades de B
    try:
        _start_user_session(user)
    except Exception:
        session['user_id'] = user['id']
        session['is_admin'] = bool(user.get('is_admin', 0))
        session['nom'] = user.get('nom') or ''


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            next_path = request.full_path[:-1] if request.full_path.endswith('?') else request.full_path
            login_lang = (
                (session.get('bridge_lang') or '').strip().lower()
                or request.accept_languages.best_match(['ca', 'es', 'en'])
                or 'ca'
            )
            return redirect(url_for('login', next=_safe_next_path(next_path, '/'), source='calc', lang=login_lang))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = query('SELECT * FROM usuaris WHERE id=?', [session['user_id']], one=True)
        if not user or not user['is_admin']:
            flash('Accés restringit a administradors.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def _closest_passpartu_taula_tolerancia(amplada, alcada, prefix='1PAS', tolerancia=2.0):
    """Min-contain amb tolerància sobre la taula passpartout. Filtra per
    prefix de referència (1PAS, DOBPAS…). Considera ambdues orientacions
    i selecciona la fila amb àrea més propera a la sol·licitada (la més
    ajustada físicament) entre les que cobreixen amb la tolerància donada."""
    rows = [dict(r) for r in query('SELECT referencia, preu_cost FROM passpartout')]
    rows = [r for r in rows if r['referencia'].upper().startswith(prefix.upper())]
    candidates = []
    tol_w = max(0.0, amplada - tolerancia)
    tol_h = max(0.0, alcada - tolerancia)
    for r in rows:
        if _row_get(r, 'preu_cost') is None:
            continue
        rw, rh = _parse_dims(r['referencia'])
        if rw is None:
            continue
        if (rw >= tol_w and rh >= tol_h) or (rh >= tol_w and rw >= tol_h):
            candidates.append((r, rw * rh))
    if not candidates:
        return None
    target_area = amplada * alcada
    return min(candidates, key=lambda c: abs(c[1] - target_area))[0]


def calcular_cost_passpartu(amplada, alcada, tipus='simple', finestres_extra=0):
    """Calcula cost i PVD del passpartú.

    Tipus 'simple':
      1. Min-contain amb tolerància sobre la taula 1PAS (per àrea més
         propera) → preu_cost.
      2. Fórmula NOMÉS com a fallback si cap fila no cobreix la mida.

    Tipus 'doble':
      1. Calcular el cost del simple per a la mateixa mida (recursivament,
         sense incloure-hi finestres extra) → cost_simple.
      2. cost_doble_base = cost_simple · 2.
      3. Si hi ha fila DOBPAS a la taula i ratio (àrea_taula/àrea_sol) ≤
         encolat_ratio_max (default 1.40) → usar preu_cost de taula.
      4. Si la fila és massa gran o no n'hi ha → usar simple · 2.
    Així el doble queda sempre coherent amb el simple per a mides on la
    taula salta a una talla força més gran que la sol·licitada.

    Les finestres extra (cost_extra) sempre s'afegeixen al final.
    Retorna dict {'cost', 'pvd', 'origen': 'taula'|'formula'|'simple_x2', 'ref'}."""
    cost_hora    = float(get_config_value('cost_hora_taller', '25'))
    temps_simple = float(get_config_value('passpartu_temps_simple', '9'))
    temps_doble  = float(get_config_value('passpartu_temps_doble', '16'))
    temps_extra  = float(get_config_value('passpartu_temps_finestra', '3.5'))
    cost_cm2     = float(get_config_value('passpartu_cost_cm2', '0.000620'))
    minim_mat    = float(get_config_value('passpartu_minim_material', '0.50'))
    marge_pas    = float(get_config_value('marge_admin_passpartu_pct', '60'))
    tolerancia   = float(get_config_value('passpartu_tolerancia_cm', '2'))
    ratio_max    = float(get_config_value('encolat_ratio_max', '1.40'))

    cost_extra = finestres_extra * (temps_extra * cost_hora / 60)

    if tipus == 'doble':
        # 1) Cost del simple sense extres (recursió controlada — el simple
        # no torna a entrar en aquest branch).
        simple = calcular_cost_passpartu(amplada, alcada, tipus='simple', finestres_extra=0)
        cost_simple_x2 = round(float(simple.get('cost') or 0) * 2, 4)

        # 2) Mirar la taula DOBPAS (només si el ratio està dins del threshold)
        fila = _closest_passpartu_taula_tolerancia(amplada, alcada, prefix='DOBPAS', tolerancia=tolerancia)
        use_taula = False
        if fila and _row_get(fila, 'preu_cost') is not None:
            rw, rh = _parse_dims(fila['referencia'])
            if rw and rh:
                area_sol = max(1.0, float(amplada) * float(alcada))
                ratio = (rw * rh) / area_sol
                if ratio <= ratio_max:
                    use_taula = True

        if use_taula:
            cost = round(float(fila['preu_cost']) + cost_extra, 4)
            origen = 'taula'
            ref = fila['referencia']
        else:
            cost = round(cost_simple_x2 + cost_extra, 4)
            origen = 'simple_x2'
            ref = f'dobpas-{amplada}x{alcada}'

        pvd = round(cost * (1 + marge_pas / 100), 4)
        return {'cost': cost, 'pvd': pvd, 'origen': origen, 'ref': ref}

    # Tipus 'simple'
    fila = _closest_passpartu_taula_tolerancia(amplada, alcada, prefix='1PAS', tolerancia=tolerancia)
    if fila and _row_get(fila, 'preu_cost') is not None:
        cost = round(float(fila['preu_cost']) + cost_extra, 4)
        origen = 'taula'
        ref = fila['referencia']
    else:
        area     = amplada * alcada
        cost_mat = max(area * cost_cm2, minim_mat)
        cost_mo  = temps_simple * cost_hora / 60
        cost = round(cost_mat + cost_mo + cost_extra, 4)
        origen = 'formula'
        ref = f'pas-{amplada}x{alcada}'

    pvd = round(cost * (1 + marge_pas / 100), 4)
    return {'cost': cost, 'pvd': pvd, 'origen': origen, 'ref': ref}


def _cost_muntatge(amplada, alcada, cost_cm2, temps_base, temps_var_cm2):
    """Base function shared by foam and laminat. Returns workshop cost (material + labor)."""
    cost_hora = float(get_config_value('cost_hora_taller', '25'))
    area = amplada * alcada
    mat = area * cost_cm2
    temps = temps_base + (area * temps_var_cm2)
    mo = temps * cost_hora / 60
    return round(mat + mo, 4)


def _closest_encolat_taula_tolerancia(amplada, alcada, prefix, tolerancia=2.0):
    """Min-contain amb tolerància sobre encolat_pro filtrat per prefix
    (ENC o PRO). Considera ambdues orientacions i selecciona la fila amb
    àrea més propera a la sol·licitada (la més ajustada físicament) entre
    les que cobreixen amb la tolerància donada."""
    rows = [dict(r) for r in query('SELECT referencia, preu_cost FROM encolat_pro')]
    rows = [r for r in rows if r['referencia'].upper().startswith(prefix.upper())]
    candidates = []
    tol_w = max(0.0, amplada - tolerancia)
    tol_h = max(0.0, alcada - tolerancia)
    for r in rows:
        if _row_get(r, 'preu_cost') is None:
            continue
        rw, rh = _parse_dims(r['referencia'])
        if rw is None:
            continue
        if (rw >= tol_w and rh >= tol_h) or (rh >= tol_w and rw >= tol_h):
            candidates.append((r, rw * rh))
    if not candidates:
        return None
    target_area = amplada * alcada
    return min(candidates, key=lambda c: abs(c[1] - target_area))[0]


def calcular_cost_foam(amplada, alcada):
    """Encolat en foam adhesiu (ProEco és àlies del mateix producte) — lògica
    híbrida amb threshold:
      1. Min-contain amb tolerància sobre la taula (per àrea més propera).
      2. Si hi ha fila i la seva àrea no excedeix 'encolat_ratio_max' (default
         1.40) vegades l'àrea sol·licitada → usa preu_cost de taula.
      3. Si la fila és massa gran (ratio > threshold) o no n'hi ha → fórmula.
    Així stock pre-tallat segueix guanyant per a mides properes (ENC30x40 per
    a 30×40 → ratio 1.0), però per a mides força més petites que la fila
    stockada (58×140 → ENC80x150 → ratio 1.48) caiem a fórmula."""
    marge      = float(get_config_value('marge_admin_encolat_pct', '60'))
    tolerancia = float(get_config_value('encolat_tolerancia_cm', '2'))
    ratio_max  = float(get_config_value('encolat_ratio_max', '1.40'))
    cost_cm2   = float(get_config_value('foam_cost_cm2', '0.001143'))
    temps_base = float(get_config_value('foam_temps_base_min', '9'))
    temps_var  = float(get_config_value('foam_temps_var_cm2', '0.0015'))

    cost_formula = _cost_muntatge(amplada, alcada, cost_cm2, temps_base, temps_var)
    area_sol = max(1.0, float(amplada) * float(alcada))

    fila = _closest_encolat_taula_tolerancia(amplada, alcada, prefix='ENC', tolerancia=tolerancia)
    use_taula = False
    if fila and _row_get(fila, 'preu_cost') is not None:
        rw, rh = _parse_dims(fila['referencia'])
        if rw and rh:
            ratio = (rw * rh) / area_sol
            if ratio <= ratio_max:
                use_taula = True

    if use_taula:
        cost = round(float(fila['preu_cost']), 4)
        origen = 'taula'
        ref = fila['referencia']
    else:
        cost = cost_formula
        origen = 'formula'
        ref = f'foam-{amplada}x{alcada}'

    pvd = round(cost * (1 + marge / 100), 4)
    return {'cost': cost, 'pvd': pvd, 'preu': pvd, 'origen': origen, 'ref': ref}


def calcular_cost_laminat(amplada, alcada, tipus='semibrillo'):
    """Laminat Protter — lògica híbrida amb threshold (vegeu calcular_cost_foam):
      1. Min-contain amb tolerància sobre encolat_pro (refs PRO%).
      2. Si fila trobada i la seva àrea ≤ encolat_ratio_max × àrea sol·licitada
         (default 1.40) → usar preu_cost de taula (amb diferencial mate).
      3. Si massa gran o no hi ha → fórmula.
    tipus: 'semibrillo' | 'mate'."""
    marge = float(get_config_value('marge_admin_encolat_pct', '60'))
    cost_cm2_semi = float(get_config_value('laminat_semibrillo_cost_cm2', '0.000504'))
    cost_cm2_mate = float(get_config_value('laminat_mate_cost_cm2', '0.000685'))
    temps_base    = float(get_config_value('laminat_temps_base_min', '12'))
    temps_var     = float(get_config_value('laminat_temps_var_cm2', '0.0012'))
    cost_cm2 = cost_cm2_mate if tipus == 'mate' else cost_cm2_semi
    tolerancia = float(get_config_value('encolat_tolerancia_cm', '2'))
    ratio_max  = float(get_config_value('encolat_ratio_max', '1.40'))

    cost_formula = _cost_muntatge(amplada, alcada, cost_cm2, temps_base, temps_var)
    area_sol = max(1.0, float(amplada) * float(alcada))

    fila = _closest_encolat_taula_tolerancia(amplada, alcada, prefix='PRO', tolerancia=tolerancia)
    use_taula = False
    if fila and _row_get(fila, 'preu_cost') is not None:
        rw, rh = _parse_dims(fila['referencia'])
        if rw and rh:
            ratio = (rw * rh) / area_sol
            if ratio <= ratio_max:
                use_taula = True

    if use_taula:
        cost_base = float(fila['preu_cost'])
        if tipus == 'mate':
            area = amplada * alcada
            extra_mat = area * (cost_cm2_mate - cost_cm2_semi)
            cost = round(cost_base + extra_mat, 4)
        else:
            cost = cost_base
        pvd = round(cost * (1 + marge / 100), 4)
        return {'cost': cost, 'pvd': pvd, 'preu': pvd, 'tipus': tipus, 'origen': 'taula', 'ref': fila['referencia']}

    # Fallback fórmula
    pvd = round(cost_formula * (1 + marge / 100), 4)
    return {'cost': cost_formula, 'pvd': pvd, 'preu': pvd, 'tipus': tipus, 'origen': 'formula', 'ref': f'laminat-{tipus}-{amplada}x{alcada}'}


def calcular_cost_protter(amplada, alcada, tipus='semibrillo'):
    """Protter = foam adhesiu + laminat.
    Suma els costs independents de foam i laminat.
    tipus: 'semibrillo' | 'mate'"""
    marge = float(get_config_value('marge_admin_encolat_pct', '60'))
    c_foam = calcular_cost_foam(amplada, alcada)['cost']
    c_lam = calcular_cost_laminat(amplada, alcada, tipus=tipus)['cost']
    cost = round(c_foam + c_lam, 4)
    pvd = round(cost * (1 + marge / 100), 4)
    return {'cost': cost, 'pvd': pvd, 'preu': pvd, 'tipus': tipus,
            'origen': 'composicio', 'ref': f'protter-{tipus}-{amplada}x{alcada}'}


def _closest_vidre_taula(amplada, alcada, prefix=''):
    """Min-contain sobre taula vidres filtrat per prefix.
    prefix: ''=vidre simple (exclou DV-/MIR-), 'DV-'=doble vidre, 'MIR-'=mirall."""
    rows = [dict(r) for r in query('SELECT * FROM vidres')]
    if prefix == '':
        # Exclou DV- i MIR-
        rows = [r for r in rows if not (r['referencia'].upper().startswith('DV-') or r['referencia'].upper().startswith('MIR-'))]
        return _find_closest(rows, amplada, alcada)
    return _find_closest(rows, amplada, alcada, prefix=prefix)


def _closest_vidre_taula_tolerancia(amplada, alcada, prefix='', tolerancia=2.0):
    """Min-contain amb tolerància configurable: un vidre 60×60 pot cobrir
    una obra 62×62 si la tolerància és >= 2 cm. Considera ambdues
    orientacions del vidre (W×H i H×W).

    Selecciona la fila amb àrea més propera a la sol·licitada (la més
    ajustada físicament); ignora files sense preu_cost."""
    rows = [dict(r) for r in query('SELECT referencia, preu_cost FROM vidres')]
    if prefix == '':
        rows = [r for r in rows
                if not (r['referencia'].upper().startswith('DV-')
                        or r['referencia'].upper().startswith('MIR-'))]
    else:
        rows = [r for r in rows if r['referencia'].upper().startswith(prefix.upper())]

    candidates = []
    tol_w = max(0.0, amplada - tolerancia)
    tol_h = max(0.0, alcada - tolerancia)
    for r in rows:
        if _row_get(r, 'preu_cost') is None:
            continue
        rw, rh = _parse_dims(r['referencia'])
        if rw is None:
            continue
        # Cobertura amb tolerància en qualsevol orientació
        if (rw >= tol_w and rh >= tol_h) or (rh >= tol_w and rw >= tol_h):
            candidates.append((r, rw * rh))
    if not candidates:
        return None
    target_area = amplada * alcada
    return min(candidates, key=lambda c: abs(c[1] - target_area))[0]


def calcular_cost_vidre(amplada, alcada):
    """Vidre simple tallat a mida — lògica híbrida:
    1. Min-contain amb tolerància sobre taula vidres (exclou DV-/MIR-)
       que considera orientacions girades.
    2. Fórmula: material × cm² + temps tall (base + lineal × perímetre).
    3. Si hi ha fila vàlida, agafa el min(taula, fórmula). Si no, usa
       només la fórmula."""
    cost_cm2   = float(get_config_value('vidre_cost_cm2', '0.002880'))
    t_base     = float(get_config_value('vidre_temps_base_min', '3'))
    t_lineal   = float(get_config_value('vidre_temps_lineal_m', '0.5'))
    cost_hora  = float(get_config_value('cost_hora_taller', '25'))
    marge      = float(get_config_value('marge_admin_vidres_pct', '60'))
    tolerancia = float(get_config_value('vidre_tolerancia_cm', '2'))

    # Càlcul de fórmula (sempre disponible com a fallback / comparació)
    area = amplada * alcada
    perimetre_m = 2 * (amplada + alcada) / 100
    mat = area * cost_cm2
    temps = t_base + (t_lineal * perimetre_m)
    mo = temps * cost_hora / 60
    cost_formula = mat + mo

    fila = _closest_vidre_taula_tolerancia(amplada, alcada, prefix='', tolerancia=tolerancia)
    if fila and _row_get(fila, 'preu_cost') is not None:
        cost_taula = float(fila['preu_cost'])
        if cost_taula <= cost_formula:
            cost = round(cost_taula, 4)
            origen = 'taula'
            ref = fila['referencia']
        else:
            cost = round(cost_formula, 4)
            origen = 'formula'
            ref = f'vidre-{amplada}x{alcada}'
    else:
        cost = round(cost_formula, 4)
        origen = 'formula'
        ref = f'vidre-{amplada}x{alcada}'

    pvd = round(cost * (1 + marge / 100), 4)
    return {'cost': cost, 'pvd': pvd, 'preu': pvd, 'origen': origen, 'ref': ref}


def calcular_cost_doble_vidre(amplada, alcada):
    """Doble vidre per FÓRMULA PURA (sense lookup de taula).

    El doble vidre és matemàticament determinista i no necessita taula:
      cost_material  = àrea · cost_cm² · 2     (dues làmines)
      MO_tall        = (t_base + t_lineal · perim_m) · cost_hora / 60
                       (es talla un cop: dues làmines tallades alhora)
      MO_muntatge_dv = vidre_dv_muntatge_eur   (alineació + sellat)
      cost_total     = material + MO_tall + MO_muntatge
      PVD            = cost_total · (1 + marge_admin/100)

    Substitueix la lògica híbrida prèvia (taula + fallback fórmula) que
    havia generat saltsd quan la taula DV-* tenia "forats" (cas 80×80
    → DV-80X120 saltava a un preu inflat ~22%). Manté la coherència
    amb el patró del vidre simple per fórmula.

    L'estructura del dict retornat manté els camps que els consumidors
    ja llegeixen ('cost', 'pvd', 'preu', 'origen', 'ref'); 'origen' és
    sempre 'formula'."""
    cost_cm2     = float(get_config_value('vidre_cost_cm2', '0.002880'))
    t_base       = float(get_config_value('vidre_temps_base_min', '3'))
    t_lineal     = float(get_config_value('vidre_temps_lineal_m', '0.5'))
    cost_hora    = float(get_config_value('cost_hora_taller', '25'))
    marge        = float(get_config_value('marge_admin_vidres_pct', '60'))
    mo_muntatge  = float(get_config_value('vidre_dv_muntatge_eur', '1.30'))

    area = float(amplada) * float(alcada)
    perimetre_m = 2 * (amplada + alcada) / 100

    cost_material = area * cost_cm2 * 2
    mo_tall       = (t_base + t_lineal * perimetre_m) * cost_hora / 60
    cost          = round(cost_material + mo_tall + mo_muntatge, 4)
    pvd           = round(cost * (1 + marge / 100), 4)

    return {
        'cost': cost, 'pvd': pvd, 'preu': pvd,
        'origen': 'formula',
        'ref': f'dv-{amplada}x{alcada}',
    }


def calcular_cost_mirall(amplada, alcada):
    """Mirall tallat a mida pel proveïdor (Cristaleria Abrio).
    Preu 5mm amb recàrrec energètic: 31,53 €/m² (= 0.003153 €/cm²).
    Facturació en múltiples de 6 dm². Sense MO de tall (el talla el proveïdor).

    1. Fórmula real sobre cost de factura (primer)
    2. Min-contain sobre taula MIR- com a fallback si la fórmula no aplica."""
    marge     = float(get_config_value('marge_admin_vidres_pct', '60'))
    cost_cm2  = float(get_config_value('mirall_cost_cm2', '0.003153'))
    multiplo  = float(get_config_value('mirall_multiplo_dm2', '6'))

    if amplada > 0 and alcada > 0 and cost_cm2 > 0 and multiplo > 0:
        # Àrea facturada (arrodonida al múltiple de dm² superior)
        area_real_dm2 = (amplada * alcada) / 100.0
        area_fact_dm2 = math.ceil(area_real_dm2 / multiplo) * multiplo
        area_fact_cm2 = area_fact_dm2 * 100.0
        cost = round(area_fact_cm2 * cost_cm2, 4)
        pvd = round(cost * (1 + marge / 100), 4)
        return {
            'cost': cost,
            'pvd': pvd,
            'preu': pvd,
            'area_real_cm2': round(amplada * alcada, 2),
            'area_fact_cm2': round(area_fact_cm2, 2),
            'origen': 'formula',
            'ref': f'mir-{amplada}x{alcada}',
        }

    # Fallback: min-contain sobre taula MIR-
    fila = _closest_vidre_taula(amplada, alcada, prefix='MIR-')
    if fila and _row_get(fila, 'preu_cost') is not None:
        cost = float(fila['preu_cost'])
        pvd = round(cost * (1 + marge / 100), 4)
        return {'cost': cost, 'pvd': pvd, 'preu': pvd, 'origen': 'taula_estimat', 'ref': fila['referencia']}

    return {'cost': 0.0, 'pvd': 0.0, 'preu': 0.0, 'origen': 'no_data', 'ref': f'mir-{amplada}x{alcada}'}


# â"€â"€ Routes: Auth â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route('/login', methods=['GET', 'POST'])
def login():
    next_path = _safe_next_path(request.values.get('next'), url_for('index'))
    login_source = (request.values.get('source') or session.get('bridge_source') or 'calc').strip().lower() or 'calc'
    login_lang = (
        (request.values.get('lang') or session.get('bridge_lang') or '').strip().lower()
        or request.accept_languages.best_match(['ca', 'es', 'en'])
        or 'ca'
    )
    web_return_url = _current_web_return_url()
    if request.method == 'POST':
        # Els usuaris es desen sempre en minúscules (signup + bridge). Normalitzem
        # aquí per evitar que "Juan@X.com" o " juan@x.com " fallin per case/espais.
        username = (request.form.get('username') or '').strip().lower()
        password = request.form.get('password') or ''
        try:
            user = query('SELECT * FROM usuaris WHERE lower(username)=?',
                         [username], one=True)
            credentials_ok = bool(user) and verify_pw(user['password'], password)
        except Exception as exc:
            app.logger.exception('login_db_error user=%s detail=%s', username[:3] + '***', exc)
            flash("No hem pogut validar l'accés ara mateix. Torna-ho a provar en uns segons.", 'error')
            return render_template(
                'login.html',
                web_return_url=web_return_url,
                login_next=next_path,
                login_source=login_source,
                login_lang=login_lang,
            )
        if credentials_ok:
            if not _user_is_allowed(user):
                status = _user_access_status(user)
                if status == 'pending':
                    flash("El teu accés encara està pendent de validació.", 'error')
                else:
                    flash("El teu accés està bloquejat. Contacta amb l'administrador.", 'error')
                return render_template(
                    'login.html',
                    web_return_url=web_return_url,
                    login_next=next_path,
                    login_source=login_source,
                    login_lang=login_lang,
                )
            # Migració progressiva del hash: si aquest usuari encara tenia
            # un SHA-256 legacy, aprofitem aquest login correcte per
            # rehashejar-lo amb pbkdf2. Silenciós si falla.
            maybe_upgrade_hash(user['id'], user['password'], password)
            try:
                _start_user_session(user)
                session['bridge_source'] = login_source
                session['bridge_lang'] = login_lang
                needs_setup = _needs_setup(user['id'])
            except Exception as exc:
                app.logger.exception('login_session_error user_id=%s detail=%s', _row_get(user, 'id'), exc)
                session.clear()
                flash("Hi ha hagut un error iniciant la sessió. Si continua, contacta amb el taller.", 'error')
                return render_template(
                    'login.html',
                    web_return_url=web_return_url,
                    login_next=next_path,
                    login_source=login_source,
                    login_lang=login_lang,
                )
            if needs_setup:
                return redirect(url_for('setup'))
            return redirect(next_path)
        flash('Usuari o contrasenya incorrectes.', 'error')
    return render_template(
        'login.html',
        web_return_url=web_return_url,
        login_next=next_path,
        login_source=login_source,
        login_lang=login_lang,
    )

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Restabliment de contrasenya (token per email) ──────────────────────────
_reset_cols_ready = False
def _ensure_reset_columns():
    """Assegura les columnes reset_token / reset_token_exp a usuaris."""
    global _reset_cols_ready
    if _reset_cols_ready:
        return
    for col in ('reset_token', 'reset_token_exp'):
        try:
            execute(f"ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS {col} TEXT")
        except Exception as e:
            print(f"[reset] ensure col {col}: {e}")
    _reset_cols_ready = True


def _send_reset_email(email, name, link):
    nom_visible = (name or '').strip() or 'professional'
    html = f"""\
<div style="font-family:sans-serif;max-width:560px;margin:0 auto;color:#1C1B18">
  <h2 style="color:#1A6B45;border-bottom:2px solid #1A6B45;padding-bottom:8px;margin-bottom:18px">Restablir la contrasenya</h2>
  <p style="font-size:14px;line-height:1.6">Hola {nom_visible},</p>
  <p style="font-size:14px;line-height:1.6">Has demanat posar una contrasenya nova al teu compte de la calculadora de Reus Revela. Clica el botó per fer-ho:</p>
  <p style="margin:22px 0">
    <a href="{link}" style="display:inline-block;background:#1A6B45;color:#fff;text-decoration:none;padding:12px 22px;border-radius:8px;font-weight:600;font-size:14px">Posar una contrasenya nova →</a>
  </p>
  <p style="font-size:13px;color:#6B6860;line-height:1.5">L'enllaç caduca en 2 hores. Si no has demanat aquest canvi, pots ignorar aquest correu: la teva contrasenya no canviarà.</p>
  <p style="font-size:12px;color:#9E9B94;line-height:1.5;margin-top:16px;word-break:break-all">Si el botó no funciona, copia aquest enllaç al navegador:<br>{link}</p>
  <p style="font-size:14px;line-height:1.6;margin-top:22px">Una salutació,<br><b>Equip Reus Revela</b></p>
</div>
"""
    return _send_user_email_html(email, 'Restablir la contrasenya · Reus Revela', html, log_tag='reset_email')


@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar():
    if request.method == 'POST':
        email = (request.form.get('email') or '').strip().lower()
        _ensure_reset_columns()
        try:
            if email and '@' in email:
                user = query(
                    'SELECT id, nom, username, email, is_admin, access_status FROM usuaris '
                    'WHERE lower(username)=? OR lower(email)=?', [email, email], one=True)
                if user and _user_is_allowed(user):
                    token = secrets.token_urlsafe(32)
                    exp = (datetime.now() + timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')
                    execute('UPDATE usuaris SET reset_token=?, reset_token_exp=? WHERE id=?',
                            [token, exp, user['id']])
                    dest = (_row_get(user, 'email', '') or '').strip() or (_row_get(user, 'username', '') or '')
                    link = f"{request.url_root.rstrip('/')}/recuperar/{token}"
                    _send_reset_email(dest, _row_get(user, 'nom', ''), link)
        except Exception as e:
            print(f"[reset] request error ({email}): {e}")
        # Resposta genèrica sempre (no revelem si l'email existeix).
        return render_template('recuperar.html', done=True)
    return render_template('recuperar.html', done=False)


@app.route('/recuperar/<token>', methods=['GET', 'POST'])
def recuperar_token(token):
    _ensure_reset_columns()
    user = None
    try:
        if token:
            user = query('SELECT id, nom, reset_token_exp FROM usuaris WHERE reset_token=?', [token], one=True)
    except Exception as e:
        print(f"[reset] lookup error: {e}")
    valid = False
    if user:
        exp = (_row_get(user, 'reset_token_exp', '') or '').strip()
        try:
            valid = bool(exp) and datetime.now() <= datetime.strptime(exp, '%Y-%m-%d %H:%M:%S')
        except Exception:
            valid = False
    if not valid:
        return render_template('recuperar_nou.html', valid=False)
    if request.method == 'POST':
        pw1 = request.form.get('password') or ''
        pw2 = request.form.get('password2') or ''
        if len(pw1) < 6:
            flash('La contrasenya ha de tenir com a mínim 6 caràcters.', 'error')
            return render_template('recuperar_nou.html', valid=True, token=token)
        if pw1 != pw2:
            flash('Les contrasenyes no coincideixen.', 'error')
            return render_template('recuperar_nou.html', valid=True, token=token)
        execute('UPDATE usuaris SET password=?, reset_token=NULL, reset_token_exp=NULL WHERE id=?',
                [hash_pw(pw1), user['id']])
        flash('Contrasenya actualitzada. Ja pots entrar amb la nova.', 'ok')
        return redirect(url_for('login'))
    return render_template('recuperar_nou.html', valid=True, token=token)


@app.route('/api/public/professional-signup', methods=['POST'])
def public_professional_signup():
    expected_token = os.environ.get('PUBLIC_SIGNUP_TOKEN', '').strip()
    provided_token = request.headers.get('X-Signup-Token', '').strip()

    if not expected_token or not hmac.compare_digest(provided_token, expected_token):
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    phone = (data.get('phone') or '').strip()
    business_name = (data.get('business_name') or '').strip()
    web_url = (data.get('web_url') or '').strip()
    instagram = (data.get('instagram') or '').strip()
    fiscal_id = (data.get('fiscal_id') or '').strip()
    subject = (data.get('subject') or '').strip()
    message = (data.get('message') or '').strip()
    profile_type = _clean_profile_type(data.get('profile_type'))

    if not name or not email:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400

    if '@' not in email or '.' not in email.split('@')[-1]:
        return jsonify({'ok': False, 'error': 'invalid_email'}), 400

    notes = "\n".join([
        f"Alta des de la web principal el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Nom: {name}",
        f"Email: {email}",
        f"Telèfon: {phone or '-'}",
        f"Assumpte: {subject or '-'}",
        f"Perfil: {profile_type}",
        f"Empresa: {business_name or '-'}",
        f"Web: {web_url or '-'}",
        f"Instagram: {instagram or '-'}",
        f"CIF/NIF: {fiscal_id or '-'}",
        "Missatge:",
        message or '-',
    ])

    existing = query('SELECT id, is_admin, access_status FROM usuaris WHERE username=?', [email], one=True)
    if existing:
        current_status = _user_access_status(existing)
        next_status = current_status if current_status in ('active', 'blocked') else 'pending'
        execute(
            'UPDATE usuaris SET nom=?, nom_empresa=?, profile_type=?, web_url=?, instagram=?, fiscal_id=?, notes_validacio=?, access_status=? WHERE id=?',
            [name, business_name, profile_type, web_url, instagram, fiscal_id, notes, next_status, existing['id']]
        )
        _notify_signup_email(
            action='updated',
            status=next_status,
            name=name, email=email, phone=phone,
            business_name=business_name, profile_type=profile_type,
            web_url=web_url, instagram=instagram, fiscal_id=fiscal_id,
            subject=subject, message=message,
        )
        return jsonify({'ok': True, 'action': 'updated', 'status': next_status})

    temp_password = secrets.token_urlsafe(12)
    execute(
        'INSERT INTO usuaris (username, password, nom, is_admin, nom_empresa, access_status, profile_type, web_url, instagram, fiscal_id, notes_validacio) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
        [email, hash_pw(temp_password), name, 0, business_name, 'pending', profile_type, web_url, instagram, fiscal_id, notes]
    )
    _notify_signup_email(
        action='created',
        status='pending',
        name=name, email=email, phone=phone,
        business_name=business_name, profile_type=profile_type,
        web_url=web_url, instagram=instagram, fiscal_id=fiscal_id,
        subject=subject, message=message,
    )
    # Email de confirmació al usuari (acusament de rebuda + què esperar)
    _send_signup_received_email(name, email)
    return jsonify({'ok': True, 'action': 'created', 'status': 'pending'})


def _notify_signup_email(*, action, status, name, email, phone, business_name,
                        profile_type, web_url, instagram, fiscal_id, subject, message):
    """Envia un email al admin notificant una nova alta o actualització de
    perfil pendent. Mai bloqueja: si Gmail no està configurat o la connexió
    falla, fa print i continua (l'alta ja s'ha desat a la BD).

    Destinatari: config['signup_notify_email'] o, per defecte,
    config['gmail_user'] (compte que envia). Així reusrevela rep
    l'avís al mateix correu que té configurat per a SMTP."""
    try:
        cfg = {r['clau']: r['valor'] for r in (query('SELECT clau, valor FROM config') or [])}
        # Destinatari: signup_notify_email → gmail_user → email del primer admin.
        dest = (cfg.get('signup_notify_email') or cfg.get('gmail_user') or '').strip()
        if not dest:
            arow = query("SELECT email FROM usuaris WHERE is_admin=1 AND email IS NOT NULL AND email<>'' ORDER BY id LIMIT 1", one=True)
            dest = (_row_get(arow, 'email', '') or '').strip()
        if not dest:
            print(f"[signup_notify] skip: cap destinatari (configura signup_notify_email) — alta {email} desada")
            return

        action_label = 'Nova sol·licitud d\'alta' if action == 'created' else 'Actualització de perfil'
        status_label = {
            'pending': 'Pendent de validar',
            'active':  'Actiu (ja autoritzat)',
            'blocked': 'Bloquejat',
        }.get(status, status)

        rows_html = []
        def _row(label, value):
            return f'<tr><td style="padding:6px 10px;color:#6B6860;font-size:13px;border-bottom:1px solid #F3F1EB"><b>{label}</b></td><td style="padding:6px 10px;font-size:13px;border-bottom:1px solid #F3F1EB">{value or "—"}</td></tr>'
        rows_html.append(_row('Nom', name))
        rows_html.append(_row('Email', email))
        rows_html.append(_row('Telèfon', phone))
        rows_html.append(_row('Empresa / Botiga', business_name))
        rows_html.append(_row('Tipus de perfil', profile_type))
        rows_html.append(_row('Web', web_url))
        rows_html.append(_row('Instagram', instagram))
        rows_html.append(_row('CIF / NIF', fiscal_id))
        rows_html.append(_row('Assumpte', subject))

        msg_block = ''
        if message:
            safe_msg = (message or '').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
            msg_block = (
                f'<p style="font-family:sans-serif;font-size:13px;margin-top:18px;color:#6B6860"><b>Missatge:</b></p>'
                f'<p style="font-family:sans-serif;font-size:13px;line-height:1.5;background:#fcfbf8;border:1px solid #E5E2DB;border-radius:8px;padding:11px 14px;margin-top:6px">{safe_msg}</p>'
            )

        html = f"""\
<div style="font-family:sans-serif;max-width:620px;margin:0 auto;color:#1C1B18">
  <h2 style="color:#1A6B45;border-bottom:2px solid #1A6B45;padding-bottom:8px;margin-bottom:14px">
    {action_label}
  </h2>
  <p style="font-size:14px;margin-bottom:14px"><b>Estat:</b>
    <span style="background:#FDF3E8;color:#C8873A;padding:3px 9px;border-radius:999px;font-weight:700;font-size:12px">{status_label}</span>
  </p>
  <p style="font-size:14px;margin-bottom:6px">{datetime.now().strftime('%d/%m/%Y · %H:%M')}</p>
  <table cellspacing="0" cellpadding="0" style="width:100%;border-collapse:collapse;margin-top:14px;background:#fff;border:1px solid #E5E2DB;border-radius:8px;overflow:hidden">
    {''.join(rows_html)}
  </table>
  {msg_block}
  <p style="margin-top:22px">
    <a href="https://calculadora.reusrevela.cat/admin/usuaris" style="display:inline-block;background:#1A6B45;color:#fff;text-decoration:none;padding:10px 18px;border-radius:8px;font-weight:600;font-size:13px">Gestionar a /admin/usuaris →</a>
  </p>
  <p style="font-size:11px;color:#9E9B94;margin-top:24px">Aquest correu s'envia automàticament des de la calculadora quan algú envia el formulari d'alta professional a reusrevela.cat.</p>
</div>
"""
        subject_email = f"[Alta professional] {name or email} — {status_label}"
        # Sender intel·ligent: Resend (HTTPS, funciona a Railway) o Gmail SMTP.
        ok = _send_user_email_html(dest, subject_email, html, log_tag='signup_notify')
        print(f"[signup_notify] {'OK' if ok else 'FAIL'}: dest={dest} email={email} action={action}")
    except Exception as e:
        # No bloquejar mai la resposta del signup: l'usuari s'ha desat,
        # només es perd la notificació puntual.
        print(f"[signup_notify] FAIL ({email}): {e}")


def _gmail_is_configured():
    """Retorna True si gmail_user i gmail_pass són presents a config."""
    gu = (get_config_value('gmail_user', '') or '').strip()
    gp = (get_config_value('gmail_pass', '') or '').strip()
    return bool(gu) and bool(gp)


def _resend_is_configured():
    """Retorna True si resend_api_key és present a config."""
    return bool((get_config_value('resend_api_key', '') or '').strip())


def _email_is_configured():
    """Retorna True si almenys un proveïdor d'email està configurat."""
    return _resend_is_configured() or _gmail_is_configured()


def _send_via_resend(to_addr, subject, html, log_tag='resend_email'):
    """Envia un mail via API HTTPS de Resend. Retorna True si OK.
    Resend funciona via HTTPS (no SMTP), per tant és compatible amb hosts
    que bloquegen el port 465/587 (Railway, Render, etc.)."""
    try:
        api_key = (get_config_value('resend_api_key', '') or '').strip()
        if not api_key:
            print(f"[{log_tag}] skip: resend_api_key no configurat")
            return False
        from_addr = (get_config_value('resend_from', '') or '').strip()
        if not from_addr:
            # Fallback: l'adreça d'onboarding de Resend (sense verificació de
            # domini). Bona per provar; per producció cal configurar un from
            # propi i verificar el domini al panell de Resend.
            from_addr = 'onboarding@resend.dev'
        payload = json.dumps({
            'from': from_addr,
            'to': [to_addr],
            'subject': subject,
            'html': html,
        }).encode('utf-8')
        req = urllib_request.Request(
            'https://api.resend.com/emails',
            data=payload,
            headers={
                'Authorization': 'Bearer ' + api_key,
                'Content-Type': 'application/json',
                # Cloudflare (que protegeix l'API de Resend) bloqueja el User-Agent
                # per defecte de Python (`Python-urllib/...`) amb error 1010. Posem
                # un UA descriptiu perquè la petició es vegi com a client legítim.
                'User-Agent': 'Calculadora-Marcs/1.0 (+https://calculadora.reusrevela.cat)',
                'Accept': 'application/json',
            },
            method='POST',
        )
        with urllib_request.urlopen(req, timeout=10) as resp:
            data = resp.read().decode('utf-8', errors='replace')
        print(f"[{log_tag}] OK: enviat a {to_addr} via Resend (resp={data[:140]})")
        return True
    except urllib_error.HTTPError as e:
        body = ''
        try:
            body = e.read().decode('utf-8', errors='replace')[:240]
        except Exception:
            pass
        print(f"[{log_tag}] FAIL HTTP {e.code} ({to_addr}): {body}")
        return False
    except Exception as e:
        print(f"[{log_tag}] FAIL ({to_addr}): {e}")
        return False


def _generate_temp_password(length=12):
    """Genera una contrasenya aleatòria llegible (sense 0/O, 1/l/I).
    Es fa servir per a l'email de benvinguda; l'usuari pot canviar-la
    immediatament des de /ajustos."""
    import secrets, string
    alphabet = ''.join(c for c in (string.ascii_letters + string.digits) if c not in '0Ol1I')
    return ''.join(secrets.choice(alphabet) for _ in range(length))


def _send_welcome_email(username, password, nom, to_addr=None):
    """Envia un mail al professional amb les seves dades d'alta i un mini
    tutorial de com fer servir la calculadora. Retorna True si s'envia OK."""
    nom_visible = (nom or '').strip() or 'professional'
    base_url = 'https://calculadora.reusrevela.cat'
    html = f"""\
<div style="font-family:'Helvetica Neue',Arial,sans-serif;max-width:560px;margin:0 auto;color:#1C1B18;padding:24px;background:#FBFAF7">
  <div style="text-align:center;margin:0 0 16px">
    <img src="https://reusrevela.cat/static/img/logo-reusrevela.png" alt="Reus Revela" width="56" height="56" style="border:0">
  </div>
  <h2 style="color:#1A6B45;border-bottom:2px solid #1A6B45;padding-bottom:10px;margin:0 0 18px">
    Et donem la benvinguda
  </h2>
  <p style="font-size:14px;line-height:1.6;margin:0 0 14px">Hola {nom_visible},</p>
  <p style="font-size:14px;line-height:1.6;margin:0 0 16px">
    Ja tens el teu compte a la <strong>calculadora de marcs i impressions de Reus Revela</strong>.
    Aquestes són les teves dades d'accés:
  </p>

  <table style="width:100%;max-width:480px;margin:0 0 18px;background:#F5F3EC;border-radius:10px;border-collapse:collapse">
    <tr><td style="padding:10px 14px;color:#6B6860;font-size:13px;width:120px">URL</td>
        <td style="padding:10px 14px;font-family:Consolas,monospace;font-size:13px"><a href="{base_url}" style="color:#1A6B45;text-decoration:none">{base_url}</a></td></tr>
    <tr><td style="padding:10px 14px;color:#6B6860;font-size:13px">Usuari</td>
        <td style="padding:10px 14px;font-family:Consolas,monospace;font-size:14px"><strong>{username}</strong></td></tr>
    <tr><td style="padding:10px 14px;color:#6B6860;font-size:13px">Contrasenya</td>
        <td style="padding:10px 14px;font-family:Consolas,monospace;font-size:14px"><strong>{password}</strong></td></tr>
  </table>

  <p style="font-size:12px;color:#6B6860;margin:0 0 22px">
    Recomanem que canviïs la contrasenya quan entris per primer cop. Ho pots fer des de <strong>/ajustos</strong>.
  </p>

  <h3 style="color:#1A6B45;margin:24px 0 10px;font-size:15px">Com funciona la calculadora (en cinc passos)</h3>
  <ol style="line-height:1.7;font-size:14px;padding-left:20px;margin:0 0 18px">
    <li>Introdueix les <strong>mides de la peça</strong> i el <strong>tipus</strong> (fotografia, samarreta, puzzle…).</li>
    <li>Tria els <strong>materials</strong>: motllura, vidre o passpartú i muntatge.</li>
    <li>El <strong>preu es calcula automàticament</strong> i veus el resum a la dreta.</li>
    <li>Pots <strong>desar diverses opcions</strong> (A, B, C…) per al mateix client.</li>
    <li>Quan tinguis la versió final, genera el <strong>PDF</strong>, envia per <strong>WhatsApp</strong> al client o passa la comanda al <strong>taller</strong>.</li>
  </ol>

  <h3 style="color:#1A6B45;margin:24px 0 10px;font-size:15px">Consells útils</h3>
  <ul style="line-height:1.7;font-size:14px;padding-left:20px;margin:0 0 18px">
    <li><strong>Botó € (capçalera del Resum)</strong> — alterna entre preu de cost (taller) i PVP final.</li>
    <li>Caixa <strong>"Extres"</strong> — afegir feines puntuals com desmuntar un marc antic o emmarcar una samarreta.</li>
    <li><strong>"Més accions"</strong> — generar PDF, enviar per WhatsApp i comanda al taller.</li>
    <li>A <strong>/ajustos</strong> pots configurar el teu <strong>marge comercial</strong>, les <strong>dades de l'empresa</strong> i les preferències de marca.</li>
  </ul>

  <p style="margin:24px 0 6px;font-size:13px;color:#6B6860">
    Si tens qualsevol dubte, escriu-nos a <a href="mailto:reusrevela@gmail.com" style="color:#1A6B45">reusrevela@gmail.com</a> o truca al 977 316 111.
  </p>
  <p style="margin:18px 0 0;font-size:13px;color:#1C1B18">
    Una salutació cordial,<br><strong>Equip Reus Revela</strong>
  </p>

  <hr style="border:none;border-top:1px solid #E5E2DB;margin:26px 0 14px">
  <p style="font-size:11px;color:#9E9B94;line-height:1.6;margin:0">
    Aquest és un correu automàtic. <strong>No responguis a aquesta adreça</strong>;
    les respostes no es llegeixen. Si has rebut aquest missatge per error, esborra'l
    i, si vols, fes-nos-ho saber a <a href="mailto:reusrevela@gmail.com" style="color:#9E9B94">reusrevela@gmail.com</a>.
  </p>
</div>
"""
    subject = 'Et donem la benvinguda · Calculadora Reus Revela'
    return _send_user_email_html(to_addr or username, subject, html, log_tag='welcome_email')


def _send_user_email_html(to_addr, subject, html, log_tag='user_email'):
    """Helper compartit per a emails al USUARI final. Tria proveïdor segons
    config: si resend_api_key està posat, fa servir l'API HTTPS de Resend
    (compatible amb hosts que bloquegen SMTP); si no, fallback a Gmail
    SMTP. Try/except: mai bloqueja el flux que el crida."""
    if _resend_is_configured():
        return _send_via_resend(to_addr, subject, html, log_tag=log_tag)
    try:
        cfg = {r['clau']: r['valor'] for r in (query('SELECT clau, valor FROM config') or [])}
        gmail_user = (cfg.get('gmail_user') or '').strip()
        gmail_pass = (cfg.get('gmail_pass') or '').strip()
        if not gmail_user or not gmail_pass:
            print(f"[{log_tag}] skip: cap proveïdor configurat (dest={to_addr})")
            return False
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        m = MIMEMultipart('alternative')
        m['Subject'] = subject
        m['From'] = gmail_user
        m['To'] = to_addr
        m['Reply-To'] = gmail_user
        m.attach(MIMEText(html, 'html'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=8) as s:
            s.login(gmail_user, gmail_pass)
            s.sendmail(gmail_user, [to_addr], m.as_string())
        print(f"[{log_tag}] OK: enviat a {to_addr}")
        return True
    except Exception as e:
        print(f"[{log_tag}] FAIL ({to_addr}): {e}")
        return False


def _send_signup_received_email(name, email):
    """Email al usuari confirmant que hem rebut la sol·licitud d'alta.
    Estableix expectatives perquè no es quedi 'penjat' esperant resposta."""
    nom_visible = (name or '').strip() or 'professional'
    html = f"""\
<div style="font-family:sans-serif;max-width:560px;margin:0 auto;color:#1C1B18">
  <h2 style="color:#1A6B45;border-bottom:2px solid #1A6B45;padding-bottom:8px;margin-bottom:18px">
    Hem rebut la teva sol·licitud
  </h2>
  <p style="font-size:14px;line-height:1.6">Hola {nom_visible},</p>
  <p style="font-size:14px;line-height:1.6">Gràcies per voler treballar amb <b>Reus Revela</b>. Hem rebut la teva sol·licitud d'alta professional i la revisarem en les <b>pròximes 24-48 hores hàbils</b>.</p>
  <p style="font-size:14px;line-height:1.6">Quan el teu compte estigui actiu, t'enviarem un correu separat amb:</p>
  <ul style="font-size:14px;line-height:1.7;padding-left:20px;margin:8px 0 18px">
    <li>El teu usuari i contrasenya inicial</li>
    <li>L'enllaç a la calculadora</li>
    <li>El tutorial per començar a fer pressupostos</li>
  </ul>
  <p style="font-size:14px;line-height:1.6">Si tens cap pregunta urgent, pots respondre directament a aquest correu o escriure'ns a <a href="mailto:reusrevela@gmail.com" style="color:#1A6B45">reusrevela@gmail.com</a>.</p>
  <p style="font-size:14px;line-height:1.6;margin-top:24px">Una salutació,<br><b>Equip Reus Revela</b></p>
  <p style="font-size:11px;color:#9E9B94;margin-top:24px;border-top:1px solid #E5E2DB;padding-top:12px">Aquest correu s'envia automàticament en rebre el formulari d'alta professional a reusrevela.cat.</p>
</div>
"""
    return _send_user_email_html(
        email,
        'Hem rebut la teva sol·licitud — Reus Revela',
        html,
        log_tag='signup_received',
    )


def _send_user_activation_email(name, email, temp_password):
    """Email al usuari quan l'admin l'activa: usuari + contrasenya inicial +
    enllaços a la calc i al tutorial. Important: la contrasenya va en clar
    (és l'única manera). El missatge demana que la canviï al primer login."""
    nom_visible = (name or '').strip() or 'professional'
    safe_pass = (temp_password or '').replace('<', '&lt;').replace('>', '&gt;')
    html = f"""\
<div style="font-family:sans-serif;max-width:580px;margin:0 auto;color:#1C1B18">
  <h2 style="color:#1A6B45;border-bottom:2px solid #1A6B45;padding-bottom:8px;margin-bottom:18px">
    Ja pots fer servir la calculadora
  </h2>
  <p style="font-size:14px;line-height:1.6">Hola {nom_visible},</p>
  <p style="font-size:14px;line-height:1.6">El teu compte a <b>Reus Revela</b> ja està actiu. Aquestes són les teves dades d'accés:</p>

  <table style="width:100%;border-collapse:collapse;background:#fcfbf8;border:1px solid #E5E2DB;border-radius:8px;overflow:hidden;margin:14px 0">
    <tr>
      <td style="padding:11px 14px;border-bottom:1px solid #F3F1EB;width:130px;color:#6B6860;font-size:13px"><b>Usuari</b></td>
      <td style="padding:11px 14px;border-bottom:1px solid #F3F1EB;font-family:Consolas,monospace;font-size:14px">{email}</td>
    </tr>
    <tr>
      <td style="padding:11px 14px;color:#6B6860;font-size:13px"><b>Contrasenya inicial</b></td>
      <td style="padding:11px 14px;font-family:Consolas,monospace;font-size:14px"><b>{safe_pass}</b></td>
    </tr>
  </table>

  <p style="font-size:13px;color:#6B6860;line-height:1.5;background:#FDF3E8;border:1px solid #E8C89A;border-radius:8px;padding:11px 14px;margin:14px 0">
    ⚠ Et recomanem canviar la contrasenya al primer accés. Pots fer-ho des de la secció <b>Ajustos</b> dins la calculadora.
  </p>

  <p style="font-size:14px;margin:22px 0 14px"><b>Enllaços útils:</b></p>
  <p style="margin:8px 0">
    <a href="https://calculadora.reusrevela.cat/calculadora" style="display:inline-block;background:#1A6B45;color:#fff;text-decoration:none;padding:11px 20px;border-radius:8px;font-weight:600;font-size:14px">Anar a la calculadora →</a>
  </p>
  <p style="margin:8px 0">
    <a href="https://reusrevela.cat/tutorial" style="display:inline-block;background:transparent;color:#1A6B45;text-decoration:none;padding:10px 18px;border:1px solid #1A6B45;border-radius:8px;font-weight:600;font-size:14px">📘 Veure el tutorial</a>
  </p>

  <p style="font-size:14px;line-height:1.6;margin-top:24px">El tutorial t'explica com fer servir la calculadora, com gestionar pressupostos i què fer si tens problemes per entrar (per exemple, si has d'esborrar les cookies).</p>

  <p style="font-size:14px;line-height:1.6">Si tens cap dubte o problema, respon a aquest correu o escriu-nos a <a href="mailto:reusrevela@gmail.com" style="color:#1A6B45">reusrevela@gmail.com</a>.</p>

  <p style="font-size:14px;line-height:1.6;margin-top:24px">Una salutació,<br><b>Equip Reus Revela</b></p>
  <p style="font-size:11px;color:#9E9B94;margin-top:24px;border-top:1px solid #E5E2DB;padding-top:12px">Aquest correu s'envia automàticament en activar el teu compte professional.</p>
</div>
"""
    return _send_user_email_html(
        email,
        'El teu compte ja està actiu — Reus Revela',
        html,
        log_tag='user_activation',
    )


@app.route('/api/public/bridge-login', methods=['POST'])
def public_bridge_login():
    provided_token = request.headers.get('X-Bridge-Token', '').strip()

    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    lang = (data.get('lang') or 'ca').strip().lower()
    source = (data.get('source') or 'web_private').strip().lower()
    service = (data.get('service') or '').strip().lower()
    next_path = _safe_next_path(data.get('next') or '', '/calculadora' if service == 'frames' else '/')

    if not username or not password:
        return jsonify({'ok': False, 'error': 'missing_credentials'}), 400

    # lower(username) per consistència amb /login i bridge-refresh: un
    # usuari antic desat amb majúscules a la BD podia fallar aquí tot
    # entrant bé per la resta d'endpoints.
    user = query('SELECT * FROM usuaris WHERE lower(username)=?', [username], one=True)
    if not user or not verify_pw(user['password'], password):
        return jsonify({'ok': False, 'error': 'invalid_credentials'}), 401

    if not _user_is_allowed(user):
        return jsonify({'ok': False, 'error': _user_access_status(user)}), 403

    # Migració progressiva del hash: aprofitem aquest login via bridge
    # per rehashejar si encara era un SHA-256 legacy.
    maybe_upgrade_hash(user['id'], user['password'], password)

    token = _build_bridge_token({
        'uid': user['id'],
        'next': next_path,
        'lang': lang,
        'source': source,
        'service': service,
        'iat': int(time.time()),
    })
    host = request.host_url.rstrip('/')
    return jsonify({
        'ok': True,
        'redirect_url': f'{host}{url_for("bridge_auth")}?token={token}',
        'next': next_path,
    })


@app.route('/api/public/bridge-refresh', methods=['POST'])
def public_bridge_refresh():
    """Emet un token de bridge curt de durada per a un usuari que ja ha
    validat credencials a la web pública (reusrevela-web).

    Es protegeix amb X-Bridge-Token (el mateix que bridge-login): si algú
    té aquest secret, ja pot emetre tokens a lliure voluntat via
    bridge-login. Per tant aquest endpoint no obre cap superfície nova,
    només permet refrescar el token d'SSO (p.ex. per entrar a la
    calculadora des de l'àrea privada passats els 120s de vida del token
    original) sense demanar la contrasenya una altra vegada.
    """
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or data.get('email') or '').strip().lower()
    lang = (data.get('lang') or 'ca').strip().lower()
    source = (data.get('source') or 'web_private').strip().lower()
    service = (data.get('service') or '').strip().lower()
    next_path = _safe_next_path(data.get('next') or '', '/calculadora' if service == 'frames' else '/')

    if not username:
        return jsonify({'ok': False, 'error': 'missing_username'}), 400

    user = query('SELECT id, is_admin, access_status FROM usuaris WHERE lower(username)=?', [username], one=True)
    if not user:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    if not _user_is_allowed(user):
        return jsonify({'ok': False, 'error': _user_access_status(user)}), 403

    token = _build_bridge_token({
        'uid': user['id'],
        'next': next_path,
        'lang': lang,
        'source': source,
        'service': service,
        'iat': int(time.time()),
    })
    host = request.host_url.rstrip('/')
    return jsonify({
        'ok': True,
        'redirect_url': f'{host}{url_for("bridge_auth")}?token={token}',
        'next': next_path,
    })


@app.route('/api/public/professional-summary', methods=['POST'])
def public_professional_summary():
    try:
        provided_token = request.headers.get('X-Bridge-Token', '').strip()

        if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
            return jsonify({'ok': False, 'error': 'unauthorized'}), 401

        data = request.get_json(silent=True) or {}
        username = (data.get('username') or '').strip().lower()
        if not username:
            return jsonify({'ok': False, 'error': 'not_found'}), 404

        user = query(
            'SELECT id, nom, nom_empresa, profile_type, access_status, '
            'marge, marge_pro_pct, marge_impressio, marge_impressio_pro_pct, '
            'imp_tram1, imp_tram2, imp_tram3, imp_tram4, imp_tram5, imp_tram6, '
            'margins_json, baryta_actiu FROM usuaris WHERE lower(username)=?',
            [username],
            one=True,
        )
        if not user:
            return jsonify({'ok': False, 'error': 'not_found'}), 404

        margins = _load_user_commercial_margins(user)

        recent_quotes_rows = query(
            '''SELECT id, data, num_pressupost, client_nom, preu_final, pagat, entregat, pendent
               FROM comandes
               WHERE user_id = ?
               ORDER BY data DESC
               LIMIT 5''',
            [user['id']],
        ) or []

        recent_quotes = []
        for row in recent_quotes_rows:
            recent_quotes.append({
                'id': row['id'],
                'date': row['data'] or '',
                'num_pressupost': row['num_pressupost'] or '',
                'client_nom': row['client_nom'] or '',
                'preu_final': float(row['preu_final'] or 0),
                'pagat': bool(row['pagat']),
                'entregat': bool(row['entregat']),
                'pendent': bool(row['pendent']),
            })

        marge_pro_pct = _row_get(user, 'marge_pro_pct')
        marge_impressio_pro_pct = _row_get(user, 'marge_impressio_pro_pct')

        def _opt_float(key):
            v = _row_get(user, key)
            return float(v) if v is not None else None

        return jsonify({
            'ok': True,
            'name': user['nom'] or '',
            'business_name': user['nom_empresa'] or '',
            'is_admin': bool(_row_get(user, 'is_admin', False)),
            'profile_type': _clean_profile_type(user['profile_type']),
            'access_status': _user_access_status(user),
            'recent_quotes': recent_quotes,
            # Marges del calc (font autoritativa). El web hauria de mostrar aquests
            # valors en lloc del JSON local, que ara fa només de cache.
            'margins': margins,
            'marge': float(user['marge']) if user['marge'] is not None else 60.0,
            'marge_pro_pct': float(marge_pro_pct) if marge_pro_pct is not None else None,
            'marge_efectiu': _get_marge_value(user),
            'marge_impressio': float(user['marge_impressio']) if user['marge_impressio'] is not None else 0.0,
            'marge_impressio_pro_pct': float(marge_impressio_pro_pct) if marge_impressio_pro_pct is not None else None,
            'marge_impressio_efectiu': _get_marge_impressio_value(user),
            # Trams d'impressió per àrea (PVD→PVP). Cada valor pot ser null;
            # el consumidor cau al default global (imp_tram{N}_marge_default)
            # quan no hi ha valor.
            'imp_tram1': _opt_float('imp_tram1'),
            'imp_tram2': _opt_float('imp_tram2'),
            'imp_tram3': _opt_float('imp_tram3'),
            'imp_tram4': _opt_float('imp_tram4'),
            'imp_tram5': _opt_float('imp_tram5'),
            'imp_tram6': _opt_float('imp_tram6'),
            # Límits dels trams (cm²) i defaults globals — exposats perquè
            # el consumidor (web) no hagi de duplicar-los i s'adapti
            # automàticament si l'admin canvia algun valor a /admin/config.
            'imp_tram_limits': [
                float(get_config_value('imp_tram1_area', '900')),
                float(get_config_value('imp_tram2_area', '2000')),
                float(get_config_value('imp_tram3_area', '4200')),
                float(get_config_value('imp_tram4_area', '6000')),
                float(get_config_value('imp_tram5_area', '14400')),
            ],
            'imp_tram_defaults': [
                float(get_config_value('imp_tram1_marge_default', '80')),
                float(get_config_value('imp_tram2_marge_default', '75')),
                float(get_config_value('imp_tram3_marge_default', '70')),
                float(get_config_value('imp_tram4_marge_default', '60')),
                float(get_config_value('imp_tram5_marge_default', '50')),
                float(get_config_value('imp_tram6_marge_default', '45')),
            ],
            # Activacions per usuari (papers premium amb pricing per trams)
            'baryta_actiu': bool(_row_get(user, 'baryta_actiu', 0)),
        })
    except Exception as exc:
        print(f'professional-summary error: {exc}')
        return jsonify({
            'ok': False,
            'error': 'internal_error',
            'name': '',
            'business_name': '',
            'profile_type': 'professional',
            'access_status': 'pending',
            'recent_quotes': [],
        }), 200


@app.route('/api/public/commercial-settings-sync', methods=['POST'])
def public_commercial_settings_sync():
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip().lower()
    if not username:
        return jsonify({'ok': False, 'error': 'missing_username'}), 400

    user = query('SELECT id FROM usuaris WHERE lower(username)=?', [username], one=True)
    if not user:
        return jsonify({'ok': False, 'error': 'user_not_found'}), 404

    margins = _normalize_commercial_margins(
        data.get('margins') if isinstance(data.get('margins'), dict) else data,
        frame_margin=data.get('marge'),
        print_margin=data.get('marge_impressio'),
    )
    marge = margins['frames']
    marge_impressio = margins['prints']
    execute(
        'UPDATE usuaris SET marge=?, marge_impressio=?, margins_json=? WHERE id=?',
        [marge, marge_impressio, json.dumps(margins, ensure_ascii=True), user['id']]
    )
    return jsonify({'ok': True, 'username': username, 'marge': marge, 'marge_impressio': marge_impressio, 'margins': margins})


@app.route('/api/public/pricing', methods=['GET'])
def public_pricing():
    """Exposa les tarifes professionals (sense marge aplicat) perquè
    la web les pugui consumir i aplicar-hi el marge propi de cada client.

    Autenticació: capçalera X-Bridge-Token (mateix token que la resta
    d'endpoints de bridge).

    Retorna:
      impressio     — taula completa de còpies fotogràfiques (ref, preu, descripcio)
      laminate_only — preus de laminat sol per mida (ref, preu)
      encolat_pro   — preus de muntatge encolat i protter per mida (ref, preu, tipus)
    """
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    # Còpies fotogràfiques
    impressio_rows = query('SELECT referencia, preu, descripcio FROM impressio ORDER BY preu') or []
    impressio = [
        {'ref': r['referencia'], 'preu': float(r['preu'] or 0), 'descripcio': r['descripcio'] or ''}
        for r in impressio_rows
    ]

    # Laminat sol (hardcoded en constants, igual que a la calculadora)
    laminate_only = [
        {'ref': ref, 'preu': float(preu)}
        for ref, preu in sorted(LAMINATE_ONLY_PRICES.items(),
                                key=lambda x: float(x[1]))
    ]

    # Encolat professional i protter (taula encolat_pro, prefix ENC / PRO)
    encolat_rows = query('SELECT referencia, preu, preu_cost FROM encolat_pro ORDER BY preu') or []
    encolat_pro = []
    for r in encolat_rows:
        ref = r['referencia'] or ''
        tipus = 'protter' if ref.upper().startswith('PRO') else 'encolat'
        mida = ref[3:] if ref.upper().startswith('PRO') or ref.upper().startswith('ENC') else ref
        pc = _row_get(r, 'preu_cost')
        pvd = calcular_pvd(pc, 'encolat') if pc is not None else None
        # NO exposar preu_cost al Repo B — és dada interna de l'admin.
        # Només enviem el PVD (preu "taller" que paga el professional).
        encolat_pro.append({
            'ref': ref,
            'mida': mida,
            'preu': pvd if pvd is not None else float(r['preu'] or 0),
            'tipus': tipus,
        })

    # Papers premium amb pricing per trams (cost·multiplicador segons àrea).
    # Exposem el cost_cm2 i els trams perquè la web pugui replicar el càlcul
    # sense haver de fer una crida per cada mida sol·licitada.
    papers_trams = {}
    for paper_id in ('baryta',):
        actiu = get_config_value(f'imp_{paper_id}_trams_actius', '0') == '1'
        if not actiu:
            continue
        trams = []
        for i in range(1, 7):
            max_str = get_config_value(f'imp_{paper_id}_t{i}_max', None)
            mult_str = get_config_value(f'imp_{paper_id}_t{i}_mult', None)
            if mult_str is None:
                continue
            try:
                mult = float(mult_str)
            except (TypeError, ValueError):
                continue
            try:
                max_area = float(max_str) if max_str not in (None, '') else None
            except (TypeError, ValueError):
                max_area = None
            trams.append({'max_area': max_area, 'mult': mult})
        try:
            cost_cm2 = float(get_config_value(f'imp_{paper_id}_cost_cm2', '0'))
        except (TypeError, ValueError):
            cost_cm2 = 0.0
        papers_trams[paper_id] = {'cost_cm2': cost_cm2, 'trams': trams}

    return jsonify({
        'ok': True,
        'impressio': impressio,
        'laminate_only': laminate_only,
        'encolat_pro': encolat_pro,
        'papers_trams': papers_trams,
    })


@app.route('/api/public/impressio-price', methods=['GET'])
def public_impressio_price():
    """Retorna el preu d'impressió calculat per a unes dimensions exactes,
    usant la mateixa lògica híbrida (taula + fórmula) que /api/closest.
    Autenticació: X-Bridge-Token.
    Params: w, h (cm), paper (lustre|silk|baryta, default lustre).
    """
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    w = float(request.args.get('w', 0))
    h = float(request.args.get('h', 0))
    paper = (request.args.get('paper') or 'lustre').strip().lower()
    if paper not in ('lustre', 'silk', 'baryta', 'poster_mate'):
        paper = 'lustre'
    if w <= 0 or h <= 0:
        return jsonify({'ok': False, 'error': 'w and h must be positive'}), 400

    result = _imp_closest(w, h, paper=paper)
    if not result:
        return jsonify({'ok': False, 'error': 'no pricing data available'}), 404

    return jsonify({
        'ok': True,
        'ref': result.get('ref', ''),
        'preu': result.get('preu', 0),
        'origen': result.get('origen', ''),
        'area': result.get('area', 0),
    })


@app.route('/api/public/clients-habituals', methods=['GET'])
def public_clients_habituals():
    """Llista de clients habituals actius per a consum des de reusrevela-web.
    Autenticació: X-Bridge-Token."""
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    rows = query("""
        SELECT c.id, c.nom, c.nif, c.tipus, c.telefon, c.email, c.usuari_id,
               u.nom_empresa AS empresa
        FROM clients_externs c
        LEFT JOIN usuaris u ON c.usuari_id = u.id
        WHERE c.actiu = TRUE
        ORDER BY c.nom
    """) or []
    return jsonify({
        'ok': True,
        'clients': [
            {
                'id': _row_get(r, 'id'),
                'nom': _row_get(r, 'nom') or '',
                'nif': _row_get(r, 'nif') or '',
                'tipus': _row_get(r, 'tipus') or 'pvp',
                'telefon': _row_get(r, 'telefon') or '',
                'email': _row_get(r, 'email') or '',
                'empresa': _row_get(r, 'empresa') or '',
                'dropbox_url': _row_get(r, 'dropbox_url') or '',
            }
            for r in rows
        ],
    })


@app.route('/api/public/clients-habituals/save', methods=['POST'])
def public_clients_habituals_save():
    """Crea o actualitza un client habitual (clients_externs) des de la web.
    Auth: X-Bridge-Token. Body: {nom/name, email, telefon/phone, nif, tipus,
    client_id (opcional)}. Dedup soft per email."""
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    telefon = (data.get('telefon') or data.get('phone') or '').strip()
    nif = (data.get('nif') or '').strip()
    tipus = (data.get('tipus') or 'pvp').strip() or 'pvp'
    client_id = (data.get('client_id') or '').strip()
    if not (nom or email or telefon):
        return jsonify({'ok': False, 'error': 'missing_identity'}), 400
    if not nom:
        nom = email or telefon

    def _out(row):
        return {
            'id': _row_get(row, 'id'),
            'nom': _row_get(row, 'nom') or '',
            'nif': _row_get(row, 'nif') or '',
            'email': _row_get(row, 'email') or '',
            'telefon': _row_get(row, 'telefon') or '',
            'tipus': _row_get(row, 'tipus') or 'pvp',
        }

    try:
        if client_id and str(client_id).isdigit():
            existing = query('SELECT id FROM clients_externs WHERE id=?', [int(client_id)], one=True)
            if not existing:
                return jsonify({'ok': False, 'error': 'client_not_found'}), 404
            execute('UPDATE clients_externs SET nom=?, nif=?, email=?, telefon=?, tipus=? WHERE id=?',
                    [nom, nif, email, telefon, tipus, int(client_id)])
            row = query('SELECT * FROM clients_externs WHERE id=?', [int(client_id)], one=True)
        else:
            existing = query('SELECT id FROM clients_externs WHERE LOWER(email)=? AND actiu=TRUE',
                             [email.lower()], one=True) if email else None
            if existing:
                execute('UPDATE clients_externs SET nom=?, nif=?, email=?, telefon=?, tipus=? WHERE id=?',
                        [nom, nif, email, telefon, tipus, existing['id']])
                row = query('SELECT * FROM clients_externs WHERE id=?', [existing['id']], one=True)
            else:
                execute('INSERT INTO clients_externs (nom, nif, email, telefon, tipus, actiu) '
                        'VALUES (?, ?, ?, ?, ?, TRUE)', [nom, nif, email, telefon, tipus])
                row = query('SELECT * FROM clients_externs WHERE nom=? ORDER BY id DESC LIMIT 1', [nom], one=True)
        return jsonify({'ok': True, 'client': _out(row)})
    except Exception as exc:
        print(f'clients_habituals_save error: {exc}')
        return jsonify({'ok': False, 'error': 'internal_error'}), 500


def _pro_clients_lookup_user_id(username):
    """Resol username (case-insensitive) → usuaris.id. Retorna None si no
    existeix o si l'usuari està bloquejat."""
    if not username:
        return None
    row = query(
        'SELECT id, access_status FROM usuaris WHERE LOWER(username)=?',
        [str(username).strip().lower()], one=True
    )
    if not row:
        return None
    status = _user_access_status(row) if row else ''
    if status == 'blocked':
        return None
    return row['id']


def _pro_clients_row_to_dict(row):
    if not row:
        return None
    def _g(key, default=None):
        try:
            return row[key]
        except (KeyError, IndexError, TypeError):
            return default
    return {
        'id': _g('id'),
        'name': _g('nom') or '',
        'company': _g('empresa') or '',
        'email': _g('email') or '',
        'phone': _g('telefon') or '',
        'city': _g('poblacio') or '',
        'notes': _g('notes') or '',
        'source': _g('source') or 'private_area',
        'last_order_ref': _g('last_order_ref') or '',
        'order_count': int(_g('order_count') or 0),
        'updated_at': str(_g('updated_at') or ''),
        'created_at': str(_g('created_at') or ''),
    }


@app.route('/api/public/pro-clients/save', methods=['POST'])
def public_pro_clients_save():
    """Crea o actualitza un client privat d'un distribuïdor (web → calc).

    Auth: X-Bridge-Token.
    Body JSON: { username, name, company, email, phone, city, notes,
                 source, last_order_ref, client_id (opcional per update) }
    Cal almenys un de {name, email, phone} no buit.
    Retorna: { ok, client: {...} } o { ok: false, error: <code> }."""
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    user_id = _pro_clients_lookup_user_id(username)
    if not user_id:
        return jsonify({'ok': False, 'error': 'user_not_found'}), 404

    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not (name or email or phone):
        return jsonify({'ok': False, 'error': 'missing_identity'}), 400
    if not name:
        name = email or phone

    company = (data.get('company') or '').strip()
    city = (data.get('city') or '').strip()
    notes = (data.get('notes') or '').strip()
    source = (data.get('source') or 'private_area').strip() or 'private_area'
    last_order_ref = (data.get('last_order_ref') or '').strip()
    client_id = (data.get('client_id') or '').strip()

    try:
        if client_id and str(client_id).isdigit():
            # Update si pertany a l'usuari (mai d'un altre distribuïdor)
            existing = query(
                'SELECT id FROM pro_clients WHERE id=? AND pro_user_id=?',
                [int(client_id), user_id], one=True
            )
            if not existing:
                return jsonify({'ok': False, 'error': 'client_not_found'}), 404
            execute(
                'UPDATE pro_clients SET nom=?, empresa=?, email=?, telefon=?, '
                'poblacio=?, notes=?, source=?, last_order_ref=?, '
                'updated_at=CURRENT_TIMESTAMP WHERE id=? AND pro_user_id=?',
                [name, company, email, phone, city, notes, source, last_order_ref,
                 int(client_id), user_id]
            )
            row = query('SELECT * FROM pro_clients WHERE id=?', [int(client_id)], one=True)
        else:
            # Dedup soft: si ja existeix un client del mateix usuari amb el
            # mateix email (no buit), l'actualitzem en comptes de duplicar.
            existing = None
            if email:
                existing = query(
                    'SELECT id FROM pro_clients WHERE pro_user_id=? AND LOWER(email)=?',
                    [user_id, email.lower()], one=True
                )
            if existing:
                execute(
                    'UPDATE pro_clients SET nom=?, empresa=?, email=?, telefon=?, '
                    'poblacio=?, notes=?, source=?, last_order_ref=?, '
                    'updated_at=CURRENT_TIMESTAMP WHERE id=? AND pro_user_id=?',
                    [name, company, email, phone, city, notes, source, last_order_ref,
                     existing['id'], user_id]
                )
                row = query('SELECT * FROM pro_clients WHERE id=?', [existing['id']], one=True)
            else:
                execute(
                    'INSERT INTO pro_clients (pro_user_id, nom, empresa, email, telefon, '
                    'poblacio, notes, source, last_order_ref) '
                    'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    [user_id, name, company, email, phone, city, notes, source, last_order_ref]
                )
                # Recuperem l'últim del mateix usuari (en PG fariem RETURNING,
                # però mantenim portabilitat amb SQLite).
                row = query(
                    'SELECT * FROM pro_clients WHERE pro_user_id=? '
                    'ORDER BY id DESC LIMIT 1', [user_id], one=True
                )
        return jsonify({'ok': True, 'client': _pro_clients_row_to_dict(row)})
    except Exception as exc:
        print(f'pro_clients_save error: {exc}')
        return jsonify({'ok': False, 'error': 'internal_error'}), 500


@app.route('/api/public/pro-clients', methods=['GET'])
def public_pro_clients_list():
    """Llista els clients privats d'un distribuïdor (ordenats per
    updated_at DESC). Auth: X-Bridge-Token. Query: ?username=&limit=."""
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    username = (request.args.get('username') or '').strip()
    user_id = _pro_clients_lookup_user_id(username)
    if not user_id:
        return jsonify({'ok': False, 'error': 'user_not_found'}), 404

    try:
        limit = int(request.args.get('limit') or 50)
    except (TypeError, ValueError):
        limit = 50
    limit = max(1, min(limit, 200))

    rows = query(
        'SELECT * FROM pro_clients WHERE pro_user_id=? '
        'ORDER BY updated_at DESC, id DESC LIMIT ?',
        [user_id, limit]
    ) or []
    return jsonify({
        'ok': True,
        'clients': [_pro_clients_row_to_dict(r) for r in rows],
    })


@app.route('/api/public/pro-clients/<int:client_id>', methods=['GET'])
def public_pro_clients_get(client_id):
    """Obté un client privat (només si pertany al distribuïdor).
    Auth: X-Bridge-Token. Query: ?username=."""
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    username = (request.args.get('username') or '').strip()
    user_id = _pro_clients_lookup_user_id(username)
    if not user_id:
        return jsonify({'ok': False, 'error': 'user_not_found'}), 404

    row = query(
        'SELECT * FROM pro_clients WHERE id=? AND pro_user_id=?',
        [client_id, user_id], one=True
    )
    if not row:
        return jsonify({'ok': False, 'error': 'client_not_found'}), 404
    return jsonify({'ok': True, 'client': _pro_clients_row_to_dict(row)})


@app.route('/api/public/compute', methods=['GET'])
def public_compute():
    """Calcula un preu base (sense marge del client, sense IVA) per un producte concret.

    Auth: X-Bridge-Token (mateix que la resta de bridge endpoints).

    Query params:
      kind        = 'impressio' | 'laminate' | 'protter' | 'frame'
      width_cm    = int (obligatori)
      height_cm   = int (obligatori)
      paper       = string opcional (impressió: 'lustre','silk','fine_art',...)
      finish      = string opcional ('none','laminate','protter','foam')
      moldura_id  = string opcional (només si kind=frame)
      qty         = int opcional, defecte 1

    Resposta: {ok, kind, width_cm, height_cm, base_price, vat_rate, breakdown}
    """
    provided_token = request.headers.get('X-Bridge-Token', '').strip()
    if not _bridge_token_ok(provided_token) or not _bridge_ip_allowed():
        return jsonify({'ok': False, 'error': 'forbidden'}), 403

    kind = (request.args.get('kind') or '').strip().lower()
    if kind not in ('impressio', 'laminate', 'protter', 'frame'):
        return jsonify({'ok': False, 'error': 'unknown_kind'}), 400

    try:
        w = float(request.args.get('width_cm', 0))
        h = float(request.args.get('height_cm', 0))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'invalid_size'}), 400
    if w <= 0 or h <= 0:
        return jsonify({'ok': False, 'error': 'invalid_size'}), 400

    try:
        qty = max(1, int(request.args.get('qty', 1)))
    except (TypeError, ValueError):
        qty = 1

    paper = (request.args.get('paper') or '').strip().lower()
    finish = (request.args.get('finish') or 'none').strip().lower()
    moldura_id = (request.args.get('moldura_id') or '').strip()

    breakdown = {}
    base_price = 0.0

    if kind == 'impressio':
        # Impressió: usa la taula amb min-contain (no té preu_cost, ja porta marge)
        imp = _imp_closest(w, h)
        if not imp:
            return jsonify({'ok': False, 'error': 'impressio_not_found'}), 404
        base_price = float(imp.get('preu', 0))
        breakdown['impressio'] = base_price
        # Acabats opcionals
        if finish == 'laminate':
            lam = _laminate_only_closest(w, h)
            if lam:
                breakdown['finish'] = float(lam.get('preu', 0))
                base_price += breakdown['finish']
        elif finish == 'protter':
            prot = calcular_cost_laminat(w, h, tipus='semibrillo')
            breakdown['finish'] = prot['pvd']
            base_price += prot['pvd']
        elif finish == 'foam':
            foam = calcular_cost_foam(w, h)
            breakdown['finish'] = foam['pvd']
            base_price += foam['pvd']

    elif kind == 'laminate':
        lam = _laminate_only_closest(w, h)
        if not lam:
            return jsonify({'ok': False, 'error': 'laminate_not_found'}), 404
        base_price = float(lam.get('preu', 0))
        breakdown['laminate'] = base_price

    elif kind == 'protter':
        r = calcular_cost_laminat(w, h, tipus='semibrillo')
        base_price = r['pvd']
        breakdown['protter'] = r['pvd']
        breakdown['origen'] = r['origen']

    elif kind == 'frame':
        if not moldura_id:
            return jsonify({'ok': False, 'error': 'missing_moldura_id'}), 400
        m = query('SELECT preu_taller, preu_cost, gruix, merma_pct, minim_cm FROM moldures WHERE LOWER(referencia)=LOWER(?)', [moldura_id], one=True)
        if not m:
            return jsonify({'ok': False, 'error': 'moldura_not_found'}), 404
        pc = _row_get(m, 'preu_cost')
        gruix = float(_row_get(m, 'gruix') or 0)
        merma = float(_row_get(m, 'merma_pct') or 10.0)
        minim = float(_row_get(m, 'minim_cm') or 100.0)
        if pc is not None:
            marc = calcular_preu_marc(w, h, gruix, float(pc), merma_pct=merma, minim_cm=minim)
            if not marc:
                return jsonify({'ok': False, 'error': 'compute_failed'}), 500
            base_price = marc['pvd']
            breakdown['moldura'] = marc['pvd']
            breakdown['cost'] = marc['cost']
        else:
            # Fallback a preu_taller com €/cm lineal
            preu_cm = float(_row_get(m, 'preu_taller') or 0)
            perimetre = 2 * (w + h)
            longitud = (perimetre + gruix * 8) * (1 + merma / 100)
            longitud = max(longitud, minim)
            base_price = round(longitud * preu_cm / 100, 4)
            breakdown['moldura'] = base_price

    # Apply quantity
    total = round(base_price * qty, 4)

    return jsonify({
        'ok': True,
        'kind': kind,
        'width_cm': w,
        'height_cm': h,
        'qty': qty,
        'base_price': total,
        'vat_rate': 0.21,
        'breakdown': breakdown,
    })


@app.route('/auth/bridge')
def bridge_auth():
    payload = _read_bridge_token(request.args.get('token'))
    if not payload:
        flash("L'accés unificat ha caducat o no és vàlid. Torna a iniciar sessió.", 'error')
        return redirect(url_for('login'))

    user = query('SELECT * FROM usuaris WHERE id=?', [payload.get('uid')], one=True)
    if not _user_is_allowed(user):
        flash("No hem pogut validar aquest accés. Contacta amb administració si cal.", 'error')
        return redirect(url_for('login'))

    _start_user_session(user)
    session['bridge_source'] = str(payload.get('source') or '').strip().lower()
    session['bridge_lang'] = str(payload.get('lang') or 'ca').strip().lower()
    target = _safe_next_path(payload.get('next'), '/')
    if _needs_setup(user['id']) and target != url_for('logout'):
        return redirect(url_for('setup'))
    return redirect(target)

# â"€â"€ Routes: App principal â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

@app.route('/ajuda')
@login_required
def ajuda():
    return render_template('ajuda.html')

@app.route('/setup')
@login_required
def setup():
    return render_template('setup.html')

@app.route('/api/setup-done', methods=['POST'])
@login_required
def setup_done():
    try:
        execute('UPDATE usuaris SET setup_done=1 WHERE id=?', [session['user_id']])
    except:
        pass
    return jsonify({'ok': True})

@app.route('/')
@login_required
def index():
    try:
        u = query('SELECT setup_done FROM usuaris WHERE id=?', [session['user_id']], one=True)
        if u and not bool(_row_get(u, 'setup_done', 0)):
            return redirect(url_for('setup'))
    except:
        pass
    web_url = _current_web_return_url()
    if web_url:
        return redirect(web_url)
    return redirect(url_for('calculadora'))


def _parse_comanda_date(value):
    """Interpreta el camp `data` (text lliure) com a data; None si no es pot."""
    s = str(value or '').strip()
    if not s:
        return None
    s = s.split(' ')[0].split('T')[0]
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return None


def _dashboard_counts(urgent_days=21):
    """Comptadors del tauler derivats de l'estat EXISTENT de `comandes`
    (marcador [ACCEPTAT] a observacions, flags pagat/entregat, fd_albara).
    NOMÉS lectura: no modifica ni el model ni les dades. Agrupa per
    sessio_id (un pressupost/comanda lògic, encara que tingui opcions A/B)."""
    is_admin = _is_admin_session()
    sql = "SELECT sessio_id, id, observacions, pagat, entregat, fd_albara, data FROM comandes"
    args = []
    if not is_admin:
        sql += " WHERE user_id=?"
        args.append(session.get('user_id'))
    rows = query(sql, args) or []
    today = datetime.now().date()
    groups = {}
    for r in rows:
        sid = _row_get(r, 'sessio_id') or ('id:' + str(_row_get(r, 'id')))
        g = groups.get(sid)
        if g is None:
            g = {'accept': False, 'pagat': False, 'entregat': False, 'albara': False, 'date': None}
            groups[sid] = g
        if '[ACCEPTAT]' in str(_row_get(r, 'observacions', '') or ''):
            g['accept'] = True
        if _row_get(r, 'pagat', 0):
            g['pagat'] = True
        if _row_get(r, 'entregat', 0):
            g['entregat'] = True
        if str(_row_get(r, 'fd_albara', '') or '').strip():
            g['albara'] = True
        d = _parse_comanda_date(_row_get(r, 'data'))
        if d and (g['date'] is None or d < g['date']):
            g['date'] = d
    c = {'pressupostos': 0, 'comandes': 0, 'pendents_albara': 0,
         'pendents_cobrament': 0, 'pendents_entrega': 0, 'entregats': 0, 'urgents': 0}
    for g in groups.values():
        if not g['accept']:
            c['pressupostos'] += 1
            continue
        c['comandes'] += 1
        if not g['albara']:
            c['pendents_albara'] += 1
        if not g['pagat']:
            c['pendents_cobrament'] += 1
        if not g['entregat']:
            c['pendents_entrega'] += 1
            if g['date'] and (today - g['date']).days >= urgent_days:
                c['urgents'] += 1
        else:
            c['entregats'] += 1
    return c


@app.route('/inici')
@login_required
def inici():
    """Tauler d'inici: accessos ràpids + comptadors de pedidos/pressupostos
    per estat. NOMÉS lectura (F1)."""
    try:
        u = query('SELECT setup_done FROM usuaris WHERE id=?', [session['user_id']], one=True)
        if u and not bool(_row_get(u, 'setup_done', 0)):
            return redirect(url_for('setup'))
    except Exception:
        pass
    try:
        counts = _dashboard_counts()
    except Exception:
        app.logger.exception('dashboard_counts_failed')
        counts = {'pressupostos': 0, 'comandes': 0, 'pendents_albara': 0,
                  'pendents_cobrament': 0, 'pendents_entrega': 0, 'entregats': 0, 'urgents': 0}
    # Novetats (what's new): es mostren en entrar (post-login), com a la calculadora.
    novetats_pendents = []
    try:
        nrow = query('SELECT novetats_vistes FROM usuaris WHERE id=?', [session['user_id']], one=True)
        novetats_pendents = _novetats_pendents(_row_get(nrow, 'novetats_vistes', '') or '',
                                                bool(session.get('is_admin')))
    except Exception as e:
        print(f'[novetats] inici lectura pendents skip: {e}')
    return render_template('inici.html', counts=counts,
                           novetats_pendents=novetats_pendents,
                           novetats_ids=[n['id'] for n in novetats_pendents])


# ── Sistema de novetats (what's new) ───────────────────────────────────────
# Registre reutilitzable: per anunciar una novetat nova, afegeix-hi una entrada.
# Cada usuari veu les novetats que encara no ha tancat (guardades a
# usuaris.novetats_vistes, llista d'ids separada per comes). 'audiencia':
# 'tots' | 'admin' (només Reus Revela) | 'usuaris' (només professionals).
# 'punts_admin' són punts extra que només veuen els admins.
NOVETATS = [
    {
        'id': 'calc-tots-productes',
        'data': '17/07/2026',
        'titol': 'Ara pots pressupostar tots els productes',
        'audiencia': 'tots',
        'intro': 'La calculadora ja no és només de marcs: amb el selector de dalt («Què vols pressupostar?») pots calcular tots els productes i ajuntar-los al mateix pressupost.',
        'punts': [
            '🖼️ Marcs i 🖨️ impressió fotogràfica, com sempre.',
            '🎨 Llenços: mides de catàleg o a mida (preu coherent), amb bastidor o enrotllat.',
            '📔 Àlbums: individual (mida, plecs, coberta i extres) i packs de boda i comunió.',
            '🎓 Orles i 🎁 regals (tasses i imants).',
            '💾 Digitalització de cintes: llista de cintes amb durada i reparació, amb total sense IVA per enviar directament.',
        ],
        'punts_admin': [
            'Tot va a la mateixa cistella amb PVD/PVP segons el client, PDF, Pressupost/Albarà a Factura Directa i botó de copiar el text per WhatsApp.',
        ],
    },
    {
        'id': 'marges-per-producte',
        'data': '17/07/2026',
        'titol': 'Marge configurable per a cada producte',
        'audiencia': 'tots',
        'intro': 'A Ajustos ara pots definir un marge diferent per a cada producte, no només per als marcs.',
        'punts': [
            'Entra a Ajustos i posa el marge que vulguis a cada categoria: llenços, àlbums, digitalització, orles i regals.',
            'Cada producte de la calculadora aplica el seu marge automàticament.',
        ],
    },
    {
        'id': 'pressupost-multimarc',
        'data': '15/06/2026',
        'titol': 'Pressupost amb diversos marcs',
        'audiencia': 'tots',
        'intro': 'Ara pots ajuntar diversos marcs en un sol pressupost per al mateix client.',
        'punts': [
            'Calcula un marc i prem «+ Pressupost» a la barra inferior per afegir-lo.',
            'Repeteix amb tants marcs com vulguis: el botó flotant 🛒 mostra el total.',
            'Obre la cistella i descarrega el PDF del pressupost amb tots els marcs i el total.',
        ],
        'punts_admin': [
            'Com a admin, a més del PDF pots enviar-ho directament a Factura Directa com a Pressupost o Albarà (PVP del client o cost de taller).',
        ],
    },
]


def _novetats_pendents(vistes_raw, is_admin):
    """Retorna les novetats que aquest usuari encara no ha tancat, segons
    audiència i la llista d'ids ja vistos (text separat per comes)."""
    vistes = {x.strip() for x in (vistes_raw or '').split(',') if x.strip()}
    out = []
    for n in NOVETATS:
        aud = n.get('audiencia', 'tots')
        if aud == 'admin' and not is_admin:
            continue
        if aud == 'usuaris' and is_admin:
            continue
        if n['id'] in vistes:
            continue
        out.append(n)
    return out


@app.route('/api/novetats/vist', methods=['POST'])
@login_required
def api_novetats_vist():
    """Marca una o més novetats com a vistes per a l'usuari actual."""
    d = request.get_json(silent=True) or {}
    ids = d.get('ids') or []
    if not isinstance(ids, list):
        ids = []
    ids = [str(i).strip() for i in ids if str(i).strip()][:50]
    if not ids:
        return jsonify({'ok': True})
    try:
        row = query('SELECT novetats_vistes FROM usuaris WHERE id=?', [session['user_id']], one=True)
        vistes = {x.strip() for x in (_row_get(row, 'novetats_vistes', '') or '').split(',') if x.strip()}
        vistes.update(ids)
        valid = {n['id'] for n in NOVETATS}
        vistes = {v for v in vistes if v in valid}  # no deixem créixer amb ids obsolets
        execute('UPDATE usuaris SET novetats_vistes=? WHERE id=?',
                [','.join(sorted(vistes)), session['user_id']])
    except Exception as e:
        print(f'[novetats] marcar vist error: {e}')
    return jsonify({'ok': True})


# ── Tarifa de llenços (portada de la web reusrevela) ───────────────────────
# El `price` de cada mida és el PVD (cost/taller). El PVP es deriva aplicant
# `margin_percent`. Sense bastidor (només impressió laminada enrotllada) es
# cobra `print_only_factor` del preu de la mida muntada.
CANVAS_PRICING = {
    'margin_percent': 30.0,
    'print_only_factor': 0.65,
    'sizes': [
        {'group': 'standard', 'w': 30,  'h': 30,  'file': [40, 40],   'price': 39.50},
        {'group': 'standard', 'w': 30,  'h': 40,  'file': [40, 50],   'price': 40.00},
        {'group': 'standard', 'w': 30,  'h': 50,  'file': [40, 60],   'price': 45.00},
        {'group': 'standard', 'w': 40,  'h': 40,  'file': [50, 50],   'price': 48.00},
        {'group': 'standard', 'w': 40,  'h': 50,  'file': [50, 60],   'price': 51.00},
        {'group': 'standard', 'w': 50,  'h': 50,  'file': [60, 60],   'price': 52.00},
        {'group': 'standard', 'w': 50,  'h': 60,  'file': [60, 70],   'price': 54.00},
        {'group': 'standard', 'w': 50,  'h': 70,  'file': [60, 80],   'price': 56.00},
        {'group': 'standard', 'w': 50,  'h': 80,  'file': [60, 90],   'price': 60.00},
        {'group': 'standard', 'w': 60,  'h': 60,  'file': [70, 70],   'price': 60.00},
        {'group': 'standard', 'w': 60,  'h': 80,  'file': [70, 90],   'price': 66.00},
        {'group': 'standard', 'w': 60,  'h': 90,  'file': [70, 100],  'price': 68.00},
        {'group': 'standard', 'w': 60,  'h': 100, 'file': [70, 110],  'price': 79.00},
        {'group': 'standard', 'w': 70,  'h': 70,  'file': [80, 80],   'price': 65.00},
        {'group': 'standard', 'w': 70,  'h': 100, 'file': [80, 110],  'price': 80.00},
        {'group': 'standard', 'w': 70,  'h': 150, 'file': [80, 160],  'price': 110.00},
        {'group': 'standard', 'w': 80,  'h': 80,  'file': [90, 90],   'price': 87.00},
        {'group': 'standard', 'w': 80,  'h': 100, 'file': [90, 110],  'price': 92.00},
        {'group': 'standard', 'w': 80,  'h': 150, 'file': [90, 160],  'price': 120.00},
        {'group': 'standard', 'w': 90,  'h': 90,  'file': [100, 100], 'price': 95.00},
        {'group': 'standard', 'w': 90,  'h': 120, 'file': [100, 130], 'price': 105.00},
        {'group': 'standard', 'w': 90,  'h': 150, 'file': [100, 160], 'price': 120.00},
        {'group': 'standard', 'w': 90,  'h': 200, 'file': [100, 210], 'price': 170.00},
        {'group': 'standard', 'w': 100, 'h': 100, 'file': [110, 110], 'price': 105.00},
        {'group': 'standard', 'w': 100, 'h': 150, 'file': [110, 160], 'price': 135.00},
        {'group': 'panoramic', 'w': 40, 'h': 100, 'file': [50, 110],  'price': 60.00},
        {'group': 'panoramic', 'w': 40, 'h': 140, 'file': [50, 150],  'price': 72.50},
        {'group': 'panoramic', 'w': 60, 'h': 120, 'file': [70, 130],  'price': 80.00},
        {'group': 'panoramic', 'w': 50, 'h': 120, 'file': [60, 130],  'price': 78.00},
        {'group': 'panoramic', 'w': 50, 'h': 200, 'file': [60, 210],  'price': 105.00},
        {'group': 'panoramic', 'w': 60, 'h': 150, 'file': [70, 160],  'price': 95.00},
        {'group': 'panoramic', 'w': 60, 'h': 200, 'file': [70, 210],  'price': 105.00},
        {'group': 'panoramic', 'w': 65, 'h': 120, 'file': [75, 130],  'price': 87.00},
        {'group': 'panoramic', 'w': 65, 'h': 200, 'file': [75, 210],  'price': 125.00},
        {'group': 'panoramic', 'w': 70, 'h': 200, 'file': [80, 210],  'price': 140.00},
        {'group': 'panoramic', 'w': 80, 'h': 200, 'file': [90, 210],  'price': 155.00},
    ],
    'edit_options': [
        {'id': 'none',           'label': 'Arxiu preparat pel fotògraf', 'price': 0.0},
        {'id': 'extend_only',    'label': 'Ampliar llenç (marges bastidor)', 'price': 2.0},
        {'id': 'extend_quality', 'label': 'Ampliar i adaptar qualitat', 'price': 5.0},
        {'id': 'full_retouch',   'label': 'Retoc complet (marges + qualitat + neteja)', 'price': 15.0},
    ],
    # Mides a mida: l'arxiu és final + 5 cm per costat. Límits del rotlle.
    'constants':   {'file_margin_cm': 10},
    'constraints': {'min_cm': 30, 'max_w_cm': 100, 'max_h_cm': 200, 'step_cm': 10},
}


def _build_canvas_price_anchors(sizes):
    """Punts d'ancoratge [àrea_cm², preu PVD] per interpolar el preu de les mides
    PERSONALITZADES a partir del catàleg estàndard, de manera que una mida a mida
    sempre quedi ENTRE els presets que la voregen (mai més barata que una de més
    petita). Igual que a la web:
      1) agrupem per àrea (mitjana quan hi ha diverses formes d'igual àrea),
      2) imposem monotonia creixent amb regressió isotònica (Pool Adjacent
         Violators),
      3) retornem [[àrea, preu], ...] ordenat per àrea."""
    by_area = {}
    for s in sizes:
        by_area.setdefault(s['w'] * s['h'], []).append(s['price'])
    pts = sorted((area, sum(v) / len(v)) for area, v in by_area.items())
    areas = [a for a, _ in pts]
    vals  = [p for _, p in pts]
    wts   = [1.0] * len(vals)
    idx   = [[i] for i in range(len(vals))]
    i = 0
    while i < len(vals) - 1:
        if vals[i] > vals[i + 1] + 1e-9:
            vals[i] = (vals[i] * wts[i] + vals[i + 1] * wts[i + 1]) / (wts[i] + wts[i + 1])
            wts[i] += wts[i + 1]
            idx[i] += idx[i + 1]
            del vals[i + 1], wts[i + 1], idx[i + 1]
            if i > 0:
                i -= 1
        else:
            i += 1
    iso = [0.0] * len(areas)
    for v, group_idx in zip(vals, idx):
        for j in group_idx:
            iso[j] = round(v, 2)
    return [[areas[k], iso[k]] for k in range(len(areas))]


CANVAS_PRICING['anchors'] = _build_canvas_price_anchors(CANVAS_PRICING['sizes'])


# ── Tarifa d'àlbums (portada de la web reusrevela) ─────────────────────────
# Preus NETS = PVD (cost/taller). El PVP = PVD × (1 + margin_percent/100).
# Àlbum individual: PVD = portada + plecs×preu_plec + suplement material + extres.
# Packs (boda/comunió): preu fix (PVD).
ALBUM_PRICING = {
    'margin_percent': 40.0,
    'min_sheets': 5,
    'max_sheets': 40,
    'default_sheets': 7,
    'sizes': [
        {'id': '15x15', 'label': '15×15', 'album_cover': 0,     'booklet_cover': 11.00, 'sheet_price': 2.00},
        {'id': '20x15', 'label': '20×15', 'album_cover': 27.50, 'booklet_cover': 13.20, 'sheet_price': 3.27},
        {'id': '20x20', 'label': '20×20', 'album_cover': 27.50, 'booklet_cover': 15.40, 'sheet_price': 3.87},
        {'id': '25x25', 'label': '25×25', 'album_cover': 33.00, 'booklet_cover': 18.70, 'sheet_price': 4.48},
        {'id': '30x20', 'label': '30×20', 'album_cover': 35.20, 'booklet_cover': 16.50, 'sheet_price': 5.08},
        {'id': '30x24', 'label': '30×24', 'album_cover': 38.50, 'booklet_cover': 19.80, 'sheet_price': 5.87},
        {'id': '30x30', 'label': '30×30', 'album_cover': 32.00, 'booklet_cover': 19.80, 'sheet_price': 6.90},
        {'id': '30x40', 'label': '30×40', 'album_cover': 44.00, 'booklet_cover': 24.20, 'sheet_price': 8.95},
        {'id': '40x30', 'label': '40×30', 'album_cover': 44.00, 'booklet_cover': 24.20, 'sheet_price': 8.95},
        {'id': '35x35', 'label': '35×35', 'album_cover': 33.00, 'booklet_cover': 16.50, 'sheet_price': 9.32},
    ],
    'materials': [
        {'id': 'lino',  'label': 'Lli',          'supplement': 0.0},
        {'id': 'fusta', 'label': 'Fusta',        'supplement': 15.0},
        {'id': 'foto',  'label': 'Foto-portada', 'supplement': 0.0},
    ],
    'extras': [
        {'id': 'uvi',     'label': 'Gravació UVI',  'price': 6.0},
        {'id': 'pintura', 'label': 'Pintura caixa', 'price': 3.0},
    ],
    'wedding_packs': [
        {'id': 'boda_30x30_sol',              'label': 'Boda 30×30 sol (30 plecs, UVI, caixa foto)',            'price': 239.00},
        {'id': 'boda_30x30_repliques',        'label': 'Boda 30×30 + 2 rèpliques pares 20×20',                  'price': 379.00},
        {'id': 'boda_30x30_fusta',            'label': 'Boda 30×30 + caixa de fusta',                           'price': 274.00},
        {'id': 'boda_30x30_fusta_repliques',  'label': 'Boda 30×30 + fusta + 2 rèpliques 20×20',                'price': 414.00},
        {'id': 'boda_30x40_sol',              'label': 'Boda 30×40 sol (30 plecs, UVI, caixa foto)',            'price': 265.00},
        {'id': 'boda_30x40_repliques',        'label': 'Boda 30×40 + 2 rèpliques pares 20×30',                  'price': 425.00},
        {'id': 'boda_30x40_fusta',            'label': 'Boda 30×40 + caixa de fusta',                           'price': 305.00},
        {'id': 'boda_30x40_fusta_repliques',  'label': 'Boda 30×40 + fusta + 2 rèpliques 20×30',                'price': 465.00},
    ],
    'communion_packs': [
        {'id': 'flop_25_15', 'label': 'Flop 25|15 · Libreto 25×25 (7 f.) + caixa + mini 15×15', 'price': 87.00},
        {'id': 'flop_25_20', 'label': 'Flop 25|20 · Libreto 25×25 (7 f.) + caixa + mini 20×20', 'price': 104.00},
        {'id': 'flop_30_15', 'label': 'Flop 30|15 · Libreto 30×30 (7 f.) + caixa + mini 15×15', 'price': 110.00},
        {'id': 'flop_30_20', 'label': 'Flop 30|20 · Libreto 30×30 (7 f.) + caixa + mini 20×20', 'price': 126.00},
        {'id': 'jack_25_15', 'label': 'Jack 25|15 · Àlbum 25×25 (7 f.) + caixa + mini 15×15',   'price': 100.00},
        {'id': 'jack_25_20', 'label': 'Jack 25|20 · Àlbum 25×25 (7 f.) + caixa + mini 20×20',   'price': 104.00},
        {'id': 'jack_30_15', 'label': 'Jack 30|15 · Àlbum 30×30 (7 f.) + caixa + mini 15×15',   'price': 133.00},
        {'id': 'jack_30_20', 'label': 'Jack 30|20 · Àlbum 30×30 (7 f.) + caixa + mini 20×20',   'price': 126.00},
    ],
}


# ── Tarifa de digitalització (portada de la web reusrevela) ────────────────
# Els preus de la web són el PVD (cost/taller, net). El PVP = PVD × (1+marge).
# L'IVA 21% s'afegeix al resum, com la resta de productes.
DIGITAL_PRICING = {
    'margin_percent': 40.0,
    'min_minuts': 60,  # mínim 1 hora per client (cintes)
    'cintes': {
        'formats': [
            {'id': 'vhs',     'label': 'VHS'},
            {'id': '8mm',     'label': '8 mm'},
            {'id': 'beta',    'label': 'Beta'},
            {'id': 'minidv',  'label': 'MiniDV'},
            {'id': 'barreja', 'label': 'Barreja de formats'},
        ],
        # preu_min = € per minut (net PVD), segons el nombre de cintes.
        'tiers': [
            {'min': 1,  'max': 6,    'label': '1–6 cintes',  'preu_min': 0.20},
            {'min': 7,  'max': 15,   'label': '7–15 cintes', 'preu_min': 0.18},
            {'min': 16, 'max': None, 'label': '16 o més',    'preu_min': 0.16},
        ],
    },
    'dvd': {  # DVD → MP4, preu per unitat segons trams.
        'tiers': [
            {'min': 1,  'max': 6,    'label': '1–6 DVD',  'preu': 8.00},
            {'min': 7,  'max': 15,   'label': '7–15 DVD', 'preu': 7.00},
            {'min': 16, 'max': None, 'label': '16 o més', 'preu': 5.00},
        ],
    },
    's8_petita': {  # Súper 8 bobina petita, preu per unitat segons trams.
        'spec': 'Ø 7,5 cm · 15 m · ±3 min',
        'tiers': [
            {'min': 1, 'max': 5,    'label': '1–5 bobines', 'preu': 12.00},
            {'min': 6, 'max': None, 'label': '6 o més',     'preu': 11.00},
        ],
    },
    's8_grans': [  # Súper 8 bobina gran, preu fix per diàmetre.
        {'id': 'd12', 'label': 'Ø 12 cm · ±14 min', 'preu': 23.00},
        {'id': 'd18', 'label': 'Ø 18 cm · ±30 min', 'preu': 33.00},
        {'id': 'd20', 'label': 'Ø 20 cm · ±40 min', 'preu': 38.00},
    ],
}


# ── Tarifa d'orles (portada de la web) ─────────────────────────────────────
# Preus per unitat segons tram de quantitat (net PVD). PVP = PVD × (1+marge).
# Muntatge opcional: primer disseny (una vegada) + muntatge per alumne.
ORLAS_PRICING = {
    'margin_percent': 40.0,
    'montatge': {'primer_disseny': 25.00, 'per_alumne': 1.00},
    'products': [
        {'id': 'lustre_30x40', 'group': 'Foto Lustre 260 gr', 'label': 'Orla 30×40',              't50': 2.70, 't100': 2.55, 't150': 2.40},
        {'id': 'lustre_30x45', 'group': 'Foto Lustre 260 gr', 'label': 'Orla 30×45',              't50': 3.24, 't100': 3.06, 't150': 2.88},
        {'id': 'lustre_40x50', 'group': 'Foto Lustre 260 gr', 'label': 'Orla 40×50',              't50': 5.40, 't100': 5.10, 't150': 4.80},
        {'id': 'lustre_50x60', 'group': 'Foto Lustre 260 gr', 'label': 'Orla 50×60',              't50': 7.65, 't100': 7.22, 't150': 6.80},
        {'id': 'offset_30x40', 'group': 'Offset Matte 300 gr', 'label': 'Orla 30×40 o 32×45 (SRA3)', 't50': 1.73, 't100': 1.64, 't150': 1.55},
        {'id': 'carnet_10x15', 'group': 'Multi foto carnet',   'label': '10×15 · 8 DNI o 4 DNI + 1 cartera', 't50': 0.48, 't100': 0.43, 't150': 0.38},
        {'id': 'carnet_15x20', 'group': 'Multi foto carnet',   'label': '15×20 · 8 DNI + cartera 7×9',       't50': 0.73, 't100': 0.67, 't150': 0.64},
    ],
}


# ── Tarifa de regals personalitzats (portada de la web) ────────────────────
# Preus per unitat (net PVD). PVP = PVD × (1+marge).
REGALS_PRICING = {
    'margin_percent': 40.0,
    'tasses': [
        {'id': 'blanca',       'label': 'Tassa blanca',              'price': 7.00,  'color': False},
        {'id': 'interior_asa', 'label': 'Interior + ansa de color',  'price': 8.00,  'color': True},
        {'id': 'magica',       'label': 'Tassa màgica',              'price': 10.00, 'color': False},
    ],
    'tassa_colors': [
        {'id': 'blanc', 'label': 'Blanc'}, {'id': 'negre', 'label': 'Negre'},
        {'id': 'marro', 'label': 'Marró'}, {'id': 'blau',  'label': 'Blau'},
    ],
    'imants': [
        {'id': '7x10',  'label': 'Imant 7×10',           'price': 1.60},
        {'id': '9x13',  'label': 'Imant 9×13 / 10×15',   'price': 3.10},
        {'id': '13x18', 'label': 'Imant 13×18 / 15×20',  'price': 5.90},
    ],
}


# ── Tarifa d'impressió digital en offset (portada de la web) ───────────────
# Preus NETS = PVD. PVP = PVD × (1 + marge). IVA 21% al resum. 'mult' = unitats
# que surten per full SRA3 32×45 (informatiu). 'preu'/'preu_esp' = paper bàsic /
# paper especial. Preus amb impressió doble cara i tall/hendido inclosos.
OFFSET_PRICING = {
    'margin_percent': 40.0,
    # Full sencer SRA3 (32×45) imprès sense tall.
    'full_sheet': {
        'options': [
            {'id': '1cara',  'label': 'Full sencer · 1 cara a color (paper bàsic 300gr)',  'price': 1.82},
            {'id': '2cares', 'label': 'Full sencer · 2 cares a color (paper bàsic 300gr)', 'price': 2.42},
        ],
        'suplement': {'label': 'Suplement acabat (Plata / Daurat / Kraft / Rústic)', 'price': 1.00},
    },
    'products': [
        # Papereria (pàg. 15)
        {'id': 'rec_10x15',     'group': 'Papereria', 'label': 'Recordatoris/Invitacions 10×15',            'mult': 8,  'preu': 0.61, 'preu_esp': 0.73},
        {'id': 'rec_15x15',     'group': 'Papereria', 'label': 'Recordatoris/Invitacions 15×15 o 10×20',    'mult': 6,  'preu': 0.72, 'preu_esp': 0.92},
        {'id': 'rec_13x18',     'group': 'Papereria', 'label': 'Recordatoris/Invitacions 13×18 o 15×20',    'mult': 4,  'preu': 0.79, 'preu_esp': 1.09},
        {'id': 'targeta_visita','group': 'Papereria', 'label': 'Targetes de visita 5,5×8,5',                'mult': 25, 'preu': 0.24, 'preu_esp': 0.28},
        {'id': 'flyer_7x14',    'group': 'Papereria', 'label': 'Flyer 7×14',                                'mult': 12, 'preu': 0.48, 'preu_esp': 0.58},
        {'id': 'targeta_6x6',   'group': 'Papereria', 'label': 'Targetes 6×6',                              'mult': 35, 'preu': 0.24, 'preu_esp': 0.28},
        {'id': 'punt_5x20',     'group': 'Papereria', 'label': 'Punts de llibre 5,5×20',                    'mult': 10, 'preu': 0.55, 'preu_esp': 0.67},
        {'id': 'punt_7x22',     'group': 'Papereria', 'label': 'Punts de llibre 7×22',                      'mult': 8,  'preu': 0.61, 'preu_esp': 0.73},
        {'id': 'diptic_15x20',  'group': 'Papereria', 'label': 'Díptic 15×20 (15×40 obert)',               'mult': 2,  'preu': 1.40, 'preu_esp': 2.00},
        {'id': 'diptic_15x15',  'group': 'Papereria', 'label': 'Díptic 15×15 (15×30) / 10×20 (10×40)',      'mult': 3,  'preu': 0.96, 'preu_esp': 1.35},
        {'id': 'diptic_22x32',  'group': 'Papereria', 'label': 'Díptic 22,5×32 (full SRA3 doblegat)',       'mult': 1,  'preu': 3.63, 'preu_esp': 4.84},
        {'id': 'acordeo',       'group': 'Papereria', 'label': 'Acordeó (15×40, 4 parts de 15×10)',         'mult': 2,  'preu': 1.57, 'preu_esp': 2.18},
        {'id': 'acordeo_tapa',  'group': 'Papereria', 'label': 'Acordeó amb tapa (15×40, 4 parts de 15×10)','mult': 2,  'preu': 5.44, 'preu_esp': 6.05},
        # Calendaris i varis (pàg. 16)
        {'id': 'cal_triangle',   'group': 'Calendaris i varis', 'label': 'Calendari triangle de sobretaula',              'mult': 2,  'preu': 1.60, 'preu_esp': 2.20},
        {'id': 'cal_butxaca',    'group': 'Calendaris i varis', 'label': 'Calendari butxaca 7×10',                        'mult': 16, 'preu': 0.50, 'preu_esp': 0.50},
        {'id': 'cal_paret_12m',  'group': 'Calendaris i varis', 'label': 'Calendari paret 32×45 · 12 mesos + forat',      'mult': 1,  'preu': 2.20, 'preu_esp': 3.40},
        {'id': 'cal_paret_22x32','group': 'Calendaris i varis', 'label': 'Calendari paret 22×32 · 6 fulls + portada',     'mult': 1,  'preu': 8.50, 'preu_esp': 16.00},
        {'id': 'cal_paret_32x45','group': 'Calendaris i varis', 'label': 'Calendari paret 32×45 · 6 fulls + portada',     'mult': 1,  'preu': 15.00,'preu_esp': 22.00},
        {'id': 'cal_faldo_petit','group': 'Calendaris i varis', 'label': 'Calendari faldó petit (12×7) + imant',          'mult': 1,  'preu': 1.80, 'preu_esp': None},
        {'id': 'cal_faldo_gran', 'group': 'Calendaris i varis', 'label': 'Calendari faldó gran (33,5×23) + forat',        'mult': 1,  'preu': 3.20, 'preu_esp': 4.30},
        {'id': 'banderoles',     'group': 'Calendaris i varis', 'label': 'Banderoles 10×13 amb corda',                    'mult': 8,  'preu': 0.70, 'preu_esp': 0.80},
    ],
}


@app.route('/calculadora')
@login_required
def calculadora():
    try:
        u = query('SELECT setup_done FROM usuaris WHERE id=?', [session['user_id']], one=True)
        if u and not bool(_row_get(u, 'setup_done', 0)):
            return redirect(url_for('setup'))
    except:
        pass
    try:
        user = query('SELECT brand_color, marge_pro_pct, marge, marge_impressio_pro_pct, marge_impressio, mr_tram1_limit, mr_tram2_limit, mr_tram1_pct, mr_tram2_pct, mr_tram3_pct, mr_trams_vist, email, margins_json FROM usuaris WHERE id=?', [session['user_id']], one=True)
    except Exception as e:
        # Si la columna 'email' encara no s'ha migrat, ho intentem aplicar i reintenta;
        # en cas extrem fem fallback sense email perquè la calculadora no caigui.
        if 'email' in str(e).lower() and ('does not exist' in str(e).lower() or 'no such column' in str(e).lower()):
            try:
                execute("ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''")
            except Exception:
                pass
            try:
                user = query('SELECT brand_color, marge_pro_pct, marge, marge_impressio_pro_pct, marge_impressio, mr_tram1_limit, mr_tram2_limit, mr_tram1_pct, mr_tram2_pct, mr_tram3_pct, mr_trams_vist, email, margins_json FROM usuaris WHERE id=?', [session['user_id']], one=True)
            except Exception:
                user = query('SELECT brand_color, marge_pro_pct, marge, marge_impressio_pro_pct, marge_impressio, mr_tram1_limit, mr_tram2_limit, mr_tram1_pct, mr_tram2_pct, mr_tram3_pct, mr_trams_vist, margins_json FROM usuaris WHERE id=?', [session['user_id']], one=True)
        else:
            raise
    user_has_email = bool((_row_get(user, 'email', '') or '').strip())
    brand_color = _normalize_hex_color(_row_get(user, 'brand_color', DEFAULT_BRAND_COLOR))
    marge_pro_actiu = get_config_value('marge_pro_actiu', '1') == '1'
    marge_pro = _get_marge_value(user) if marge_pro_actiu else 0.0
    marge_imp_pro = _get_marge_impressio_value(user) if marge_pro_actiu else 0.0
    # Trams de marge per a marcs (defaults segurs: si NULL → marge_pro)
    fb = marge_pro if marge_pro_actiu else 0.0
    # Marge actual del propi user per calcular recomanats personalitzats al modal
    marge_actual_user = _row_get(user, 'marge_pro_pct') if _row_get(user, 'marge_pro_pct') is not None else _row_get(user, 'marge')
    mr_trams = {
        'tram1_limit': int(_row_get(user, 'mr_tram1_limit') or MR_TRAM_LIMITS_DEFAULT['tram1_limit']),
        'tram2_limit': int(_row_get(user, 'mr_tram2_limit') or MR_TRAM_LIMITS_DEFAULT['tram2_limit']),
        'tram1_pct':   float(_row_get(user, 'mr_tram1_pct') if _row_get(user, 'mr_tram1_pct') is not None else fb) if marge_pro_actiu else 0.0,
        'tram2_pct':   float(_row_get(user, 'mr_tram2_pct') if _row_get(user, 'mr_tram2_pct') is not None else fb) if marge_pro_actiu else 0.0,
        'tram3_pct':   float(_row_get(user, 'mr_tram3_pct') if _row_get(user, 'mr_tram3_pct') is not None else fb) if marge_pro_actiu else 0.0,
        'defaults_recomanats': get_mr_recomendats(marge_actual_user),
    }
    # Novetats pendents (tolerant si la columna encara no s'ha migrat)
    novetats_pendents = []
    try:
        nrow = query('SELECT novetats_vistes FROM usuaris WHERE id=?', [session['user_id']], one=True)
        novetats_pendents = _novetats_pendents(_row_get(nrow, 'novetats_vistes', '') or '',
                                                bool(session.get('is_admin')))
    except Exception as e:
        print(f'[novetats] lectura pendents skip: {e}')
    return render_template('calculadora.html',
                           web_return_url=_current_web_return_url(),
                           web_order_url=_current_web_order_url(),
                           color_filters=MOLDURA_COLOR_FILTERS,
                           gruix_filters=MOLDURA_GRUIX_FILTERS,
                           brand_color=brand_color,
                           brand_color_light=_mix_with_white(brand_color),
                           marge_pro_actiu=marge_pro_actiu,
                           marge_pro=marge_pro,
                           marge_impressio_pro=marge_imp_pro,
                           is_admin=1 if session.get('is_admin') else 0,
                           mr_trams=mr_trams,
                           mr_trams_vist=bool(_row_get(user, 'mr_trams_vist', 0)),
                           novetats_pendents=novetats_pendents,
                           novetats_ids=[n['id'] for n in novetats_pendents],
                           extras=get_extras_list(),
                           canvas_pricing=CANVAS_PRICING,
                           album_pricing=ALBUM_PRICING,
                           digital_pricing=DIGITAL_PRICING,
                           orlas_pricing=ORLAS_PRICING,
                           regals_pricing=REGALS_PRICING,
                           offset_pricing=OFFSET_PRICING,
                           commercial_margins=_load_user_commercial_margins(user),
                           user_has_email=user_has_email)

@app.route('/api/lookup')
@login_required
def lookup():
    ref = request.args.get('ref', '').strip()
    tipus = request.args.get('tipus', 'moldura')
    is_admin = session.get('is_admin')
    if tipus == 'moldura':
        try:
            r = query('SELECT preu_taller, preu_cost, gruix, merma_pct, minim_cm, descripcio, foto, ref2, descatalogada, notes_stock FROM moldures WHERE LOWER(referencia)=LOWER(?)', [ref], one=True)
            print(f"lookup moldura ref={ref} result={r}")
            if r:
                pvd = calcular_pvd(_row_get(r, 'preu_cost'), 'moldures')
                preu = pvd if pvd is not None else r['preu_taller']
                resp = {'ok': True, 'preu': preu, 'gruix': r['gruix'],
                        'merma_pct': _row_get(r, 'merma_pct', 10.0),
                        'minim_cm': _row_get(r, 'minim_cm', 100.0),
                        'descripcio': r['descripcio'], 'foto': _resolve_moldura_photo(ref, r['foto'], ref2=_row_get(r,'ref2','')),
                        'descatalogada': bool(_row_get(r, 'descatalogada', False)),
                        'notes_stock': _row_get(r, 'notes_stock', '') or ''}
                if is_admin:
                    resp['preu_cost'] = _row_get(r, 'preu_cost')
                return jsonify(resp)
        except Exception as e:
            print(f"lookup ERROR: {e}")
            return jsonify({'ok': False, 'error': str(e)})
    elif tipus == 'vidre':
        r = query('SELECT preu, preu_cost FROM vidres WHERE LOWER(referencia)=LOWER(?)', [ref], one=True)
        if r:
            pvd = calcular_pvd(_row_get(r, 'preu_cost'), 'vidres')
            preu = pvd if pvd is not None else r['preu']
            resp = {'ok': True, 'preu': preu}
            if is_admin: resp['preu_cost'] = _row_get(r, 'preu_cost')
            return jsonify(resp)
    elif tipus == 'passpartout':
        try:
            parts = ref.lower().replace('pas-','').split('x')
            ew, eh = float(parts[0]), float(parts[1])
            is_doble = 'doble' in ref.lower() or 'dob' in ref.lower()
            r = calcular_cost_passpartu(ew, eh, tipus='doble' if is_doble else 'simple')
            resp = {'ok': True, 'preu': r['pvd'], 'origen': r['origen']}
            if is_admin:
                resp['preu_cost'] = r['cost']
            return jsonify(resp)
        except Exception:
            return jsonify({'ok': False})
    elif tipus == 'encolat':
        r = query('SELECT preu, preu_cost FROM encolat_pro WHERE LOWER(referencia)=LOWER(?)', [ref], one=True)
        if r:
            pvd = calcular_pvd(_row_get(r, 'preu_cost'), 'encolat')
            preu = pvd if pvd is not None else r['preu']
            resp = {'ok': True, 'preu': preu}
            if is_admin: resp['preu_cost'] = _row_get(r, 'preu_cost')
            return jsonify(resp)
    elif tipus == 'impressio':
        r = query('SELECT preu, descripcio FROM impressio WHERE LOWER(referencia)=LOWER(?)', [ref], one=True)
        if r: return jsonify({'ok': True, 'preu': r['preu'], 'descripcio': r['descripcio']})
    return jsonify({'ok': False})

@app.route('/api/refs')
@login_required
def refs():
    tipus = request.args.get('tipus', 'moldura')
    tables = {'moldura': ('moldures', 'referencia'),
              'vidre': ('vidres', 'referencia'),
              'passpartout': ('passpartout', 'referencia'),
              'encolat': ('encolat_pro', 'referencia'),
              'impressio': ('impressio', 'referencia')}
    if tipus not in tables: return jsonify([])
    t, col = tables[tipus]
    rows = query(f'SELECT {col} FROM {t} ORDER BY {col}')
    col = list(rows[0].keys())[0] if rows else 'referencia'
    return jsonify([r[col] for r in rows])


@app.route('/api/pendents-albara')
@login_required
def api_pendents_albara():
    if not session.get('is_admin'):
        return jsonify({'n': 0})
    row = query('''SELECT COUNT(*) as n FROM comandes
                   WHERE observacions LIKE '%[ACCEPTAT]%'
                     AND (fd_albara IS NULL OR fd_albara='')''', one=True)
    return jsonify({'n': row['n'] if row else 0})


@app.route('/api/feedback', methods=['POST'])
@login_required
def api_feedback():
    d = request.get_json(force=True) or {}
    missatge = (d.get('missatge') or '').strip()
    if not missatge:
        return jsonify({'ok': False, 'error': 'Cal escriure un missatge'}), 400
    tipus = d.get('tipus', 'millora')
    if tipus not in ('error', 'millora', 'altre'):
        tipus = 'millora'
    pagina = (d.get('pagina') or '')[:200]
    from datetime import datetime
    execute("INSERT INTO feedback (user_id, tipus, missatge, pagina, data) VALUES (?,?,?,?,?)",
            [session['user_id'], tipus, missatge[:2000], pagina, datetime.now().strftime('%Y-%m-%d %H:%M')])
    return jsonify({'ok': True})


@app.route('/admin/feedback')
@admin_required
def admin_feedback():
    rows = query('''SELECT f.*, u.nom as usuari_nom FROM feedback f
                    JOIN usuaris u ON f.user_id=u.id
                    ORDER BY f.id DESC LIMIT 100''')
    # Mark all as read
    execute("UPDATE feedback SET llegit=1 WHERE llegit=0")
    return render_template('admin_feedback.html', feedbacks=rows)


@app.route('/api/feedback/count')
@login_required
def api_feedback_count():
    if not session.get('is_admin'):
        return jsonify({'n': 0})
    row = query("SELECT COUNT(*) as n FROM feedback WHERE llegit=0", one=True)
    return jsonify({'n': row['n'] if row else 0})


@app.route('/admin/preus-cost')
@admin_required
def admin_preus_cost():
    """Llistat de preus de cost amb filtres per taula, proveïdor i verificat."""
    taules_valides = ['moldures', 'vidres', 'encolat_pro', 'passpartout']
    taula = request.args.get('taula', 'moldures')
    if taula not in taules_valides:
        taula = 'moldures'
    proveidor = request.args.get('proveidor', '').strip()
    verificat = request.args.get('verificat', '')

    preu_orig = 'preu_taller' if taula == 'moldures' else 'preu'
    # Només 'moldures' té descripcio i proveidor; les altres (vidres, encolat_pro,
    # passpartout) només tenen referencia + preu + columnes v2.
    base_cols = 'referencia, '
    if taula == 'moldures':
        base_cols += 'descripcio, '
    cols = f'{base_cols}preu_cost, preu_cost_ant, data_cost, cost_verificat, notes_cost, {preu_orig} as preu_original'
    if taula == 'moldures':
        cols += ', proveidor, descatalogada, notes_stock'

    conditions, args = [], []
    if proveidor and taula == 'moldures':
        conditions.append("LOWER(proveidor) LIKE LOWER(?)")
        args.append(f'%{proveidor}%')
    if verificat == '1':
        conditions.append("cost_verificat = 1")
    elif verificat == '0':
        conditions.append("(cost_verificat = 0 OR cost_verificat IS NULL)")
    elif verificat == 'descatalogada' and taula == 'moldures':
        conditions.append("descatalogada = TRUE")

    where = (' WHERE ' + ' AND '.join(conditions)) if conditions else ''
    sql = f'SELECT {cols} FROM {taula}{where} ORDER BY cost_verificat ASC, referencia ASC'
    rows = query(sql, args)

    # Proveïdors per al filtre (només moldures)
    proveidors = []
    if taula == 'moldures':
        prov_rows = query("SELECT DISTINCT proveidor FROM moldures WHERE proveidor IS NOT NULL AND proveidor != '' ORDER BY proveidor")
        proveidors = [r['proveidor'] for r in (prov_rows or [])]

    # Compute PVD + estadístiques per taula
    cat_map = {'moldures': 'moldures', 'vidres': 'vidres', 'encolat_pro': 'encolat', 'passpartout': 'passpartu'}
    categoria = cat_map.get(taula, 'moldures')
    for r in rows:
        pc = _row_get(r, 'preu_cost')
        r['pvd'] = calcular_pvd(pc, categoria) if pc is not None else None

    stats = {}
    for t in taules_valides:
        s = query(f"SELECT COUNT(*) as total, SUM(CASE WHEN cost_verificat=1 THEN 1 ELSE 0 END) as verificats FROM {t}", one=True)
        stats[t] = {'total': s['total'] or 0, 'verificats': s['verificats'] or 0}

    cfg = {r['clau']: r['valor'] for r in (query("SELECT clau, valor FROM config") or [])}
    return render_template('admin_preus_cost.html', rows=rows, taula=taula,
                           proveidor=proveidor, verificat=verificat, proveidors=proveidors, stats=stats, config=cfg)


@app.route('/admin/auditoria-costos')
@admin_required
def admin_auditoria_costos():
    w = float(request.args.get('w', 60))
    h = float(request.args.get('h', 80))
    paper = (request.args.get('paper') or 'lustre').strip().lower()
    if paper not in ('lustre', 'silk', 'baryta', 'poster_mate'):
        paper = 'lustre'
    moldura_ref = (request.args.get('moldura') or '').strip()

    components = []

    # Marc
    if moldura_ref:
        mol = query('SELECT preu_taller, preu_cost, gruix, merma_pct, minim_cm, descripcio FROM moldures WHERE LOWER(referencia)=LOWER(?)', [moldura_ref], one=True)
        if mol:
            marc_res = calcular_preu_marc(w, h, _row_get(mol, 'gruix', 0), _row_get(mol, 'preu_cost'), _row_get(mol, 'merma_pct', 10), _row_get(mol, 'minim_cm', 100))
            marge_mol = float(get_config_value('marge_admin_moldures_pct', '60'))
            components.append({
                'nom': f'Marc ({moldura_ref})',
                'preu_cost_unitari': _row_get(mol, 'preu_cost'),
                'unitat': '€/cm lineal',
                'cost': marc_res['cost'] if marc_res else 0,
                'pvd': marc_res['pvd'] if marc_res else 0,
                'origen': 'taula',
                'marge_pct': marge_mol,
                'detall': f"Perímetre={2*(w+h):.0f}cm, gruix={_row_get(mol,'gruix',0)}, merma={_row_get(mol,'merma_pct',10)}%, mínim={_row_get(mol,'minim_cm',100)}cm",
            })

    # Vidre
    vidre = calcular_cost_vidre(w, h)
    marge_v = float(get_config_value('marge_admin_vidres_pct', '60'))
    components.append({
        'nom': 'Vidre simple',
        'cost': vidre['cost'], 'pvd': vidre['pvd'], 'origen': vidre['origen'], 'ref': vidre['ref'],
        'marge_pct': marge_v,
        'detall': f"cost_cm2={get_config_value('vidre_cost_cm2','0.002880')}, t_base={get_config_value('vidre_temps_base_min','3')}min, t_lineal={get_config_value('vidre_temps_lineal_m','0.5')}min/m, cost_hora={get_config_value('cost_hora_taller','25')}€/h",
    })

    # Doble vidre
    dv = calcular_cost_doble_vidre(w, h)
    components.append({
        'nom': 'Doble vidre',
        'cost': dv['cost'], 'pvd': dv['pvd'], 'origen': dv['origen'], 'ref': dv['ref'],
        'marge_pct': marge_v,
        'detall': f"vidre×2 + muntatge ({get_config_value('vidre_dv_muntatge_eur','1.30')}€)",
    })

    # Passpartú simple
    pas = calcular_cost_passpartu(w, h, tipus='simple')
    marge_p = float(get_config_value('marge_admin_passpartu_pct', '60'))
    components.append({
        'nom': 'Passpartú simple',
        'cost': pas['cost'], 'pvd': pas['pvd'], 'origen': pas['origen'], 'ref': pas['ref'],
        'marge_pct': marge_p,
        'detall': f"cost_cm2={get_config_value('passpartu_cost_cm2','0.001200')}, t_base={get_config_value('passpartu_temps_base_min','5')}min",
    })

    # Passpartú doble
    dpas = calcular_cost_passpartu(w, h, tipus='doble')
    components.append({
        'nom': 'Passpartú doble',
        'cost': dpas['cost'], 'pvd': dpas['pvd'], 'origen': dpas['origen'], 'ref': dpas['ref'],
        'marge_pct': marge_p,
        'detall': 'simple×2 + finestres extra',
    })

    # Foam
    foam = calcular_cost_foam(w, h)
    marge_e = float(get_config_value('marge_admin_encolat_pct', '60'))
    components.append({
        'nom': 'Foam (encolat)',
        'cost': foam['cost'], 'pvd': foam['pvd'], 'origen': foam['origen'], 'ref': foam['ref'],
        'marge_pct': marge_e,
        'detall': f"foam_cost_cm2={get_config_value('foam_cost_cm2','0.001143')}, t_base={get_config_value('foam_temps_base_min','9')}min",
    })

    # Laminat
    lam_semi = calcular_cost_laminat(w, h, tipus='semibrillo')
    components.append({
        'nom': 'Laminat semibrillo',
        'cost': lam_semi['cost'], 'pvd': lam_semi['pvd'], 'origen': lam_semi['origen'], 'ref': lam_semi['ref'],
        'marge_pct': marge_e,
        'detall': f"laminat_cost_cm2={get_config_value('laminat_cost_cm2','0.001000')}",
    })

    # Protter
    prot = calcular_cost_protter(w, h, tipus='semibrillo')
    components.append({
        'nom': 'Protter (foam+laminat)',
        'cost': prot['cost'], 'pvd': prot['pvd'], 'origen': prot['origen'], 'ref': prot['ref'],
        'marge_pct': marge_e,
        'detall': 'foam + laminat combinats',
    })

    # Impressió
    imp = _imp_closest(w, h, paper=paper)
    if imp:
        components.append({
            'nom': f'Impressió ({paper})',
            'cost': imp['preu'], 'pvd': imp['preu'], 'origen': imp['origen'], 'ref': imp.get('ref', ''),
            'marge_pct': 0,
            'detall': f"cost_cm2={get_config_value(f'imp_{paper}_cost_cm2', get_config_value('imp_lustre_cost_cm2','0.000703'))}, àrea={w*h:.0f}cm²",
        })

    # Mirall
    mir = calcular_cost_mirall(w, h)
    components.append({
        'nom': 'Mirall',
        'cost': mir['cost'], 'pvd': mir['pvd'], 'origen': mir['origen'], 'ref': mir['ref'],
        'marge_pct': marge_v,
        'detall': 'Mirall tallat a mida',
    })

    config_params = {
        'marge_admin_moldures_pct': get_config_value('marge_admin_moldures_pct', '60'),
        'marge_admin_vidres_pct': get_config_value('marge_admin_vidres_pct', '60'),
        'marge_admin_passpartu_pct': get_config_value('marge_admin_passpartu_pct', '60'),
        'marge_admin_encolat_pct': get_config_value('marge_admin_encolat_pct', '60'),
        'vidre_cost_cm2': get_config_value('vidre_cost_cm2', '0.002880'),
        'foam_cost_cm2': get_config_value('foam_cost_cm2', '0.001143'),
        'laminat_cost_cm2': get_config_value('laminat_cost_cm2', '0.001000'),
        'passpartu_cost_cm2': get_config_value('passpartu_cost_cm2', '0.001200'),
        'imp_lustre_cost_cm2': get_config_value('imp_lustre_cost_cm2', '0.000703'),
        'cost_hora_taller': get_config_value('cost_hora_taller', '25'),
    }

    return render_template('admin_auditoria_costos.html',
                           w=w, h=h, paper=paper, moldura_ref=moldura_ref,
                           components=components, config=config_params)


@app.route('/admin/preus-cost/update', methods=['POST'])
@admin_required
def admin_preus_cost_update():
    """Actualitza el preu de cost d'un producte i registra a l'historial."""
    d = request.get_json(force=True) or {}
    taula = d.get('taula', 'moldures')
    if taula not in ('moldures', 'vidres', 'encolat_pro', 'passpartout'):
        return jsonify({'ok': False, 'error': 'Taula no vàlida'}), 400
    referencia = (d.get('referencia') or '').strip()
    if not referencia:
        return jsonify({'ok': False, 'error': 'Falta referència'}), 400
    try:
        preu_cost_nou = round(float(d.get('preu_cost_nou', 0)), 4)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': 'Preu no vàlid'}), 400
    notes = (d.get('notes') or '')[:500]

    # Read current cost
    row = query(f'SELECT preu_cost FROM {taula} WHERE referencia=?', [referencia], one=True)
    if not row:
        return jsonify({'ok': False, 'error': f'Referència {referencia} no trobada a {taula}'}), 404
    preu_cost_ant = _row_get(row, 'preu_cost')

    from datetime import datetime
    avui = datetime.now().strftime('%Y-%m-%d')

    # Update the product
    execute(f'''UPDATE {taula} SET preu_cost_ant=?, preu_cost=?, data_cost=?,
                usuari_cost_id=?, cost_verificat=1, notes_cost=?
                WHERE referencia=?''',
            [preu_cost_ant, preu_cost_nou, avui, session['user_id'], notes, referencia])

    # Also update legacy preu/preu_taller to keep backward compat
    cat_map = {'moldures': 'moldures', 'vidres': 'vidres', 'encolat_pro': 'encolat', 'passpartout': 'passpartu'}
    pvd = calcular_pvd(preu_cost_nou, cat_map.get(taula, 'moldures'))
    if taula == 'moldures':
        execute('UPDATE moldures SET preu_taller=? WHERE referencia=?', [pvd, referencia])
    else:
        execute(f'UPDATE {taula} SET preu=? WHERE referencia=?', [pvd, referencia])

    # Insert audit trail
    execute('''INSERT INTO historial_preus_cost (taula, referencia, preu_cost_antic, preu_cost_nou, usuari_id, data, notes)
               VALUES (?,?,?,?,?,?,?)''',
            [taula, referencia, preu_cost_ant, preu_cost_nou, session['user_id'], avui, notes])

    marge_admin = float(get_config_value(f'marge_admin_{cat_map.get(taula, "moldures")}_pct', '60'))
    return jsonify({'ok': True, 'preu_cost_nou': preu_cost_nou, 'pvd': pvd,
                     'preu_cost_ant': preu_cost_ant, 'marge_aplicat': marge_admin})


@app.route('/admin/tarifes/actualitzar', methods=['GET', 'POST'])
@admin_required
def admin_tarifes_actualitzar():
    """Eina d'actualització global de tarifes per categoria.
    Dues famílies de categories:
      - Taules (moldures, vidres, passpartout) → UPDATE de la columna preu_cost
        per cada fila amb preu_cost no nul, amb snapshot (preu_cost_ant) i
        entrada a historial_preus_cost.
      - Config (mirall, encolat_foam, laminat_semibrillo, laminat_mate, protter) →
        UPDATE de les claus corresponents a la taula config. El protter és un
        composite i actualitza dues claus (foam + laminat_semibrillo), de manera
        que incrementar protter un N% equival a incrementar les dues materials."""
    if request.method == 'GET':
        return render_template('admin_tarifes_actualitzar.html')

    accio      = (request.form.get('accio') or '').strip()
    categoria  = (request.form.get('categoria') or '').strip()
    metode     = (request.form.get('metode') or 'percent').strip()
    try:
        valor = float(request.form.get('valor', '0'))
    except ValueError:
        return jsonify({'error': 'Valor no numèric'}), 400

    # Categories sobre taules físiques (cada fila té preu_cost propi).
    TAULES_CATEGORIES = ['moldures', 'vidres', 'passpartout']
    # Categories sobre claus de config (un cost per cm2 global per categoria).
    CONFIG_CATEGORIES = {
        'mirall':             ['mirall_cost_cm2'],
        'encolat_foam':       ['foam_cost_cm2'],
        'laminat_semibrillo': ['laminat_semibrillo_cost_cm2'],
        'laminat_mate':       ['laminat_mate_cost_cm2'],
        # Protter = foam + laminat semibrillo → incrementar les dues claus.
        'protter':            ['foam_cost_cm2', 'laminat_semibrillo_cost_cm2'],
    }

    if categoria not in TAULES_CATEGORIES and categoria not in CONFIG_CATEGORIES:
        return jsonify({'error': 'Categoria no vàlida'}), 400
    if accio not in ('previsualitzar', 'aplicar'):
        return jsonify({'error': "Acció ha de ser 'previsualitzar' o 'aplicar'"}), 400

    previsualitzacio = []
    if categoria in TAULES_CATEGORIES:
        rows = query(
            f"SELECT referencia, preu_cost FROM {categoria} "
            f"WHERE preu_cost IS NOT NULL ORDER BY referencia"
        ) or []
        for r in rows:
            cost_actual = float(r['preu_cost'])
            cost_nou = round(cost_actual * (1 + valor / 100), 4)
            diff_pct = round((cost_nou / cost_actual - 1) * 100, 1) if cost_actual else 0.0
            previsualitzacio.append({
                'referencia':  r['referencia'],
                'cost_actual': cost_actual,
                'cost_nou':    cost_nou,
                'diff_pct':    diff_pct,
            })
    else:
        # CONFIG_CATEGORIES: cada entrada és una clau de config (cost_cm2).
        # Tractem cada clau com una "referència" i mantenim 6 decimals
        # perquè els cost_cm2 són valors petits (~0.00x).
        for clau in CONFIG_CATEGORIES[categoria]:
            val_actual_str = get_config_value(clau, None)
            if val_actual_str is None:
                continue
            cost_actual = float(val_actual_str)
            cost_nou = round(cost_actual * (1 + valor / 100), 6)
            diff_pct = round((cost_nou / cost_actual - 1) * 100, 1) if cost_actual else 0.0
            previsualitzacio.append({
                'referencia':  clau,
                'cost_actual': cost_actual,
                'cost_nou':    cost_nou,
                'diff_pct':    diff_pct,
            })

    if accio == 'previsualitzar':
        import hashlib
        token = hashlib.md5(
            json.dumps(previsualitzacio, sort_keys=True).encode('utf-8')
        ).hexdigest()[:8]
        return jsonify({
            'previsualitzacio': previsualitzacio,
            'token': token,
            'total': len(previsualitzacio),
        })

    # accio == 'aplicar'
    ara = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    admin_id = session.get('user_id')
    notes = f'Actualització global {metode} {"+" if valor >= 0 else ""}{valor}%'
    aplicats = 0
    errors = []

    if categoria in TAULES_CATEGORIES:
        for p in previsualitzacio:
            try:
                execute(
                    f"UPDATE {categoria} SET "
                    f"preu_cost_ant = preu_cost, preu_cost = ?, data_cost = ?, "
                    f"usuari_cost_id = ?, notes_cost = ?, cost_verificat = 1 "
                    f"WHERE referencia = ?",
                    [p['cost_nou'], ara, admin_id, notes, p['referencia']],
                )
                execute(
                    "INSERT INTO historial_preus_cost "
                    "(taula, referencia, preu_cost_antic, preu_cost_nou, usuari_id, notes) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    [categoria, p['referencia'], p['cost_actual'], p['cost_nou'], admin_id, notes],
                )
                aplicats += 1
            except Exception as e:
                errors.append({'referencia': p['referencia'], 'error': str(e)[:120]})
    else:
        # Config categories: UPDATE config SET valor per cada clau.
        for p in previsualitzacio:
            try:
                execute(
                    "UPDATE config SET valor = ? WHERE clau = ?",
                    [str(p['cost_nou']), p['referencia']],
                )
                execute(
                    "INSERT INTO historial_preus_cost "
                    "(taula, referencia, preu_cost_antic, preu_cost_nou, usuari_id, notes) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    ['config', p['referencia'], p['cost_actual'], p['cost_nou'], admin_id, notes],
                )
                aplicats += 1
            except Exception as e:
                errors.append({'referencia': p['referencia'], 'error': str(e)[:120]})

    return jsonify({'ok': True, 'aplicats': aplicats, 'errors': errors})


@app.route('/admin/preus-cost/historial')
@admin_required
def admin_preus_cost_historial():
    """Historial de canvis de preu de cost per una referència."""
    referencia = request.args.get('referencia', '').strip()
    taula = request.args.get('taula', '')
    conditions, args = [], []
    if referencia:
        conditions.append("referencia=?")
        args.append(referencia)
    if taula:
        conditions.append("taula=?")
        args.append(taula)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    rows = query(f'''SELECT h.*, u.nom as usuari_nom
                     FROM historial_preus_cost h
                     LEFT JOIN usuaris u ON h.usuari_id=u.id
                     {where}
                     ORDER BY h.id DESC LIMIT 200''', args)
    return jsonify([dict(r) for r in (rows or [])])


@app.route('/api/moldura-options')
@login_required
def moldura_options():
    rows = query("""SELECT referencia, gruix, descripcio, foto, ref2,
                           descatalogada, notes_stock
                    FROM moldures
                    ORDER BY referencia""")
    return jsonify(_serialize_moldures(rows))

@app.route('/api/marge')
@login_required
def get_marge():
    u = query('SELECT marge, marge_pro_pct, marge_impressio, marge_impressio_pro_pct, nom_empresa, empresa_adreca, empresa_tel, margins_json, brand_color, brand_color_secondary, brand_color_menu FROM usuaris WHERE id=?', [session['user_id']], one=True)
    # Prioritat clara: marge_pro_pct > marge (legacy) > 60.
    # margins_json NO s'ha de fer servir per al marge general — només per a
    # categories específiques (albums, canvas…) si se n'afegeixen en el futur.
    marge_pro_actiu = get_config_value('marge_pro_actiu', '0') == '1'
    if marge_pro_actiu:
        marge = _get_marge_value(u)
    else:
        marge = 0.0
    # marge_impressio es calcula sempre, independentment del flag, perquè
    # marge_impressio_pro_pct és un valor específic de la impressió que no
    # hauria de quedar a 0 només pel fet que el flag d'overrides està off.
    marge_imp = _get_marge_impressio_value(u)
    nom_emp = u['nom_empresa'] if u and u['nom_empresa'] else ''
    brand_color = _normalize_hex_color(_row_get(u, 'brand_color', DEFAULT_BRAND_COLOR))
    brand_color_secondary = _normalize_hex_color(
        _row_get(u, 'brand_color_secondary', DEFAULT_BRAND_SECONDARY_COLOR),
        DEFAULT_BRAND_SECONDARY_COLOR,
    )
    brand_color_menu = _normalize_hex_color(
        _row_get(u, 'brand_color_menu', brand_color),
        brand_color,
    )
    margins = _load_user_commercial_margins(u)
    cfg_rows = query("SELECT clau, valor FROM config WHERE clau LIKE 'empresa_%'")
    cfg = {r['clau']: r['valor'] for r in (cfg_rows or [])}
    if not nom_emp:
        nom_emp = cfg.get('empresa_nom', 'Objectiu Emmarcació')
    user_adreca = _row_get(u, 'empresa_adreca', '') or ''
    user_tel    = _row_get(u, 'empresa_tel', '') or ''
    return jsonify({
        'marge': marge,
        'marge_impressio': marge_imp,
        'margins': margins,
        'brand_color': brand_color,
        'empresa_nom': nom_emp,
        'empresa_adreca': user_adreca if user_adreca else cfg.get('empresa_adreca',''),
        'empresa_tel':    user_tel    if user_tel    else cfg.get('empresa_tel',''),
        # Descomptes per combinació de productes (en %). El front els
        # aplica al subtotal PVP abans del descompte manual del client.
        'combo_desc': {
            'marc_imp_protter': float(get_config_value('combo_desc_marc_imp_protter', '6')),
            'marc_imp_foam':    float(get_config_value('combo_desc_marc_imp_foam', '5')),
            'marc_imp':         float(get_config_value('combo_desc_marc_imp', '3')),
            'marc_suport':      float(get_config_value('combo_desc_marc_suport', '3')),
            'minim_pvp':        float(get_config_value('combo_desc_minim_pvp', '80')),
        },
    })


@app.route('/api/logo', methods=['POST'])
@login_required
def upload_logo():
    f = request.files.get('logo')
    if not f: return jsonify({'ok': False})
    import base64
    data = f.read()
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'png'
    mime = 'image/png' if ext == 'png' else 'image/jpeg'
    b64 = 'data:' + mime + ';base64,' + base64.b64encode(data).decode()
    # Save per user
    try:
        execute('UPDATE usuaris SET logo_b64=? WHERE id=?', [b64, session['user_id']])
    except:
        pass
    return jsonify({'ok': True})

@app.route('/api/logo', methods=['GET'])
@login_required
def get_logo():
    r = query('SELECT logo_b64 FROM usuaris WHERE id=?', [session['user_id']], one=True)
    return jsonify({'url': _row_get(r, 'logo_b64', '') or ''})


# ── Marca per a pressupostos PVD (Reus Revela) ─────────────────────────────
# Els pressupostos a preu PVD (taller) surten amb la marca Reus Revela; a PVP
# (client final) surten amb la marca actual (Objectiu Fotògrafs). Nom, adreça i
# logo de la marca PVD es guarden a config (global, admin).
@app.route('/api/pvd-brand', methods=['GET'])
@admin_required
def get_pvd_brand():
    def _cfg(k):
        r = query("SELECT valor FROM config WHERE clau=?", [k], one=True)
        return (r['valor'] if r else '') or ''
    return jsonify({
        'nom': _cfg('pvd_brand_nom') or 'Reus Revela',
        'adreca': _cfg('pvd_brand_adreca'),
        'logo': _cfg('pvd_brand_logo_b64'),
    })

@app.route('/api/pvd-brand', methods=['POST'])
@admin_required
def save_pvd_brand():
    d = request.get_json(silent=True) or {}
    nom = (d.get('nom') or '').strip()
    adreca = (d.get('adreca') or '').strip()
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('pvd_brand_nom', ?)", [nom])
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('pvd_brand_adreca', ?)", [adreca])
    return jsonify({'ok': True})

@app.route('/api/pvd-brand/logo', methods=['POST'])
@admin_required
def upload_pvd_brand_logo():
    f = request.files.get('logo')
    if not f:
        return jsonify({'ok': False, 'error': 'missing_file'}), 400
    import base64
    data = f.read()
    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in (f.filename or '') else 'png'
    mime = 'image/png' if ext == 'png' else 'image/jpeg'
    b64 = 'data:' + mime + ';base64,' + base64.b64encode(data).decode()
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('pvd_brand_logo_b64', ?)", [b64])
    return jsonify({'ok': True})

@app.route('/static/logo-preview')
@login_required
def logo_preview():
    import base64
    r = query("SELECT valor FROM config WHERE clau='empresa_logo'", one=True)
    if not r or not r['valor']: return '', 404
    data_url = r['valor']
    if data_url.startswith('data:'):
        header, b64 = data_url.split(',', 1)
        mime = header.split(':')[1].split(';')[0]
        data = base64.b64decode(b64)
        from flask import Response
        return Response(data, mimetype=mime)
    return '', 404

@app.route('/api/empresa', methods=['POST'])
@admin_required
def api_empresa():
    # Nota: aquest endpoint és només per admins i actualitza el config global
    # que actua de valors per defecte. Els clients no-admin guarden les seves
    # dades pròpies a usuaris.empresa_adreca / empresa_tel via /api/desar-marge.
    d = request.json or {}
    nom    = d.get('nom','')
    adreca = d.get('adreca','')
    tel    = d.get('tel','')
    execute("UPDATE config SET valor=? WHERE clau='empresa_nom'",    [nom])
    execute("UPDATE config SET valor=? WHERE clau='empresa_adreca'", [adreca])
    execute("UPDATE config SET valor=? WHERE clau='empresa_tel'",    [tel])
    return jsonify({'ok': True})


# ── Trams de marge per a marcs (config per usuari) ────────────────────────
@app.route('/api/marcs-trams', methods=['GET'])
@login_required
def api_marcs_trams_get():
    u = query('SELECT mr_tram1_limit, mr_tram2_limit, mr_tram1_pct, mr_tram2_pct, mr_tram3_pct, mr_trams_vist, marge_pro_pct, marge FROM usuaris WHERE id=?', [session['user_id']], one=True)
    if not u:
        return jsonify({'ok': False, 'error': 'user_not_found'}), 404
    marge_actual = _row_get(u, 'marge_pro_pct') if _row_get(u, 'marge_pro_pct') is not None else _row_get(u, 'marge')
    fb = float(marge_actual or 60)
    return jsonify({
        'ok': True,
        'tram1_limit': int(_row_get(u, 'mr_tram1_limit') or MR_TRAM_LIMITS_DEFAULT['tram1_limit']),
        'tram2_limit': int(_row_get(u, 'mr_tram2_limit') or MR_TRAM_LIMITS_DEFAULT['tram2_limit']),
        'tram1_pct':   float(_row_get(u, 'mr_tram1_pct') if _row_get(u, 'mr_tram1_pct') is not None else fb),
        'tram2_pct':   float(_row_get(u, 'mr_tram2_pct') if _row_get(u, 'mr_tram2_pct') is not None else fb),
        'tram3_pct':   float(_row_get(u, 'mr_tram3_pct') if _row_get(u, 'mr_tram3_pct') is not None else fb),
        'vist': bool(_row_get(u, 'mr_trams_vist', 0)),
        'marge_actual': float(marge_actual) if marge_actual is not None else None,
        'defaults_recomanats': get_mr_recomendats(marge_actual),
    })


@app.route('/api/marcs-trams', methods=['POST'])
@login_required
def api_marcs_trams_post():
    d = request.get_json(silent=True) or {}
    def _f(key, default=None):
        try:
            v = float(d.get(key)) if d.get(key) is not None else default
        except (TypeError, ValueError):
            return default
        return v
    t1 = _f('tram1_pct')
    t2 = _f('tram2_pct')
    t3 = _f('tram3_pct')
    for v in (t1, t2, t3):
        if v is None or v < 0 or v > 500:
            return jsonify({'ok': False, 'error': 'invalid_pct'}), 400
    # Llegir valors antics per audit log
    old = query('SELECT mr_tram1_pct, mr_tram2_pct, mr_tram3_pct FROM usuaris WHERE id=?', [session['user_id']], one=True)
    execute('UPDATE usuaris SET mr_tram1_pct=?, mr_tram2_pct=?, mr_tram3_pct=?, mr_trams_vist=1 WHERE id=?',
            [t1, t2, t3, session['user_id']])
    print(f"[mr_trams] user_id={session['user_id']} {dict(old) if old else {}} → t1={t1} t2={t2} t3={t3}")
    return jsonify({'ok': True, 'tram1_pct': t1, 'tram2_pct': t2, 'tram3_pct': t3})


@app.route('/api/marcs-trams/vist', methods=['POST'])
@login_required
def api_marcs_trams_vist():
    """Marca l'avís com a vist sense canviar valors (per al botó 'Més tard' del modal)."""
    execute('UPDATE usuaris SET mr_trams_vist=1 WHERE id=?', [session['user_id']])
    return jsonify({'ok': True})


@app.route('/api/desar-marge', methods=['POST'])
@login_required
def desar_marge():
    d = request.get_json(silent=True) or {}
    m = float(d.get('marge', 60))
    mi = float(d.get('marge_impressio', 100))
    ne = d.get('nom_empresa', '')
    nf = d.get('nom_fiscal', '')
    fi = d.get('fiscal_id', '')
    ea = d.get('empresa_adreca', '')
    et = d.get('empresa_tel', '')
    em = (d.get('email', '') or '').strip()
    brand_color = _normalize_hex_color(d.get('brand_color', DEFAULT_BRAND_COLOR))
    brand_color_secondary = _normalize_hex_color(
        d.get('brand_color_secondary', DEFAULT_BRAND_SECONDARY_COLOR),
        DEFAULT_BRAND_SECONDARY_COLOR,
    )
    brand_color_menu = _normalize_hex_color(
        d.get('brand_color_menu', brand_color),
        brand_color,
    )
    margins = _normalize_commercial_margins(
        d.get('margins') if isinstance(d.get('margins'), dict) else d,
        frame_margin=m,
        print_margin=mi,
    )
    execute(
        'UPDATE usuaris SET marge=?, marge_impressio=?, nom_empresa=?, nom_fiscal=?, fiscal_id=?, empresa_adreca=?, empresa_tel=?, email=?, margins_json=?, brand_color=?, brand_color_secondary=?, brand_color_menu=? WHERE id=?',
        [
            margins['frames'],
            margins['prints'],
            ne,
            nf,
            fi,
            ea,
            et,
            em,
            json.dumps(margins, ensure_ascii=True),
            brand_color,
            brand_color_secondary,
            brand_color_menu,
            session['user_id'],
        ]
    )
    if ne: session['empresa_nom'] = ne
    session['brand_color'] = brand_color
    session['brand_color_secondary'] = brand_color_secondary
    session['brand_color_menu'] = brand_color_menu
    _sync_private_commercial_settings(margins['frames'], margins['prints'], margins=margins)
    return jsonify({'ok': True, 'margins': margins})

# â"€â"€ Routes: Guardar comanda i historial â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route('/guardar', methods=['POST'])
@login_required
def guardar():
    d = request.json
    comanda_id = d.get('comanda_id')  # if set → UPDATE existing row

    # Snapshot dels marges del professional en el moment de guardar.
    # marge_pro_snap congela el % aplicat perquè futurs canvis a usuaris.marge_pro_pct
    # no alterin la reimpressió de pressupostos antics.
    marge_pro_actiu = get_config_value('marge_pro_actiu', '1') == '1'
    usuari = query('SELECT marge_pro_pct, marge FROM usuaris WHERE id=?', [session['user_id']], one=True)
    if marge_pro_actiu:
        marge_pro_snap = _get_marge_value(usuari)
    else:
        marge_pro_snap = 0.0

    qty = float(d.get('quantitat') or 1) or 1.0
    cost_produccio = float(d.get('cost_produccio') or 0)
    # pvd_unitari: cost per unit amb marge admin aplicat (el que retorna /api/closest com a 'pvd').
    # cost_produccio ja ve amb marge admin aplicat (suma de CLOSEST[k].preu × qty), per tant:
    pvd_unitari = round(cost_produccio / qty, 4) if qty else 0.0
    # cost_unitari: cost workshop sense marge admin. Requereix desglòs per categoria que el
    # payload actual no envia; queda NULL fins que el frontend enviï el detall.
    cost_unitari = None

    # Common field values
    # client_extern_id: opcional. Si arriba, normalitzem a int o None.
    try:
        client_extern_id = int(d.get('client_extern_id') or 0) or None
    except (TypeError, ValueError):
        client_extern_id = None

    vals_comuns = [
        d.get('client_nom',''), d.get('client_tel',''),
        d.get('pre_marc',''), d.get('marc_principal',''),
        d.get('amplada',0), d.get('alcada',0), d.get('copia',0),
        d.get('encolat',''), d.get('vidre',''), d.get('passpartout',''),
        d.get('passpartu_ref',''), d.get('impressio',''),
        1 if d.get('revers_peu') else 0, d.get('revers_peu_preu', 0),
        d.get('tipus_peca','fotografia'), d.get('tipus_peca_detall',''),
        d.get('final_amplada',0), d.get('final_alcada',0),
        d.get('marge',60), d.get('descompte',0), d.get('quantitat',1),
        d.get('preu_net',0), d.get('preu_final',0),
        cost_produccio,
        d.get('entrega',0), d.get('pendent',0),
        d.get('observacions',''), d.get('opcio_nom','Opció A'), d.get('lang','ca'),
        cost_unitari, pvd_unitari, marge_pro_snap,
        client_extern_id,
    ]

    if comanda_id:
        # Verify ownership before updating
        existing = query('SELECT id, num_pressupost, sessio_id FROM comandes WHERE id=? AND user_id=?',
                         [comanda_id, session['user_id']], one=True)
        if existing:
            execute('''UPDATE comandes SET
                client_nom=?, client_tel=?, pre_marc=?, marc_principal=?,
                amplada=?, alcada=?, copia=?,
                encolat=?, vidre=?, passpartout=?, passpartu_ref=?, impressio=?,
                revers_peu=?, revers_peu_preu=?,
                tipus_peca=?, tipus_peca_detall=?, final_amplada=?, final_alcada=?,
                marge=?, descompte=?, quantitat=?,
                preu_net=?, preu_final=?, cost_produccio=?,
                entrega=?, pendent=?, observacions=?, opcio_nom=?, lang=?,
                cost_unitari=?, pvd_unitari=?, marge_pro_snap=?,
                client_extern_id=?
                WHERE id=? AND user_id=?''',
                vals_comuns + [comanda_id, session['user_id']])
            return jsonify({'ok': True, 'id': existing['id'],
                            'sessio_id': existing['sessio_id'],
                            'num': existing['num_pressupost']})
        # If not found (wrong user or deleted), fall through to INSERT

    sessio_id = d.get('sessio_id') or secrets.token_hex(8)
    num_pressupost = generar_num_pressupost()
    cid = execute('''INSERT INTO comandes
        (user_id, data, client_nom, client_tel,
         pre_marc, marc_principal, amplada, alcada, copia,
         encolat, vidre, passpartout, passpartu_ref, impressio,
         revers_peu, revers_peu_preu,
         tipus_peca, tipus_peca_detall, final_amplada, final_alcada,
         marge, descompte, quantitat,
         preu_net, preu_final, cost_produccio, entrega, pendent, observacions,
         sessio_id, opcio_nom, num_pressupost, lang,
         cost_unitari, pvd_unitari, marge_pro_snap, client_extern_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
        [session['user_id'], datetime.now().strftime('%d/%m/%Y %H:%M')] +
        vals_comuns[:27] + [sessio_id] + [vals_comuns[27]] + [num_pressupost] + [vals_comuns[28]] +
        vals_comuns[29:32] + [vals_comuns[32]]
    )
    return jsonify({'ok': True, 'id': cid, 'sessio_id': sessio_id, 'num': num_pressupost})


@app.route('/api/desar-cistella', methods=['POST'])
@login_required
def api_desar_cistella():
    """Desa la cistella multi-producte com una comanda a l'historial: una fila
    per línia, totes sota el mateix sessio_id (queda agrupada com un pressupost).
    Cada línia: {text, quantity, preu_net (PVP), cost_produccio (PVD)}."""
    d = request.get_json(force=True) or {}
    items = d.get('items')
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'La cistella és buida.'}), 400
    client_nom = (d.get('client_nom') or '').strip()
    client_tel = (d.get('client_tel') or '').strip()
    client_extern_id = (d.get('client_extern_id') or '').strip() or None
    observacions = (d.get('observacions') or '').strip()
    lang = (d.get('lang') or 'ca').strip() or 'ca'
    sessio_id = secrets.token_hex(8)
    num_pressupost = generar_num_pressupost()
    data_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    n = 0
    for i, it in enumerate(items):
        if not isinstance(it, dict):
            continue
        text = (str(it.get('text') or 'Producte')).strip()[:300]
        try:
            qty = float(it.get('quantity') or 1) or 1
        except (TypeError, ValueError):
            qty = 1
        try:
            pvp = float(it.get('preu_net') or 0)
        except (TypeError, ValueError):
            pvp = 0.0
        try:
            pvd = float(it.get('cost_produccio') or 0)
        except (TypeError, ValueError):
            pvd = 0.0
        execute(
            '''INSERT INTO comandes
               (user_id, data, client_nom, client_tel, marc_principal, quantitat,
                preu_net, preu_final, cost_produccio, entrega, pendent, observacions,
                sessio_id, opcio_nom, num_pressupost, lang, client_extern_id, tipus_peca)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            [session['user_id'], data_str, client_nom, client_tel, text, qty,
             pvp, round(pvp * 1.21, 2), pvd, 0, round(pvp * 1.21, 2),
             observacions if i == 0 else '',
             sessio_id, f'Línia {i + 1}', num_pressupost, lang, client_extern_id, 'producte'])
        n += 1
    return jsonify({'ok': True, 'sessio_id': sessio_id, 'num': num_pressupost, 'n': n})


@app.route('/sessio/<sessio_id>/pagat', methods=['POST'])
@login_required
def marcar_pagat(sessio_id):
    c = _get_comanda_by_sessio_for_session(sessio_id)
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    pagat = (request.json or {}).get('pagat', 1)
    execute('UPDATE comandes SET pagat=? WHERE sessio_id=?', [pagat, sessio_id])
    return jsonify({'ok': True})

@app.route('/sessio/<sessio_id>/entregat', methods=['POST'])
@login_required
def marcar_entregat(sessio_id):
    c = _get_comanda_by_sessio_for_session(sessio_id)
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    entregat = (request.json or {}).get('entregat', 1)
    execute('UPDATE comandes SET entregat=? WHERE sessio_id=?', [entregat, sessio_id])
    # Sincronitza el cicle d'estats amb el flag d'entrega (sense perdre estat manual).
    if int(entregat or 0) == 1:
        execute("UPDATE comandes SET estat='entregat' WHERE sessio_id=?", [sessio_id])
    return jsonify({'ok': True})


@app.route('/sessio/<sessio_id>/estat', methods=['POST'])
@login_required
def marcar_estat(sessio_id):
    """F2: fixa l'estat del cicle de la comanda (per sessio_id, com pagat/
    entregat). Valida contra el conjunt d'estats. No toca pagat ni el càlcul."""
    c = _get_comanda_by_sessio_for_session(sessio_id)
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    data = request.get_json(silent=True) or request.form or {}
    nou = str(data.get('estat', '') or '').strip().lower()
    if nou not in COMANDA_ESTAT_KEYS:
        return jsonify({'ok': False, 'error': 'estat_invalid'}), 400
    execute('UPDATE comandes SET estat=? WHERE sessio_id=?', [nou, sessio_id])
    # Mantenir coherent el flag d'entrega amb l'estat final.
    if nou == 'entregat':
        execute('UPDATE comandes SET entregat=1 WHERE sessio_id=?', [sessio_id])
    return jsonify({'ok': True, 'estat': nou, 'label': COMANDA_ESTAT_LABELS.get(nou, nou)})

@app.route('/comanda/<int:cid>/liquidar', methods=['POST'])
@login_required
def liquidar_comanda(cid):
    c = _get_comanda_for_session(cid, fields='id, user_id')
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    execute('UPDATE comandes SET entrega=preu_final, pendent=0, pagat=1 WHERE id=?', [cid])
    return jsonify({'ok': True})


@app.route('/comanda/<int:cid>/a-compte', methods=['POST'])
@login_required
def a_compte_comanda(cid):
    """Marca un import rebut A COMPTE (senyal) per a una opció: desa l'entrega
    parcial i recalcula el pendent (= total − entrega). Si cobreix el total,
    marca com a pagat. El botó de WhatsApp de l'historial ja reflecteix
    l'entrega i el pendent (així es confirma la senyal al client)."""
    c = _get_comanda_for_session(cid, fields='id, user_id, preu_final')
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    data = request.get_json(silent=True) or {}
    try:
        entrega = float(data.get('entrega') or 0)
    except (TypeError, ValueError):
        entrega = 0.0
    total = float(_row_get(c, 'preu_final', 0) or 0)
    entrega = round(max(0.0, min(entrega, total)), 2)
    pendent = round(total - entrega, 2)
    pagat = 1 if pendent <= 0.01 else 0
    execute('UPDATE comandes SET entrega=?, pendent=?, pagat=? WHERE id=?',
            [entrega, pendent, pagat, cid])
    return jsonify({'ok': True, 'entrega': entrega, 'pendent': pendent, 'pagat': bool(pagat)})


@app.route('/admin/eliminar-tot', methods=['POST'])
@admin_required
def eliminar_tot():
    uid = request.json.get('user_id')
    sessio_ids = request.json.get('sessio_ids', [])
    if sessio_ids:
        for sid in sessio_ids:
            execute('DELETE FROM comandes WHERE sessio_id=?', [sid])
    elif uid:
        execute('DELETE FROM comandes WHERE user_id=?', [uid])
    return jsonify({'ok': True})

@app.route('/comanda/<int:cid>/eliminar', methods=['POST'])
@login_required
def eliminar_comanda(cid):
    c = _get_comanda_for_session(cid, fields='id, user_id')
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'})
    execute('DELETE FROM comandes WHERE id=?', [cid])
    return jsonify({'ok': True})

@app.route('/comanda/<int:cid>/acceptar', methods=['POST'])
@login_required
def acceptar_comanda(cid):
    c = _get_comanda_for_session(cid, fields='id, user_id')
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    estat = (request.json or {}).get('estat', 'acceptat')
    execute('UPDATE comandes SET observacions = CASE WHEN observacions IS NULL OR observacions=\'\' THEN ? ELSE observacions || \' | \' || ? END WHERE id=?',
            [f'[{estat.upper()}]', f'[{estat.upper()}]', cid])
    return jsonify({'ok': True})


# ── Enviar fotografia al laboratori d'impressió ─────────────────────────
def _format_lab_template(template, comanda, default=''):
    """Substitueix {num_pressupost} {client} {mida} {format} a la plantilla."""
    if not template:
        return default
    num = _row_get(comanda, 'num_pressupost', '') or f"#{_row_get(comanda, 'id', '')}"
    client = _row_get(comanda, 'client_nom', '') or 'sense nom'
    fw = _row_get(comanda, 'final_amplada', 0) or _row_get(comanda, 'amplada', 0) or 0
    fh = _row_get(comanda, 'final_alcada', 0) or _row_get(comanda, 'alcada', 0) or 0
    try:
        mida = f"{int(float(fw))}×{int(float(fh))} cm"
    except (TypeError, ValueError):
        mida = f"{fw}×{fh} cm"
    impr = (_row_get(comanda, 'impressio', '') or '').strip() or '—'
    encolat = (_row_get(comanda, 'encolat', '') or '').strip() or ''
    fmt = impr if impr != '—' else ''
    if encolat and encolat != '-':
        fmt = (fmt + ' / ' + encolat).strip(' /')
    return template.format(
        num_pressupost=num,
        client=client,
        mida=mida,
        format=fmt or '—',
    )


@app.route('/admin/comanda/<int:cid>/lab-send', methods=['GET', 'POST'])
@admin_required
def admin_lab_send(cid):
    """Form per enviar la fotografia d'una comanda al laboratori (Fase 1: només email).
    GET  → mostra form amb pujada de fitxer + previsualització.
    POST → adjunta el fitxer i envia via Gmail SMTP. Registra a lab_sends."""
    comanda = query('SELECT * FROM comandes WHERE id=?', [cid], one=True)
    if not comanda:
        flash('Pressupost no trobat.', 'error')
        return redirect(url_for('historial'))

    cfg = {r['clau']: r['valor'] for r in (query('SELECT * FROM config') or [])}
    historial = query(
        'SELECT id, canal, destinacio, filename, mida_kb, ok, error, link, sent_at FROM lab_sends '
        'WHERE comanda_id=? ORDER BY id DESC',
        [cid],
    ) or []

    if request.method == 'POST':
        canal = (request.form.get('canal') or 'email').strip()
        dest = (request.form.get('destinacio') or '').strip()
        notes = (request.form.get('notes') or '').strip()
        f = request.files.get('foto')

        if canal != 'email':
            flash("Aquest canal encara no està disponible (Fase 2). Tria Email.", 'error')
            return redirect(url_for('admin_lab_send', cid=cid))
        if not f or not f.filename:
            flash("Has d'adjuntar un fitxer.", 'error')
            return redirect(url_for('admin_lab_send', cid=cid))
        if not dest:
            flash('Falta l\'adreça destinatària.', 'error')
            return redirect(url_for('admin_lab_send', cid=cid))

        data = f.read()
        size_kb = max(1, len(data) // 1024)
        filename = f.filename or 'foto.jpg'

        gmail_user = cfg.get('gmail_user','')
        gmail_pass = cfg.get('gmail_pass','')
        if not gmail_user or not gmail_pass:
            flash("Falta configurar gmail_user / gmail_pass a /admin/config (secció Gmail).", 'error')
            return redirect(url_for('admin_lab_send', cid=cid))

        assumpte_t = cfg.get('lab_assumpte_template') or 'Comanda Laboratori — {num_pressupost} · {mida} {format}'
        cos_t = cfg.get('lab_cos_template') or "Bon dia,\n\nAdjunto la fotografia per imprimir corresponent al pressupost {num_pressupost}.\n\nClient: {client}\nMida d'impressió: {mida}\nFormat / acabat: {format}\n\nGràcies!"
        assumpte = _format_lab_template(assumpte_t, comanda, default='Foto per imprimir')
        cos = _format_lab_template(cos_t, comanda, default='Adjunto la fotografia per imprimir.')
        if notes:
            cos += "\n\nNotes addicionals:\n" + notes

        ok, error = 1, None
        try:
            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            from email.mime.base import MIMEBase
            from email import encoders
            msg = MIMEMultipart()
            msg['From'] = gmail_user
            msg['To'] = dest
            msg['Subject'] = assumpte
            msg.attach(MIMEText(cos, 'plain', 'utf-8'))
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=8) as s:
                s.login(gmail_user, gmail_pass)
                s.sendmail(gmail_user, [dest], msg.as_string())
        except Exception as e:
            ok = 0
            error = str(e)[:300]

        execute(
            'INSERT INTO lab_sends (comanda_id, canal, destinacio, filename, mida_kb, ok, error, link, sent_at, user_id) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            [cid, 'email', dest, filename, size_kb, ok, error, None,
             datetime.now().strftime('%Y-%m-%d %H:%M:%S'), session.get('user_id')],
        )

        if ok:
            flash(f'Fitxer enviat a {dest} ({size_kb} kB).', 'ok')
        else:
            flash(f'Error enviant: {error}', 'error')
        return redirect(url_for('admin_lab_send', cid=cid))

    return render_template('admin_lab_send.html', comanda=comanda, config=cfg, historial=historial)


# ── Catàleg de motllures ──────────────────────────────────────────────────
@app.route('/cataleg')
@login_required
def cataleg():
    q = request.args.get('q', '').strip()
    color = request.args.get('color', '').strip().lower()
    gruix = request.args.get('gruix', '').strip().lower()
    sql = """SELECT referencia, gruix, descripcio, foto, ref2
             FROM moldures"""
    args = []
    if q:
        sql += """ WHERE LOWER(referencia) LIKE ?
                   OR LOWER(COALESCE(descripcio, '')) LIKE ?"""
        args = [f'%{q.lower()}%', f'%{q.lower()}%']
    sql += " ORDER BY referencia"
    moldures = _serialize_moldures(query(sql, args))
    moldures = [
        m for m in moldures
        if _matches_moldura_color(m.get('descripcio', ''), color)
        and _matches_moldura_gruix(m.get('gruix', 0), gruix)
    ]
    total = query("SELECT COUNT(*) as n FROM moldures", one=True)
    return render_template('cataleg.html', moldures=moldures, q=q,
                           total=total['n'] if total else 0,
                           color=color, gruix=gruix,
                           color_filters=MOLDURA_COLOR_FILTERS,
                           gruix_filters=MOLDURA_GRUIX_FILTERS)

@app.route('/admin/ensure-clients-externs')
@admin_required
def admin_ensure_clients_externs():
    """Emergència: crea/posa al dia la taula clients_externs + índexs + FK
    a comandes. Pensat per quan run-migrations es queda penjat i el panell
    /admin/clients-externs torna 500.

    Versió defensiva: cada sentència s'executa amb el seu propi commit/
    rollback explícit, sense lock_timeout (que en alguns casos enverina la
    connexió quan una ALTER triga més del previst). Si tot peta, retorna
    igualment HTML amb el traceback per poder diagnosticar."""
    import traceback
    resultats = []
    try:
        db = get_db()
    except Exception:
        return ("<h2>clients_externs</h2><pre>get_db() failed:\n"
                + traceback.format_exc() + "</pre>"), 500

    sentencies = [
        """CREATE TABLE IF NOT EXISTS clients_externs (
            id SERIAL PRIMARY KEY,
            nom VARCHAR(255) NOT NULL,
            nif VARCHAR(30),
            fd_contact_id VARCHAR(100) NOT NULL UNIQUE,
            actiu BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        # Columnes afegides posteriorment a la taula original. El panell
        # /admin/clients-externs fa SELECT d'aquestes columnes, així que la
        # taula creada abans amb només el CREATE TABLE de dalt provoca un
        # error "column does not exist" (HTTP 500). Aquestes ALTERs són
        # IF NOT EXISTS, per tant idempotents.
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS tipus VARCHAR(20) DEFAULT 'pvp'",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS telefon VARCHAR(50)",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS nom_comercial VARCHAR(255)",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS email VARCHAR(255)",
        "ALTER TABLE clients_externs ALTER COLUMN fd_contact_id DROP NOT NULL",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS usuari_id INTEGER",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS dropbox_url TEXT",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS recarrec_equiv BOOLEAN DEFAULT FALSE",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_nom ON clients_externs(nom)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_nif ON clients_externs(nif)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_actiu ON clients_externs(actiu)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_tipus ON clients_externs(tipus)",
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS client_extern_id INTEGER",
        "CREATE INDEX IF NOT EXISTS idx_comandes_client_extern ON comandes(client_extern_id)",
    ]
    for sql in sentencies:
        try:
            execute(sql)
            try:
                db.commit()
            except Exception:
                pass
            resultats.append(f"OK: {sql[:80].strip()}…")
        except Exception as e:
            # Important: fer rollback per sortir de l'estat "transaction
            # aborted" que deixa Postgres després d'una sentència fallida.
            # Sense això, totes les sentències següents també fallarien.
            try:
                db.rollback()
            except Exception:
                pass
            resultats.append(f"SKIP: {sql[:80].strip()}… — {type(e).__name__}: {str(e)[:200]}")
    return "<h2>clients_externs</h2><pre>" + "\n".join(resultats) + "</pre>"


@app.route('/admin/clients-externs-debug')
@admin_required
def admin_clients_externs_debug():
    """Diagnòstic: llista columnes actuals i nombre de files de
    clients_externs. Útil quan /admin/clients-externs encara peta després
    d'executar /admin/ensure-clients-externs."""
    import traceback
    try:
        if USE_PG:
            cols = query("""
                SELECT column_name, data_type, is_nullable, column_default
                FROM information_schema.columns
                WHERE table_name = 'clients_externs'
                ORDER BY ordinal_position
            """) or []
        else:
            cols = query("PRAGMA table_info(clients_externs)") or []
        try:
            count_row = query("SELECT COUNT(*) AS n FROM clients_externs", one=True)
            count = count_row['n'] if count_row else 0
        except Exception as e:
            count = f"(no es pot comptar: {e})"
        lines = [f"Files: {count}", "", "Columnes:"]
        for c in cols:
            lines.append(f"  - {dict(c)}")
        return "<h2>clients_externs (debug)</h2><pre>" + "\n".join(lines) + "</pre>"
    except Exception:
        return ("<h2>clients_externs (debug)</h2><pre>ERROR:\n"
                + traceback.format_exc() + "</pre>"), 500


@app.route('/admin/ensure-pro-clients')
@admin_required
def admin_ensure_pro_clients():
    """Emergència: només crea la taula pro_clients + índexs, sense passar
    pel loop sencer de /admin/run-migrations. Pensat per al moment del
    rollout de la unificació de clients web↔calc."""
    resultats = []
    db = get_db()
    if USE_PG:
        try:
            execute("SET lock_timeout = '3000ms'")
            db.commit()
        except Exception as e:
            try: db.rollback()
            except Exception: pass
            resultats.append(f"SKIP lock_timeout: {str(e)[:120]}")
    sentencies = [
        """CREATE TABLE IF NOT EXISTS pro_clients (
            id SERIAL PRIMARY KEY,
            pro_user_id INTEGER NOT NULL,
            nom VARCHAR(255) NOT NULL,
            empresa VARCHAR(255),
            email VARCHAR(255),
            telefon VARCHAR(50),
            poblacio VARCHAR(120),
            notes TEXT,
            source VARCHAR(50) DEFAULT 'private_area',
            last_order_ref VARCHAR(120),
            order_count INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user ON pro_clients(pro_user_id)",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user_email ON pro_clients(pro_user_id, email)",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user_updated ON pro_clients(pro_user_id, updated_at DESC)",
    ]
    for sql in sentencies:
        try:
            execute(sql)
            resultats.append(f"OK: {sql[:80].strip()}…")
        except Exception as e:
            try: db.rollback()
            except Exception: pass
            resultats.append(f"SKIP: {str(e)[:120]}")
    return "<h2>pro_clients</h2><pre>" + "\n".join(resultats) + "</pre>"


@app.route('/admin/run-migrations')
@admin_required
def admin_run_migrations():
    """TEMPORAL: únic punt d'inicialització/migració de la BD.

    GET sense ?confirm=1 → pàgina de confirmació (ràpida, no toca BD).
    GET amb ?confirm=1   → executa la migració (init_db + seeds + ALTERs +…).

    El gate ?confirm=1 evita que retries automatitzats del navegador o
    Cloudflare disparin la feina pesada en bucle quan la primera petició
    timeoutejà.
    """
    if request.args.get('confirm') != '1':
        return (
            '<h2>⚠️ Migracions manuals</h2>'
            '<p>Aquesta operació pot trigar fins a uns minuts i potencialment '
            'donar timeout (Cloudflare 524). Cal executar-la només manualment, '
            'i NO recarregar la pestanya si triga.</p>'
            '<p><b>Si tot el que vols és inserir els 7 passpartous</b>, '
            'usa <a href="/admin/seed-passpartous">/admin/seed-passpartous</a> '
            '(és instantani).</p>'
            '<p><a href="/admin/run-migrations?confirm=1" '
            'style="display:inline-block;padding:10px 18px;background:#B23A3A;'
            'color:#fff;text-decoration:none;border-radius:6px">'
            'Executar migracions ara</a></p>'
            '<p><a href="/admin">← Tornar a /admin</a></p>'
        )

    resultats = []
    db = get_db()

    # PG: si una ALTER queda esperant un lock (autovacuum, transaccions
    # llargues, etc.) cau ràpid (3s) en lloc de drenar el statement_timeout
    # complet. Sense això, 3 ALTERs blocades poden esgotar el budget de
    # gunicorn i deixar el loop a mig fer (i.e. amb taules noves sense crear).
    if USE_PG:
        try:
            execute("SET lock_timeout = '3000ms'")
            db.commit()
            resultats.append("OK: lock_timeout=3s")
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            resultats.append(f"SKIP lock_timeout: {str(e)[:120]}")

    # 1) init_db() (CREATE TABLE + ALTER + seeds lleugers)
    try:
        init_db()
        resultats.append("OK: init_db()")
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        resultats.append(f"SKIP init_db(): {str(e)[:120]}")

    # Pas A: CREATE TABLE / CREATE INDEX primer. Aquestes operacions són
    # ràpides (taules buides o ja existents amb IF NOT EXISTS) i no requereixen
    # un ACCESS EXCLUSIVE lock sobre taules grans existents, per tant no es
    # queden bloquejades. Les fem ABANS que les ALTERs perquè si una ALTER
    # esgota el budget de gunicorn, almenys les taules noves ja existeixen.
    creates_primers = [
        # Moviments d'stock de marcs (control d'inventari per cm lineals)
        """CREATE TABLE IF NOT EXISTS moviments_stock_marc (
            id SERIAL PRIMARY KEY,
            referencia TEXT NOT NULL,
            data TEXT,
            tipus TEXT NOT NULL,
            cm REAL,
            motiu TEXT,
            albara_num TEXT,
            usuari_id INTEGER,
            stock_resultant REAL
        )""",
        "CREATE INDEX IF NOT EXISTS idx_mov_stock_ref ON moviments_stock_marc(referencia, data)",
        # Historial de preus de cost
        """CREATE TABLE IF NOT EXISTS historial_preus_cost (
            id SERIAL PRIMARY KEY,
            taula VARCHAR(50) NOT NULL,
            referencia VARCHAR(50) NOT NULL,
            preu_cost_antic DECIMAL(8,4) NOT NULL,
            preu_cost_nou DECIMAL(8,4) NOT NULL,
            data TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            usuari_id INTEGER NOT NULL,
            notes TEXT
        )""",
        # Historial d'enviaments a laboratori (impressió fotogràfica)
        """CREATE TABLE IF NOT EXISTS lab_sends (
            id SERIAL PRIMARY KEY,
            comanda_id INTEGER NOT NULL,
            canal VARCHAR(20) NOT NULL,
            destinacio TEXT,
            filename TEXT,
            mida_kb INTEGER,
            ok INTEGER DEFAULT 0,
            error TEXT,
            link TEXT,
            sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            user_id INTEGER
        )""",
        # Audit log (accions d'admin sobre comptes d'usuari)
        """CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            actor_user_id INTEGER NOT NULL,
            actor_username VARCHAR(120),
            target_user_id INTEGER,
            target_username VARCHAR(120),
            action VARCHAR(60) NOT NULL,
            details TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        # Clients externs (no usuaris de la plataforma) per a albarans ràpids.
        # fd_contact_id cached perquè saltem la cerca de FD a cada albarà.
        # tipus: 'taller' (preu taller) o 'pvp' (tarifa PVP).
        """CREATE TABLE IF NOT EXISTS clients_externs (
            id SERIAL PRIMARY KEY,
            nom VARCHAR(255) NOT NULL,
            nif VARCHAR(30),
            fd_contact_id VARCHAR(100) UNIQUE,
            tipus VARCHAR(20) DEFAULT 'pvp',
            telefon VARCHAR(50),
            email VARCHAR(255),
            actiu BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        # ALTER abans dels índexs: la primera vegada que s'executa això sobre
        # una taula vella (creada amb /admin/ensure-clients-externs), la
        # columna `tipus` encara no existeix, així que el seu CREATE INDEX
        # fallaria si anés primer.
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS tipus VARCHAR(20) DEFAULT 'pvp'",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS telefon VARCHAR(50)",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS nom_comercial VARCHAR(255)",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS email VARCHAR(255)",
        "ALTER TABLE clients_externs ALTER COLUMN fd_contact_id DROP NOT NULL",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS usuari_id INTEGER",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS dropbox_url TEXT",
        "ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS recarrec_equiv BOOLEAN DEFAULT FALSE",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_nom ON clients_externs(nom)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_nif ON clients_externs(nif)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_actiu ON clients_externs(actiu)",
        "CREATE INDEX IF NOT EXISTS idx_clients_externs_tipus ON clients_externs(tipus)",
        # FK a comandes (nul·lable: les comandes existents queden a NULL).
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS client_extern_id INTEGER",
        "CREATE INDEX IF NOT EXISTS idx_comandes_client_extern ON comandes(client_extern_id)",
        # Clients privats dels distribuïdors (els seus clients finals).
        # Separats de clients_externs, que és la llista del laboratori per a
        # albarans FD. Cada distribuïdor (pro_user_id) té els seus, no es
        # comparteixen entre distribuïdors.
        """CREATE TABLE IF NOT EXISTS pro_clients (
            id SERIAL PRIMARY KEY,
            pro_user_id INTEGER NOT NULL,
            nom VARCHAR(255) NOT NULL,
            empresa VARCHAR(255),
            email VARCHAR(255),
            telefon VARCHAR(50),
            poblacio VARCHAR(120),
            notes TEXT,
            source VARCHAR(50) DEFAULT 'private_area',
            last_order_ref VARCHAR(120),
            order_count INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user ON pro_clients(pro_user_id)",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user_email ON pro_clients(pro_user_id, email)",
        "CREATE INDEX IF NOT EXISTS idx_pro_clients_user_updated ON pro_clients(pro_user_id, updated_at DESC)",
    ]
    for sql in creates_primers:
        try:
            execute(sql)
            resultats.append(f"OK: {sql[:80].strip()}…")
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            resultats.append(f"SKIP: {str(e)[:120]}")

    alteracions = [
        # Moldures
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS preu_cost DECIMAL(8,4)",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS merma_pct DECIMAL(4,2) DEFAULT 10.00",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS minim_cm DECIMAL(6,1) DEFAULT 100.0",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS preu_cost_ant DECIMAL(8,4)",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS data_cost DATE",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS usuari_cost_id INTEGER",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS notes_cost TEXT",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS cost_verificat INTEGER DEFAULT 0",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS descatalogada BOOLEAN DEFAULT FALSE",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS notes_stock TEXT",
        # Control d'stock de marcs (cm lineals)
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS stock_cm REAL",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS stock_min_cm REAL",
        # Vidres
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS preu_cost DECIMAL(8,4)",
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS preu_cost_ant DECIMAL(8,4)",
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS data_cost DATE",
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS usuari_cost_id INTEGER",
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS notes_cost TEXT",
        "ALTER TABLE vidres ADD COLUMN IF NOT EXISTS cost_verificat INTEGER DEFAULT 0",
        # Encolat
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS preu_cost DECIMAL(8,4)",
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS preu_cost_ant DECIMAL(8,4)",
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS data_cost DATE",
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS usuari_cost_id INTEGER",
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS notes_cost TEXT",
        "ALTER TABLE encolat_pro ADD COLUMN IF NOT EXISTS cost_verificat INTEGER DEFAULT 0",
        # Passpartout
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS preu_cost DECIMAL(8,4)",
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS preu_cost_ant DECIMAL(8,4)",
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS data_cost DATE",
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS usuari_cost_id INTEGER",
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS notes_cost TEXT",
        "ALTER TABLE passpartout ADD COLUMN IF NOT EXISTS cost_verificat INTEGER DEFAULT 0",
        # Comandes
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS cost_unitari DECIMAL(8,4)",
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS pvd_unitari DECIMAL(8,4)",
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS marge_admin_snap DECIMAL(5,2)",
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS marge_pro_snap DECIMAL(5,2)",
        # F2 — cicle d'estats de comanda. Columna + backfill IDEMPOTENT des dels
        # flags existents (entregat / [ACCEPTAT]). No toca pagat (és cobrament).
        # El WHERE estat='' fa que re-executar no sobreescrigui estats ja fixats.
        "ALTER TABLE comandes ADD COLUMN IF NOT EXISTS estat TEXT DEFAULT ''",
        "UPDATE comandes SET estat = CASE WHEN entregat=1 THEN 'entregat' WHEN observacions LIKE '%[ACCEPTAT]%' THEN 'produccio' ELSE 'nou' END WHERE estat IS NULL OR estat=''",
        # Usuaris
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS marge_pro_pct DECIMAL(5,2)",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS marge_impressio_pro_pct DECIMAL(5,2)",
        # Trams de marge per impressions (PVD→PVP segons àrea de la foto)
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram1 REAL",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram2 REAL",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram3 REAL",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram4 REAL",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram5 REAL",
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram6 REAL",
        # NOTA: les CREATE TABLE / CREATE INDEX d'historial_preus_cost, lab_sends,
        # audit_log, clients_externs i les seves dependències (comandes.client_extern_id)
        # ara s'executen a `creates_primers` ABANS d'aquest bloc — així no es
        # bloquegen mai darrere d'ALTERs lentes.

        # Migració de l'antic límit del tram 4 (7500 cm² → 6000 cm²).
        # Només actua si el valor és exactament '7500' per no sobreescriure
        # personalitzacions explícites. Idempotent.
        "UPDATE config SET valor='6000' WHERE clau='imp_tram4_area' AND valor='7500'",
        # Threshold ratio per a la lògica híbrida d'encolat (foam/laminat).
        # Si la fila de taula és més de 1.40× l'àrea sol·licitada, fem fórmula.
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('encolat_ratio_max', '1.40')",
        # Costos €/cm² per a la fórmula d'impressions per àrea (vegeu _imp_closest).
        # Reusa el threshold encolat_ratio_max — mateix patró taula/fórmula.
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_lustre_cost_cm2', '0.000703')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_silk_cost_cm2', '0.000756')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_matte_cost_cm2', '0.000447')",
        # Descomptes per combinació de productes (combo packs).
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('combo_desc_marc_imp_protter', '6')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('combo_desc_marc_imp_foam', '5')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('combo_desc_marc_imp', '3')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('combo_desc_marc_suport', '3')",
        # Mínim de subtotal PVP perquè s'apliqui el descompte combo (€).
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('combo_desc_minim_pvp', '80')",
        # MO de muntatge del doble vidre en € (substitueix l'antic
        # vidre_dv_muntatge_min, que queda inutilitzat).
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('vidre_dv_muntatge_eur', '1.30')",
        # ── Paper Hahnemühle Photo Rag Baryta ────────────────────────
        # Sistema de marge per trams (en lloc de marge admin fix). El
        # helper _imp_closest detecta el flag imp_{paper}_trams_actius i,
        # si està a '1', salta la lògica de taula i aplica el múltiple
        # del tram corresponent. Cost real del paper (HM8152) + tinta.
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_cost_cm2', '0.005351')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_trams_actius', '1')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t1_max', '300')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t1_mult', '9.5')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t2_max', '900')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t2_mult', '6.5')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t3_max', '2000')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t3_mult', '4.3')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t4_max', '4000')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t4_mult', '3.5')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t5_max', '8000')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t5_mult', '3.2')",
        "INSERT OR IGNORE INTO config (clau, valor) VALUES ('imp_baryta_t6_mult', '3.1')",
        # Flag per usuari: cada distribuïdor habilita el Baryta manualment.
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS baryta_actiu BOOLEAN DEFAULT FALSE",
        # Email de contacte (separat de username, que pot ser un alies o email
        # de login). El welcome email s'envia aquí si està informat.
        "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''",
    ]

    resultats = []
    db = get_db()
    for sql in alteracions:
        try:
            execute(sql)
            resultats.append(f"OK: {sql[:80].strip()}…")
        except Exception as e:
            # A PG, una sentència fallida deixa la transacció en estat aborted
            # i tota la resta del loop fallaria. Rollback explícit per recuperar.
            try:
                db.rollback()
            except Exception:
                pass
            resultats.append(f"SKIP: {str(e)[:120]}")

    # Operacions pesades: totes mogudes aquí des de init_db() per no penjar
    # l'arrencada dels workers a PG. Cadascuna amb el seu try/except + rollback.
    for nom, fn in [
        ("_seed_proeco_preus",      lambda: _seed_proeco_preus(db, use_pg=USE_PG)),
        ("_seed_intermol_moldures", lambda: _seed_intermol_moldures(db, use_pg=USE_PG)),
        ("_fix_ref2_errors",        lambda: _fix_ref2_errors(db, use_pg=USE_PG)),
        ("_run_v2_price_backfill",  lambda: _run_v2_price_backfill(db)),
        ("_run_mr_trams_backfill",  lambda: _run_mr_trams_backfill(db)),
        ("_seed_impressio_tarifa_granformat", lambda: _seed_impressio_tarifa_granformat(db, use_pg=USE_PG)),
    ]:
        try:
            fn()
            try:
                db.commit()
            except Exception:
                pass
            resultats.append(f"OK: {nom}")
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            resultats.append(f"SKIP {nom}: {str(e)[:120]}")

    # Catch-up migrations per a DBs desplegades amb schema intermedi.
    # Només s'executen a PG (SQLite no té tipus estrictes i aquestes fallen
    # allà però el SKIP és inofensiu).
    if USE_PG:
        # BOOLEAN → INTEGER: a PG cal DROP DEFAULT, ALTER TYPE amb USING, SET DEFAULT.
        # Si la columna ja és INTEGER les sentències fallen i es registren com a SKIP.
        bool_to_int = []
        for t in ('moldures', 'vidres', 'encolat_pro', 'passpartout'):
            bool_to_int += [
                f"ALTER TABLE {t} ALTER COLUMN cost_verificat DROP DEFAULT",
                f"ALTER TABLE {t} ALTER COLUMN cost_verificat TYPE INTEGER "
                f"USING (CASE WHEN cost_verificat THEN 1 ELSE 0 END)",
                f"ALTER TABLE {t} ALTER COLUMN cost_verificat SET DEFAULT 0",
            ]
        for sql in bool_to_int:
            try:
                execute(sql)
                resultats.append(f"OK: {sql[:80].strip()}…")
            except Exception as e:
                try:
                    db.rollback()
                except Exception:
                    pass
                resultats.append(f"SKIP: {str(e)[:120]}")

        # historial_preus_cost: renoms per alinear noms antics → nous.
        renames = [
            "ALTER TABLE historial_preus_cost RENAME COLUMN taula_origen TO taula",
            "ALTER TABLE historial_preus_cost RENAME COLUMN preu_cost_ant TO preu_cost_antic",
            "ALTER TABLE historial_preus_cost RENAME COLUMN data_canvi TO data",
            "ALTER TABLE historial_preus_cost RENAME COLUMN motiu TO notes",
        ]
        for sql in renames:
            try:
                execute(sql)
                resultats.append(f"OK: {sql[:80].strip()}…")
            except Exception as e:
                try:
                    db.rollback()
                except Exception:
                    pass
                resultats.append(f"SKIP: {str(e)[:120]}")

    return "<br>".join(resultats)


_PASSPARTOUS_SEED = [
    ('P001', 'Blanc Cru',  '', 'Blanc cru'),
    ('P003', 'Negre',      '', 'Negre'),
    ('P008', 'Groc',       '', 'Groc'),
    ('P016', 'Blanc',      '', 'Blanc'),
    ('P076', 'Marron',     '', 'Marró'),
    ('P077', 'Gris Clar',  '', 'Gris clar'),
    ('P078', 'Gris Fosc',  '', 'Gris fosc'),
]


@app.route('/admin/seed-passpartous')
@admin_required
def admin_seed_passpartous():
    """Seed ràpid només dels 7 passpartous de color. Pensat per no fer
    timeout: insereix una a una via cursor amb commit/rollback per fila."""
    db = get_db()
    # Reset connection state in case a prior aborted transaction is around
    try: db.rollback()
    except Exception: pass
    inserted, skipped = [], []
    for ref, color, textura, descripcio in _PASSPARTOUS_SEED:
        try:
            if USE_PG:
                cur = db.cursor()
                cur.execute(
                    "INSERT INTO passpartout (referencia, color, textura, descripcio) "
                    "VALUES (%s, %s, %s, %s) ON CONFLICT (referencia) DO NOTHING",
                    [ref, color, textura, descripcio],
                )
                db.commit()
            else:
                db.execute(
                    "INSERT OR IGNORE INTO passpartout (referencia, color, textura, descripcio) "
                    "VALUES (?, ?, ?, ?)",
                    [ref, color, textura, descripcio],
                )
                db.commit()
            inserted.append(ref)
        except Exception as e:
            try: db.rollback()
            except Exception: pass
            skipped.append(f'{ref}: {str(e)[:120]}')
    # Comptar quantes hi ha realment a la BD
    try:
        rows = query("SELECT referencia, color FROM passpartout WHERE color IS NOT NULL AND color <> '' ORDER BY referencia") or []
    except Exception as e:
        rows = []
    out = ['<h2>Seed passpartous</h2>']
    out.append(f'<p>Inserides/processades: {len(inserted)} de {len(_PASSPARTOUS_SEED)}</p>')
    if skipped:
        out.append('<p><b>Errors:</b><br>' + '<br>'.join(skipped) + '</p>')
    out.append(f'<p><b>Total amb color a la BD ara mateix:</b> {len(rows)}</p>')
    out.append('<ul>' + ''.join(f'<li>{r["referencia"]} — {r["color"]}</li>' for r in rows) + '</ul>')
    out.append('<p><a href="/admin/passpartous">→ Tornar a /admin/passpartous</a></p>')
    return '\n'.join(out)


@app.route('/admin/normalitzar-costos')
@admin_required
def admin_normalitzar_costos():
    """TEMPORAL: actualitza preu_cost de passpartú (1PAS, DOBPAS) i encolat
    foam (ENC) amb costos calculats per fórmula. Útil per re-baselinar la
    taula després de canviar paràmetres de cost (cm², temps, hora)."""
    COST_CM2_PAS  = 0.000620
    TEMPS_PAS     = 9.0       # minuts
    COST_CM2_FOAM = 0.001143
    TEMPS_FOAM    = 9.0       # minuts base
    TEMPS_VAR     = 0.0015    # minuts per cm²
    COST_HORA     = 25.0

    results, errors = [], []

    # ── Passpartú simple (1PAS...) ────────────────────────────
    files = query("SELECT referencia FROM passpartout WHERE UPPER(referencia) LIKE '1PAS%'") or []
    for f in files:
        ref = _row_get(f, 'referencia')
        try:
            parts = ref.upper().replace('1PAS', '').split('X')
            w, h  = float(parts[0]), float(parts[1])
            cost  = round(w * h * COST_CM2_PAS + TEMPS_PAS * COST_HORA / 60, 4)
            execute("UPDATE passpartout SET preu_cost = ? WHERE referencia = ?", (cost, ref))
            results.append(f"{ref}: {cost}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    # ── Passpartú doble = simple × 2 ──────────────────────────
    files = query("SELECT referencia FROM passpartout WHERE UPPER(referencia) LIKE 'DOBPAS%'") or []
    for f in files:
        ref = _row_get(f, 'referencia')
        try:
            ref_simple = ref.upper().replace('DOBPAS', '1PAS')
            simple = query(
                "SELECT preu_cost FROM passpartout WHERE UPPER(referencia) = ?",
                (ref_simple,), one=True
            )
            if simple and _row_get(simple, 'preu_cost') is not None:
                cost = round(float(_row_get(simple, 'preu_cost')) * 2.0, 4)
                execute("UPDATE passpartout SET preu_cost = ? WHERE referencia = ?", (cost, ref))
                results.append(f"{ref}: {cost}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    # ── Encolat foam (ENC...) ─────────────────────────────────
    files = query("SELECT referencia FROM encolat_pro WHERE UPPER(referencia) LIKE 'ENC%'") or []
    for f in files:
        ref = _row_get(f, 'referencia')
        try:
            parts = ref.upper().replace('ENCC', '').replace('ENC', '').split('X')
            w, h  = float(parts[0]), float(parts[1])
            area  = w * h
            temps = TEMPS_FOAM + area * TEMPS_VAR
            cost  = round(area * COST_CM2_FOAM + temps * COST_HORA / 60, 4)
            execute("UPDATE encolat_pro SET preu_cost = ? WHERE referencia = ?", (cost, ref))
            results.append(f"{ref}: {cost}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    return '<br>'.join([
        f"<b>OK: {len(results)} files actualitzades</b>",
        *results,
        f"<b>Errors: {len(errors)}</b>",
        *errors,
    ])


@app.route('/admin/revisar-taules')
@admin_required
def admin_revisar_taules():
    """TEMPORAL: dump JSON de les referències amb preu_cost ordenat
    ascendentment per a passpartú (1PAS%, DOBPAS%), encolat foam (ENC%),
    vidres simples (excloent DV-/MIR-) i doble vidre (DV-%). Útil per
    detectar inconsistències a les taules abans de calibrar marges."""
    def _rows(sql):
        return [
            {'ref': _row_get(r, 'referencia'), 'cost': float(_row_get(r, 'preu_cost') or 0)}
            for r in (query(sql) or [])
        ]

    return jsonify({
        'passpartu': _rows(
            "SELECT referencia, preu_cost FROM passpartout "
            "WHERE UPPER(referencia) LIKE '1PAS%' AND preu_cost IS NOT NULL "
            "ORDER BY preu_cost ASC"
        ),
        'doble_pas': _rows(
            "SELECT referencia, preu_cost FROM passpartout "
            "WHERE UPPER(referencia) LIKE 'DOBPAS%' AND preu_cost IS NOT NULL "
            "ORDER BY preu_cost ASC"
        ),
        'encolat': _rows(
            "SELECT referencia, preu_cost FROM encolat_pro "
            "WHERE UPPER(referencia) LIKE 'ENC%' AND preu_cost IS NOT NULL "
            "ORDER BY preu_cost ASC"
        ),
        'vidre': _rows(
            "SELECT referencia, preu_cost FROM vidres "
            "WHERE UPPER(referencia) NOT LIKE 'DV-%' "
            "AND UPPER(referencia) NOT LIKE 'MIR-%' "
            "AND preu_cost IS NOT NULL "
            "ORDER BY preu_cost ASC"
        ),
        'doble_vidre': _rows(
            "SELECT referencia, preu_cost FROM vidres "
            "WHERE UPPER(referencia) LIKE 'DV-%' AND preu_cost IS NOT NULL "
            "ORDER BY preu_cost ASC"
        ),
    })


@app.route('/admin/normalitzar-vidres')
@admin_required
def admin_normalitzar_vidres():
    """TEMPORAL: recalcula preu_cost i preu de les mides noves de vidre
    usant la fórmula real (cm² × 0.002880 + (3 + 0.5·perim_m)·25/60),
    amb marge fix 1.60. Només toca les mides afegides recentment."""
    MIDES = [
        (65, 65), (65, 80), (65, 90), (65, 100),
        (70, 70), (70, 80), (70, 90), (80, 80),
        (80, 100), (90, 120), (100, 150),
    ]

    results, errors = [], []
    for w, h in MIDES:
        area    = w * h
        perim_m = 2 * (w + h) / 100
        cost    = round(area * 0.002880 + (3 + 0.5 * perim_m) * 25 / 60, 4)
        preu    = round(cost * 1.60, 3)
        ref     = f'{w}x{h}'
        try:
            execute(
                "UPDATE vidres SET preu_cost = ?, preu = ? "
                "WHERE LOWER(referencia) = LOWER(?)",
                (cost, preu, ref),
            )
            results.append(f"{ref}: cost={cost} preu={preu}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    return '<br>'.join([
        f"<b>OK: {len(results)} mides actualitzades</b>",
        *results,
        f"<b>Errors: {len(errors)}</b>",
        *errors,
    ])


@app.route('/admin/normalitzar-vidres-tots')
@admin_required
def admin_normalitzar_vidres_tots():
    """TEMPORAL: recalcula preu_cost i preu de TOTES les mides de vidre
    simple (excloent DV-% i MIR-%) usant la fórmula real."""
    rows = query(
        "SELECT referencia FROM vidres "
        "WHERE UPPER(referencia) NOT LIKE 'DV-%' "
        "AND UPPER(referencia) NOT LIKE 'MIR-%'"
    ) or []

    results, errors = [], []
    for r in rows:
        ref = _row_get(r, 'referencia')
        try:
            parts = ref.upper().split('X')
            w, h  = float(parts[0]), float(parts[1])
            area  = w * h
            perim = 2 * (w + h) / 100
            cost  = round(area * 0.002880 + (3 + 0.5 * perim) * 25 / 60, 4)
            preu  = round(cost * 1.60, 3)
            execute(
                "UPDATE vidres SET preu_cost = ?, preu = ? WHERE referencia = ?",
                (cost, preu, ref),
            )
            results.append(f"{ref}: cost={cost} preu={preu}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    return '<br>'.join([
        f"<b>OK: {len(results)} mides actualitzades</b>",
        *results,
        f"<b>Errors: {len(errors)}</b>",
        *errors,
    ])


@app.route('/admin/normalitzar-doble-vidre')
@admin_required
def admin_normalitzar_doble_vidre():
    """TEMPORAL: recalcula preu_cost i preu de totes les mides DV-:
        cost = vidre_simple_cost · 2 + MO_muntatge
        MO_muntatge = 5 min · 25€/h / 60 = 2.0833 €
        preu = cost · 1.60
    Cada DV-WxH busca la fila simple WxH (sense prefix) i hi calcula a sobre.
    """
    MO_MUNTATGE = round(5 * 25 / 60, 4)
    rows = query("SELECT referencia FROM vidres WHERE UPPER(referencia) LIKE 'DV-%'") or []

    results, errors = [], []
    for r in rows:
        ref = _row_get(r, 'referencia')
        try:
            ref_simple = ref.upper().replace('DV-', '')
            simple = query(
                "SELECT preu_cost FROM vidres WHERE UPPER(referencia) = ?",
                (ref_simple,), one=True,
            )
            if simple and _row_get(simple, 'preu_cost') is not None:
                cost_simple = float(_row_get(simple, 'preu_cost'))
                cost = round(cost_simple * 2 + MO_MUNTATGE, 4)
                preu = round(cost * 1.60, 3)
                execute(
                    "UPDATE vidres SET preu_cost = ?, preu = ? WHERE referencia = ?",
                    (cost, preu, ref),
                )
                results.append(f"{ref}: {cost} € (simple {cost_simple} × 2 + {MO_MUNTATGE})")
            else:
                errors.append(f"{ref}: no s'ha trobat la mida simple {ref_simple}")
        except Exception as e:
            errors.append(f"{ref}: {e}")

    return '<br>'.join([
        f"<b>OK: {len(results)} mides actualitzades</b>",
        *results,
        f"<b>Errors: {len(errors)}</b>",
        *errors,
    ])


@app.route('/admin/seed-imp-trams')
@admin_required
def admin_seed_imp_trams():
    """TEMPORAL: garanteix que les columnes imp_tram1..6 existeixen i
    omple els valors segons els grups acordats (admin/Fotoimag/La Capsa →
    trams alts; pros mitjans; inactius → defaults). Idempotent —
    sobreescriu els valors si ja n'hi havia.

    Cada ALTER i cada UPDATE es fa amb cursor + commit propi i try/except
    aïllats perquè un error a una transacció no avorti la resta."""
    db = get_db()
    schema_actions, errors = [], []

    # 1) ALTERs idempotents per garantir que les columnes hi són
    alter_stmts = [
        ("imp_tram1", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram1 REAL"),
        ("imp_tram2", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram2 REAL"),
        ("imp_tram3", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram3 REAL"),
        ("imp_tram4", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram4 REAL"),
        ("imp_tram5", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram5 REAL"),
        ("imp_tram6", "ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS imp_tram6 REAL"),
    ]
    sqlite_fallback = "ALTER TABLE usuaris ADD COLUMN {col} REAL"
    for col, stmt in alter_stmts:
        try:
            cur = db.cursor()
            cur.execute(stmt)
            db.commit()
            schema_actions.append(f"OK {col}: {stmt}")
        except Exception as e:
            try: db.rollback()
            except Exception: pass
            # Fallback per SQLite (no suporta IF NOT EXISTS a ADD COLUMN)
            try:
                cur = db.cursor()
                cur.execute(sqlite_fallback.format(col=col))
                db.commit()
                schema_actions.append(f"OK {col} (sqlite fallback)")
            except Exception as e2:
                # Probablement la columna ja existeix (cas SQLite quan ja està)
                try: db.rollback()
                except Exception: pass
                schema_actions.append(f"SKIP {col}: {str(e)[:80]} / {str(e2)[:80]}")

    # 2) UPDATEs per grup, cada un amb el seu commit
    GRUPS = [
        ('alt',     [1, 7, 9],                            [140, 130, 120, 100, 90, 80]),
        ('mig',     [3, 8, 13, 17, 19],                   [100, 95, 90, 75, 65, 55]),
        ('default', [4, 5, 6, 10, 11, 12, 14, 16, 18],    [80, 75, 70, 60, 50, 45]),
    ]
    sql_pg = (
        "UPDATE usuaris SET imp_tram1=%s, imp_tram2=%s, imp_tram3=%s, "
        "imp_tram4=%s, imp_tram5=%s, imp_tram6=%s WHERE id=%s"
    )
    sql_sqlite = (
        "UPDATE usuaris SET imp_tram1=?, imp_tram2=?, imp_tram3=?, "
        "imp_tram4=?, imp_tram5=?, imp_tram6=? WHERE id=?"
    )
    use_pg = USE_PG
    results = []
    for nom, ids, valors in GRUPS:
        for uid in ids:
            try:
                cur = db.cursor()
                cur.execute(sql_pg if use_pg else sql_sqlite, valors + [uid])
                db.commit()
                results.append(f"id={uid} grup={nom} → {valors}")
            except Exception as e:
                try: db.rollback()
                except Exception: pass
                errors.append(f"id={uid}: {str(e)[:160]}")

    out = []
    out.append(f"<b>Schema ({len(schema_actions)} accions)</b>")
    out += schema_actions
    out.append(f"<b>UPDATEs OK ({len(results)})</b>")
    out += results
    out.append(f"<b>Errors ({len(errors)})</b>")
    out += errors
    return '<br>'.join(out)


@app.route('/admin/auditoria-moldures')
@admin_required
def admin_auditoria_moldures():
    """TEMPORAL: dump JSON amb dades per a auditoria de marges de moldures.
    Retorna {detall, resum} on:
      - detall: una fila per moldura amb ratio actual i pvd nou (cost·1.60)
      - resum:  agregat per proveïdor (count, ratios mig/min/max)
    Filtra files amb preu_cost o preu_taller nuls/zero per evitar div/0."""
    detall_rows = query("""
        SELECT
            referencia,
            preu_taller,
            preu_cost,
            gruix,
            proveidor,
            ROUND(CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC), 3) AS ratio_actual,
            ROUND((CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC) - 1) * 100, 1) AS marge_actual_pct,
            ROUND(CAST(preu_cost AS NUMERIC) * 1.60, 4) AS pvd_nou,
            ROUND((CAST(preu_cost AS NUMERIC) * 1.60 / CAST(preu_taller AS NUMERIC) - 1) * 100, 1) AS diff_pvd_vs_preu_taller_pct
        FROM moldures
        WHERE preu_cost IS NOT NULL AND preu_cost <> 0
          AND preu_taller IS NOT NULL AND preu_taller <> 0
        ORDER BY proveidor, referencia
    """) or []

    resum_rows = query("""
        SELECT
            proveidor,
            COUNT(*) AS total_refs,
            ROUND(AVG(CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC)), 3) AS ratio_mig,
            ROUND(AVG((CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC) - 1) * 100), 1) AS marge_mig_pct,
            ROUND(MIN((CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC) - 1) * 100), 1) AS marge_min_pct,
            ROUND(MAX((CAST(preu_taller AS NUMERIC) / CAST(preu_cost AS NUMERIC) - 1) * 100), 1) AS marge_max_pct
        FROM moldures
        WHERE preu_cost IS NOT NULL AND preu_cost <> 0
          AND preu_taller IS NOT NULL AND preu_taller <> 0
        GROUP BY proveidor
        ORDER BY proveidor
    """) or []

    def _to_float(v):
        try: return float(v) if v is not None else None
        except (TypeError, ValueError): return None

    detall = [{
        'referencia': _row_get(r, 'referencia'),
        'preu_taller': _to_float(_row_get(r, 'preu_taller')),
        'preu_cost': _to_float(_row_get(r, 'preu_cost')),
        'gruix': _row_get(r, 'gruix'),
        'proveidor': _row_get(r, 'proveidor'),
        'ratio_actual': _to_float(_row_get(r, 'ratio_actual')),
        'marge_actual_pct': _to_float(_row_get(r, 'marge_actual_pct')),
        'pvd_nou': _to_float(_row_get(r, 'pvd_nou')),
        'diff_pvd_vs_preu_taller_pct': _to_float(_row_get(r, 'diff_pvd_vs_preu_taller_pct')),
    } for r in detall_rows]

    resum = [{
        'proveidor': _row_get(r, 'proveidor'),
        'total_refs': _row_get(r, 'total_refs'),
        'ratio_mig': _to_float(_row_get(r, 'ratio_mig')),
        'marge_mig_pct': _to_float(_row_get(r, 'marge_mig_pct')),
        'marge_min_pct': _to_float(_row_get(r, 'marge_min_pct')),
        'marge_max_pct': _to_float(_row_get(r, 'marge_max_pct')),
    } for r in resum_rows]

    return jsonify({'detall': detall, 'resum': resum})


@app.route('/admin/auditoria-marges')
@admin_required
def admin_auditoria_marges():
    """TEMPORAL: anàlisi de marges de moldures, agregat per proveïdor en
    Python (no per SQL — més portable entre PG/SQLite). Retorna
    {detall, resum} on:
      - detall: per moldura, marge actual + pvd nou (cost·1.60)
      - resum:  per proveïdor, total refs + marge mig/min/max"""
    rows = query("""
        SELECT referencia, preu_taller, preu_cost, gruix, proveidor
        FROM moldures
        WHERE preu_cost IS NOT NULL AND preu_cost > 0
          AND preu_taller IS NOT NULL AND preu_taller > 0
        ORDER BY proveidor, referencia
    """) or []

    detall = []
    for r in rows:
        pt = float(_row_get(r, 'preu_taller') or 0)
        pc = float(_row_get(r, 'preu_cost') or 0)
        if pt <= 0 or pc <= 0:
            continue
        detall.append({
            'ref':              _row_get(r, 'referencia'),
            'proveidor':        _row_get(r, 'proveidor'),
            'gruix':            _row_get(r, 'gruix'),
            'preu_taller':      pt,
            'preu_cost':        pc,
            'marge_actual_pct': round((pt / pc - 1) * 100, 1),
            'pvd_nou':          round(pc * 1.60, 4),
            'diff_pct':         round((pc * 1.60 / pt - 1) * 100, 1),
        })

    from collections import defaultdict
    proveidors = defaultdict(list)
    for d in detall:
        proveidors[d['proveidor']].append(d['marge_actual_pct'])

    resum = []
    for prov, marges in sorted(proveidors.items(), key=lambda kv: (kv[0] or '')):
        resum.append({
            'proveidor':  prov,
            'total_refs': len(marges),
            'marge_mig':  round(sum(marges) / len(marges), 1),
            'marge_min':  round(min(marges), 1),
            'marge_max':  round(max(marges), 1),
        })

    return jsonify({'detall': detall, 'resum': resum})


@app.route('/admin/auditoria-vidre-protter')
@admin_required
def admin_auditoria_vidre_protter():
    """TEMPORAL: taula HTML comparant per a un conjunt de mides
    representatives el cost i PVD de vidre vs protter (i els components
    foam/laminat per separat). Mostra també la mà d'obra de fórmula
    per a foam i laminat."""
    MIDES = [(30,40),(40,50),(46,61),(50,70),(58,140),
             (60,80),(70,100),(80,120),(90,150),(100,150)]
    MIDES = sorted(MIDES, key=lambda wh: wh[0]*wh[1])

    # Temps base + variables per a calcular MO de fórmula explícitament
    cost_hora    = float(get_config_value('cost_hora_taller', '25'))
    foam_tb      = float(get_config_value('foam_temps_base_min', '9'))
    foam_tv      = float(get_config_value('foam_temps_var_cm2', '0.0015'))
    lam_tb       = float(get_config_value('laminat_temps_base_min', '12'))
    lam_tv       = float(get_config_value('laminat_temps_var_cm2', '0.0012'))

    def _mo(temps_base, temps_var, area):
        return round(((temps_base + area * temps_var) * cost_hora) / 60, 4)

    def _marge_pct(cost, pvd):
        if not cost or cost <= 0: return None
        return round((float(pvd) / float(cost) - 1) * 100, 1)

    rows = []
    for w, h in MIDES:
        area = w * h
        v = calcular_cost_vidre(w, h) or {}
        p = calcular_cost_protter(w, h, tipus='semibrillo') or {}
        f = calcular_cost_foam(w, h) or {}
        l = calcular_cost_laminat(w, h, tipus='semibrillo') or {}

        v_pvd  = float(v.get('pvd') or 0)
        v_cost = float(v.get('cost') or 0)
        p_pvd  = float(p.get('pvd') or 0)
        p_cost = float(p.get('cost') or 0)
        f_pvd  = float(f.get('pvd') or 0)
        l_pvd  = float(l.get('pvd') or 0)
        diff_eur = round(p_pvd - v_pvd, 2)
        diff_pct = round((p_pvd - v_pvd) / v_pvd * 100, 1) if v_pvd > 0 else None

        rows.append({
            'w': w, 'h': h, 'area': area,
            'v_ref': v.get('ref') or '—', 'v_origen': v.get('origen') or '—',
            'v_cost': v_cost, 'v_pvd': v_pvd, 'v_marge': _marge_pct(v_cost, v_pvd),
            'p_ref': p.get('ref') or '—', 'p_origen': p.get('origen') or '—',
            'p_cost': p_cost, 'p_pvd': p_pvd, 'p_marge': _marge_pct(p_cost, p_pvd),
            'f_pvd': f_pvd, 'f_origen': f.get('origen') or '—',
            'l_pvd': l_pvd, 'l_origen': l.get('origen') or '—',
            'mo_foam': _mo(foam_tb, foam_tv, area),
            'mo_lam':  _mo(lam_tb,  lam_tv,  area),
            'diff_eur': diff_eur, 'diff_pct': diff_pct,
        })

    # Render HTML inline (route temporal)
    def _eur(v): return f'{v:.2f} €' if isinstance(v, (int, float)) and v > 0 else ('0,00 €' if isinstance(v, (int, float)) else '—')
    def _pct(v): return f'{v:.1f} %' if isinstance(v, (int, float)) else '—'
    def _diff(v):
        if not isinstance(v, (int, float)): return '—'
        s = '+' if v > 0 else ''
        return f'{s}{v:.2f} €'

    html = ['<!DOCTYPE html><html><head><meta charset="UTF-8">',
            '<title>Auditoria vidre vs protter</title>',
            '<style>',
            'body{font-family:system-ui,sans-serif;background:#F8F7F4;color:#1C1B18;margin:0;padding:1.5rem 2rem;max-width:1480px}',
            'h1{font-size:24px;margin:0 0 .25rem}',
            'p.muted{color:#6B6860;font-size:13px;margin:0 0 1.5rem}',
            'table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #E5E2DB;border-radius:8px;overflow:hidden;font-size:12px}',
            'th{background:#F5F4F1;text-align:left;padding:8px 8px;font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:#6B6860;border-bottom:1px solid #E5E2DB;white-space:nowrap}',
            'td{padding:8px;border-bottom:1px solid #F3F1EB;font-family:Consolas,monospace}',
            'td.r{text-align:right}',
            'tr:last-child td{border-bottom:none}',
            'tr.protter-cheaper td{background:#E8F3EE}',
            'tr.vidre-cheaper td{background:#FAEAEA}',
            '.tag{display:inline-block;font-size:10px;padding:1px 6px;border-radius:4px;font-weight:700;background:#EEF2FF;color:#4F46E5}',
            '.tag.formula{background:#FDF3E8;color:#C8873A}',
            '</style></head><body>',
            '<h1>Auditoria vidre vs protter</h1>',
            '<p class="muted">Comparativa per a 10 mides representatives. PVD = preu venda al pro (cost · marge admin). Marge = (PVD/cost − 1)·100. MO foam/laminat es calcula sempre amb la fórmula (independentment de si la funció ha agafat taula o fórmula com a cost).</p>',
            '<table><thead><tr>',
            '<th>Mida</th><th class="r">Àrea</th>',
            '<th colspan="3">Vidre</th>',
            '<th colspan="3">Protter (foam+lam)</th>',
            '<th class="r">Diff PVD</th><th class="r">Diff %</th>',
            '<th class="r">Foam PVD</th><th class="r">Lam PVD</th>',
            '<th class="r">MO foam</th><th class="r">MO lam</th>',
            '</tr><tr>',
            '<th></th><th></th>',
            '<th class="r">Cost</th><th class="r">PVD</th><th class="r">Marge</th>',
            '<th class="r">Cost</th><th class="r">PVD</th><th class="r">Marge</th>',
            '<th></th><th></th><th></th><th></th><th></th><th></th>',
            '</tr></thead><tbody>']

    for r in rows:
        diff = r['diff_eur']
        if diff < 0: cls = 'protter-cheaper'
        elif diff > 0: cls = 'vidre-cheaper'
        else: cls = ''
        v_origen_tag = f'<span class="tag {("formula" if r["v_origen"]=="formula" else "")}">{r["v_origen"]}</span>'
        p_origen_tag = f'<span class="tag {("formula" if r["p_origen"]=="formula" else "")}">{r["p_origen"]}</span>'
        html.append(
            f'<tr class="{cls}">'
            f'<td>{r["w"]}×{r["h"]}</td>'
            f'<td class="r">{r["area"]}</td>'
            f'<td class="r">{_eur(r["v_cost"])}</td>'
            f'<td class="r">{_eur(r["v_pvd"])}<br>{v_origen_tag}<br><small style="color:#9E9B94">{r["v_ref"]}</small></td>'
            f'<td class="r">{_pct(r["v_marge"])}</td>'
            f'<td class="r">{_eur(r["p_cost"])}</td>'
            f'<td class="r">{_eur(r["p_pvd"])}<br><small style="color:#9E9B94">f:{r["f_origen"]} l:{r["l_origen"]}</small></td>'
            f'<td class="r">{_pct(r["p_marge"])}</td>'
            f'<td class="r"><strong>{_diff(diff)}</strong></td>'
            f'<td class="r">{_pct(r["diff_pct"])}</td>'
            f'<td class="r">{_eur(r["f_pvd"])}</td>'
            f'<td class="r">{_eur(r["l_pvd"])}</td>'
            f'<td class="r">{_eur(r["mo_foam"])}</td>'
            f'<td class="r">{_eur(r["mo_lam"])}</td>'
            f'</tr>'
        )
    html.append('</tbody></table>')
    html.append('<p class="muted" style="margin-top:1rem">Files verdes: protter més barat que vidre. Files vermelles: vidre més barat. Tags <span class="tag">taula</span>/<span class="tag formula">formula</span> indiquen l\'origen del cost en cada càlcul.</p>')
    html.append('</body></html>')
    return ''.join(html)


@app.route('/admin/auditoria-general')
@admin_required
def admin_auditoria_general():
    """TEMPORAL: auditoria completa del càlcul per a un conjunt de
    mides representatives. Per cada mida calcula vidre, doble vidre,
    passpartú simple/doble, foam, laminat, protter i impressió;
    per a cadascú mostra ref/origen/cost/PVD/marge i destaca anomalies.

    Comprovacions de coherència:
      - DV cost ≈ vidre cost × 2 + 2,08 € (±0,10)
      - DOBPAS cost ≈ 1PAS cost × 2 (±0,05)
      - Protter cost ≈ foam cost + laminat cost (±0,10)
      - Marge admin: dins de 55-65% (avís) o esperat (60%)
      - Origen 'taula' amb àrea_taula/àrea_sol > 1,40 (incoherent
        amb el threshold del PR #126)"""
    MIDES = [(20,25),(30,40),(40,50),(46,61),(50,70),
             (60,80),(70,100),(80,120),(90,150),(100,150)]
    MIDES = sorted(MIDES, key=lambda wh: wh[0]*wh[1])

    DV_OFFSET = 2.0833    # 5 min · 25/60 (vegeu /admin/normalitzar-doble-vidre)
    RATIO_MAX = float(get_config_value('encolat_ratio_max', '1.40'))
    MARGE_OK_LO, MARGE_OK_HI = 55, 65

    usuari_actual = query('SELECT * FROM usuaris WHERE id=?', [session.get('user_id')], one=True)

    def _ratio_taula(ref, w, h):
        try:
            rw, rh = _parse_dims(ref or '')
            if rw and rh and w and h:
                return (rw * rh) / max(1.0, float(w) * float(h))
        except Exception:
            pass
        return None

    def _marge_pct(cost, pvd):
        if not cost or float(cost) <= 0:
            return None
        return (float(pvd) / float(cost) - 1) * 100

    def _comp(d, w, h):
        """Estructura uniforme per a tots els components."""
        if not isinstance(d, dict):
            return None
        cost = float(d.get('cost') or 0)
        pvd  = float(d.get('pvd')  or 0)
        ref  = d.get('ref') or '—'
        origen = d.get('origen') or '—'
        ratio  = _ratio_taula(ref, w, h) if origen == 'taula' else None
        return {
            'ref': ref, 'origen': origen,
            'cost': cost, 'pvd': pvd,
            'marge': _marge_pct(cost, pvd),
            'ratio': ratio,
        }

    rows = []
    for w, h in MIDES:
        area = w * h
        v   = _comp(calcular_cost_vidre(w, h), w, h)
        dv  = _comp(calcular_cost_doble_vidre(w, h), w, h)
        pas = _comp(calcular_cost_passpartu(w, h, tipus='simple'), w, h)
        dpas= _comp(calcular_cost_passpartu(w, h, tipus='doble'), w, h)
        fm  = _comp(calcular_cost_foam(w, h), w, h)
        lm  = _comp(calcular_cost_laminat(w, h, tipus='semibrillo'), w, h)
        prt = _comp(calcular_cost_protter(w, h, tipus='semibrillo'), w, h)

        # Impressió: _imp_closest retorna {ref, preu, origen, area} (sense cost
        # separat). Calculem el tram per saber el % de marge.
        imp_raw = _imp_closest(w, h) or {}
        tram_info = get_marge_impressio_tram(area, usuari_actual) or {}
        imp = {
            'ref': imp_raw.get('ref') or '—',
            'origen': imp_raw.get('origen') or '—',
            'pvd': float(imp_raw.get('preu') or 0),
            'tram': tram_info.get('tram'),
            'marge': float(tram_info.get('marge') or 0),
            'ratio': _ratio_taula(imp_raw.get('ref') or '', w, h) if imp_raw.get('origen') == 'taula' else None,
        }

        # Comprovacions de coherència
        warn = []
        # 1) DV ≈ V·2 + offset
        if v and dv and v['cost'] > 0:
            esperat_dv = v['cost'] * 2 + DV_OFFSET
            if abs(dv['cost'] - esperat_dv) > 0.10:
                warn.append(f"DV cost {dv['cost']:.2f}€ ≠ esperat {esperat_dv:.2f}€ (Δ={dv['cost']-esperat_dv:+.2f})")
        # 2) DOBPAS ≈ 1PAS · 2
        if pas and dpas and pas['cost'] > 0:
            esperat_dpas = pas['cost'] * 2
            if abs(dpas['cost'] - esperat_dpas) > 0.05:
                warn.append(f"DOBPAS cost {dpas['cost']:.2f}€ ≠ 1PAS·2 = {esperat_dpas:.2f}€ (Δ={dpas['cost']-esperat_dpas:+.2f})")
        # 3) Protter ≈ foam + laminat
        if prt and fm and lm:
            esperat_prt = fm['cost'] + lm['cost']
            if abs(prt['cost'] - esperat_prt) > 0.10:
                warn.append(f"Protter cost {prt['cost']:.2f}€ ≠ foam+lam = {esperat_prt:.2f}€ (Δ={prt['cost']-esperat_prt:+.2f})")
        # 4) Marges fora de rang 55-65%
        for lbl, c in [('vidre', v), ('DV', dv), ('1PAS', pas), ('DOBPAS', dpas),
                       ('foam', fm), ('laminat', lm), ('protter', prt)]:
            if c and isinstance(c.get('marge'), float):
                if not (MARGE_OK_LO <= c['marge'] <= MARGE_OK_HI):
                    warn.append(f"Marge {lbl} {c['marge']:.1f}% fora de {MARGE_OK_LO}-{MARGE_OK_HI}%")
        # 5) Origen taula amb ratio > threshold
        for lbl, c in [('vidre', v), ('DV', dv), ('1PAS', pas), ('DOBPAS', dpas),
                       ('foam', fm), ('laminat', lm), ('impressió', imp)]:
            if c and c.get('origen') == 'taula' and c.get('ratio') and c['ratio'] > RATIO_MAX:
                warn.append(f"{lbl} taula ratio {c['ratio']:.2f} > {RATIO_MAX}")

        rows.append({
            'w': w, 'h': h, 'area': area,
            'v': v, 'dv': dv, 'pas': pas, 'dpas': dpas,
            'fm': fm, 'lm': lm, 'prt': prt, 'imp': imp,
            'warn': warn,
        })

    # ---------- Render ----------
    def _eur(v):
        if v is None or not isinstance(v, (int, float)): return '—'
        return f'{v:.2f} €'
    def _pct(v):
        if v is None or not isinstance(v, (int, float)): return '—'
        return f'{v:.1f} %'
    def _origen_tag(c):
        if not c: return '—'
        o = c.get('origen') or '—'
        cls = 'tag-formula' if o == 'formula' else ('tag-taula' if o == 'taula' else 'tag-other')
        ref = c.get('ref') or ''
        return f'<span class="tag {cls}">{o}</span><br><small style="color:#9E9B94">{ref}</small>'

    def _row_color(c):
        """Tornar el color de fons d'una cel·la segons marge/ratio."""
        if not c: return ''
        m = c.get('marge')
        r = c.get('ratio')
        if isinstance(m, (int, float)) and not (MARGE_OK_LO <= m <= MARGE_OK_HI):
            return 'bad'
        if c.get('origen') == 'taula' and isinstance(r, (int, float)) and r > RATIO_MAX:
            return 'warn'
        return 'ok'

    html = ['<!DOCTYPE html><html><head><meta charset="UTF-8">',
            '<title>Auditoria general del càlcul</title>',
            '<style>',
            'body{font-family:system-ui,sans-serif;background:#F8F7F4;color:#1C1B18;margin:0;padding:1.5rem 2rem;max-width:1900px}',
            'h1{font-size:24px;margin:0 0 .25rem}',
            'h3{font-size:14px;margin:1.25rem 0 .35rem;color:#1A6B45}',
            'p.muted{color:#6B6860;font-size:13px;margin:0 0 1.5rem;line-height:1.55}',
            'table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #E5E2DB;border-radius:8px;overflow:hidden;font-size:11px;margin-bottom:1.5rem}',
            'th{background:#F5F4F1;text-align:left;padding:6px 6px;font-weight:700;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:#6B6860;border-bottom:1px solid #E5E2DB;white-space:nowrap}',
            'td{padding:6px;border-bottom:1px solid #F3F1EB;font-family:Consolas,monospace;vertical-align:top}',
            'td.r{text-align:right}',
            'tr:last-child td{border-bottom:none}',
            'td.ok{background:#F4FBF7}',
            'td.warn{background:#FDF3E8}',
            'td.bad{background:#FAEAEA;color:#B84040;font-weight:700}',
            '.tag{display:inline-block;font-size:9px;padding:1px 5px;border-radius:3px;font-weight:700}',
            '.tag-taula{background:#EEF2FF;color:#4F46E5}',
            '.tag-formula{background:#FDF3E8;color:#C8873A}',
            '.tag-other{background:#F1ECE3;color:#6B6860}',
            '.alert{background:#FAEAEA;border:1px solid #E8B4B4;color:#B84040;padding:.4rem .6rem;font-size:11px;border-radius:6px;margin:.4rem 0}',
            '.alert-row{font-size:10px;color:#B84040;line-height:1.4}',
            '</style></head><body>',
            '<h1>Auditoria general del càlcul</h1>',
            f'<p class="muted">Auditoria de {len(MIDES)} mides per a tots els components. ',
            f'Threshold ratio (encolat_ratio_max) = {RATIO_MAX}. Marge esperat = {MARGE_OK_LO}-{MARGE_OK_HI}%. ',
            'Cel·les <strong>verdes</strong>: marge i origen OK. <strong>Groc</strong>: avís (taula amb ratio alt). <strong>Vermell</strong>: error (marge fora de rang).</p>']

    # Capçalera
    html.append('<table><thead><tr>')
    html.append('<th>Mida</th><th class="r">Àrea</th>')
    for lbl in ['Vidre','DV','1PAS','DOBPAS','Foam','Laminat','Protter','Impressió']:
        html.append(f'<th>{lbl}<br>ref · origen</th><th class="r">PVD</th><th class="r">Marge</th>')
    html.append('</tr></thead><tbody>')

    for r in rows:
        cells = []
        cells.append(f'<td>{r["w"]}×{r["h"]}</td>')
        cells.append(f'<td class="r">{r["area"]}</td>')
        for key in ['v','dv','pas','dpas','fm','lm','prt','imp']:
            c = r.get(key)
            cls = _row_color(c)
            cells.append(f'<td>{_origen_tag(c)}</td>')
            cells.append(f'<td class="r {cls}">{_eur(c["pvd"]) if c else "—"}</td>')
            if key == 'imp':
                # Impressió mostra tram + marge
                if c:
                    tram_str = (f'tram {c["tram"]}<br>' if c.get('tram') else '')
                    cells.append(f'<td class="r {cls}">{tram_str}{_pct(c["marge"])}</td>')
                else:
                    cells.append('<td class="r">—</td>')
            else:
                cells.append(f'<td class="r {cls}">{_pct(c["marge"]) if c else "—"}</td>')
        html.append('<tr>' + ''.join(cells) + '</tr>')
        # Files d'avís per a aquesta mida
        if r['warn']:
            warn_cell = '<br>'.join(f'⚠ {msg}' for msg in r['warn'])
            html.append(f'<tr><td colspan="26" class="alert-row">{warn_cell}</td></tr>')

    html.append('</tbody></table>')
    html.append('</body></html>')
    return ''.join(html)


@app.route('/admin/dump-pro')
@admin_required
def admin_dump_pro():
    """TEMPORAL: dump JSON de totes les files PRO* d'encolat_pro
    (laminat Protter), ordenades per referencia."""
    rows = query(
        "SELECT referencia, preu, preu_cost FROM encolat_pro "
        "WHERE UPPER(referencia) LIKE 'PRO%' ORDER BY referencia"
    ) or []
    return jsonify([
        {
            'ref': _row_get(r, 'referencia'),
            'preu': float(_row_get(r, 'preu') or 0),
            'preu_cost': float(_row_get(r, 'preu_cost') or 0)
                if _row_get(r, 'preu_cost') is not None else None,
        }
        for r in rows
    ])


def _au_f(v):
    """Converteix a float; None/buit/invàlid → None."""
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _au_median(vals):
    s = sorted(vals)
    n = len(s)
    if n == 0:
        return None
    mid = n // 2
    return s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2


def _au_outliers(pairs, taula, camp, add, factor=8.0, min_rows=10):
    """Marca valors molt lluny de la mediana (possibles errors de tecleig)."""
    vals = [v for _, v in pairs if v is not None and v > 0]
    if len(vals) < min_rows:
        return
    med = _au_median(vals)
    if not med or med <= 0:
        return
    hi, lo = med * factor, med / factor
    for ref, v in pairs:
        if v is None or v <= 0:
            continue
        if v > hi:
            add('warning', taula, ref, camp,
                f'Valor molt alt ({v:g}) respecte la mediana ({med:g}). Possible error de tecleig.', v)
        elif v < lo:
            add('info', taula, ref, camp,
                f'Valor molt baix ({v:g}) respecte la mediana ({med:g}). Revisa-ho.', v)


def _auditoria_tarifes():
    """Repassa els preus del catàleg (moldures, vidres, passpartout, encolat_pro)
    i retorna (findings, resum). Cada finding: {sev, taula, ref, camp, missatge,
    valor}. sev ∈ {critical, warning, info}. NOMÉS LECTURA: no modifica dades."""
    findings = []

    def add(sev, taula, ref, camp, missatge, valor=None):
        findings.append({'sev': sev, 'taula': taula, 'ref': (ref or '—'),
                         'camp': camp, 'missatge': missatge, 'valor': valor})

    # ── MOLDURES ────────────────────────────────────────────────────
    try:
        mols = [dict(r) for r in (query('SELECT * FROM moldures') or [])]
    except Exception:
        mols = []
    costos_mol = []
    ref2_map = {}
    for m in mols:
        ref = (m.get('referencia') or '').strip()
        descat = bool(m.get('descatalogada'))
        pt = _au_f(m.get('preu_taller'))
        pc = _au_f(m.get('preu_cost'))
        gruix = _au_f(m.get('gruix'))
        merma = _au_f(m.get('merma_pct'))
        minim = _au_f(m.get('minim_cm'))
        sense_pt = pt is None or pt <= 0
        sense_pc = pc is None or pc <= 0
        if sense_pt and sense_pc:
            if not descat:
                add('critical', 'moldures', ref, 'preu',
                    'Sense preu de cap mena (ni preu_taller ni preu_cost).')
        else:
            if sense_pc and not descat:
                add('warning', 'moldures', ref, 'preu_cost',
                    'Sense cost v2 (preu_cost): es calcula amb el preu_taller legacy.')
            if pc is not None and pt is not None and pt > 0 and pc > pt:
                add('warning', 'moldures', ref, 'preu_cost',
                    f'preu_cost ({pc:g}) > preu_taller ({pt:g}): cost per sobre del preu de venda.', pc)
        if (gruix is None or gruix <= 0) and not descat:
            add('warning', 'moldures', ref, 'gruix',
                'Gruix buit o 0: afecta el consum de motllura i el càlcul del preu.')
        if merma is not None and (merma < 0 or merma > 30):
            add('warning', 'moldures', ref, 'merma_pct',
                f'Merma atípica ({merma:g}%). Rang esperat 0–30%.', merma)
        if minim is not None and (minim < 10 or minim > 600):
            add('info', 'moldures', ref, 'minim_cm',
                f'Mínim atípic ({minim:g} cm). Rang esperat 10–600 cm.', minim)
        r2 = (m.get('ref2') or '').strip().lower()
        if r2:
            ref2_map.setdefault(r2, []).append(ref)
        if pc is not None and pc > 0 and not descat:
            costos_mol.append((ref, pc))
    for r2, refs in ref2_map.items():
        if len(refs) > 1:
            add('info', 'moldures', ', '.join(refs), 'ref2',
                f'Referència alternativa duplicada ("{r2}") en {len(refs)} motllures.')
    _au_outliers(costos_mol, 'moldures', 'preu_cost', add)

    # ── VIDRES / PASSPARTOUT / ENCOLAT_PRO ──────────────────────────
    # NOTA: aquestes taules es tarifen PER MIDA (una fila per talla), a
    # diferència de moldures (€/100cm lineal, normalitzat). Per això:
    #   - NO fem outliers per mediana plana: els formats petits sortirien
    #     sempre "baixos" i els grans "alts" (falsos positius).
    #   - A passpartout NOMÉS són files de preu les tarifes 1PAS/DOBPAS.
    #     Les altres files (P0xx) són el catàleg de COLORS i han d'anar sense
    #     preu (no és cap anomalia). Una referència d'una altra família
    #     (PROECO, ENC, MIR…) dins passpartout es marca com a mal ubicada.
    PASS_FOREIGN = ('PROECO', 'ENC', 'MIR', 'DV', 'LAM', 'MOL', 'M')
    for taula in ('vidres', 'passpartout', 'encolat_pro'):
        try:
            rows = [dict(r) for r in (query(f'SELECT * FROM {taula}') or [])]
        except Exception:
            rows = []
        for r in rows:
            ref = (r.get('referencia') or '').strip()
            refU = ref.upper()
            preu = _au_f(r.get('preu'))
            pc = _au_f(r.get('preu_cost'))
            if taula == 'passpartout':
                es_tarifa = refU.startswith('1PAS') or refU.startswith('DOBPAS')
                if not es_tarifa:
                    # Referència d'una altra família mal ubicada a passpartús.
                    if refU.startswith(PASS_FOREIGN):
                        add('warning', taula, ref, 'referencia',
                            'Referència d\'una altra família dins la taula de passpartús: sembla mal ubicada.')
                    # Els colors (P0xx) no tenen preu — és correcte, no avisem.
                    continue
            sense_preu = preu is None or preu <= 0
            sense_pc = pc is None or pc <= 0
            if sense_preu and sense_pc:
                add('critical', taula, ref, 'preu', 'Sense preu (ni preu ni preu_cost).')
            else:
                if sense_pc:
                    add('warning', taula, ref, 'preu_cost',
                        'Sense preu_cost (model v2): es fa servir el preu legacy.')
                if pc is not None and preu is not None and preu > 0 and pc > preu:
                    add('warning', taula, ref, 'preu_cost',
                        f'preu_cost ({pc:g}) > preu ({preu:g}): cost per sobre del preu de venda.', pc)

    ordre = {'critical': 0, 'warning': 1, 'info': 2}
    findings.sort(key=lambda f: (ordre.get(f['sev'], 9), f['taula'], str(f['ref'])))
    resum = {
        'total': len(findings),
        'critical': sum(1 for f in findings if f['sev'] == 'critical'),
        'warning': sum(1 for f in findings if f['sev'] == 'warning'),
        'info': sum(1 for f in findings if f['sev'] == 'info'),
        'moldures': len(mols),
    }
    return findings, resum


@app.route('/admin/auditoria-tarifes')
@admin_required
def admin_auditoria_tarifes():
    """Repàs de qualitat de dades del catàleg de preus: buits, incoherències
    preu↔cost, gruix/merma atípics, duplicats i valors fora de rang."""
    findings, resum = _auditoria_tarifes()
    return render_template('admin_auditoria_tarifes.html', findings=findings, resum=resum)


@app.route('/admin/auditoria-preus')
@admin_required
def admin_auditoria_preus():
    """TEMPORAL: compara els preus que retorna la calc per impressió
    (via _imp_closest + get_marge_impressio_tram) amb el que faria
    reusrevela-web (replicant _resolve_print_margin_from_trams sobre
    les dades que rep via /api/public/professional-summary). Si tots
    dos camins donen el mateix número, la cadena de dades és coherent."""
    MIDES = [
        (13, 18), (20, 30), (30, 40), (40, 50), (50, 70),
        (60, 80), (70, 100), (80, 120), (90, 150), (100, 200),
    ]
    USUARIS = [
        (1, 'admin (alt)'),
        (3, 'Foto Focus (mig)'),
    ]

    def _summary_data_for(usuari):
        """Replica el JSON que retornaria /api/public/professional-summary
        per a aquest usuari (només els camps que el web fa servir)."""
        return {
            'imp_tram1': _row_get(usuari, 'imp_tram1'),
            'imp_tram2': _row_get(usuari, 'imp_tram2'),
            'imp_tram3': _row_get(usuari, 'imp_tram3'),
            'imp_tram4': _row_get(usuari, 'imp_tram4'),
            'imp_tram5': _row_get(usuari, 'imp_tram5'),
            'imp_tram6': _row_get(usuari, 'imp_tram6'),
            'imp_tram_limits': [
                float(get_config_value('imp_tram1_area', '900')),
                float(get_config_value('imp_tram2_area', '2000')),
                float(get_config_value('imp_tram3_area', '4200')),
                float(get_config_value('imp_tram4_area', '6000')),
                float(get_config_value('imp_tram5_area', '14400')),
            ],
            'imp_tram_defaults': [
                float(get_config_value('imp_tram1_marge_default', '80')),
                float(get_config_value('imp_tram2_marge_default', '75')),
                float(get_config_value('imp_tram3_marge_default', '70')),
                float(get_config_value('imp_tram4_marge_default', '60')),
                float(get_config_value('imp_tram5_marge_default', '50')),
                float(get_config_value('imp_tram6_marge_default', '45')),
            ],
        }

    def _simulate_web_resolution(width, height, summary):
        """Replica _resolve_print_margin_from_trams de reusrevela-web."""
        raw_limits = summary.get('imp_tram_limits')
        if isinstance(raw_limits, list) and len(raw_limits) >= 5:
            try:
                limits = [float(v) for v in raw_limits[:5]] + [float('inf')]
            except (TypeError, ValueError):
                limits = [900, 2000, 4200, 6000, 14400, float('inf')]
        else:
            limits = [900, 2000, 4200, 6000, 14400, float('inf')]
        raw_defaults = summary.get('imp_tram_defaults')
        if isinstance(raw_defaults, list) and len(raw_defaults) >= 6:
            try:
                defaults = [float(v) for v in raw_defaults[:6]]
            except (TypeError, ValueError):
                defaults = [80, 75, 70, 60, 50, 45]
        else:
            defaults = [80, 75, 70, 60, 50, 45]
        area = max(0.0, float(width or 0)) * max(0.0, float(height or 0))
        for i, lim in enumerate(limits):
            if area <= lim:
                v = summary.get(f'imp_tram{i + 1}')
                if v is not None:
                    try: return {'marge': float(v), 'tram': i + 1, 'source': 'user'}
                    except (TypeError, ValueError): pass
                return {'marge': float(defaults[i]), 'tram': i + 1, 'source': 'default'}
        return {'marge': float(defaults[-1]), 'tram': 6, 'source': 'default'}

    seccions = []
    for uid, label in USUARIS:
        usuari = query('SELECT * FROM usuaris WHERE id=?', [uid], one=True)
        if not usuari:
            seccions.append({'label': f'{label} — id={uid}', 'rows': [], 'error': 'usuari no trobat'})
            continue
        summary = _summary_data_for(usuari)
        files = []
        for w, h in MIDES:
            r = _imp_closest(w, h)
            if not r:
                files.append({'w': w, 'h': h, 'area': w * h, 'ref': '—', 'pvd': 0, 'tram_calc': None,
                              'marge_calc': None, 'pvp_calc': None, 'tram_web': None,
                              'marge_web': None, 'pvp_web': None, 'match': None,
                              'error': 'sense match a impressio'})
                continue
            pvd = float(r.get('preu') or 0)

            tram_calc = get_marge_impressio_tram(w * h, usuari)
            pvp_calc = round(pvd * (1 + tram_calc['marge'] / 100), 4)

            tram_web = _simulate_web_resolution(w, h, summary)
            pvp_web = round(pvd * (1 + tram_web['marge'] / 100), 4)

            match = abs(pvp_calc - pvp_web) <= 0.01
            files.append({
                'w': w, 'h': h, 'area': w * h, 'ref': r['ref'], 'pvd': pvd,
                'tram_calc': tram_calc['tram'], 'marge_calc': tram_calc['marge'],
                'pvp_calc': pvp_calc,
                'tram_web': tram_web['tram'], 'marge_web': tram_web['marge'],
                'pvp_web': pvp_web, 'match': match, 'error': None,
            })
        seccions.append({'label': f'{label} — id={uid}', 'rows': files, 'error': None})

    # HTML inline (route temporal — no val la pena un template separat)
    html = ['<!DOCTYPE html><html><head><meta charset="UTF-8">',
            '<title>Auditoria preus impressió</title>',
            '<style>',
            'body{font-family:system-ui,sans-serif;background:#F8F7F4;color:#1C1B18;margin:0;padding:1.5rem 2rem;max-width:1240px}',
            'h1{font-size:24px;margin:0 0 .25rem}',
            'h2{font-size:16px;margin:1.5rem 0 .5rem;color:#1A6B45}',
            'p.muted{color:#6B6860;font-size:13px;margin:0 0 1.5rem}',
            'table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #E5E2DB;border-radius:8px;overflow:hidden;font-size:12px}',
            'th{background:#F5F4F1;text-align:left;padding:8px 10px;font-weight:700;font-size:11px;text-transform:uppercase;color:#6B6860;border-bottom:1px solid #E5E2DB}',
            'td{padding:8px 10px;border-bottom:1px solid #F3F1EB}',
            'td.r{text-align:right;font-family:Consolas,monospace}',
            'tr:last-child td{border-bottom:none}',
            'tr.match-no td{background:#FAEAEA}',
            '.tag{display:inline-block;font-size:11px;padding:2px 7px;border-radius:4px;font-weight:700}',
            '.ok{background:#E8F3EE;color:#1A6B45}',
            '.bad{background:#FAEAEA;color:#B84040}',
            '.err{background:#FDF3E8;color:#C8873A}',
            '</style></head><body>',
            '<h1>Auditoria preus impressió</h1>',
            '<p class="muted">Compara <code>/api/closest</code> (camí directe) vs reusrevela-web (camí simulat sobre <code>/api/public/professional-summary</code>). Si tots dos camins donen el mateix PVP, la cadena de dades és coherent.</p>']
    for s in seccions:
        html.append(f'<h2>{s["label"]}</h2>')
        if s['error']:
            html.append(f'<p class="muted">⚠ {s["error"]}</p>')
            continue
        html.append('<table><thead><tr>'
                    '<th>Mida</th><th class="r">Àrea</th><th>Ref</th>'
                    '<th class="r">PVD</th>'
                    '<th class="r">Tram calc</th><th class="r">Marge calc</th><th class="r">PVP calc</th>'
                    '<th class="r">Tram web</th><th class="r">Marge web</th><th class="r">PVP web</th>'
                    '<th>Match</th></tr></thead><tbody>')

        def _eur(v):
            return f'{v:.4f} €' if isinstance(v, (int, float)) else '—'

        def _pct(v):
            return f'{v:.1f} %' if isinstance(v, (int, float)) else '—'

        def _num(v):
            return str(v) if v is not None else '—'

        for row in s['rows']:
            if row.get('error'):
                cls = 'match-no'
                tag = f'<span class="tag err">{row["error"]}</span>'
            elif row['match']:
                cls = 'match-yes'
                tag = '<span class="tag ok">✓ OK</span>'
            else:
                cls = 'match-no'
                diff = (row['pvp_calc'] or 0) - (row['pvp_web'] or 0)
                tag = f'<span class="tag bad">⚠ Δ{diff:+.4f}€</span>'

            cells = [
                f'{row["w"]}×{row["h"]}',
                f'<span class="r">{row["area"]}</span>',
                row['ref'],
                f'<span class="r">{_eur(row["pvd"])}</span>',
                f'<span class="r">{_num(row["tram_calc"])}</span>',
                f'<span class="r">{_pct(row["marge_calc"])}</span>',
                f'<span class="r">{_eur(row["pvp_calc"])}</span>',
                f'<span class="r">{_num(row["tram_web"])}</span>',
                f'<span class="r">{_pct(row["marge_web"])}</span>',
                f'<span class="r">{_eur(row["pvp_web"])}</span>',
                tag,
            ]
            html.append(f'<tr class="{cls}">' + ''.join(f'<td>{c}</td>' for c in cells) + '</tr>')
        html.append('</tbody></table>')
    html.append('</body></html>')
    return ''.join(html)


@app.route('/admin/marcar-descatalogades')
@admin_required
def admin_marcar_descatalogades():
    """TEMPORAL: afegeix les columnes 'descatalogada' i 'notes_stock' a
    moldures (idempotent), i marca com a descatalogades les refs dels
    proveïdors INTERBAJA / INTRERBAJA."""
    db = get_db()
    accions, errors = [], []

    # 1) Garantir columnes (idempotent — IF NOT EXISTS és PG; SQLite cau a try/except)
    schema_stmts = [
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS descatalogada BOOLEAN DEFAULT FALSE",
        "ALTER TABLE moldures ADD COLUMN IF NOT EXISTS notes_stock TEXT",
    ]
    sqlite_fallback = [
        "ALTER TABLE moldures ADD COLUMN descatalogada INTEGER DEFAULT 0",
        "ALTER TABLE moldures ADD COLUMN notes_stock TEXT",
    ]
    for i, stmt in enumerate(schema_stmts):
        try:
            cur = db.cursor()
            cur.execute(stmt)
            db.commit()
            accions.append(stmt)
        except Exception as e:
            try: db.rollback()
            except Exception: pass
            try:
                cur = db.cursor()
                cur.execute(sqlite_fallback[i])
                db.commit()
                accions.append(sqlite_fallback[i] + ' (sqlite)')
            except Exception as e2:
                errors.append(f"{stmt}: {str(e)[:120]} / fallback: {str(e2)[:120]}")

    # 2) Marcar descatalogades
    nota = "Motllura descatalogada d'Intermolduras. Només stock residual, no reposar."
    try:
        execute(
            "UPDATE moldures SET descatalogada = ?, notes_stock = ? "
            "WHERE proveidor IN ('INTERBAJA', 'INTRERBAJA')",
            (True, nota),
        )
        rows = query(
            "SELECT referencia, proveidor FROM moldures "
            "WHERE proveidor IN ('INTERBAJA', 'INTRERBAJA') "
            "ORDER BY proveidor, referencia"
        ) or []
        accions.append(f"UPDATE descatalogada=TRUE: {len(rows)} files")
        for r in rows:
            accions.append(f"  · {_row_get(r, 'proveidor')} / {_row_get(r, 'referencia')}")
    except Exception as e:
        errors.append(f"UPDATE: {str(e)[:160]}")

    out = [f"<b>Accions ({len(accions)})</b>"] + accions
    if errors:
        out += [f"<b>Errors ({len(errors)})</b>"] + errors
    return '<br>'.join(out)


@app.route('/admin/cataleg')
@admin_required
def admin_cataleg():
    q = request.args.get('q', '').strip()
    proveidor = request.args.get('proveidor', '').strip()
    moldures = _serialize_moldures(_query_moldures(q=q, proveidor=proveidor))
    proveidors = query("SELECT DISTINCT proveidor FROM moldures WHERE proveidor!='' ORDER BY proveidor")
    total = query("SELECT COUNT(*) as n FROM moldures", one=True)
    pdf_params = {}
    if q:
        pdf_params['q'] = q
    if proveidor:
        pdf_params['proveidor'] = proveidor
    pdf_url = url_for('admin_cataleg_pdf')
    if pdf_params:
        pdf_url += '?' + urlencode(pdf_params)
    return render_template('admin_cataleg.html', moldures=moldures, q=q,
                           proveidor=proveidor, proveidors=proveidors,
                           total=total['n'] if total else 0,
                           pdf_url=pdf_url)


@app.route('/admin/cataleg/pdf')
@admin_required
def admin_cataleg_pdf():
    q = request.args.get('q', '').strip()
    proveidor = request.args.get('proveidor', '').strip()
    moldures = _serialize_moldures(_query_moldures(q=q, proveidor=proveidor))
    pdf = crear_pdf_cataleg_admin(moldures, q=q, proveidor=proveidor)
    return send_file(pdf, mimetype='application/pdf', as_attachment=False,
                     download_name='cataleg-motllures-admin.pdf')

@app.route('/admin/cataleg/nou', methods=['GET','POST'])
@admin_required
def admin_moldura_nova():
    if request.method == 'POST':
        d = request.form
        existing = query('SELECT referencia FROM moldures WHERE referencia=?', [d['referencia']], one=True)
        if existing:
            return render_template('admin_moldura_form.html', error='Ja existeix una motllura amb aquesta referència.', moldura=d, nova=True)
        foto = d.get('foto', '').strip()
        try:
            uploaded_photo = _save_moldura_photo(request.files.get('foto_fitxer'), d['referencia'])
        except ValueError as e:
            return render_template('admin_moldura_form.html', error=str(e), moldura=d, nova=True)
        if uploaded_photo:
            foto = uploaded_photo
        execute("""INSERT INTO moldures (referencia,preu_taller,gruix,cost,proveidor,ref2,ubicacio,descripcio,foto,stock_min_cm)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                [d['referencia'], float(d.get('preu_taller',0)),
                 float(d.get('gruix',0)), float(d.get('cost',0)),
                 d.get('proveidor',''), d.get('ref2',''),
                 d.get('ubicacio',''), d.get('descripcio',''), foto,
                 _m_a_cm(d.get('stock_min_m'))])
        return redirect(url_for('admin_cataleg'))
    return render_template('admin_moldura_form.html', moldura={}, nova=True, error=None)

@app.route('/admin/cataleg/<ref>/editar', methods=['GET','POST'])
@admin_required
def admin_moldura_editar(ref):
    moldura = query('SELECT * FROM moldures WHERE referencia=?', [ref], one=True)
    if not moldura:
        return redirect(url_for('admin_cataleg'))
    if request.method == 'POST':
        d = request.form
        foto = d.get('foto', '').strip()
        try:
            uploaded_photo = _save_moldura_photo(request.files.get('foto_fitxer'), ref)
        except ValueError as e:
            moldura_data = dict(moldura)
            moldura_data.update(d)
            moldura_data['foto'] = foto or _resolve_moldura_photo(ref, _row_get(moldura, 'foto', ''))
            return render_template('admin_moldura_form.html', error=str(e), moldura=moldura_data, nova=False)
        if uploaded_photo:
            foto = uploaded_photo
        execute("""UPDATE moldures SET preu_taller=?,gruix=?,cost=?,proveidor=?,
                   ref2=?,ubicacio=?,descripcio=?,foto=?,stock_min_cm=? WHERE referencia=?""",
                [float(d.get('preu_taller',0)), float(d.get('gruix',0)),
                 float(d.get('cost',0)), d.get('proveidor',''),
                 d.get('ref2',''), d.get('ubicacio',''),
                 d.get('descripcio',''), foto, _m_a_cm(d.get('stock_min_m')), ref])
        return redirect(url_for('admin_cataleg'))
    return render_template('admin_moldura_form.html', moldura=_serialize_moldura(moldura), nova=False, error=None)

@app.route('/admin/cataleg/<ref>/stock', methods=['GET','POST'])
@admin_required
def admin_moldura_stock(ref):
    """Gestió d'stock d'una motllura: activar control, registrar compres
    (entrada), ajustar el valor i veure l'historial de moviments."""
    moldura = query('SELECT * FROM moldures WHERE referencia=?', [ref], one=True)
    if not moldura:
        return redirect(url_for('admin_cataleg'))
    error = None
    if request.method == 'POST':
        accio = (request.form.get('accio') or '').strip()
        motiu = (request.form.get('motiu') or '').strip()
        cm = _m_a_cm(request.form.get('valor_m'))
        if accio == 'desactivar':
            execute('UPDATE moldures SET stock_cm=NULL WHERE referencia=?', [ref])
            return redirect(url_for('admin_moldura_stock', ref=ref))
        if cm is None or cm < 0:
            error = 'Introdueix un valor en metres vàlid.'
        elif accio in ('entrada', 'ajust'):
            res = _aplica_moviment_stock(
                ref, cm, accio, motiu or ('Compra' if accio == 'entrada' else 'Ajust manual'),
                usuari_id=session.get('user_id'), activar=True)
            if res is None:
                error = 'No s\'ha pogut aplicar el moviment.'
            else:
                return redirect(url_for('admin_moldura_stock', ref=ref))
        else:
            error = 'Acció no vàlida.'
    moviments = query(
        'SELECT * FROM moviments_stock_marc WHERE LOWER(referencia)=LOWER(?) ORDER BY id DESC LIMIT 200',
        [ref]) or []
    return render_template('admin_moldura_stock.html',
                           moldura=_serialize_moldura(moldura),
                           moviments=moviments, error=error)

@app.route('/admin/cataleg/<ref>/eliminar', methods=['POST'])
@admin_required
def admin_moldura_eliminar(ref):
    execute('DELETE FROM moldures WHERE referencia=?', [ref])
    return jsonify({'ok': True})

@app.route('/admin/cataleg/<ref>/toggle-actiu', methods=['POST'])
@admin_required
def admin_moldura_toggle(ref):
    m = query('SELECT actiu FROM moldures WHERE referencia=?', [ref], one=True)
    nou = 0 if (m and _row_get(m, 'actiu', 1)) else 1
    try:
        execute('UPDATE moldures SET actiu=? WHERE referencia=?', [nou, ref])
    except:
        pass
    return jsonify({'ok': True, 'actiu': nou})


@app.route('/admin/moldura/descatalogada', methods=['POST'])
@admin_required
def admin_moldura_descatalogada():
    """Toggle del flag 'descatalogada'. Accepta 'ref' i opcionalment 'descatalogada'
    (true/false) per fixar un valor concret. Si no s'envia, fa toggle del valor actual.
    Retorna {ok, ref, descatalogada, notes_stock}."""
    payload = request.get_json(silent=True) or request.form
    ref = (payload.get('ref') or '').strip()
    if not ref:
        return jsonify({'ok': False, 'error': 'missing_ref'}), 400

    m = query(
        'SELECT descatalogada, notes_stock FROM moldures WHERE LOWER(referencia)=LOWER(?)',
        [ref], one=True,
    )
    if not m:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    actual = bool(_row_get(m, 'descatalogada', False))
    if 'descatalogada' in payload:
        raw = payload.get('descatalogada')
        nou = str(raw).lower() in ('1', 'true', 'yes', 'on') if not isinstance(raw, bool) else raw
    else:
        nou = not actual

    try:
        execute(
            'UPDATE moldures SET descatalogada = ? WHERE LOWER(referencia)=LOWER(?)',
            [bool(nou), ref],
        )
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:160]}), 500

    return jsonify({
        'ok': True,
        'ref': ref,
        'descatalogada': bool(nou),
        'notes_stock': _row_get(m, 'notes_stock', '') or '',
    })

# â"€â"€ API: buscar moldura per ref exacte (autocomplete) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route('/admin/cataleg/api/cerca')
@admin_required
def api_cerca_moldura():
    q = request.args.get('q','').strip()
    if not q: return jsonify([])
    rows = query("""SELECT referencia, preu_taller, gruix, descripcio, proveidor,
                           descatalogada, notes_stock
                    FROM moldures WHERE LOWER(referencia) LIKE ? OR LOWER(descripcio) LIKE ?
                    ORDER BY referencia LIMIT 20""",
                 [f'%{q.lower()}%', f'%{q.lower()}%'])
    return jsonify([dict(r) for r in rows])

@app.route('/admin/taules-preus')
@admin_required
def admin_taules_preus():
    """Mostra totes les taules de preus i simula lookups per mides personalitzades."""
    w = request.args.get('w', 0, type=float)
    h = request.args.get('h', 0, type=float)

    vidres_rows   = [dict(r) for r in query('SELECT referencia, preu FROM vidres ORDER BY referencia')   or []]
    encolat_rows  = [dict(r) for r in query('SELECT referencia, preu FROM encolat_pro ORDER BY referencia') or []]

    lookup = {}
    if w > 0 and h > 0:
        def sim(rows, prefix=None):
            r = _find_closest(rows, w, h, prefix=prefix)
            return {'ref': r['referencia'], 'preu': r['preu']} if r else None
        lookup = {
            'vidre':       sim([r for r in vidres_rows if not r['referencia'].upper().startswith(('DV-','MIR-'))]),
            'doble_vidre': sim(vidres_rows, prefix='DV-'),
            'mirall':      sim(vidres_rows, prefix='MIR-'),
            'encolat':     sim(encolat_rows, prefix='ENC'),
            'protter':     sim(encolat_rows, prefix='PRO'),
        }

    def rows_html(rows, highlight_ref=None):
        lines = ['<table border="1" cellpadding="4" cellspacing="0" style="border-collapse:collapse;font-size:13px">',
                 '<tr><th>Referència</th><th>Preu (€)</th></tr>']
        for r in rows:
            bg = ' style="background:#ffe082"' if highlight_ref and r['referencia'] == highlight_ref else ''
            lines.append(f'<tr{bg}><td>{r["referencia"]}</td><td>{r["preu"]}</td></tr>')
        lines.append('</table>')
        return '\n'.join(lines)

    vid_hl   = lookup.get('vidre',   {}).get('ref') if lookup else None
    enc_hl   = lookup.get('encolat', {}).get('ref') if lookup else None
    pro_hl   = lookup.get('protter', {}).get('ref') if lookup else None

    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Taules de preus — Admin</title>
<style>body{{font-family:Arial,sans-serif;padding:2rem;background:#f5f6fa}}
h2{{margin-top:2rem}}
.lookup{{background:#e8f5e9;border:1px solid #a5d6a7;padding:1rem;border-radius:8px;margin-bottom:1.5rem;font-size:14px}}
.lookup b{{color:#1b5e20}}</style></head><body>
<h1>Taules de preus</h1>
<form method="get">
  Mida marc: <input name="w" value="{w or ''}" placeholder="Amplada" style="width:70px"> ×
             <input name="h" value="{h or ''}" placeholder="Alçada"  style="width:70px"> cm
  <button type="submit">Simular lookup</button>
</form>
"""
    if lookup:
        def lrow(k, label):
            v = lookup.get(k)
            return f'<li><b>{label}:</b> {v["ref"]} → {v["preu"]}€</li>' if v else f'<li><b>{label}:</b> no trobat</li>'
        html += f'<div class="lookup"><b>Lookup per {w}×{h} cm:</b><ul>'
        html += lrow('vidre',       'Vidre')
        html += lrow('doble_vidre', 'Doble vidre')
        html += lrow('mirall',      'Mirall')
        html += lrow('encolat',     'Encolat')
        html += lrow('protter',     'Protter')
        html += '</ul></div>'

    html += f'<h2>Vidres ({len(vidres_rows)} files)</h2>' + rows_html(vidres_rows, vid_hl)
    html += f'<h2>Encolat / Protter ({len(encolat_rows)} files)</h2>' + rows_html(encolat_rows, enc_hl or pro_hl)
    html += '</body></html>'
    return html


def _wa_phone(tel):
    """Normalitza un telèfon per a wa.me (només dígits; afegeix 34 si són 9)."""
    digits = ''.join(ch for ch in str(tel or '') if ch.isdigit())
    if len(digits) == 9:
        digits = '34' + digits
    return digits


def _wa_client_text(op, emp_nom, emp_adr, emp_tel):
    """Construeix el text de WhatsApp per al client a partir d'una comanda
    desada (una opció). Mateix contingut que el missatge de la calculadora, però
    des de les dades guardades. En català."""
    def _v(k):
        v = op.get(k)
        return '' if v is None else str(v).strip()
    def _f(k):
        try:
            return float(op.get(k) or 0)
        except (TypeError, ValueError):
            return 0.0
    L = []
    num = _v('num_pressupost')
    L.append('Hola' + (f' (Ref {num})' if num else ''))
    L.append('')
    if _v('client_nom'):
        L.append('Client: ' + _v('client_nom'))
    fw, fh = _f('final_amplada'), _f('final_alcada')
    if fw and fh:
        L.append('Mida final: %d × %d cm' % (round(fw), round(fh)))
    aw, ah = _f('amplada'), _f('alcada')
    if aw and ah:
        L.append('Mides: %d × %d cm' % (round(aw), round(ah)))
    if _v('marc_principal'):
        marc = _v('marc_principal')
        if _v('pre_marc') and _v('pre_marc') != '-':
            marc += ' (+ ' + _v('pre_marc') + ')'
        L.append('Marc: ' + marc)
    if _v('encolat') and _v('encolat') != '-':
        L.append('Muntatge: ' + _v('encolat'))
    if _v('vidre') and _v('vidre') != '-':
        L.append('Vidre: ' + _v('vidre'))
    if _v('passpartout') and _v('passpartout') != '-':
        pp = _v('passpartout')
        if _v('passpartu_ref'):
            pp += ' (' + _v('passpartu_ref') + ')'
        L.append('Interior: ' + pp)
    if _v('impressio') and _v('impressio') != '-':
        L.append('Impressió: Sí')
    if _v('observacions'):
        L.append('Notes: ' + _v('observacions'))
    L.append('')
    desc = _f('descompte')
    if desc > 0:
        L.append('Descompte aplicat: %d%%' % round(desc))
    L.append('*Total: %.2f EUR (IVA inclòs)*' % _f('preu_final'))
    entrega = _f('entrega')
    if entrega > 0:
        L.append('Pagament: %.2f EUR' % entrega)
        L.append('*Pendent: %.2f EUR*' % max(0.0, _f('pendent')))
    L.append('')
    L.append('Gràcies, ' + (emp_nom or 'Reus Revela'))
    if emp_adr:
        L.append(emp_adr)
    if emp_tel:
        L.append('Tel: ' + emp_tel)
    return '\n'.join(L)


@app.route('/historial')
@login_required
def historial():
    filtre_uid_raw = request.args.get('user_id', '').strip()
    filtre_uid = int(filtre_uid_raw) if filtre_uid_raw.isdigit() else None
    filtre_all  = filtre_uid_raw == 'all'
    filtre_albara = request.args.get('albara') == 'pendent'
    filtre_client_raw = request.args.get('client', '').strip()
    filtre_client = int(filtre_client_raw) if filtre_client_raw.isdigit() else None
    if session.get('is_admin'):
        if filtre_client:
            comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                               JOIN usuaris u ON c.user_id=u.id
                               WHERE c.client_extern_id=? ORDER BY c.id DESC''', [filtre_client])
        elif filtre_albara:
            comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                               JOIN usuaris u ON c.user_id=u.id
                               WHERE c.observacions LIKE '%[ACCEPTAT]%'
                                 AND (c.fd_albara IS NULL OR c.fd_albara='')
                               ORDER BY c.id DESC''')
        elif filtre_all:
            comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                               JOIN usuaris u ON c.user_id=u.id
                               ORDER BY c.id DESC''')
        elif filtre_uid:
            comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                               JOIN usuaris u ON c.user_id=u.id
                               WHERE c.user_id=? ORDER BY c.id DESC''', [filtre_uid])
        else:
            filtre_uid = session['user_id']
            comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                               JOIN usuaris u ON c.user_id=u.id
                               WHERE c.user_id=? ORDER BY c.id DESC''', [filtre_uid])
        usuaris_list = query('SELECT id, nom, username FROM usuaris WHERE is_admin=0 ORDER BY nom')
        clients_habituals = query('SELECT id, nom, tipus FROM clients_externs WHERE actiu=TRUE ORDER BY nom') or []
        n_pendents_albara = query('''SELECT COUNT(*) as n FROM comandes
                                    WHERE observacions LIKE '%[ACCEPTAT]%'
                                      AND (fd_albara IS NULL OR fd_albara='')''', one=True)
        n_pendents_albara = n_pendents_albara['n'] if n_pendents_albara else 0
    else:
        usuaris_list = []
        clients_habituals = []
        comandes = query('''SELECT c.*, u.nom as usuari_nom FROM comandes c
                           JOIN usuaris u ON c.user_id=u.id
                           WHERE c.user_id=? ORDER BY c.id DESC''', [session['user_id']])
    # Dades d'empresa (per a la firma del missatge de WhatsApp al client).
    _emp = query('SELECT nom_empresa, empresa_adreca, empresa_tel FROM usuaris WHERE id=?',
                 [session['user_id']], one=True)
    _emp_nom = session.get('empresa_nom') or _row_get(_emp, 'nom_empresa', '') or 'Reus Revela'
    _emp_adr = _row_get(_emp, 'empresa_adreca', '') or ''
    _emp_tel = _row_get(_emp, 'empresa_tel', '') or ''
    # Group by sessio_id
    sessions = {}
    for c in comandes:
        sid = c['sessio_id'] or str(c['id'])
        if sid not in sessions:
            sessions[sid] = []
        d = dict(c)
        # Text i telèfon per enviar aquesta opció per WhatsApp des de l'historial.
        d['wa_text'] = _wa_client_text(d, _emp_nom, _emp_adr, _emp_tel)
        d['wa_tel'] = _wa_phone(d.get('client_tel'))
        sessions[sid].append(d)
    sessio_list = list(sessions.values())
    # Add pagat/entregat flag + estat (F2) to first item of each session
    for grp in sessio_list:
        grp[0]['pagat']    = any(op.get('pagat')    for op in grp)
        grp[0]['entregat'] = any(op.get('entregat') for op in grp)
        grp[0]['estat']    = _derive_estat(grp[0])
        grp[0]['urgent']   = _comanda_es_urgent(grp[0])
    # Filtre d'estat (F2), aplicat en Python sobre els grups perquè no cal
    # tocar les branques SQL ni perdre el filtre per albarà/client/usuari.
    filtre_estat = request.args.get('estat', '').strip().lower()
    if filtre_estat == 'pendents':
        sessio_list = [g for g in sessio_list if g[0]['estat'] not in COMANDA_ESTATS_TANCATS]
    elif filtre_estat == 'entregats':
        sessio_list = [g for g in sessio_list if g[0]['estat'] == 'entregat']
    elif filtre_estat == 'urgents':
        sessio_list = [g for g in sessio_list if g[0]['urgent']]
    elif filtre_estat in COMANDA_ESTAT_KEYS:
        sessio_list = [g for g in sessio_list if g[0]['estat'] == filtre_estat]
    # Agrupar les sessions per CLIENT (nom + telèfon normalitzats): una targeta
    # per client amb tots els seus pressupostos junts. Manté l'ordre (més recent
    # primer) segons la primera sessió que apareix de cada client.
    def _norm_tel(t):
        return ''.join(ch for ch in str(t or '') if ch.isdigit())
    _cli_map = {}
    _cli_order = []
    for grp in sessio_list:
        c0 = grp[0]
        key = (str(c0.get('client_nom') or '').strip().lower(), _norm_tel(c0.get('client_tel')))
        g = _cli_map.get(key)
        if g is None:
            g = {'nom': c0.get('client_nom') or '(sense nom)',
                 'tel': c0.get('client_tel') or '',
                 'sessions': [], 'n': 0, 'pendent': 0.0}
            _cli_map[key] = g
            _cli_order.append(key)
        g['sessions'].append(grp)
        g['n'] += 1
        for op in grp:
            try:
                g['pendent'] += float(op.get('pendent') or 0)
            except (TypeError, ValueError):
                pass
    client_groups = [_cli_map[k] for k in _cli_order]
    filtre_client_nom = ''
    if filtre_client:
        _fc = query('SELECT nom FROM clients_externs WHERE id=?', [filtre_client], one=True)
        filtre_client_nom = _row_get(_fc, 'nom', '') if _fc else ''
    return render_template('historial.html', comandes=comandes, sessio_list=sessio_list,
                           client_groups=client_groups,
                           usuaris_list=usuaris_list if session.get('is_admin') else [],
                           clients_habituals=clients_habituals if session.get('is_admin') else [],
                           filtre_uid=filtre_uid, filtre_all=filtre_all if session.get('is_admin') else False,
                           filtre_albara=filtre_albara if session.get('is_admin') else False,
                           filtre_client=filtre_client, filtre_client_nom=filtre_client_nom,
                           n_pendents_albara=n_pendents_albara if session.get('is_admin') else 0,
                           estats=COMANDA_ESTATS, estat_labels=COMANDA_ESTAT_LABELS,
                           filtre_estat=filtre_estat,
                           web_return_url=_current_web_return_url())

@app.route('/pdf-comparativa/<sessio_id>')
@login_required
def pdf_comparativa(sessio_id):
    comandes = query('''SELECT * FROM comandes WHERE sessio_id=? ORDER BY id''', [sessio_id])
    if not comandes:
        return 'No trobat', 404
    if not session.get('is_admin') and comandes[0]['user_id'] != session['user_id']:
        return 'No autoritzat', 403
    pdf = crear_pdf_comparativa([dict(c) for c in comandes])
    nom = comandes[0]['client_nom'] or 'comparativa'
    return send_file(pdf, mimetype='application/pdf',
                     download_name=f"comparativa_{nom}.pdf")

@app.route('/pdf/<int:comanda_id>')
@login_required
def generar_pdf(comanda_id):
    c = query('SELECT * FROM comandes WHERE id=?', [comanda_id], one=True)
    if not c: return 'No trobat', 404
    if not session.get('is_admin') and c['user_id'] != session['user_id']:
        return 'No autoritzat', 403
    mode = (request.args.get('mode') or '').strip().lower()
    pdf = crear_pdf(dict(c), mode=mode)
    return send_file(pdf, mimetype='application/pdf',
                     download_name=f"pressupost_{c['client_nom']}_{comanda_id}.pdf")

# â"€â"€ Routes: Admin â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route('/admin')
@admin_required
def admin():
    usuaris = query('SELECT * FROM usuaris ORDER BY nom')
    config = {r['clau']: r['valor'] for r in query('SELECT * FROM config')}

    # Stats operatives per al panell. data a comandes es guarda com a text
    # 'dd/mm/YYYY HH:MM', així que filtrem per substring del mes+any.
    now = datetime.now()
    pat_mes = f"%/{now.strftime('%m')}/{now.strftime('%Y')}%"
    row = query('SELECT COUNT(*) AS n FROM comandes WHERE data LIKE ?', [pat_mes], one=True)
    def _estat(u):
        return (_row_get(u, 'access_status', '') or 'active').lower()
    stats = {
        'pressupostos_mes':   row['n'] if row else 0,
        'usuaris_total':      len(usuaris) if usuaris else 0,
        'usuaris_pendents':   sum(1 for u in (usuaris or []) if _estat(u) == 'pending'),
        'usuaris_bloquejats': sum(1 for u in (usuaris or []) if _estat(u) == 'blocked'),
    }
    return render_template('admin.html', usuaris=usuaris, config=config, stats=stats)

@app.route('/admin/usuaris')
@admin_required
def admin_usuaris():
    """Pàgina dedicada a gestió d'usuaris (clients professionals). El POST
    continua a /admin/usuari (singular) per compatibilitat amb els forms
    existents — aquí només renderitzem el llistat.

    Nota privacitat: els marges generals (marge, marge_pro_pct, marge_impressio,
    marge_impressio_pro_pct, margins_json) es consideren dades privades del
    professional i NO arriben a la UI d'admin. Cada pro els gestiona des de
    /ajustos. Els trams d'impressió (imp_tram1..6) sí es mostren aquí per a
    onboarding/suport però queden traçats al audit log si l'admin els toca."""
    # Si la columna 'email' encara no s'ha migrat (cal visitar
    # /admin/run-migrations després del deploy), provem un ALTER idempotent
    # i tornem a llegir. Així evitem el 500 i recuperem la pàgina sol.
    base_sql = """
        SELECT id, username, nom, nom_empresa, profile_type, access_status,
               web_url, instagram, fiscal_id, notes_validacio, is_admin,
               imp_tram1, imp_tram2, imp_tram3, imp_tram4, imp_tram5, imp_tram6,
               baryta_actiu, email
        FROM usuaris ORDER BY nom
    """
    try:
        usuaris = query(base_sql)
    except Exception as e:
        if 'email' in str(e).lower() and ('does not exist' in str(e).lower() or 'no such column' in str(e).lower()):
            try:
                execute("ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''")
                usuaris = query(base_sql)
                flash("S'ha afegit la columna 'email' automàticament. Si veus altres problemes, executa /admin/run-migrations.", 'ok')
            except Exception as e2:
                flash(f"Cal aplicar migracions: visita /admin/run-migrations. Detall: {e2}", 'error')
                usuaris = query(base_sql.replace(", email", ""))
        else:
            raise
    # Defaults globals per als placeholders
    imp_defaults = {
        f'tram{i}': float(get_config_value(f'imp_tram{i}_marge_default', '0'))
        for i in range(1, 7)
    }
    return render_template('admin_usuaris.html', usuaris=usuaris, imp_defaults=imp_defaults)


@app.route('/admin/usuari', methods=['POST'])
@admin_required
def admin_usuari():
    action = request.form.get('action')
    if action == 'crear':
        username = (request.form.get('username') or '').strip()
        email = (request.form.get('email') or '').strip()
        # Pre-check de duplicat per donar missatge net (en comptes d'esperar al
        # UniqueViolation de Postgres / IntegrityError de SQLite).
        if username and query('SELECT 1 FROM usuaris WHERE LOWER(username)=LOWER(?)', [username], one=True):
            flash(f"Ja existeix un usuari amb username '{username}'. Tria'n un altre o edita l'existent.", 'error')
            return redirect(url_for('admin_usuaris'))
        try:
            execute('INSERT INTO usuaris (username, password, nom, is_admin, access_status, profile_type, web_url, instagram, fiscal_id, notes_validacio, email) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
                    [username, hash_pw(request.form['password']),
                     request.form['nom'], int(request.form.get('is_admin', 0)),
                     request.form.get('access_status', 'active'),
                     request.form.get('profile_type', 'professional'),
                     request.form.get('web_url', '').strip(),
                     request.form.get('instagram', '').strip(),
                     request.form.get('fiscal_id', '').strip(),
                     request.form.get('notes_validacio', '').strip(),
                     email])
        except Exception as e:
            # Race condition o constraint inesperat — convertim en flash net.
            msg = str(e).lower()
            if 'unique' in msg or 'duplicate' in msg:
                flash(f"Ja existeix un usuari amb username '{username}'.", 'error')
            else:
                flash(f"No s'ha pogut crear l'usuari: {e}", 'error')
            return redirect(url_for('admin_usuaris'))
        _audit_log('user.create', target_username=username,
                   details=f"nom={request.form.get('nom','')} is_admin={request.form.get('is_admin','0')} email={email}")
        flash(f"Usuari '{request.form['nom']}' creat.", 'ok')
    elif action == 'eliminar':
        uid = request.form['uid']
        target = query('SELECT username FROM usuaris WHERE id=?', [uid], one=True)
        execute('DELETE FROM usuaris WHERE id=?', [uid])
        _audit_log('user.delete', target_user_id=int(uid) if str(uid).isdigit() else None,
                   target_username=_row_get(target, 'username', '') if target else '')
        flash('Usuari eliminat.', 'ok')
    elif action == 'canviar_pw':
        uid = request.form['uid']
        target = query('SELECT username FROM usuaris WHERE id=?', [uid], one=True)
        execute('UPDATE usuaris SET password=? WHERE id=?',
                [hash_pw(request.form['password']), uid])
        _audit_log('user.password_change', target_user_id=int(uid) if str(uid).isdigit() else None,
                   target_username=_row_get(target, 'username', '') if target else '')
        flash('Contrasenya actualitzada.', 'ok')
    elif action == 'actualitzar_email':
        uid = request.form['uid']
        nou_email = (request.form.get('email') or '').strip()
        target = query('SELECT username FROM usuaris WHERE id=?', [uid], one=True)
        if not target:
            flash('Usuari no trobat.', 'error')
        elif nou_email and '@' not in nou_email:
            flash(f"L'adreça «{nou_email}» no sembla un email vàlid.", 'error')
        else:
            execute('UPDATE usuaris SET email=? WHERE id=?', [nou_email, uid])
            _audit_log('user.email_change',
                       target_user_id=int(uid) if str(uid).isdigit() else None,
                       target_username=_row_get(target, 'username', '') if target else '',
                       details=f"new_email={nou_email or '(buit)'}")
            if nou_email:
                flash(f"Email actualitzat a «{nou_email}».", 'ok')
            else:
                flash("Email esborrat (el Welcome farà fallback al username si és vàlid).", 'ok')
    elif action == 'send_welcome':
        # Genera una contrasenya nova, la desa hashed i envia un mail al
        # professional amb les dades d'accés + mini tutorial de la calculadora.
        # IMPORTANT: si Gmail no està configurat, NO toquem la contrasenya
        # (per no deixar l'usuari sense accés sense rebre el mail).
        uid = request.form['uid']
        target = query('SELECT username, nom, email FROM usuaris WHERE id=?', [uid], one=True)
        if not target:
            flash('Usuari no trobat.', 'error')
        else:
            username = _row_get(target, 'username', '') or ''
            nom = _row_get(target, 'nom', '') or ''
            stored_email = (_row_get(target, 'email', '') or '').strip()
            # Destinatari: el camp email té prioritat; si no, fem servir username
            # (assumim que és un email vàlid, com a fallback per a usuaris antics).
            to_addr = stored_email if '@' in stored_email else username
            if '@' not in to_addr:
                flash(f"L'usuari «{nom or username}» no té cap email vàlid (ni al camp email ni al username). Edita'l abans d'enviar la benvinguda.", 'error')
            elif not _email_is_configured():
                flash("Cap proveïdor d'email està configurat. Posa el resend_api_key (recomanat) o gmail_user/gmail_pass a /admin/config abans de fer servir 'Welcome'. La contrasenya NO s'ha tocat.", 'error')
            else:
                new_pw = _generate_temp_password(12)
                sent = _send_welcome_email(username, new_pw, nom, to_addr=to_addr)
                if sent:
                    # Només rotem la contrasenya quan el mail ha sortit OK.
                    execute('UPDATE usuaris SET password=? WHERE id=?', [hash_pw(new_pw), uid])
                    _audit_log('user.welcome_email_sent',
                               target_user_id=int(uid) if str(uid).isdigit() else None,
                               target_username=username,
                               details='sent=yes')
                    flash(f"Mail de benvinguda enviat a {username} amb contrasenya nova.", 'ok')
                else:
                    _audit_log('user.welcome_email_sent',
                               target_user_id=int(uid) if str(uid).isdigit() else None,
                               target_username=username,
                               details='sent=no (gmail send failed, password kept)')
                    flash(f"No s'ha pogut enviar el mail a {username}. La contrasenya antiga es conserva. Comprova els logs i la config Gmail.", 'error')
    elif action == 'actualitzar_estat':
        uid = request.form['uid']
        # Llegim l'estat actual i les dades de l'usuari ABANS de l'UPDATE
        # per detectar la transició a 'active' i poder enviar email amb la
        # contrasenya regenerada al usuari.
        target = query(
            'SELECT username, nom, access_status FROM usuaris WHERE id=?',
            [uid], one=True,
        )
        old_status = _user_access_status(target) if target else ''
        new_status = request.form.get('access_status', 'active') or 'active'

        # Si l'estat passa a 'active' venint de no-actiu, regenerem la
        # contrasenya i la inclourem a l'email d'activació. Si ja era
        # actiu, no toquem la contrasenya (poden estar editant altres
        # camps del perfil).
        notify_activation = (new_status == 'active' and old_status != 'active')
        new_temp_password = None
        if notify_activation:
            new_temp_password = secrets.token_urlsafe(12)
            execute('UPDATE usuaris SET password=? WHERE id=?',
                    [hash_pw(new_temp_password), uid])

        execute('UPDATE usuaris SET access_status=?, profile_type=?, web_url=?, instagram=?, fiscal_id=?, notes_validacio=? WHERE id=?',
                [new_status,
                 request.form.get('profile_type', 'professional'),
                 request.form.get('web_url', '').strip(),
                 request.form.get('instagram', '').strip(),
                 request.form.get('fiscal_id', '').strip(),
                 request.form.get('notes_validacio', '').strip(),
                 uid])
        _audit_log('user.profile_update', target_user_id=int(uid) if str(uid).isdigit() else None,
                   target_username=_row_get(target, 'username', '') if target else '',
                   details=f"status={new_status} profile={request.form.get('profile_type','professional')}"
                           + (' [activation_email]' if notify_activation else ''))
        if notify_activation:
            user_email = (_row_get(target, 'username', '') or '').strip()
            user_name  = (_row_get(target, 'nom', '') or '').strip()
            if user_email:
                _send_user_activation_email(user_name, user_email, new_temp_password)
            flash(f"Perfil actualitzat. Compte activat i email d'accés enviat a {user_email}.", 'ok')
        else:
            flash('Perfil professional actualitzat.', 'ok')
    elif action == 'actualitzar_imp_trams':
        uid = request.form['uid']
        target = query('SELECT username FROM usuaris WHERE id=?', [uid], one=True)
        valors = []
        for i in range(1, 7):
            raw = (request.form.get(f'imp_tram{i}') or '').strip()
            valors.append(float(raw) if raw else None)
        execute(
            'UPDATE usuaris SET imp_tram1=?, imp_tram2=?, imp_tram3=?, '
            'imp_tram4=?, imp_tram5=?, imp_tram6=? WHERE id=?',
            valors + [uid],
        )
        _audit_log('user.imp_trams_update', target_user_id=int(uid) if str(uid).isdigit() else None,
                   target_username=_row_get(target, 'username', '') if target else '',
                   details=' '.join(f"t{i+1}={v}" for i, v in enumerate(valors)))
        flash('Trams d\'impressió actualitzats.', 'ok')
    elif action == 'baryta_toggle':
        uid = request.form['uid']
        target = query('SELECT username, baryta_actiu FROM usuaris WHERE id=?', [uid], one=True)
        nou = 1 if request.form.get('baryta_actiu') in ('1', 'on', 'true') else 0
        execute('UPDATE usuaris SET baryta_actiu=? WHERE id=?', [nou, uid])
        _audit_log('user.baryta_toggle', target_user_id=int(uid) if str(uid).isdigit() else None,
                   target_username=_row_get(target, 'username', '') if target else '',
                   details=f"baryta_actiu={nou}")
        flash(f"Paper Baryta {'activat' if nou else 'desactivat'}.", 'ok')
    return redirect(url_for('admin_usuaris'))


@app.route('/admin/audit-log')
@admin_required
def admin_audit_log():
    """Visor del registre d'accions administratives sobre comptes
    d'usuari (creacions, eliminacions, canvis de contrasenya/perfil).
    Pensat com a traça per saber qui ha tocat què."""
    rows = query("""
        SELECT id, actor_user_id, actor_username, target_user_id, target_username,
               action, details, created_at
        FROM audit_log
        ORDER BY id DESC
        LIMIT 200
    """) or []
    return render_template('admin_audit_log.html', rows=rows)


ADMIN_MARGE_CATEGORIES = [
    ('moldures',  'Moldures'),
    ('vidres',    'Vidres'),
    ('encolat',   'Encolat / Laminat'),
    ('passpartu', 'Passpartú'),
]


# ── Tarifes generador ─────────────────────────────────────────────────────
TARIFA_PRODUCTS = [
    ('impressio',     'Impressions fotogràfiques',  'imp'),
    ('vidre',         'Vidre simple',                'vidre'),
    ('doble_vidre',   'Doble vidre',                 'doble_vidre'),
    ('mirall',        'Mirall',                      'mirall'),
    ('passpartu',     'Passpartú simple',            'pas'),
    ('doble_pas',     'Doble passpartú',             'dobpas'),
    ('foam',          'Encolat foam',                'foam'),
    ('laminat_semi',  'Laminat SemiBrillo',          'lamS'),
    ('laminat_mate',  'Laminat Mate',                'lamM'),
    ('protter_semi',  'Protter SemiBrillo',          'protS'),
    ('protter_mate',  'Protter Mate',                'protM'),
]


def _tarifa_default_sizes(product):
    """Retorna llista de (w, h) per defecte d'un producte, llegint la BD."""
    sizes = []
    if product == 'impressio':
        rows = query("SELECT referencia FROM impressio") or []
    elif product == 'vidre':
        rows = query("SELECT referencia FROM vidres "
                     "WHERE UPPER(referencia) NOT LIKE 'DV-%' "
                     "AND UPPER(referencia) NOT LIKE 'MIR-%'") or []
    elif product == 'doble_vidre':
        rows = query("SELECT referencia FROM vidres WHERE UPPER(referencia) LIKE 'DV-%'") or []
    elif product == 'mirall':
        rows = query("SELECT referencia FROM vidres WHERE UPPER(referencia) LIKE 'MIR-%'") or []
    elif product == 'passpartu':
        rows = query("SELECT referencia FROM passpartout WHERE UPPER(referencia) LIKE '1PAS%'") or []
    elif product == 'doble_pas':
        rows = query("SELECT referencia FROM passpartout WHERE UPPER(referencia) LIKE 'DOBPAS%'") or []
    elif product in ('foam',):
        rows = query("SELECT referencia FROM encolat_pro WHERE UPPER(referencia) LIKE 'ENC%'") or []
    elif product in ('laminat_semi', 'laminat_mate', 'protter_semi', 'protter_mate'):
        rows = query("SELECT referencia FROM encolat_pro WHERE UPPER(referencia) LIKE 'PRO%'") or []
    else:
        rows = []
    seen = set()
    for r in rows:
        ref = _row_get(r, 'referencia') or ''
        rw, rh = _parse_dims(ref)
        if rw and rh and (rw, rh) not in seen:
            seen.add((rw, rh))
            sizes.append((rw, rh))
    sizes.sort(key=lambda wh: wh[0] * wh[1])
    return sizes


def _tarifa_compute_one(product, w, h, usuari):
    """Computa cost/PVD/PVP per a un sol producte+mida.
    Retorna {ref, w, h, pvd, pvp, origen} o None si error."""
    try:
        if product == 'impressio':
            r = _imp_closest(w, h) or {}
            pvd = float(r.get('preu') or 0)
            ref = r.get('ref') or f'imp-{w}x{h}'
            origen = r.get('origen') or '—'
            tram = get_marge_impressio_tram(w * h, usuari) or {}
            pvp = round(pvd * (1 + float(tram.get('marge') or 0) / 100), 2)
        else:
            if product == 'vidre':
                d = calcular_cost_vidre(w, h)
            elif product == 'doble_vidre':
                d = calcular_cost_doble_vidre(w, h)
            elif product == 'mirall':
                d = calcular_cost_mirall(w, h)
            elif product == 'passpartu':
                d = calcular_cost_passpartu(w, h, tipus='simple')
            elif product == 'doble_pas':
                d = calcular_cost_passpartu(w, h, tipus='doble')
            elif product == 'foam':
                d = calcular_cost_foam(w, h)
            elif product == 'laminat_semi':
                d = calcular_cost_laminat(w, h, tipus='semibrillo')
            elif product == 'laminat_mate':
                d = calcular_cost_laminat(w, h, tipus='mate')
            elif product == 'protter_semi':
                d = calcular_cost_protter(w, h, tipus='semibrillo')
            elif product == 'protter_mate':
                d = calcular_cost_protter(w, h, tipus='mate')
            else:
                return None
            d = d or {}
            pvd = float(d.get('pvd') or 0)
            ref = d.get('ref') or '—'
            origen = d.get('origen') or '—'
            # PVP = PVD · (1 + marge_user/100). Marge legacy o pro_pct.
            user_marge = _get_marge_value(usuari) if usuari else 0.0
            pvp = round(pvd * (1 + user_marge / 100), 2)
        return {
            'ref': ref, 'w': w, 'h': h,
            'mida_label': f'{w}×{h} cm',
            'pvd': round(pvd, 2),
            'pvp': pvp,
            'origen': origen,
        }
    except Exception as e:
        print(f'tarifa_compute_one error {product} {w}×{h}: {e}')
        return None


def _tarifa_parse_custom_sizes(text):
    """Parseja un text 'WxH, WxH, ...' a llista de tuples (w, h)."""
    import re
    result = []
    for token in (text or '').replace('\n', ',').split(','):
        token = token.strip().lower().replace('×', 'x')
        if not token: continue
        m = re.match(r'(\d+)\s*x\s*(\d+)', token)
        if m:
            try:
                result.append((int(m.group(1)), int(m.group(2))))
            except ValueError:
                continue
    return result


# Mapeig producte → (taula catàleg, filtre prefix, columna preu, categoria marge).
# Categoria marge serveix per derivar PVD = preu_cost · (1 + marge_admin_<cat>/100)
# als productes on la taula desa preu_cost en lloc del PVD final (com fa
# impressio.preu, que ja és PVD-equivalent directe).
# protter_* es deixa fora del floor: és composite (foam + laminat) i no
# té una sola fila de catàleg per a una mida.
TARIFA_CATALOG_SPEC = {
    'impressio':    ('impressio',    None,      'preu',      None),
    'vidre':        ('vidres',       'NOT_DV_MIR', 'preu_cost', 'vidres'),
    'doble_vidre':  ('vidres',       'DV-',     'preu_cost', 'vidres'),
    'mirall':       ('vidres',       'MIR-',    'preu_cost', 'vidres'),
    'passpartu':    ('passpartout',  '1PAS',    'preu_cost', 'passpartu'),
    'doble_pas':    ('passpartout',  'DOBPAS',  'preu_cost', 'passpartu'),
    'foam':         ('encolat_pro',  'ENC',     'preu_cost', 'encolat'),
    'laminat_semi': ('encolat_pro',  'PRO',     'preu_cost', 'encolat'),
    'laminat_mate': ('encolat_pro',  'PRO',     'preu_cost', 'encolat'),
}


def _tarifa_catalog_price_map(prod_key):
    """Per a un producte, retorna {(w, h): pvd_catalog} amb les files exactes
    del catàleg. pvd_catalog és el PVD que aplicaria el sistema si aquesta
    mida fos l'exacta del catàleg. Útil com a sòl al floor de tarifa."""
    spec = TARIFA_CATALOG_SPEC.get(prod_key)
    if not spec:
        return {}
    table, prefix, col, categoria = spec
    if prefix == 'NOT_DV_MIR':
        sql = (f"SELECT referencia, {col} FROM {table} "
               "WHERE UPPER(referencia) NOT LIKE 'DV-%' "
               "AND UPPER(referencia) NOT LIKE 'MIR-%'")
    elif prefix:
        sql = f"SELECT referencia, {col} FROM {table} WHERE UPPER(referencia) LIKE '{prefix}%'"
    else:
        sql = f"SELECT referencia, {col} FROM {table}"
    rows = query(sql) or []
    result = {}
    marge = 0.0
    if categoria:
        try:
            marge = float(get_config_value(f'marge_admin_{categoria}_pct', '60'))
        except (TypeError, ValueError):
            marge = 60.0
    for r in rows:
        rw, rh = _parse_dims(r.get('referencia') or '')
        if not (rw and rh):
            continue
        v = r.get(col)
        if v is None:
            continue
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        # impressio.preu ja és PVD-equivalent; la resta de taules desen
        # preu_cost i el PVD s'aplica multiplicant pel marge admin.
        pvd_catalog = v if categoria is None else round(v * (1 + marge / 100), 2)
        result[(rw, rh)] = pvd_catalog
    return result


def _tarifa_collect_data(products, custom_sizes_per_product, usuari):
    """Per a cada producte seleccionat, collect [{mida, pvd, pvp, ...}].

    Dues garanties addicionals (sobre el càlcul base de _tarifa_compute_one):

    1. Ref única per a mides addicionals. Mides del catàleg conserven la
       seva ref original. Mides custom reben IMP{W}x{H} a impressió per
       evitar col·lisions amb el catàleg (bug reportat: 9×13 i 10×15
       acabaven amb la mateixa ref `IMP10x15`).

    2. Floor "no baixar mai del catàleg actual": si una mida coincideix
       exactament amb una fila del catàleg, el PVD final ha de ser ≥
       PVD_catàleg × tarifa_floor_factor (default 1.03). Red de
       seguretat tècnica per la directriu comercial 'la tarifa nova
       mai és per sota de l'actual'. S'aplica a tots els productes amb
       catàleg directe (no a protter, que és composite).
    """
    try:
        floor_factor = float(get_config_value('tarifa_floor_factor', '1.03'))
    except (TypeError, ValueError):
        floor_factor = 1.03

    result = []
    for prod_key, label, _short in TARIFA_PRODUCTS:
        if prod_key not in products:
            continue
        catalog_sizes = _tarifa_default_sizes(prod_key)
        catalog_set = set(catalog_sizes)
        sizes = list(catalog_sizes)
        seen = set(catalog_set)
        for wh in custom_sizes_per_product.get(prod_key, []):
            if wh not in seen:
                sizes.append(wh)
                seen.add(wh)
        sizes.sort(key=lambda wh: wh[0] * wh[1])

        catalog_pvd_map = _tarifa_catalog_price_map(prod_key)

        rows = []
        for w, h in sizes:
            r = _tarifa_compute_one(prod_key, w, h, usuari)
            if not r:
                continue
            if prod_key == 'impressio' and (w, h) not in catalog_set:
                r['ref'] = f'IMP{int(w)}x{int(h)}'
            # Floor: si la mida està al catàleg, no baixar del PVD del
            # catàleg × factor. Recalculem PVP a partir del PVD ajustat.
            catalog_pvd = catalog_pvd_map.get((w, h))
            if catalog_pvd is not None and floor_factor > 0:
                floor_pvd = round(catalog_pvd * floor_factor, 2)
                if (r.get('pvd') or 0) < floor_pvd:
                    r['pvd'] = floor_pvd
                    if prod_key == 'impressio':
                        tram = get_marge_impressio_tram(w * h, usuari) or {}
                        marge_pvp = float(tram.get('marge') or 0)
                    else:
                        marge_pvp = _get_marge_value(usuari) if usuari else 0.0
                    r['pvp'] = round(floor_pvd * (1 + marge_pvp / 100), 2)
                    r['origen'] = 'floor'
            rows.append(r)
        result.append({'key': prod_key, 'label': label, 'rows': rows})
    return result


def _tarifa_build_pdf(data, vista, user_label, brand_color):
    """Genera un PDF A4 portrait amb una secció per producte."""
    from reportlab.platypus import PageBreak
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    dark = colors.HexColor('#1C1B18')
    green = colors.HexColor(brand_color or '#1A6B45')
    border = colors.HexColor('#E5E2DB')
    light = colors.HexColor('#F5F4F1')
    muted = colors.HexColor('#6B6860')

    def p(txt, *, bold=False, size=9, color=dark, align=0):
        return Paragraph(str(txt), ParagraphStyle(
            name=f't-{size}-{int(bold)}-{align}',
            fontName='DejaVu-Bold' if bold else 'DejaVu',
            fontSize=size, leading=size + 2, textColor=color, alignment=align,
        ))

    avui = datetime.now().strftime('%d/%m/%Y')
    story = [
        p(f"Tarifa {vista}", bold=True, size=18, color=green),
        Spacer(1, 1*mm),
        p(f"Reus Revela · {avui}" + (f" · {user_label}" if user_label else ''),
          size=9, color=muted),
        Spacer(1, 5*mm),
    ]

    for i, sec in enumerate(data):
        if i > 0:
            story.append(PageBreak())
        story.append(p(sec['label'], bold=True, size=14, color=dark))
        story.append(Spacer(1, 2*mm))
        if not sec['rows']:
            story.append(p('— Sense mides per a aquest producte —', size=9, color=muted))
            continue
        cells = [[p('Referència', bold=True, size=8, color=colors.white, align=TA_CENTER),
                  p('Mida', bold=True, size=8, color=colors.white, align=TA_CENTER),
                  p('Preu sense IVA', bold=True, size=8, color=colors.white, align=TA_CENTER),
                  p('Preu amb IVA (21%)', bold=True, size=8, color=colors.white, align=TA_CENTER)]]
        col = 'pvp' if vista.upper() == 'PVP' else 'pvd'
        for r in sec['rows']:
            preu = float(r.get(col) or 0)
            cells.append([
                p(r['ref'], size=8),
                p(r['mida_label'], size=8, align=TA_RIGHT),
                p(f'{preu:.2f} €', size=8, align=TA_RIGHT),
                p(f'{preu * 1.21:.2f} €', size=8, align=TA_RIGHT),
            ])
        table = Table(cells, repeatRows=1, colWidths=[55*mm, 30*mm, 40*mm, 40*mm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), dark),
            ('BOX', (0, 0), (-1, -1), 0.4, border),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, border),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(table)

    story.append(Spacer(1, 6*mm))
    story.append(p(f'Preus vigents a {avui}. Sense IVA llevat que s\'indiqui.',
                   size=8, color=muted, align=TA_CENTER))
    doc.build(story)
    buf.seek(0)
    return buf


def _tarifa_build_excel(data, vista, user_label):
    """Genera un Excel amb un full per producte."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    wb.remove(wb.active)
    avui = datetime.now().strftime('%d/%m/%Y')

    head_fill = PatternFill('solid', fgColor='1C1B18')
    head_font = Font(bold=True, color='FFFFFF', size=10)
    info_font = Font(italic=True, color='6B6860')
    title_font = Font(bold=True, size=14, color='1A6B45')
    thin = Side(border_style='thin', color='E5E2DB')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    col = 'pvp' if vista.upper() == 'PVP' else 'pvd'
    for sec in data:
        sheet_name = sec['label'][:30]
        ws = wb.create_sheet(title=sheet_name)
        ws['A1'] = f'Tarifa {vista} · {sec["label"]}'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        ws['A2'] = f'Reus Revela · {avui}' + (f' · {user_label}' if user_label else '')
        ws['A2'].font = info_font
        ws.merge_cells('A2:D2')

        # Headers a la fila 4
        for idx, label in enumerate(['Referència', 'Mida', 'Preu sense IVA', 'Preu amb IVA (21%)'], start=1):
            c = ws.cell(row=4, column=idx, value=label)
            c.fill = head_fill; c.font = head_font; c.border = bd
            c.alignment = Alignment(horizontal='center', vertical='center')

        for i, r in enumerate(sec['rows'], start=5):
            preu = float(r.get(col) or 0)
            ws.cell(row=i, column=1, value=r['ref']).border = bd
            ws.cell(row=i, column=2, value=r['mida_label']).border = bd
            cell3 = ws.cell(row=i, column=3, value=preu); cell3.border = bd; cell3.number_format = '#,##0.00 €'
            cell4 = ws.cell(row=i, column=4, value=preu * 1.21); cell4.border = bd; cell4.number_format = '#,##0.00 €'
            ws.cell(row=i, column=2).alignment = Alignment(horizontal='right')

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22

    if not wb.sheetnames:
        wb.create_sheet('Tarifa').append(['Sense dades'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _tarifa_save_config(payload):
    """Desa la última configuració del generador a config['tarifes_config']."""
    try:
        execute(
            "INSERT OR REPLACE INTO config (clau, valor) VALUES (?, ?)",
            ['tarifes_config', json.dumps(payload, ensure_ascii=False)],
        )
    except Exception as e:
        print(f'tarifes_save_config skip: {e}')


def _tarifa_load_config():
    raw = get_config_value('tarifes_config', '')
    if not raw:
        return {}
    try:
        return json.loads(raw) or {}
    except Exception:
        return {}


@app.route('/admin/tarifes', methods=['GET', 'POST'])
@admin_required
def admin_tarifes():
    """Generador de tarifes (PVD/PVP) amb sortida online, PDF o Excel."""
    usuaris = query("SELECT id, username, nom, nom_empresa FROM usuaris "
                    "WHERE access_status='active' ORDER BY nom") or []
    saved = _tarifa_load_config()

    if request.method == 'GET':
        return render_template(
            'admin_tarifes.html',
            usuaris=usuaris, products=TARIFA_PRODUCTS,
            saved=saved, result=None,
        )

    # POST → processar
    vista = (request.form.get('vista') or 'PVD').upper()
    if vista not in ('PVD', 'PVP'):
        vista = 'PVD'
    user_id = request.form.get('user_id') or ''
    products_sel = request.form.getlist('products')
    custom_global = _tarifa_parse_custom_sizes(request.form.get('custom_sizes', ''))
    fmt = (request.form.get('format') or 'online').lower()

    # Cada producte selecciona també les seves mides custom (si el form les porta).
    # MVP: usem el mateix conjunt custom per a tots els productes.
    custom_per_product = {p: list(custom_global) for p, _, _ in TARIFA_PRODUCTS}

    usuari = None
    user_label = ''
    if vista == 'PVP' and user_id:
        usuari = query('SELECT * FROM usuaris WHERE id=?', [user_id], one=True)
        if usuari:
            user_label = (_row_get(usuari, 'nom_empresa') or _row_get(usuari, 'nom') or '').strip()

    data = _tarifa_collect_data(products_sel, custom_per_product, usuari)

    # Desar config (MVP: només els productes seleccionats i el text de mides)
    _tarifa_save_config({
        'vista': vista, 'user_id': user_id,
        'products': products_sel, 'custom_sizes': request.form.get('custom_sizes', ''),
    })

    if fmt == 'pdf':
        brand = _normalize_hex_color(get_config_value('brand_color', DEFAULT_BRAND_COLOR))
        pdf = _tarifa_build_pdf(data, vista, user_label, brand)
        return send_file(pdf, mimetype='application/pdf', as_attachment=True,
                         download_name=f'tarifa-{vista.lower()}-{datetime.now().strftime("%Y%m%d")}.pdf')
    if fmt == 'excel':
        xlsx = _tarifa_build_excel(data, vista, user_label)
        return send_file(xlsx, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'tarifa-{vista.lower()}-{datetime.now().strftime("%Y%m%d")}.xlsx')

    # Online
    return render_template(
        'admin_tarifes.html',
        usuaris=usuaris, products=TARIFA_PRODUCTS,
        saved={'vista': vista, 'user_id': user_id,
               'products': products_sel,
               'custom_sizes': request.form.get('custom_sizes', '')},
        result={'vista': vista, 'user_label': user_label, 'data': data},
    )


@app.route('/admin/config', methods=['GET', 'POST'])
@admin_required
def admin_config():
    if request.method == 'POST':
        execute("UPDATE config SET valor=? WHERE clau='marge_defecte'",
                [request.form.get('marge', 60)])
        # Toggle marge pro
        mpa = '1' if request.form.get('marge_pro_actiu') else '0'
        execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('marge_pro_actiu', ?)", [mpa])
        # Marges admin per categoria (cost → PVD).
        for cat, _ in ADMIN_MARGE_CATEGORIES:
            val = request.form.get(f'marge_admin_{cat}_pct')
            if val is not None and val != '':
                execute(
                    "INSERT OR REPLACE INTO config (clau, valor) VALUES (?, ?)",
                    [f'marge_admin_{cat}_pct', str(val)],
                )
        # Descomptes combo per combinació de productes + mínim de subtotal.
        for clau in ('combo_desc_marc_imp_protter', 'combo_desc_marc_imp_foam',
                     'combo_desc_marc_imp', 'combo_desc_marc_suport',
                     'combo_desc_minim_pvp'):
            val = request.form.get(clau)
            if val is not None and val != '':
                execute(
                    "INSERT OR REPLACE INTO config (clau, valor) VALUES (?, ?)",
                    [clau, str(val).strip()],
                )
        # Gmail SMTP: desem si l'admin omple els camps. El password és type=password
        # i mai s'echo-eja al GET, així que enviar-lo buit (cas normal) NO esborra
        # el valor anterior — només es desa quan el camp porta contingut.
        gu = (request.form.get('gmail_user') or '').strip()
        gp = (request.form.get('gmail_pass') or '').strip().replace(' ', '')
        if gu:
            execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('gmail_user', ?)", [gu])
        if gp:
            execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('gmail_pass', ?)", [gp])
        # Resend (HTTPS API). El from es desa sempre que estigui informat;
        # l'api_key només si l'admin n'ha posat un de nou (mateix patró que Gmail).
        rk = (request.form.get('resend_api_key') or '').strip()
        rf = (request.form.get('resend_from') or '').strip()
        if rk:
            execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('resend_api_key', ?)", [rk])
        if rf:
            execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('resend_from', ?)", [rf])
        # Adreça de respostes (reply-to). Es desa sempre que el camp arribi,
        # perquè es pugui també esborrar deixant-lo buit.
        rr = request.form.get('resend_reply_to')
        if rr is not None:
            execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('resend_reply_to', ?)", [rr.strip()])
        # Marca i responsable legal del mailing (poden diferir d'empresa_nom:
        # marca de cara al client vs nom fiscal). Buit = fallback a empresa_nom.
        for clau in ('mailing_marca', 'mailing_responsable'):
            val = request.form.get(clau)
            if val is not None:
                execute('INSERT OR REPLACE INTO config (clau, valor) VALUES (?, ?)', [clau, val.strip()])
        # Laboratori d'impressió (Fase 1: només email).
        for clau in ('lab_email_dest', 'lab_canal_default', 'lab_assumpte_template', 'lab_cos_template'):
            val = request.form.get(clau)
            if val is not None:
                execute('INSERT OR REPLACE INTO config (clau, valor) VALUES (?, ?)', [clau, val.strip()])
        flash('Configuració desada.', 'ok')
        return redirect(url_for('admin_config'))

    config = {r['clau']: r['valor'] for r in query('SELECT * FROM config')}
    return render_template(
        'admin_config.html',
        config=config,
        admin_marges=ADMIN_MARGE_CATEGORIES,
    )


@app.route('/admin/extras', methods=['GET', 'POST'])
@admin_required
def admin_extras():
    if request.method == 'POST':
        # Form sends N parallel arrays (one entry per extra row). We
        # reconstruct the list zipping them together. The `delete` checkbox
        # marks rows to drop; rows without a name are also dropped.
        keys         = request.form.getlist('extra_key[]')
        names        = request.form.getlist('extra_name[]')
        descs        = request.form.getlist('extra_description[]')
        prices       = request.form.getlist('extra_price_pvd[]')
        margins      = request.form.getlist('extra_margin_pct[]')
        modes        = request.form.getlist('extra_mode[]')
        piece_types  = request.form.getlist('extra_piece_types[]')
        actius       = request.form.getlist('extra_actiu[]')   # checkbox per row → indices of checked rows
        deletes      = request.form.getlist('extra_delete[]')  # idem
        actius_set   = set(actius)
        deletes_set  = set(deletes)
        updated = []
        for i, name in enumerate(names):
            idx = str(i)
            if idx in deletes_set:
                continue
            name = (name or '').strip()
            if not name:
                continue
            key = (keys[i] if i < len(keys) else '').strip() or 'extra_' + str(int(time.time())) + '_' + str(i)
            updated.append({
                'key': key,
                'name': name,
                'description': descs[i] if i < len(descs) else '',
                'price_pvd': prices[i] if i < len(prices) else '0',
                'margin_pct': margins[i] if i < len(margins) else '',
                'mode': modes[i] if i < len(modes) else 'manual',
                'piece_types': piece_types[i] if i < len(piece_types) else '',
                'actiu': idx in actius_set,
                'ordre': i + 1,
            })
        save_extras_list(updated)
        flash('Extres desats.', 'ok')
        return redirect(url_for('admin_extras'))

    return render_template('admin_extras.html', extras=get_extras_list())

# ── Factura Directa ───────────────────────────────────────────────────────
_FD_TOKEN   = os.environ.get('FACTURADIRECTA_TOKEN', '')
_FD_COMPANY = os.environ.get('FACTURADIRECTA_COMPANY', '')
_FD_BASE    = 'https://app.facturadirecta.com/api'
# Sèries de numeració a FacturaDirecta. L'API EXIGEIX que la propietat
# docNumber estigui present (400 'must have required property docNumber' si
# falta), però la SÈRIE dins seu és opcional: si la deixem buida, FD aplica la
# sèrie PER DEFECTE del compte per a aquell tipus de document. Per això només
# forcem una sèrie concreta si l'usuari l'ha configurada explícitament.
#  · Albarà: 'AL' funciona al compte actual (sèrie existent) → per defecte.
#  · Pressupost: buit per defecte → FD tria la sèrie per defecte d'estimates,
#    així no cal saber-ne el codi. Es pot forçar amb FD_ESTIMATE_SERIES.
_FD_ALBARA_SERIES   = os.environ.get('FD_ALBARA_SERIES', 'AL')
_FD_ESTIMATE_SERIES = os.environ.get('FD_ESTIMATE_SERIES', '')

# IVA i recàrrec d'equivalència. A FacturaDirecta el recàrrec d'equivalència és
# una taxa que s'afegeix a la línia JUNT amb l'IVA (per a clients revenedors en
# règim de RE). 5,2% correspon a l'IVA 21%. El codi es pot ajustar amb FD_RE_CODE
# si el compte usa un identificador diferent.
_FD_IVA_CODE = os.environ.get('FD_IVA_CODE', 'S_IVA_21')
_FD_RE_CODE  = os.environ.get('FD_RE_CODE', 'S_REQ_52')

def _fd_line_tax(recarrec=False):
    """Llista de codis de taxa per a una línia FD: IVA 21% i, si el client està
    en règim de recàrrec d'equivalència, també el recàrrec (5,2%)."""
    tax = [_FD_IVA_CODE]
    if recarrec:
        tax.append(_FD_RE_CODE)
    return tax

def _fd_docnumber(series):
    """Construeix el docNumber per a l'API de FD. Sempre ha d'anar present;
    si no forcem cap sèrie (buida), FD aplica la sèrie per defecte del compte."""
    s = (series or '').strip()
    return {'series': s} if s else {}

def _fd_headers():
    return {
        'facturadirecta-api-key': _FD_TOKEN,
        'Content-Type': 'application/json',
        'Accept': 'application/json',
    }

def _fd_get(path, timeout=10):
    url = f'{_FD_BASE}/{_FD_COMPANY}/{path}'
    req = urllib_request.Request(url, headers=_fd_headers())
    try:
        with urllib_request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode())
    except urllib_error.HTTPError as e:
        return {'_error': e.code, '_msg': e.read().decode()[:500]}
    except Exception as e:
        # Xarxa/timeout/SSL/JSON: mai propaguem. Amb 1 worker de gunicorn, una
        # excepció no controlada o un cuelgue pot deixar tota l'app sense
        # respondre (502). Retornem un error net perquè el caller el gestioni.
        return {'_error': 'net', '_msg': str(e)[:300]}

def _fd_post(path, data, timeout=10):
    url = f'{_FD_BASE}/{_FD_COMPANY}/{path}'
    body = json.dumps(data).encode()
    req = urllib_request.Request(url, data=body, headers=_fd_headers(), method='POST')
    try:
        with urllib_request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read().decode())
    except urllib_error.HTTPError as e:
        return {'_error': e.code, '_msg': e.read().decode()[:500]}
    except Exception as e:
        return {'_error': 'net', '_msg': str(e)[:300]}

def _fd_cerca_contacte(nom=None, nif=None):
    """Cerca un contacte a FD per NIF o per nom."""
    if nif:
        res = _fd_get(f'contacts?fiscalId={urllib_quote(nif)}')
        items = res.get('items') or res.get('data') or (res if isinstance(res, list) else [])
        if items:
            return items[0]
    if nom:
        res = _fd_get(f'contacts?search={urllib_quote(nom)}')
        items = res.get('items') or res.get('data') or (res if isinstance(res, list) else [])
        if items:
            return items[0]
    return None

def _fd_extract_contact_id(r):
    """Extreu l'ID d'un contacte de la resposta de l'API de FD."""
    if not r:
        return ''
    c = r.get('content') or {}
    m = c.get('main') or {}
    return (r.get('id') or r.get('uuid') or r.get('contactId') or r.get('_id') or
            c.get('uuid') or c.get('id') or c.get('contactId') or
            m.get('id') or m.get('uuid') or '')

def _fd_crear_contacte(nom, nif=None, telefon=None):
    main = {
        'name': nom, 'country': 'ES', 'currency': 'EUR',
        'accounts': {'client': '430000', 'clientCredit': '438000'},
    }
    if nif:     main['fiscalId'] = nif
    if telefon: main['phone']    = telefon
    res = _fd_post('contacts', {'content': {'type': 'contact', 'main': main}})
    if '_error' in (res or {}):
        return res
    # Si la resposta ja porta ID, la retornem directament
    if _fd_extract_contact_id(res):
        return res
    # Si no té ID però tenim NIF, busquem per NIF (cerca exacta, mai per nom)
    if nif:
        trobat = _fd_cerca_contacte(nif=nif)
        if trobat:
            return trobat
    # Retornem la resposta original (sense ID); l'error es gestionarà a dalt
    print(f'FD crear_contacte: resposta sense ID (nom={nom}, nif={nif}): {json.dumps(res, ensure_ascii=False)}')
    return res

def _fd_crear_albara(contact_id, linies, notes='', data_doc=None):
    if not data_doc:
        data_doc = datetime.now().strftime('%Y-%m-%d')
    main = {
        'contact':   contact_id,
        'currency':  'EUR',
        'baseState': 'pending',
        'docNumber': _fd_docnumber(_FD_ALBARA_SERIES),
        'lines':     linies,
    }
    if data_doc:
        main['date'] = data_doc
    if notes:
        main['notes'] = notes
    return _fd_post('deliveryNotes', {'content': {'type': 'deliveryNote', 'main': main}})


def _fd_crear_estimate(contact_id, linies, notes='', data_doc=None):
    """Crea un PRESSUPOST (estimate) a FacturaDirecta. Mateix patro que
    l'albara pero amb type/endpoint 'estimate'/'estimates'. L'API EXIGEIX que
    docNumber estigui present (400 'must have required property docNumber' si
    falta), pero la serie dins seu es opcional: si la deixem buida FD aplica la
    serie per defecte del compte. Es pot forçar amb FD_ESTIMATE_SERIES."""
    if not data_doc:
        data_doc = datetime.now().strftime('%Y-%m-%d')
    main = {
        'contact':   contact_id,
        'currency':  'EUR',
        'baseState': 'pending',  # estat inicial del pressupost (requerit per l'API)
        'docNumber': _fd_docnumber(_FD_ESTIMATE_SERIES),
        'date':      data_doc,
        'lines':     linies,
    }
    if notes:
        main['notes'] = notes
    res = _fd_post('estimates', {'content': {'type': 'estimate', 'main': main}})
    # Auto-correcció: si haviem forçat una serie que no existeix al compte, FD
    # respon 400 queixant-se de docNumber/series. Reintentem un cop deixant que
    # FD triï la serie per defecte (docNumber buit), sense haver de saber-ne el codi.
    if (isinstance(res, dict) and res.get('_error') == 400
            and (_FD_ESTIMATE_SERIES or '').strip()):
        msg = (res.get('_msg') or '').lower()
        if 'docnumber' in msg or 'series' in msg:
            main['docNumber'] = {}
            res = _fd_post('estimates', {'content': {'type': 'estimate', 'main': main}})
    return res


def _fd_crear_document(doc_type, contact_id, linies, notes='', data_doc=None):
    """Despatxa segons el tipus de document FD demanat:
    'pressupost' -> estimate · qualsevol altre (per defecte) -> albara."""
    if str(doc_type).strip().lower() in ('pressupost', 'estimate', 'presupuesto'):
        return _fd_crear_estimate(contact_id, linies, notes=notes, data_doc=data_doc)
    return _fd_crear_albara(contact_id, linies, notes=notes, data_doc=data_doc)


def _fd_linies_de_comandes(comandes, recarrec=False):
    """Construeix les linies FD (a cost de produccio, una per fila de comanda)
    i les notes a partir d'una llista de files de `comandes`. Compartit per
    l'albara de sessio i pel document conjunt de diversos pressupostos."""
    linies = []
    notes_parts = []
    for com in comandes:
        marc         = (_row_get(com, 'marc_principal', '') or '').strip()
        pre_marc     = (_row_get(com, 'pre_marc', '') or '').strip()
        passpartout  = (_row_get(com, 'passpartout', '') or '').strip()
        vidre        = (_row_get(com, 'vidre', '') or '').strip()
        opcio_nom    = (_row_get(com, 'opcio_nom', '') or '').strip()
        num_pres     = (_row_get(com, 'num_pressupost', '') or '').strip()
        observacions = (_row_get(com, 'observacions', '') or '').strip()
        quantitat    = int(_row_get(com, 'quantitat', 1) or 1)
        cost_prod    = float(_row_get(com, 'cost_produccio', 0) or 0)
        client_nom   = (_row_get(com, 'client_nom', '') or '').strip()

        revers_peu   = str(_row_get(com, 'revers_peu', '') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        encolat      = (_row_get(com, 'encolat', '') or '').strip()
        impressio    = (_row_get(com, 'impressio', '') or '').strip()

        parts_opc = []
        if passpartout and passpartout != 'cap': parts_opc.append(passpartout)
        if vidre:                                parts_opc.append(vidre)
        if pre_marc and pre_marc != '-':         parts_opc.append(f'+ {pre_marc}')
        if encolat and encolat != '-':           parts_opc.append(encolat)
        if revers_peu:                           parts_opc.append('Revers amb peu')
        if impressio and impressio != '-':       parts_opc.append(impressio)
        desc_marc = f'Marc {marc}' if marc else 'Emmarcació'
        if parts_opc:
            desc_marc += f' · {", ".join(parts_opc)}'
        if opcio_nom and opcio_nom != 'Opció A':
            desc_marc += f' ({opcio_nom})'
        if client_nom:
            desc_marc = f'[{client_nom}] ' + desc_marc

        # cost_produccio és el total de totes les unitats; el dividim per qty.
        unit_cost = round(cost_prod / quantitat, 2) if quantitat > 0 else round(cost_prod, 2)
        linies.append({
            'text':      desc_marc,
            'quantity':  float(quantitat),
            'unitPrice': unit_cost,
            'tax':       _fd_line_tax(recarrec),
        })
        if num_pres and num_pres not in notes_parts:
            notes_parts.append(f'Pressupost: {num_pres}')
        if observacions:
            notes_parts.append(f'Obs: {observacions}')
    return linies, notes_parts


def _fd_extract_contacts_list(r):
    """FD pot retornar la llista en diverses claus. Aquesta funció és
    una variant de _fd_extract_contact_id però per a llistes."""
    if not r:
        return []
    if isinstance(r, list):
        return r
    for key in ('items', 'contacts', 'data', 'content', 'results'):
        v = r.get(key) if isinstance(r, dict) else None
        if isinstance(v, list):
            return v
    return []


# ── Clients externs (clients freqüents que no són usuaris de la calc) ─────
@app.route('/admin/fd/contacts')
@admin_required
def admin_fd_contacts_search():
    """Cerca contactes a FacturaDirecta per importar-los com a clients
    externs locals. Mínim 2 caràcters per evitar carregar tota la llista."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'fd_not_configured'}), 503
    q = (request.args.get('q') or '').strip()
    if len(q) < 2:
        return jsonify({'ok': True, 'results': []})
    try:
        r = _fd_get(f'contacts?search={urllib_quote(q)}')
        if isinstance(r, dict) and '_error' in r:
            return jsonify({'ok': False, 'error': f"FD {r.get('_error')}: {r.get('_msg','')}"}), 502
        contacts = _fd_extract_contacts_list(r)
        results = []
        for c in contacts:
            cid = _fd_extract_contact_id(c) or c.get('id') or c.get('uuid')
            if not cid:
                continue
            main = (c.get('content') or {}).get('main') or {}
            nom = c.get('name') or main.get('name') or ''
            nom_comercial = main.get('commercialName') or main.get('tradeName') or main.get('company') or c.get('commercialName') or ''
            nif = c.get('fiscalId') or main.get('fiscalId') or ''
            email = main.get('email') or c.get('email') or ''
            telefon = main.get('phone') or main.get('mobile') or c.get('phone') or ''
            poblacio = main.get('city') or main.get('town') or ''
            results.append({
                'fd_id': str(cid), 'nom': nom, 'nom_comercial': nom_comercial, 'nif': nif,
                'email': email, 'telefon': telefon, 'poblacio': poblacio,
            })
        return jsonify({'ok': True, 'results': results})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


def _fd_get_bounded(path, max_bytes=1_000_000, timeout=6):
    """GET a FD llegint com a MÀXIM max_bytes (evita OOM → 502 amb 1 worker).
    Retorna un dict amb l'estat i el JSON parsejat (o el cap del cos si falla)."""
    url = f'{_FD_BASE}/{_FD_COMPANY}/{path}'
    out = {'status': None, 'ctype': '', 'truncated': False, 'parsed': None,
           'parse_err': None, 'http_error': None, 'head': '', 'exc': None}
    try:
        req = urllib_request.Request(url, headers=_fd_headers())
        try:
            with urllib_request.urlopen(req, timeout=timeout) as resp:
                out['status'] = getattr(resp, 'status', 200)
                out['ctype'] = resp.headers.get('Content-Type', '')
                raw = resp.read(max_bytes + 1)
        except urllib_error.HTTPError as e:
            out['http_error'] = e.code
            out['head'] = e.read(2000).decode('utf-8', 'replace')
            return out
        out['truncated'] = len(raw) > max_bytes
        raw = raw[:max_bytes]
        text = raw.decode('utf-8', 'replace')
        out['head'] = text[:1500]
        if not out['truncated']:
            try:
                out['parsed'] = json.loads(text)
            except Exception as e:
                out['parse_err'] = str(e)[:200]
    except Exception as e:
        out['exc'] = str(e)[:300]
    return out

def _fd_product_row(p):
    """Normalitza un producte de FD als camps que ens interessen per mapejar."""
    content = p.get('content') or {}
    main = content.get('main') or {}
    sales = main.get('sales') or {}
    return {
        'id':      p.get('uuid') or content.get('uuid') or main.get('uuid') or main.get('id') or '',
        'sku':     main.get('sku') or main.get('reference') or '',
        'name':    main.get('name') or main.get('title') or p.get('name') or '',
        'price':   sales.get('price', main.get('price', '')),
        'tax':     sales.get('tax', main.get('tax', '')),
        'account': sales.get('account', ''),
    }


def _fd_write(path, data, method='PUT', max_bytes=200_000, timeout=8):
    """Escriptura a FD (PUT/PATCH/POST) amb lectura acotada de la resposta.
    Mai propaga: retorna dict amb status/http_error/head/exc."""
    url = f'{_FD_BASE}/{_FD_COMPANY}/{path}'
    out = {'status': None, 'http_error': None, 'head': '', 'exc': None}
    try:
        body = json.dumps(data).encode()
        req = urllib_request.Request(url, data=body, headers=_fd_headers(), method=method)
        try:
            with urllib_request.urlopen(req, timeout=timeout) as resp:
                out['status'] = getattr(resp, 'status', None)
                out['head'] = resp.read(max_bytes).decode('utf-8', 'replace')
        except urllib_error.HTTPError as e:
            out['http_error'] = e.code
            out['head'] = e.read(2000).decode('utf-8', 'replace')
    except Exception as e:
        out['exc'] = str(e)[:300]
    return out


@app.route('/admin/fd/products')
@admin_required
def admin_fd_products_list():
    """Inspecció (només lectura) de productes de FD. NO modifica res.
    FD pagina amb limit/offset (total ~970). Aquest endpoint auto-pagina i
    retorna tot el catàleg (o filtrat). Params:
      ?prefix=A → només productes amb sku/nom que comencen per 'A' (hojas àlbum),
                  ?prefix=F → còpies Lustre, etc.
      ?q=text   → search=text a FD
      ?qs=...   → mode debug: una sola crida amb aquesta query crua (sense auto-paginar)
    Lectura acotada per pàgina (mai OOM)."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'fd_not_configured'}), 503
    prefix = (request.args.get('prefix') or '').strip().upper()
    q      = (request.args.get('q')  or '').strip()
    qs     = (request.args.get('qs') or '').strip()
    search = ('search=' + urllib_quote(q) + '&') if q else ''

    # Mode debug: una sola crida amb query crua.
    if qs:
        res = _fd_get_bounded('products?' + qs)
        if res['http_error'] is not None:
            return jsonify({'ok': False, 'reason': 'http_error', 'http_status': res['http_error'], 'head': res['head']}), 200
        if res['parsed'] is None:
            return jsonify({'ok': False, 'reason': 'truncated_or_not_json', 'head': res['head'], 'exc': res['exc']}), 200
        items = _fd_extract_contacts_list(res['parsed'])
        return jsonify({'ok': True, 'mode': 'debug', 'count': len(items),
                        'products': [_fd_product_row(p) for p in items if isinstance(p, dict)]})

    # Mode auto: paginar amb limit/offset fins esgotar el total.
    LIMIT = 200
    offset = 0
    total = None
    pagination = None
    all_rows = []
    for _ in range(60):  # seguretat: 60 × 200 = 12000 >> 970
        res = _fd_get_bounded(f'products?{search}limit={LIMIT}&offset={offset}')
        if res['http_error'] is not None:
            return jsonify({'ok': False, 'reason': 'http_error', 'http_status': res['http_error'],
                            'head': res['head'], 'offset': offset}), 200
        if res['exc'] is not None:
            return jsonify({'ok': False, 'reason': 'exception', 'error': res['exc'], 'offset': offset}), 200
        parsed = res['parsed']
        if parsed is None:
            return jsonify({'ok': False, 'reason': 'truncated_or_not_json',
                            'truncated': res['truncated'], 'head': res['head'], 'offset': offset}), 200
        if pagination is None and isinstance(parsed, dict):
            pagination = parsed.get('pagination')
        items = _fd_extract_contacts_list(parsed)
        if not items:
            break
        for p in items:
            if isinstance(p, dict):
                all_rows.append(_fd_product_row(p))
        if isinstance(parsed, dict):
            total = (parsed.get('pagination') or {}).get('total', total)
        offset += len(items)
        if total is not None and offset >= total:
            break
        if len(items) < LIMIT:
            break

    rows = all_rows
    if prefix:
        rows = [r for r in all_rows
                if str(r.get('sku') or '').upper().startswith(prefix)
                or str(r.get('name') or '').upper().startswith(prefix)]
    return jsonify({'ok': True, 'total_fd': total, 'fetched': len(all_rows),
                    'returned': len(rows), 'prefix': prefix,
                    'pagination_first_page': pagination, 'products': rows})


@app.route('/admin/fd/product-write-test')
@admin_required
def admin_fd_product_write_test():
    """PROVA (no-op) per descobrir com s'actualitza un producte a FD. Llegeix el
    producte i el re-escriu IDÈNTIC (mateix contingut → no canvia res), provant
    el mètode indicat. Serveix per confirmar endpoint/mètode abans de cap push.
      ?id=pro_...   (per defecte A1313)   ?method=PUT|PATCH|POST  (per defecte PUT)
    Retorna l'estat de l'escriptura (status/http_error/resposta)."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'fd_not_configured'}), 503
    pid = (request.args.get('id') or 'pro_0fb33689-ca29-409a-bf7b-324ecdb9ab5b').strip()  # A1313
    method = (request.args.get('method') or 'PUT').strip().upper()
    if method not in ('PUT', 'PATCH', 'POST'):
        method = 'PUT'
    # 1) Llegim el producte sencer.
    g = _fd_get_bounded(f'products/{pid}')
    if g['http_error'] is not None:
        return jsonify({'ok': False, 'step': 'get', 'http_status': g['http_error'],
                        'head': g['head'], 'pid': pid}), 200
    if g['parsed'] is None:
        return jsonify({'ok': False, 'step': 'get', 'reason': 'no_json',
                        'head': g['head'], 'pid': pid}), 200
    obj = g['parsed']
    content = obj.get('content') if isinstance(obj, dict) else None
    if not content:
        return jsonify({'ok': False, 'step': 'get', 'reason': 'no_content',
                        'keys': list(obj.keys()) if isinstance(obj, dict) else None}), 200
    price_before = ((content.get('main') or {}).get('sales') or {}).get('price')
    # 2) Re-escrivim EXACTAMENT el mateix contingut (no-op) amb el mètode provat.
    w = _fd_write(f'products/{pid}', {'content': content}, method=method)
    return jsonify({'ok': (w['http_error'] is None and w['exc'] is None),
                    'pid': pid, 'method': method, 'price_before': price_before,
                    'write_status': w['status'], 'write_http_error': w['http_error'],
                    'write_exc': w['exc'], 'write_head': (w['head'] or '')[:800]})


# ── Sincronització de preus de fulla d'àlbum: web → FacturaDirecta ──────────
def _fetch_web_album_prices(timeout=8):
    """Llegeix els preus de fulla d'àlbum de la web (font de veritat)."""
    url = _main_site_url() + '/api/album-prices'
    try:
        req = urllib_request.Request(url, headers={
            'User-Agent': 'calculadora-marcs-bridge/1.0 (+https://calculadora.reusrevela.cat)'})
        with urllib_request.urlopen(req, timeout=timeout) as r:
            return json.loads(r.read(500_000).decode('utf-8', 'replace'))
    except Exception as e:
        return {'_error': str(e)[:200]}

def _album_size_to_sku(size_id):
    """'30x40' → 'A3040' (2 dígits per costat). None si no és WxH."""
    try:
        w, h = str(size_id).lower().replace(' ', '').split('x')
        if not (w.isdigit() and h.isdigit()):
            return None
        return 'A' + w.zfill(2) + h.zfill(2)
    except Exception:
        return None

def _fd_products_by_prefix(prefix, limit=200, max_pages=20):
    """dict {sku_upper: row} dels productes FD amb sku que comença per prefix."""
    out = {}
    offset = 0
    for _ in range(max_pages):
        res = _fd_get_bounded(f'products?limit={limit}&offset={offset}')
        if res['http_error'] is not None:
            return out, f"FD {res['http_error']}: {(res['head'] or '')[:200]}"
        if res['parsed'] is None:
            return out, 'resposta FD no JSON'
        items = _fd_extract_contacts_list(res['parsed'])
        if not items:
            break
        for p in items:
            if not isinstance(p, dict):
                continue
            row = _fd_product_row(p)
            sku = str(row.get('sku') or '')
            if sku.upper().startswith(prefix.upper()):
                out[sku.upper()] = row
        total = (res['parsed'].get('pagination') or {}).get('total') if isinstance(res['parsed'], dict) else None
        offset += len(items)
        if total is not None and offset >= total:
            break
        if len(items) < limit:
            break
    return out, None

def _album_sync_compare():
    """Compara preus de fulla d'àlbum (web) amb els productes A#### de FD.
    Retorna (rows, error). Cada row: size, label, sku, web_price, fd_price, fd_id, status."""
    web = _fetch_web_album_prices()
    if not isinstance(web, dict) or web.get('_error') or not web.get('sizes'):
        msg = web.get('_error') if isinstance(web, dict) else 'resposta inesperada'
        return None, f"No s'han pogut llegir els preus de la web: {msg}"
    fd_map, err = _fd_products_by_prefix('A')
    if err:
        return None, f"No s'han pogut llegir productes de FD: {err}"
    rows = []
    for s in web['sizes']:
        sku = _album_size_to_sku(s.get('id'))
        try:
            web_price = round(float(s.get('sheet_price')), 2)
        except Exception:
            continue
        fd = fd_map.get((sku or '').upper()) if sku else None
        if not sku:
            status, fd_price, fd_id = 'skip', None, None
        elif not fd:
            status, fd_price, fd_id = 'missing', None, None
        else:
            fd_id = fd.get('id')
            try:
                fd_price = round(float(fd.get('price')), 2)
            except Exception:
                fd_price = None
            status = 'ok' if fd_price == web_price else 'diff'
        rows.append({'size': s.get('id'), 'label': s.get('label'), 'sku': sku,
                     'web_price': web_price, 'fd_price': fd_price, 'fd_id': fd_id, 'status': status})
    return rows, None

def _album_sync_apply(rows):
    """Aplica a FD: PUT preu web a les diferències, POST crea els que falten.
    Només toca SKUs A#### derivats de les mides de la web (allow-list implícita)."""
    results = []
    for r in rows:
        if r['status'] == 'diff' and r.get('fd_id'):
            g = _fd_get_bounded(f"products/{r['fd_id']}")
            if not isinstance(g.get('parsed'), dict):
                results.append({**r, 'result': f"error_get {g.get('http_error') or g.get('exc') or '?'}"})
                continue
            content = g['parsed'].get('content') or {}
            main = content.get('main') or {}
            sales = main.get('sales') or {}
            sales['price'] = r['web_price']
            main['sales'] = sales
            content['main'] = main
            w = _fd_write(f"products/{r['fd_id']}", {'content': content}, method='PUT')
            ok = w['http_error'] is None and w['exc'] is None
            results.append({**r, 'result': 'updated' if ok else f"error_put {w['http_error'] or w['exc']}"})
        elif r['status'] == 'missing' and r.get('sku'):
            main = {'sku': r['sku'], 'name': r['sku'], 'title': r['sku'], 'currency': 'EUR',
                    'sales': {'account': '700000', 'price': r['web_price'],
                              'tax': ['S_IVA_21'], 'description': f"Fulla àlbum {r['size']}"}}
            w = _fd_write('products', {'content': {'type': 'product', 'main': main}}, method='POST')
            ok = w['http_error'] is None and w['exc'] is None
            results.append({**r, 'result': 'created' if ok else f"error_post {w['http_error'] or w['exc']}"})
    return results

def _album_sync_html(rows, results=None):
    color = {'ok': '#1A6B45', 'diff': '#C8873A', 'missing': '#B84040', 'skip': '#9E9B94'}
    trs = []
    for r in rows:
        fd = '—' if r['fd_price'] is None else f"{r['fd_price']:.2f} €"
        res_cell = ''
        if results is not None:
            match = next((x for x in results if x.get('sku') == r['sku']), None)
            res_cell = f"<td><strong>{match['result']}</strong></td>" if match else '<td>—</td>'
        trs.append(
            f"<tr><td>{r['size']}</td><td><code>{r['sku'] or '—'}</code></td>"
            f"<td style='text-align:right'>{r['web_price']:.2f} €</td>"
            f"<td style='text-align:right'>{fd}</td>"
            f"<td style='color:{color.get(r['status'],'#000')};font-weight:700'>{r['status']}</td>{res_cell}</tr>")
    n_diff = sum(1 for r in rows if r['status'] == 'diff')
    n_missing = sum(1 for r in rows if r['status'] == 'missing')
    res_head = '<th>Resultat</th>' if results is not None else ''
    btn = ''
    if results is None and (n_diff or n_missing):
        btn = (f"<form method='post' action='/admin/fd/album-sync/apply' "
               f"onsubmit=\"return confirm('Aplicar a FacturaDirecta: {n_diff} preus a actualitzar i {n_missing} productes a crear?');\">"
               f"<button type='submit' style='margin-top:1rem;padding:10px 18px;background:#1A6B45;color:#fff;border:none;border-radius:8px;font-weight:700;cursor:pointer'>"
               f"Sincronitzar a FacturaDirecta ({n_diff} canvis · {n_missing} nous)</button></form>")
    elif results is None:
        btn = "<p style='color:#1A6B45;font-weight:700;margin-top:1rem'>✔ Tot quadra, res a sincronitzar.</p>"
    title = 'Resultat de la sincronització' if results is not None else 'Sincronització fulles d\'àlbum · web → FacturaDirecta'
    return (
        "<!doctype html><meta charset='utf-8'><title>Album sync FD</title>"
        "<div style='font-family:system-ui,sans-serif;max-width:760px;margin:2rem auto;padding:0 1rem'>"
        f"<h2>{title}</h2>"
        f"<p style='color:#6B6860'>Diferències: <strong>{n_diff}</strong> · Falten a FD: <strong>{n_missing}</strong>. "
        "El preu de la web mana; el botó actualitza FD perquè coincideixi.</p>"
        "<table style='border-collapse:collapse;width:100%' border='1' cellpadding='7'>"
        f"<tr style='background:#F5F4F1'><th>Mida</th><th>SKU FD</th><th>Web</th><th>FD</th><th>Estat</th>{res_head}</tr>"
        + ''.join(trs) + "</table>" + btn + "</div>")


@app.route('/admin/fd/album-sync')
@admin_required
def admin_fd_album_sync():
    """Informe comparatiu de preus de fulla d'àlbum web ↔ FD (només lectura)."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat (variables d\'entorn)', 503
    rows, err = _album_sync_compare()
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><h2>Album sync</h2><p style='color:#B84040'>{err}</p></div>", 200
    return _album_sync_html(rows)


@app.route('/admin/fd/album-sync/apply', methods=['POST'])
@admin_required
def admin_fd_album_sync_apply():
    """Aplica la sincronització: PUT preus diferents, POST crea els que falten."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat', 503
    rows, err = _album_sync_compare()
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><p style='color:#B84040'>{err}</p></div>", 200
    results = _album_sync_apply(rows)
    return _album_sync_html(rows, results=results)


# ── Sincronització general de preus → FacturaDirecta (multi-família) ─────────
# Famílies: àlbum (origen web), vidres i miralls (origen taula `vidres` del calc).
_SYNC_FAMILIES = {
    'album':   {'titol': "Fulles d'àlbum", 'prefix': 'A',   'desc': 'Fulla àlbum', 'origen': 'web'},
    'cristal': {'titol': 'Vidres',         'prefix': 'VID', 'desc': 'Vidre',       'origen': 'calc'},
    'mirall':  {'titol': 'Miralls',        'prefix': 'MIR', 'desc': 'Mirall',      'origen': 'calc'},
}

def _sync_family_source(family):
    """Retorna (rows, error). rows: [{'label','sku','price'}] des de l'origen."""
    if family == 'album':
        web = _fetch_web_album_prices()
        if not isinstance(web, dict) or web.get('_error') or not web.get('sizes'):
            return None, f"web album: {web.get('_error') if isinstance(web, dict) else 'error'}"
        rows = []
        for s in web['sizes']:
            sku = _album_size_to_sku(s.get('id'))
            if not sku:
                continue
            try:
                rows.append({'label': s.get('label') or s.get('id'), 'sku': sku,
                             'price': round(float(s['sheet_price']), 2)})
            except Exception:
                continue
        return rows, None
    if family in ('cristal', 'mirall'):
        # IMPORTANT: el preu viu del calc NO és la columna vidres.preu, sinó el
        # que retorna calcular_cost_vidre/mirall (cost × marge_admin_vidres_pct,
        # amb fórmula). Calculem el preu real per cada mida de la tarifa.
        try:
            allrows = query('SELECT referencia, preu_cost FROM vidres') or []
        except Exception as e:
            return None, f"taula vidres: {str(e)[:150]}"
        rows = []
        for r in allrows:
            ref = (_row_get(r, 'referencia') or '').strip()
            if not ref:
                continue
            up = ref.upper()
            cost_col = _row_get(r, 'preu_cost')
            if family == 'cristal':
                if up.startswith('DV-') or up.startswith('MIR-'):
                    continue
                w, h = _parse_dims(ref)
                if not w or not h:
                    continue
                d = calcular_cost_vidre(w, h) or {}
                sku = 'VID' + ref
            else:  # mirall (preu per fórmula, sense cost editable per mida)
                if not up.startswith('MIR-'):
                    continue
                w, h = _parse_dims(ref)
                if not w or not h:
                    continue
                d = calcular_cost_mirall(w, h) or {}
                sku = 'MIR' + ref[4:]
            preu = d.get('preu')
            if preu is None:
                continue
            row = {'label': ref, 'ref': ref, 'sku': sku, 'price': round(float(preu), 2)}
            if family == 'cristal':
                try:
                    row['cost'] = round(float(cost_col), 4) if cost_col is not None else None
                except Exception:
                    row['cost'] = None
            rows.append(row)
        return rows, None
    return None, f"família desconeguda: {family}"

def _sync_compare(family):
    cfg = _SYNC_FAMILIES.get(family)
    if not cfg:
        return None, f"família desconeguda: {family}"
    src, err = _sync_family_source(family)
    if err:
        return None, err
    fd_map, err = _fd_products_by_prefix(cfg['prefix'])
    if err:
        return None, err
    rows = []
    for s in src:
        fd = fd_map.get(s['sku'].upper())
        if not fd:
            status, fd_price, fd_id = 'missing', None, None
        else:
            fd_id = fd.get('id')
            try:
                fd_price = round(float(fd.get('price')), 2)
            except Exception:
                fd_price = None
            status = 'ok' if fd_price == s['price'] else 'diff'
        rows.append({**s, 'fd_price': fd_price, 'fd_id': fd_id, 'status': status})
    return rows, None

def _sync_apply(family, rows):
    cfg = _SYNC_FAMILIES.get(family) or {}
    desc = cfg.get('desc', '')
    results = []
    for r in rows:
        if r['status'] == 'diff' and r.get('fd_id'):
            g = _fd_get_bounded(f"products/{r['fd_id']}")
            if not isinstance(g.get('parsed'), dict):
                results.append({**r, 'result': f"error_get {g.get('http_error') or g.get('exc')}"})
                continue
            content = g['parsed'].get('content') or {}
            main = content.get('main') or {}
            sales = main.get('sales') or {}
            sales['price'] = r['price']
            main['sales'] = sales
            content['main'] = main
            w = _fd_write(f"products/{r['fd_id']}", {'content': content}, method='PUT')
            ok = w['http_error'] is None and w['exc'] is None
            results.append({**r, 'result': 'updated' if ok else f"error_put {w['http_error'] or w['exc']}"})
        elif r['status'] == 'missing':
            main = {'sku': r['sku'], 'name': r['sku'], 'title': r['sku'], 'currency': 'EUR',
                    'sales': {'account': '700000', 'price': r['price'], 'tax': ['S_IVA_21'],
                              'description': f"{desc} {r['label']}"}}
            w = _fd_write('products', {'content': {'type': 'product', 'main': main}}, method='POST')
            ok = w['http_error'] is None and w['exc'] is None
            results.append({**r, 'result': 'created' if ok else f"error_post {w['http_error'] or w['exc']}"})
    return results

def _fmt_cost(c):
    if c is None:
        return ''
    s = f"{c:.4f}".rstrip('0').rstrip('.')
    return s or '0'

def _sync_html(family, rows, results=None):
    cfg = _SYNC_FAMILIES.get(family, {})
    color = {'ok': '#1A6B45', 'diff': '#C8873A', 'missing': '#B84040'}
    is_cristal = (family == 'cristal')
    editable = is_cristal and results is None  # editor de cost només a cristal (vista, no resultats)
    nav = ' · '.join(
        (f"<strong>{c['titol']}</strong>" if k == family
         else f"<a href='/admin/fd/sync?family={k}'>{c['titol']}</a>")
        for k, c in _SYNC_FAMILIES.items())
    trs = []
    for r in rows:
        fd = '—' if r['fd_price'] is None else f"{r['fd_price']:.2f} €"
        rescell = ''
        if results is not None:
            m = next((x for x in results if x.get('sku') == r['sku']), None)
            rescell = f"<td><strong>{m['result']}</strong></td>" if m else "<td>—</td>"
        costcell = ''
        if is_cristal:
            cv = _fmt_cost(r.get('cost'))
            if editable:
                costcell = (f"<td><input type='number' step='0.0001' min='0' "
                            f"name='cost_{r.get('ref', '')}' value='{cv}' style='width:88px'> €</td>")
            else:
                costcell = f"<td style='text-align:right'>{(cv + ' €') if cv else '—'}</td>"
        trs.append(
            f"<tr><td>{r['label']}</td><td><code>{r['sku']}</code></td>{costcell}"
            f"<td style='text-align:right'>{r['price']:.2f} €</td>"
            f"<td style='text-align:right'>{fd}</td>"
            f"<td style='color:{color.get(r['status'], '#000')};font-weight:700'>{r['status']}</td>{rescell}</tr>")
    n_diff = sum(1 for r in rows if r['status'] == 'diff')
    n_missing = sum(1 for r in rows if r['status'] == 'missing')
    reshead = '<th>Resultat</th>' if results is not None else ''
    costhead = '<th>Cost</th>' if is_cristal else ''
    btnstyle = ("margin-top:1rem;padding:10px 18px;background:#1A6B45;color:#fff;"
                "border:none;border-radius:8px;font-weight:700;cursor:pointer")
    table_html = (
        "<table style='border-collapse:collapse;width:100%' border='1' cellpadding='7'>"
        f"<tr style='background:#F5F4F1'><th>Mida</th><th>SKU FD</th>{costhead}"
        f"<th>Preu</th><th>FD</th><th>Estat</th>{reshead}</tr>"
        + ''.join(trs) + "</table>")
    if editable:
        intro = ("Edita el <strong>cost</strong> de cada mida; el preu = cost × marge i es desa "
                 "(amb historial) i es sincronitza a FacturaDirecta.")
        body = (f"<form method='post' action='/admin/fd/sync/save-cost?family=cristal' "
                f"onsubmit=\"return confirm('Desar els costos editats i sincronitzar a FacturaDirecta?');\">"
                + table_html
                + f"<button type='submit' style='{btnstyle}'>Desar costos i sincronitzar "
                f"({n_diff} dif · {n_missing} nous)</button></form>")
    else:
        intro = "El preu del calculador/web mana; el botó actualitza FD perquè coincideixi."
        btn = ''
        if results is None and (n_diff or n_missing):
            btn = (f"<form method='post' action='/admin/fd/sync/apply?family={family}' "
                   f"onsubmit=\"return confirm('Aplicar a FacturaDirecta ({cfg.get('titol','')}): "
                   f"{n_diff} preus a actualitzar i {n_missing} productes a crear?');\">"
                   f"<button type='submit' style='{btnstyle}'>"
                   f"Sincronitzar {cfg.get('titol','')} ({n_diff} canvis · {n_missing} nous)</button></form>")
        elif results is None:
            btn = "<p style='color:#1A6B45;font-weight:700;margin-top:1rem'>✔ Tot quadra, res a sincronitzar.</p>"
        body = table_html + btn
    return (
        "<!doctype html><meta charset='utf-8'><title>Sync preus FD</title>"
        "<div style='font-family:system-ui,sans-serif;max-width:880px;margin:2rem auto;padding:0 1rem'>"
        "<h2>Sincronització de preus → FacturaDirecta</h2>"
        f"<p style='font-size:15px'>{nav}</p><h3>{cfg.get('titol', family)}</h3>"
        f"<p style='color:#6B6860'>Diferències: <strong>{n_diff}</strong> · Falten a FD: <strong>{n_missing}</strong>. {intro}</p>"
        + body + "</div>")


@app.route('/admin/fd/sync')
@admin_required
def admin_fd_sync():
    """Eina general de sincronització de preus → FD (àlbum, vidres, miralls)."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat (variables d\'entorn)', 503
    family = (request.args.get('family') or 'album').strip().lower()
    if family not in _SYNC_FAMILIES:
        family = 'album'
    rows, err = _sync_compare(family)
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><h2>Sync FD · {family}</h2><p style='color:#B84040'>{err}</p></div>", 200
    return _sync_html(family, rows)


@app.route('/admin/fd/sync/apply', methods=['POST'])
@admin_required
def admin_fd_sync_apply():
    """Aplica la sincronització d'una família: PUT diferències, POST crea faltants."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat', 503
    family = (request.args.get('family') or '').strip().lower()
    if family not in _SYNC_FAMILIES:
        return 'família desconeguda', 400
    rows, err = _sync_compare(family)
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><p style='color:#B84040'>{err}</p></div>", 200
    results = _sync_apply(family, rows)
    return _sync_html(family, rows, results=results)


def _update_vidre_cost(ref, new_cost, notes='Editat des de sync FD'):
    """Actualitza el preu_cost d'una fila de vidres (amb historial), com fa
    /admin/preus-cost/update. Recalcula també la columna preu (cost × marge)."""
    row = query('SELECT preu_cost FROM vidres WHERE referencia=?', [ref], one=True)
    if not row:
        return False
    ant = _row_get(row, 'preu_cost')
    avui = datetime.now().strftime('%Y-%m-%d')
    execute('''UPDATE vidres SET preu_cost_ant=?, preu_cost=?, data_cost=?,
               usuari_cost_id=?, cost_verificat=1, notes_cost=? WHERE referencia=?''',
            [ant, new_cost, avui, session.get('user_id'), notes, ref])
    pvd = calcular_pvd(new_cost, 'vidres')
    execute('UPDATE vidres SET preu=? WHERE referencia=?', [pvd, ref])
    execute('''INSERT INTO historial_preus_cost
               (taula, referencia, preu_cost_antic, preu_cost_nou, usuari_id, data, notes)
               VALUES (?,?,?,?,?,?,?)''',
            ['vidres', ref, ant, new_cost, session.get('user_id'), avui, notes])
    return True


@app.route('/admin/fd/sync/save-cost', methods=['POST'])
@admin_required
def admin_fd_sync_save_cost():
    """Editor de cost de vidre: desa els preu_cost editats (amb historial) i
    després sincronitza els preus calculats a FD. Només per a la família cristal."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat', 503
    family = (request.args.get('family') or '').strip().lower()
    if family != 'cristal':
        return 'Editor de cost només disponible per a vidres', 400
    try:
        refs = query("SELECT referencia, preu_cost FROM vidres "
                     "WHERE UPPER(referencia) NOT LIKE 'DV-%' "
                     "AND UPPER(referencia) NOT LIKE 'MIR-%'") or []
    except Exception as e:
        return f"<p style='color:#B84040'>Error vidres: {str(e)[:150]}</p>", 500
    canvis = 0
    for r in refs:
        ref = (_row_get(r, 'referencia') or '').strip()
        if not ref:
            continue
        raw = request.form.get('cost_' + ref)
        if raw is None or str(raw).strip() == '':
            continue
        try:
            nou = round(float(str(raw).replace(',', '.')), 4)
        except Exception:
            continue
        if nou < 0:
            continue
        ant = _row_get(r, 'preu_cost')
        try:
            antf = round(float(ant), 4) if ant is not None else None
        except Exception:
            antf = None
        if antf is not None and abs(antf - nou) < 0.00005:
            continue  # sense canvi
        if _update_vidre_cost(ref, nou):
            canvis += 1
    rows, err = _sync_compare('cristal')
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><p style='color:#B84040'>{err}</p></div>", 200
    results = _sync_apply('cristal', rows)
    return _sync_html('cristal', rows, results=results)


# ── Creació de productes d'orla a FacturaDirecta (idempotent) ────────────────
def _orles_fd_catalog():
    """Productes d'orla a crear a FD (un per mida+tram). Tarifa Canva pàg.17."""
    tier_labels = ['+50', '+100', '+150']
    prods = [
        ("Orla Lustre 30x40", "ORLA LUS 30x40", {'+50': 2.70, '+100': 2.55, '+150': 2.40}),
        ("Orla Lustre 30x45", "ORLA LUS 30x45", {'+50': 3.24, '+100': 3.06, '+150': 2.88}),
        ("Orla Lustre 40x50", "ORLA LUS 40x50", {'+50': 5.40, '+100': 5.10, '+150': 4.80}),
        ("Orla Lustre 50x60", "ORLA LUS 50x60", {'+50': 7.65, '+100': 7.22, '+150': 6.80}),
        ("Orla Offset 30x40/32x45 SRA3", "ORLA OFF 30x40", {'+50': 1.73, '+100': 1.64, '+150': 1.55}),
        ("Carnet orla 10x15 (8 DNI o 4 DNI+cartera)", "ORLA CARNET 10x15", {'+50': 0.48, '+100': 0.43, '+150': 0.38}),
        ("Carnet orla 15x20 (8 DNI+cartera)", "ORLA CARNET 15x20", {'+50': 0.73, '+100': 0.67, '+150': 0.64}),
    ]
    items = []
    for name, sku, prices in prods:
        for tl in tier_labels:
            items.append({'sku': f"{sku} {tl}", 'name': f"{name} ({tl})", 'price': prices[tl]})
    items.append({'sku': 'ORLA MONTATGE', 'name': 'Montatge orla (per alumne)', 'price': 1.00})
    items.append({'sku': 'ORLA PRIMER DISSENY', 'name': 'Primer disseny orla', 'price': 25.00})
    return items


def _orles_create_html(catalog, results=None):
    n_falten = sum(1 for it in catalog if not it.get('exists'))
    trs = []
    for it in catalog:
        if results is not None:
            m = next((x for x in results if x['sku'] == it['sku']), None)
            estat = m['result'] if m else '—'
        else:
            estat = 'ja existeix' if it.get('exists') else 'a crear'
        col = '#1A6B45' if estat in ('ja existeix', 'creat') else ('#B84040' if estat.startswith('error') else '#C8873A')
        trs.append(f"<tr><td><code>{it['sku']}</code></td><td>{it['name']}</td>"
                   f"<td style='text-align:right'>{it['price']:.2f} €</td>"
                   f"<td style='color:{col};font-weight:700'>{estat}</td></tr>")
    btnstyle = ("margin-top:1rem;padding:10px 18px;background:#1A6B45;color:#fff;"
                "border:none;border-radius:8px;font-weight:700;cursor:pointer")
    if results is None and n_falten:
        btn = (f"<form method='post' onsubmit=\"return confirm('Crear {n_falten} productes d\\'orla a FacturaDirecta?');\">"
               f"<button type='submit' style='{btnstyle}'>Crear els {n_falten} que falten</button></form>")
    elif results is None:
        btn = "<p style='color:#1A6B45;font-weight:700;margin-top:1rem'>✔ Ja existeixen tots.</p>"
    else:
        btn = "<p style='margin-top:1rem'><a href='/admin/fd/orles-create'>↻ Refrescar</a></p>"
    return ("<!doctype html><meta charset='utf-8'><title>Crear orles FD</title>"
            "<div style='font-family:system-ui,sans-serif;max-width:780px;margin:2rem auto;padding:0 1rem'>"
            "<h2>Crear productes d'orla a FacturaDirecta</h2>"
            f"<p style='color:#6B6860'>Total: {len(catalog)} · A crear: <strong>{n_falten}</strong>. "
            "Compte 700000 · IVA 21%. Idempotent: només crea els SKU que no existeixin.</p>"
            "<table style='border-collapse:collapse;width:100%' border='1' cellpadding='7'>"
            "<tr style='background:#F5F4F1'><th>SKU</th><th>Nom</th><th>Preu</th><th>Estat</th></tr>"
            + ''.join(trs) + "</table>" + btn + "</div>")


@app.route('/admin/fd/orles-create', methods=['GET', 'POST'])
@admin_required
def admin_fd_orles_create():
    """Crea (idempotentment) els productes d'orla a FD. GET = previsualització,
    POST = crea els que falten."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return 'FacturaDirecta no configurat', 503
    catalog = _orles_fd_catalog()
    existing, err = _fd_products_by_prefix('ORLA')
    if err:
        return f"<div style='font-family:sans-serif;margin:2rem'><p style='color:#B84040'>{err}</p></div>", 200
    existing_skus = set(existing.keys())  # ja en majúscules
    for it in catalog:
        it['exists'] = it['sku'].upper() in existing_skus
    if request.method == 'GET':
        return _orles_create_html(catalog)
    # POST: crear els que falten
    results = []
    for it in catalog:
        if it['exists']:
            results.append({**it, 'result': 'ja existeix'})
            continue
        main = {'sku': it['sku'], 'name': it['name'], 'title': it['sku'], 'currency': 'EUR',
                'sales': {'account': '700000', 'price': it['price'], 'tax': ['S_IVA_21'],
                          'description': it['name']}}
        w = _fd_write('products', {'content': {'type': 'product', 'main': main}}, method='POST')
        ok = w['http_error'] is None and w['exc'] is None
        results.append({**it, 'result': 'creat' if ok else f"error {w['http_error'] or w['exc']}"})
    return _orles_create_html(catalog, results=results)


@app.route('/admin/clients-externs')
@admin_required
def admin_clients_externs():
    """Llistat de clients habituals (taller + pvp)."""
    _ensure_recarrec_column()
    _ensure_nom_comercial_column()
    filtre = (request.args.get('tipus') or '').strip().lower()
    if filtre in ('taller', 'pvp'):
        clients = query("""
            SELECT c.id, c.nom, c.nom_comercial, c.nif, c.fd_contact_id, c.tipus, c.telefon, c.email,
                   c.actiu, c.created_at, c.usuari_id, c.dropbox_url, c.recarrec_equiv,
                   u.username AS usuari_username, u.nom AS usuari_nom
            FROM clients_externs c
            LEFT JOIN usuaris u ON c.usuari_id = u.id
            WHERE c.tipus = ? ORDER BY c.actiu DESC, c.nom
        """, [filtre]) or []
    else:
        filtre = ''
        clients = query("""
            SELECT c.id, c.nom, c.nom_comercial, c.nif, c.fd_contact_id, c.tipus, c.telefon, c.email,
                   c.actiu, c.created_at, c.usuari_id, c.dropbox_url, c.recarrec_equiv,
                   u.username AS usuari_username, u.nom AS usuari_nom
            FROM clients_externs c
            LEFT JOIN usuaris u ON c.usuari_id = u.id
            ORDER BY c.actiu DESC, c.nom
        """) or []
    usuaris = query('SELECT id, username, nom, nom_empresa FROM usuaris ORDER BY nom') or []
    return render_template('admin_clients_externs.html', clients=clients, filtre=filtre, usuaris=usuaris)


@app.route('/admin/clients-externs/import-fd', methods=['POST'])
@admin_required
def admin_clients_externs_import_fd():
    """Importa un contacte de FD com a client habitual (només referència local)."""
    payload = request.get_json(silent=True) or request.form
    fd_id = (payload.get('fd_contact_id') or '').strip()
    nom   = (payload.get('nom') or '').strip()
    nif   = (payload.get('nif') or '').strip()
    tipus = (payload.get('tipus') or 'pvp').strip().lower()
    if tipus not in ('taller', 'pvp'):
        tipus = 'pvp'

    if not fd_id or not nom:
        return jsonify({'ok': False, 'error': 'missing_data'}), 400

    existing = query(
        'SELECT id FROM clients_externs WHERE fd_contact_id = ?',
        [fd_id], one=True,
    )
    if existing:
        return jsonify({'ok': False, 'error': 'already_exists',
                        'id': _row_get(existing, 'id')}), 409

    telefon = (payload.get('telefon') or '').strip() or None
    email = (payload.get('email') or '').strip() or None
    new_id = execute(
        "INSERT INTO clients_externs (nom, nif, fd_contact_id, tipus, telefon, email, actiu) "
        "VALUES (?, ?, ?, ?, ?, ?, ?)",
        [nom, nif or None, fd_id, tipus, telefon, email, True],
    )
    print(f"[clients_externs] import: id={new_id} fd_id={fd_id} nom={nom} tipus={tipus}")
    return jsonify({'ok': True, 'id': new_id})


@app.route('/admin/clients-externs/<int:client_id>/toggle', methods=['POST'])
@admin_required
def admin_clients_externs_toggle(client_id):
    """Toggle actiu/desactiu d'un client extern."""
    row = query('SELECT actiu FROM clients_externs WHERE id=?', [client_id], one=True)
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    actual = bool(_row_get(row, 'actiu', True))
    nou = not actual
    execute('UPDATE clients_externs SET actiu=? WHERE id=?', [nou, client_id])
    print(f"[clients_externs] toggle: id={client_id} actiu={nou}")
    return jsonify({'ok': True, 'actiu': nou})


@app.route('/admin/clients-externs/crear', methods=['POST'])
@admin_required
def admin_clients_externs_crear():
    """Crea un client habitual manualment (sense FD)."""
    payload = request.get_json(silent=True) or request.form
    nom = (payload.get('nom') or '').strip()
    if not nom:
        return jsonify({'ok': False, 'error': 'El nom és obligatori'}), 400
    nom_comercial = (payload.get('nom_comercial') or '').strip() or None
    nif = (payload.get('nif') or '').strip() or None
    tipus = (payload.get('tipus') or 'pvp').strip().lower()
    if tipus not in ('taller', 'pvp'):
        tipus = 'pvp'
    telefon = (payload.get('telefon') or '').strip() or None
    email = (payload.get('email') or '').strip() or None
    fd_id = (payload.get('fd_contact_id') or '').strip() or None

    usuari_id = payload.get('usuari_id')
    if usuari_id:
        try: usuari_id = int(usuari_id)
        except (TypeError, ValueError): usuari_id = None
    else:
        usuari_id = None
    dropbox_url = (payload.get('dropbox_url') or '').strip() or None
    recarrec = _parse_bool(payload.get('recarrec_equiv'))
    _ensure_recarrec_column()
    _ensure_nom_comercial_column()
    new_id = execute(
        "INSERT INTO clients_externs (nom, nom_comercial, nif, fd_contact_id, tipus, telefon, email, usuari_id, dropbox_url, recarrec_equiv, actiu) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [nom, nom_comercial, nif, fd_id, tipus, telefon, email, usuari_id, dropbox_url, recarrec, True],
    )
    print(f"[clients_externs] crear: id={new_id} nom={nom} tipus={tipus} usuari_id={usuari_id}")
    return jsonify({'ok': True, 'id': new_id})


@app.route('/admin/clients-externs/<int:client_id>/editar', methods=['POST'])
@admin_required
def admin_clients_externs_editar(client_id):
    """Edita un client habitual existent."""
    row = query('SELECT id FROM clients_externs WHERE id=?', [client_id], one=True)
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    payload = request.get_json(silent=True) or request.form
    nom = (payload.get('nom') or '').strip()
    if not nom:
        return jsonify({'ok': False, 'error': 'El nom és obligatori'}), 400
    nom_comercial = (payload.get('nom_comercial') or '').strip() or None
    nif = (payload.get('nif') or '').strip() or None
    tipus = (payload.get('tipus') or 'pvp').strip().lower()
    if tipus not in ('taller', 'pvp'):
        tipus = 'pvp'
    telefon = (payload.get('telefon') or '').strip() or None
    email = (payload.get('email') or '').strip() or None
    usuari_id = payload.get('usuari_id')
    if usuari_id:
        try: usuari_id = int(usuari_id)
        except (TypeError, ValueError): usuari_id = None
    else:
        usuari_id = None
    dropbox_url = (payload.get('dropbox_url') or '').strip() or None
    recarrec = _parse_bool(payload.get('recarrec_equiv'))
    _ensure_recarrec_column()
    _ensure_nom_comercial_column()
    execute(
        "UPDATE clients_externs SET nom=?, nom_comercial=?, nif=?, tipus=?, telefon=?, email=?, usuari_id=?, dropbox_url=?, recarrec_equiv=? WHERE id=?",
        [nom, nom_comercial, nif, tipus, telefon, email, usuari_id, dropbox_url, recarrec, client_id],
    )
    print(f"[clients_externs] editar: id={client_id} nom={nom} tipus={tipus} usuari_id={usuari_id}")
    return jsonify({'ok': True})


@app.route('/admin/clients-externs/<int:client_id>/eliminar', methods=['POST'])
@admin_required
def admin_clients_externs_eliminar(client_id):
    """Elimina un client habitual (si no té comandes vinculades)."""
    linked = query('SELECT id FROM comandes WHERE client_extern_id=? LIMIT 1', [client_id], one=True)
    if linked:
        return jsonify({'ok': False, 'error': 'Aquest client té comandes vinculades. Desactiva\'l en lloc d\'eliminar-lo.'}), 409
    execute('DELETE FROM clients_externs WHERE id=?', [client_id])
    print(f"[clients_externs] eliminar: id={client_id}")
    return jsonify({'ok': True})


@app.route('/api/clients-externs')
@login_required
def api_clients_externs():
    """Llista de clients externs actius — alimenta el cercador del pressupost.
    Els NO-admin no gestionen clients tipus 'taller' (PVD), així que només se'ls
    retornen els clients PVP."""
    _ensure_recarrec_column()
    _ensure_nom_comercial_column()
    where_extra = '' if session.get('is_admin') else " AND c.tipus <> 'taller'"
    sql_re = ("""
        SELECT c.id, c.nom, c.nom_comercial, c.nif, c.tipus, c.telefon, c.email, c.usuari_id,
               c.recarrec_equiv, u.nom_empresa AS empresa
        FROM clients_externs c
        LEFT JOIN usuaris u ON c.usuari_id = u.id
        WHERE c.actiu = TRUE""" + where_extra + " ORDER BY c.nom")
    sql_plain = ("""
        SELECT c.id, c.nom, c.nom_comercial, c.nif, c.tipus, c.telefon, c.email, c.usuari_id,
               u.nom_empresa AS empresa
        FROM clients_externs c
        LEFT JOIN usuaris u ON c.usuari_id = u.id
        WHERE c.actiu = TRUE""" + where_extra + " ORDER BY c.nom")
    try:
        rows = query(sql_re) or []
    except Exception as e:
        # La columna recarrec_equiv encara no existeix a la BD: no trenquem el
        # cercador de clients; tornem sense el camp (default False).
        print(f"[clients_externs] recarrec_equiv absent, fallback: {e}")
        rows = query(sql_plain) or []
    return jsonify({
        'ok': True,
        'clients': [
            {
                'id': _row_get(r, 'id'),
                'nom': _row_get(r, 'nom') or '',
                'nom_comercial': _row_get(r, 'nom_comercial') or '',
                'nif': _row_get(r, 'nif') or '',
                'tipus': _row_get(r, 'tipus') or 'pvp',
                'telefon': _row_get(r, 'telefon') or '',
                'email': _row_get(r, 'email') or '',
                'usuari_id': _row_get(r, 'usuari_id'),
                'recarrec_equiv': bool(_row_get(r, 'recarrec_equiv', False)),
                'empresa': _row_get(r, 'empresa') or '',
            }
            for r in rows
        ],
    })


def _resolve_client_extern_fd_id(client_extern_id):
    """Si client_extern_id apunta a un client actiu amb fd_contact_id,
    retorna (contact_id, nom). Si està desactivat o no existeix, retorna
    (None, None) — el caller fa fallback al flux antic."""
    if not client_extern_id:
        return None, None
    row = query(
        'SELECT nom, nif, fd_contact_id, actiu FROM clients_externs WHERE id=?',
        [client_extern_id], one=True,
    )
    if not row:
        return None, None
    if not _row_get(row, 'actiu', False):
        print(f"[clients_externs] WARN: client_extern_id={client_extern_id} desactivat — fallback al flux antic")
        return None, None
    return _row_get(row, 'fd_contact_id') or None, _row_get(row, 'nom') or None


def _parse_bool(v):
    """Interpreta un valor divers (checkbox, JSON, form) com a booleà."""
    if isinstance(v, bool):
        return v
    return str(v or '').strip().lower() in ('1', 'true', 'yes', 'on', 't')


_nom_comercial_col_ready = False
def _ensure_nom_comercial_column():
    """Assegura que existeix clients_externs.nom_comercial (nom del comerç, per
    reconèixer el client a les llistes internes). Idempotent i cached."""
    global _nom_comercial_col_ready
    if _nom_comercial_col_ready:
        return
    try:
        execute("ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS nom_comercial VARCHAR(255)")
        _nom_comercial_col_ready = True
    except Exception as e:
        print(f"[nom_comercial] ensure column skip: {e}")


_recarrec_col_ready = False
def _ensure_recarrec_column():
    """Assegura que existeix clients_externs.recarrec_equiv. Les migracions no
    sempre s'han executat a l'arrencada (com passa amb altres columnes), i
    referenciar-la abans peta la consulta. Idempotent i cached per procés."""
    global _recarrec_col_ready
    if _recarrec_col_ready:
        return
    try:
        execute("ALTER TABLE clients_externs ADD COLUMN IF NOT EXISTS recarrec_equiv BOOLEAN DEFAULT FALSE")
        _recarrec_col_ready = True
    except Exception as e:
        print(f"[recarrec] ensure column skip: {e}")


def _client_extern_recarrec(client_extern_id):
    """True si el client habitual està en règim de recàrrec d'equivalència.
    Tolerant si la columna encara no s'ha migrat (retorna False)."""
    if not client_extern_id:
        return False
    try:
        row = query('SELECT recarrec_equiv FROM clients_externs WHERE id=?', [client_extern_id], one=True)
    except Exception:
        return False
    return bool(_row_get(row, 'recarrec_equiv', False)) if row else False


# ── Enviament a client final (tarifes NACEX, vàlides fins 31/12/2026) ────────
# Preus NETS (sense IVA). Origen: Reus (Tarragona). El client final paga el
# net × (1 + marge_enviament_pct) + 21% IVA. Zona deduïda pel codi postal.
ENVIAMENT_TARIFES = {
    'pluspack': {
        'nom': 'Pluspack (econòmic, 1-2 dies)',
        'max_kg': 20,
        'max_sum_cm': 150,
        'zones': {
            # (kg_màx_del_tram, preu_net)
            'provincial': [(2, 7.10), (5, 8.41), (10, 9.00), (15, 9.73), (20, 10.46)],
            'regional':   [(2, 7.31), (5, 8.64), (10, 9.22), (15, 9.95), (20, 10.70)],
            'nacional':   [(2, 7.83), (5, 9.51), (10, 11.21), (15, 11.92), (20, 12.64)],
            'portugal':   [(2, 7.83), (5, 9.51), (10, 11.21), (15, 11.92), (20, 12.64)],
        },
    },
    'peninsular': {
        'nom': 'Peninsular (urgent, dia següent 19h)',
        'max_kg': 40,
        'max_sum_cm': 100,
        'zones': {
            'provincial': [(2, 9.11), (5, 10.21), (10, 12.87), (15, 15.78), (20, 18.68), (25, 21.59), (30, 24.49), (35, 27.39), (40, 30.29)],
            'regional':   [(2, 10.76), (5, 13.26), (10, 17.85), (15, 23.33), (20, 28.82), (25, 34.29), (30, 39.76), (35, 45.26), (40, 50.75)],
            'nacional':   [(2, 11.49), (5, 14.50), (10, 20.52), (15, 26.52), (20, 32.51), (25, 38.52), (30, 44.52), (35, 50.52), (40, 56.50)],
        },
        # Fracció addicional per cada 5 kg que superi el màxim (només peninsular)
        'fraccio_5kg': {'provincial': 2.90, 'regional': 5.49, 'nacional': 6.01},
    },
}

# Codis de província (2 dígits del CP) per a la deducció de zona des de Reus.
_ENV_CP_REGIONAL = {'08', '17', '25'}              # Barcelona, Girona, Lleida (resta Catalunya)
_ENV_CP_NO_PENINSULAR = {'07', '35', '38', '51', '52'}  # Balears, Las Palmas, Tenerife, Ceuta, Melilla

# Tarifa de manipulació/embalatge (ingrés del taller), per tram de pes (≈ volum).
# Es suma al cost net del transport, abans d'IVA.
ENVIAMENT_MANIPULACIO = [
    (1, 1.00),     # ≤1 kg — sobre / fotos petites
    (5, 2.50),     # 1-5 kg — paquet petit
    (15, 5.00),    # 5-15 kg — voluminós
    (float('inf'), 8.00),  # >15 kg — gran / fràgil (marcs amb vidre)
]


def _enviament_manipulacio(pes_kg):
    try:
        pes = float(pes_kg or 0)
    except (TypeError, ValueError):
        pes = 0.0
    if pes <= 0:
        pes = 0.001
    for upper, fee in ENVIAMENT_MANIPULACIO:
        if pes <= upper:
            return fee
    return ENVIAMENT_MANIPULACIO[-1][1]


def _enviament_zona_from_cp(cp, pais='ES'):
    """Retorna 'provincial'|'regional'|'nacional'|'portugal'|'no_cobert'|None."""
    pais = (pais or 'ES').strip().upper()
    if pais in ('PT', 'PORTUGAL'):
        return 'portugal'
    if pais in ('AD', 'ANDORRA'):
        return 'nacional'  # "Nacional Peninsular + Andorra"
    digits = re.sub(r'\D', '', str(cp or ''))
    if len(digits) < 2:
        return None
    pref = digits[:2]
    if pref in _ENV_CP_NO_PENINSULAR:
        return 'no_cobert'
    if pref == '43':
        return 'provincial'
    if pref in _ENV_CP_REGIONAL:
        return 'regional'
    try:
        n = int(pref)
    except ValueError:
        return None
    return 'nacional' if 1 <= n <= 52 else None


def _enviament_preu_net(tarifa_key, zona, pes_kg):
    """Preu net del tram segons pes. Si supera el màxim, aplica fracció +5kg
    (peninsular) o retorna None (pluspack no admet excés)."""
    t = ENVIAMENT_TARIFES.get(tarifa_key)
    if not t or zona not in t['zones']:
        return None
    try:
        pes = float(pes_kg or 0)
    except (TypeError, ValueError):
        pes = 0.0
    if pes <= 0:
        pes = 0.001
    brackets = t['zones'][zona]
    if pes <= t['max_kg']:
        for upper, price in brackets:
            if pes <= upper:
                return price
        return brackets[-1][1]
    fr = (t.get('fraccio_5kg') or {}).get(zona)
    if fr is None:
        return None
    extra = pes - t['max_kg']
    n = math.ceil(extra / 5.0)
    return round(brackets[-1][1] + n * fr, 2)


def calcular_enviament(tarifa_key, cp, pes_kg, pais='ES', sum_cm=None):
    """Calcula el cost d'enviament a client final. Retorna dict amb ok/error."""
    tarifa_key = (tarifa_key or 'pluspack').strip().lower()
    t = ENVIAMENT_TARIFES.get(tarifa_key)
    if not t:
        return {'ok': False, 'error': 'Tarifa d\'enviament no vàlida.'}
    zona = _enviament_zona_from_cp(cp, pais)
    if zona == 'no_cobert':
        return {'ok': False, 'error': 'Destí no cobert per aquesta tarifa terrestre (Balears/Canàries/Ceuta/Melilla). Cal enviament especial.'}
    if not zona:
        return {'ok': False, 'error': 'Codi postal o país no vàlid.'}
    if zona not in t['zones']:
        return {'ok': False, 'error': f"La tarifa «{t['nom']}» no cobreix aquest destí."}
    net = _enviament_preu_net(tarifa_key, zona, pes_kg)
    if net is None:
        return {'ok': False, 'error': f"Pes fora de rang per a «{t['nom']}» (màxim {t['max_kg']} kg)."}
    avisos = []
    try:
        sc = float(sum_cm) if sum_cm not in (None, '') else None
    except (TypeError, ValueError):
        sc = None
    if sc and sc > t['max_sum_cm']:
        avisos.append(f"La suma ample+llarg+alt ({sc:.0f} cm) supera el màxim de {t['max_sum_cm']} cm; pot comportar recàrrec per excés de mides.")
    manip = _enviament_manipulacio(pes_kg)
    base = round(net + manip, 2)        # transport net + manipulació (sense IVA)
    iva = round(base * 0.21, 2)
    total = round(base + iva, 2)
    return {
        'ok': True, 'zona': zona, 'tarifa': tarifa_key, 'tarifa_nom': t['nom'],
        'net': round(net, 2), 'manipulacio': round(manip, 2), 'base': base, 'iva': iva, 'total': total,
        'avisos': avisos,
    }


def estimar_pes_marc(w_cm, h_cm, te_vidre=True, doble_vidre=False):
    """Estimació orientativa del pes d'un marc embalat (kg), editable per l'usuari."""
    try:
        area = (float(w_cm) / 100.0) * (float(h_cm) / 100.0)
    except (TypeError, ValueError):
        return 0.0
    if area <= 0:
        return 0.0
    vidre = (10.0 if doble_vidre else 5.0) if te_vidre else 0.0  # kg/m²
    estructura = 6.0  # motllura + fons + embalatge (kg/m² aprox)
    return round(area * (vidre + estructura) + 0.5, 1)


@app.route('/api/calcular-enviament', methods=['POST'])
@login_required
def api_calcular_enviament():
    d = request.get_json(force=True) or {}
    res = calcular_enviament(
        d.get('tarifa') or 'pluspack',
        d.get('cp') or '',
        d.get('pes') or 0,
        pais=(d.get('pais') or 'ES'),
        sum_cm=d.get('sum_cm'),
    )
    return jsonify(res)


@app.route('/api/crear-albara', methods=['POST'])
@login_required
def api_crear_albara():
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'Factura Directa no configurat (variables d\'entorn)'}), 503

    d = request.get_json(force=True) or {}
    client_nom   = (d.get('client_nom') or '').strip()
    client_tel   = (d.get('client_tel') or '').strip()
    cost_prod    = float(d.get('cost_produccio') or 0)
    preu_net     = float(d.get('preu_net') or 0)
    preu_final   = float(d.get('preu_final') or 0)
    marc         = (d.get('marc') or '').strip()
    opcions_text = (d.get('opcions') or '').strip()
    observacions = (d.get('observacions') or '').strip()
    num_pressupost = (d.get('num_pressupost') or '').strip()
    quantitat    = float(d.get('quantitat') or 1)
    revers_peu   = bool(d.get('revers_peu'))
    mode_preu    = (d.get('mode_preu') or 'pvp').strip().lower()  # 'pvp' or 'cost'

    # Només l'admin (Reus Revela) pot crear albarans a FD
    user = query('SELECT nom, nom_empresa, nom_fiscal, fiscal_id, is_admin FROM usuaris WHERE id=?',
                 [session['user_id']], one=True)
    is_admin = bool(_row_get(user, 'is_admin', 0))
    if not is_admin:
        return jsonify({'ok': False, 'error': 'Només l\'administrador pot crear albarans a Factura Directa.'}), 403

    # El contacte FD és el CLIENT (nom del client de la calculadora)
    # El NIF del client és opcional — si s'envia, s'usa per cercar exacte a FD
    client_extern_id = (d.get('client_extern_id') or '').strip() or None
    nif_client = (d.get('client_nif') or '').strip()
    nom_fd     = client_nom
    recarrec   = _client_extern_recarrec(client_extern_id)

    # 1) PRIORITAT: si s'ha triat un client habitual (importat) ja enllaçat amb un
    #    contacte real de Factura Directa, fem servir aquest contacte directament.
    #    Així l'albarà va a l'"usuari real" de FD i no creem un duplicat per nom.
    contact_id = None
    fd_id, fd_nom = _resolve_client_extern_fd_id(client_extern_id)
    if fd_id:
        contact_id = fd_id
        if fd_nom:
            nom_fd = fd_nom

    # 2) Fallback (sense client habitual enllaçat): cercar per NIF o crear nou.
    if not contact_id:
        if not nom_fd:
            return jsonify({'ok': False, 'error': 'Cal omplir el nom del client abans de crear l\'albarà.'}), 400

        # Buscar per NIF (exacte) o crear nou — no busquem per nom per evitar falsos positius
        contacte = _fd_cerca_contacte(nif=nif_client) if nif_client else None
        if not contacte:
            contacte = _fd_crear_contacte(nom_fd, nif=nif_client or None, telefon=client_tel or None)
        if '_error' in (contacte or {}):
            return jsonify({'ok': False, 'error': f'Error contacte FD {contacte.get("_error")}: {contacte.get("_msg","")}'}), 500

        contact_id = _fd_extract_contact_id(contacte)
        if not contact_id:
            print(f'FD contacte sense ID (api_crear_albara): {json.dumps(contacte, ensure_ascii=False)}')
            return jsonify({'ok': False, 'error': f'Contacte FD sense ID. Resposta: {json.dumps(contacte, ensure_ascii=False)}'}), 500

    # Mides del marc (molt important al detall de l'albarà)
    final_w = float(d.get('final_amplada') or 0)
    final_h = float(d.get('final_alcada') or 0)
    peca_w  = float(d.get('amplada') or 0)
    peca_h  = float(d.get('alcada') or 0)

    def _fmt_cm(w, h):
        if w and h:
            return f'{w:g}×{h:g} cm'
        return ''

    mida_marc = _fmt_cm(final_w, final_h)
    mida_foto = _fmt_cm(peca_w, peca_h)

    # Línies de l'albarà
    desc_marc = f'Marc {marc}' if marc else 'Emmarcació'
    parts = []
    if mida_marc:
        parts.append(f'Mida {mida_marc}')
    if mida_foto and mida_foto != mida_marc:
        parts.append(f'foto {mida_foto}')
    if opcions_text:
        parts.append(opcions_text)
    if revers_peu:
        parts.append('Revers amb peu')
    if parts:
        desc_marc += f' · {", ".join(parts)}'

    # mode_preu: 'pvp' uses preu_net (PVP sense IVA), 'cost' uses cost_produccio
    base_total = preu_net if mode_preu == 'pvp' else cost_prod
    unit_price = round(base_total / quantitat, 2) if quantitat > 0 else round(base_total, 2)
    linies = [{
        'text':      desc_marc,
        'quantity':  float(quantitat),
        'unitPrice': unit_price,
        'tax':       _fd_line_tax(recarrec),
    }]

    # Línia d'enviament a client final (NACEX), si s'ha activat a la calc.
    env = d.get('enviament')
    env = env if isinstance(env, dict) else None
    if env:
        try:
            env_base = float(env.get('base') or 0)  # net × (1+marge), sense IVA
        except (TypeError, ValueError):
            env_base = 0.0
        if env_base > 0:
            env_text = f"Enviament {env.get('tarifa_nom') or 'NACEX'}"
            if env.get('poblacio'):
                env_text += f" a {env.get('poblacio')}"
            if env.get('cp'):
                env_text += f" (CP {env.get('cp')})"
            linies.append({
                'text':      env_text,
                'quantity':  1.0,
                'unitPrice': round(env_base, 2),
                'tax':       _fd_line_tax(recarrec),
            })

    notes_parts = []
    if num_pressupost:
        notes_parts.append(f'Pressupost: {num_pressupost}')
    if observacions:
        notes_parts.append(f'Obs: {observacions}')
    if env:
        dest = [p for p in [env.get('nom'), env.get('adreca'), env.get('poblacio'),
                            env.get('cp'), env.get('tel')] if p]
        if dest:
            notes_parts.append('Enviament a: ' + ', '.join(str(x) for x in dest))
    notes = ' | '.join(notes_parts)

    # doc_type: 'albara' (per defecte) crea un albarà i descompta stock;
    # 'pressupost' crea un estimate a FD i NO toca stock ni marca fd_albara.
    doc_type = (d.get('doc_type') or 'albara').strip().lower()
    es_pressupost = doc_type in ('pressupost', 'estimate', 'presupuesto')

    doc = _fd_crear_document(doc_type, contact_id, linies, notes=notes)
    if '_error' in (doc or {}):
        etiqueta = 'pressupost' if es_pressupost else 'albarà'
        return jsonify({'ok': False, 'error': f'Error {etiqueta} FD {doc.get("_error")}: {doc.get("_msg","")}'}), 500

    num_doc = doc.get('number') or doc.get('documentNumber') or doc.get('id', '—')

    if es_pressupost:
        return jsonify({'ok': True, 'doc_type': 'pressupost', 'pressupost': num_doc,
                        'numero': num_doc, 'contact': nom_fd, 'avisos_stock': []})

    # Descompte automàtic d'stock de marcs (best-effort: no fa fallar l'albarà).
    avisos_stock = []
    try:
        pre_marc = (d.get('pre_marc') or '').strip()
        avisos_stock = _descompta_stock_albara(
            marc, pre_marc, final_w, final_h, quantitat, num_doc, session.get('user_id'))
    except Exception as e:
        print('descompte stock albarà err:', e)

    return jsonify({'ok': True, 'doc_type': 'albara', 'albara': num_doc, 'numero': num_doc,
                    'contact': nom_fd, 'avisos_stock': avisos_stock})


@app.route('/api/albara-de-comanda', methods=['POST'])
@admin_required
def api_albara_de_comanda():
    """Crea un albarà FD a partir d'una sessió de comanda d'un client professional."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'Factura Directa no configurat (variables d\'entorn)'}), 503

    d = request.get_json(force=True) or {}
    sessio_id = (d.get('sessio_id') or '').strip()
    if not sessio_id:
        return jsonify({'ok': False, 'error': 'Falta sessio_id'}), 400

    # Read all lines of this session
    comandes = query(
        '''SELECT c.*, u.nom as usuari_nom, u.nom_empresa, u.nom_fiscal, u.fiscal_id, u.empresa_tel
           FROM comandes c JOIN usuaris u ON c.user_id=u.id
           WHERE c.sessio_id=? ORDER BY c.id''', [sessio_id])
    if not comandes:
        return jsonify({'ok': False, 'error': 'Sessió no trobada'}), 404

    c0 = comandes[0]

    # ── Client extern (cached fd_contact_id) ───────────────────────────
    # Si la comanda (qualsevol fila de la sessió) té client_extern_id,
    # saltem la cerca/creació de FD i fem servir el contact_id cached.
    client_extern_id = _row_get(c0, 'client_extern_id')
    cached_fd_id, cached_nom = _resolve_client_extern_fd_id(client_extern_id)
    contact_id = cached_fd_id
    nom_fd = cached_nom or ''
    is_client_extern = bool(contact_id)

    if not contact_id:
        # Fallback al flux antic: contacte FD = professional propietari de la sessió.
        nom_fiscal   = (_row_get(c0, 'nom_fiscal', '') or '').strip()
        nom_empresa  = (_row_get(c0, 'nom_empresa', '') or '').strip()
        usuari_nom   = (_row_get(c0, 'usuari_nom', '') or '').strip()
        fiscal_id    = (_row_get(c0, 'fiscal_id', '') or '').strip()
        empresa_tel  = (_row_get(c0, 'empresa_tel', '') or '').strip()

        nom_fd = nom_fiscal or nom_empresa or usuari_nom
        if not nom_fd:
            return jsonify({'ok': False, 'error': 'El client no té nom fiscal ni nom d\'empresa configurat.'}), 400

        contacte = _fd_cerca_contacte(nif=fiscal_id) if fiscal_id else None
        if not contacte:
            contacte = _fd_crear_contacte(nom_fd, nif=fiscal_id or None, telefon=empresa_tel or None)
        if '_error' in (contacte or {}):
            return jsonify({'ok': False, 'error': f'Error contacte FD {contacte.get("_error")}: {contacte.get("_msg","")}'}), 500

        contact_id = _fd_extract_contact_id(contacte)
        if not contact_id:
            print(f'FD contacte sense ID (api_albara_de_comanda): {json.dumps(contacte, ensure_ascii=False)}')
            return jsonify({'ok': False, 'error': f'Contacte FD sense ID. Resposta: {json.dumps(contacte, ensure_ascii=False)}'}), 500
    else:
        print(f"[albara_de_comanda] usant fd_contact_id cached: {contact_id} per a client extern {client_extern_id}")

    # Build albaran lines — one per comanda row (helper compartit)
    linies, notes_parts = _fd_linies_de_comandes(comandes, recarrec=_client_extern_recarrec(client_extern_id))
    notes = ' | '.join(notes_parts)

    albara = _fd_crear_albara(contact_id, linies, notes=notes)
    if '_error' in (albara or {}):
        return jsonify({'ok': False, 'error': f'Error albarà FD {albara.get("_error")}: {albara.get("_msg","")}'}), 500

    num_albara = albara.get('number') or albara.get('documentNumber') or albara.get('id', '—')

    # Store albaran number on the session rows
    execute("UPDATE comandes SET fd_albara=? WHERE sessio_id=?", [str(num_albara), sessio_id])

    return jsonify({'ok': True, 'albara': num_albara, 'contact': nom_fd})


@app.route('/api/crear-doc-conjunt', methods=['POST'])
@admin_required
def api_crear_doc_conjunt():
    """Crea UN sol albara o pressupost a FacturaDirecta amb les linies de
    DIVERSOS pressupostos (sessions) del MATEIX client, seleccionats a
    /historial. doc_type: 'albara' (per defecte) o 'pressupost'."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'Factura Directa no configurat (variables d\'entorn)'}), 503

    d = request.get_json(force=True) or {}
    sessio_ids = d.get('sessio_ids')
    sessio_ids = [str(s).strip() for s in sessio_ids if str(s).strip()] if isinstance(sessio_ids, list) else []
    doc_type = (d.get('doc_type') or 'albara').strip().lower()
    es_pressupost = doc_type in ('pressupost', 'estimate', 'presupuesto')
    if not sessio_ids:
        return jsonify({'ok': False, 'error': 'Cap pressupost seleccionat.'}), 400

    ph = ','.join(['?'] * len(sessio_ids))
    comandes = query(
        f'''SELECT c.*, u.nom as usuari_nom, u.nom_empresa, u.nom_fiscal, u.fiscal_id, u.empresa_tel
            FROM comandes c JOIN usuaris u ON c.user_id=u.id
            WHERE c.sessio_id IN ({ph}) ORDER BY c.client_nom, c.id''', sessio_ids) or []
    if not comandes:
        return jsonify({'ok': False, 'error': 'Pressupostos no trobats.'}), 404

    # Tots han de ser del mateix client (per client extern enllacat o, si no,
    # pel nom del client). Si n'hi ha de barrejats, parem.
    def _client_key(c):
        cei = _row_get(c, 'client_extern_id')
        return ('ext', cei) if cei else ('nom', (_row_get(c, 'client_nom', '') or '').strip().lower())
    if len({_client_key(c) for c in comandes}) > 1:
        return jsonify({'ok': False, 'error': 'Els pressupostos seleccionats no son tots del mateix client. Filtra per client o selecciona nomes els d\'un mateix client.'}), 400

    c0 = comandes[0]
    # Contacte FD: client extern enllacat -> contacte cached; si no, crear/
    # cercar per nom del client final.
    client_extern_id = _row_get(c0, 'client_extern_id')
    contact_id, nom_fd = _resolve_client_extern_fd_id(client_extern_id)
    nom_fd = nom_fd or ''
    if not contact_id:
        nom_fd = (_row_get(c0, 'client_nom', '') or '').strip()
        if not nom_fd:
            return jsonify({'ok': False, 'error': 'Els pressupostos no tenen nom de client per crear el contacte a FD.'}), 400
        contacte = _fd_crear_contacte(nom_fd)
        if '_error' in (contacte or {}):
            return jsonify({'ok': False, 'error': f'Error contacte FD {contacte.get("_error")}: {contacte.get("_msg","")}'}), 500
        contact_id = _fd_extract_contact_id(contacte)
        if not contact_id:
            return jsonify({'ok': False, 'error': 'Contacte FD sense ID.'}), 500

    linies, notes_parts = _fd_linies_de_comandes(comandes, recarrec=_client_extern_recarrec(client_extern_id))
    notes = ' | '.join(notes_parts)

    doc = _fd_crear_document(doc_type, contact_id, linies, notes=notes)
    if '_error' in (doc or {}):
        etiqueta = 'pressupost' if es_pressupost else 'albarà'
        return jsonify({'ok': False, 'error': f'Error {etiqueta} FD {doc.get("_error")}: {doc.get("_msg","")}'}), 500
    num_doc = doc.get('number') or doc.get('documentNumber') or doc.get('id', '—')

    # Nomes l'albara marca les sessions com a albarades.
    if not es_pressupost:
        execute(f"UPDATE comandes SET fd_albara=? WHERE sessio_id IN ({ph})",
                [str(num_doc)] + sessio_ids)

    return jsonify({'ok': True, 'doc_type': 'pressupost' if es_pressupost else 'albara',
                    'numero': num_doc, 'contact': nom_fd, 'n_pressupostos': len(sessio_ids)})


def _resolve_fd_contact_for_request(d):
    """Resol el contacte FD per a un client de la calculadora a partir del
    payload (client_extern_id enllaçat -> contacte cached; si no, cercar per
    NIF o crear pel nom). Retorna (contact_id, nom_fd, error_response_or_None)."""
    client_extern_id = (d.get('client_extern_id') or '').strip() or None
    nif_client = (d.get('client_nif') or '').strip()
    client_tel = (d.get('client_tel') or '').strip()
    nom_fd = (d.get('client_nom') or '').strip()

    contact_id, fd_nom = _resolve_client_extern_fd_id(client_extern_id)
    if contact_id:
        return contact_id, (fd_nom or nom_fd), None
    if not nom_fd:
        return None, '', (jsonify({'ok': False, 'error': 'Cal omplir el nom del client.'}), 400)
    contacte = _fd_cerca_contacte(nif=nif_client) if nif_client else None
    if not contacte:
        contacte = _fd_crear_contacte(nom_fd, nif=nif_client or None, telefon=client_tel or None)
    if '_error' in (contacte or {}):
        return None, '', (jsonify({'ok': False, 'error': f'Error contacte FD {contacte.get("_error")}: {contacte.get("_msg","")}'}), 500)
    cid = _fd_extract_contact_id(contacte)
    if not cid:
        return None, '', (jsonify({'ok': False, 'error': 'Contacte FD sense ID.'}), 500)
    return cid, nom_fd, None


def _linies_de_cistella(items, mode_preu, recarrec=False):
    """Construeix les linies FD a partir de la cistella de marcs del pressupost
    multi-marc. Cada item: {text, quantity, preu_net, cost_produccio}. Segons
    mode_preu fa servir el PVP net o el cost de produccio."""
    linies = []
    for it in (items or []):
        if not isinstance(it, dict):
            continue
        text = (str(it.get('text') or 'Emmarcació')).strip()[:300]
        try:
            qty = float(it.get('quantity') or 1) or 1
        except (TypeError, ValueError):
            qty = 1
        try:
            base = float(it.get('preu_net') or 0) if mode_preu == 'pvp' else float(it.get('cost_produccio') or 0)
        except (TypeError, ValueError):
            base = 0.0
        unit = round(base / qty, 2) if qty > 0 else round(base, 2)
        linies.append({'text': text, 'quantity': qty, 'unitPrice': unit, 'tax': _fd_line_tax(recarrec)})
    return linies


@app.route('/api/crear-doc-marcs', methods=['POST'])
@admin_required
def api_crear_doc_marcs():
    """Crea UN sol document a FacturaDirecta (albara o pressupost) amb DIVERSOS
    marcs (cistella del pressupost multi-marc), per a un mateix client."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'Factura Directa no configurat (variables d\'entorn)'}), 503
    d = request.get_json(force=True) or {}
    items = d.get('items')
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'El pressupost no té cap marc.'}), 400
    doc_type = (d.get('doc_type') or 'albara').strip().lower()
    es_pressupost = doc_type in ('pressupost', 'estimate', 'presupuesto')
    mode_preu = (d.get('mode_preu') or 'pvp').strip().lower()

    contact_id, nom_fd, err = _resolve_fd_contact_for_request(d)
    if err is not None:
        return err

    recarrec = _client_extern_recarrec((d.get('client_extern_id') or '').strip() or None)
    linies = _linies_de_cistella(items, mode_preu, recarrec=recarrec)
    if not linies:
        return jsonify({'ok': False, 'error': 'No s\'han pogut construir les línies del pressupost.'}), 400

    notes_parts = []
    if (d.get('num_pressupost') or '').strip():
        notes_parts.append(f"Pressupost: {d.get('num_pressupost').strip()}")
    if (d.get('observacions') or '').strip():
        notes_parts.append(f"Obs: {d.get('observacions').strip()}")
    notes = ' | '.join(notes_parts)

    doc = _fd_crear_document(doc_type, contact_id, linies, notes=notes)
    if '_error' in (doc or {}):
        etiqueta = 'pressupost' if es_pressupost else 'albarà'
        return jsonify({'ok': False, 'error': f'Error {etiqueta} FD {doc.get("_error")}: {doc.get("_msg","")}'}), 500
    num_doc = doc.get('number') or doc.get('documentNumber') or doc.get('id', '—')
    return jsonify({'ok': True, 'doc_type': 'pressupost' if es_pressupost else 'albara',
                    'numero': num_doc, 'contact': nom_fd, 'n_marcs': len(linies)})


@app.route('/api/albara-individual', methods=['POST'])
@admin_required
def api_albara_individual():
    """Crea un albarà FD per una comanda individual (no per sessió sencera)."""
    if not _FD_TOKEN or not _FD_COMPANY:
        return jsonify({'ok': False, 'error': 'Factura Directa no configurat (variables d\'entorn)'}), 503

    d = request.get_json(force=True) or {}
    comanda_id = d.get('comanda_id')
    if not comanda_id:
        return jsonify({'ok': False, 'error': 'Falta comanda_id'}), 400

    mode_preu = (d.get('mode_preu') or 'pvp').strip().lower()  # 'pvp' (default) or 'cost'

    com = query(
        '''SELECT c.*, u.nom as usuari_nom, u.nom_empresa, u.nom_fiscal, u.fiscal_id, u.empresa_tel, u.is_admin as owner_is_admin
           FROM comandes c JOIN usuaris u ON c.user_id=u.id
           WHERE c.id=?''', [comanda_id], one=True)
    if not com:
        return jsonify({'ok': False, 'error': 'Comanda no trobada'}), 404

    # ── Client extern (cached fd_contact_id) ───────────────────────────
    # Si la comanda té client_extern_id apuntant a un client actiu,
    # saltem la cerca de FD i forcem mode_preu=cost (PVD) per disseny.
    client_extern_id = _row_get(com, 'client_extern_id')
    recarrec = _client_extern_recarrec(client_extern_id)
    cached_fd_id, cached_nom = _resolve_client_extern_fd_id(client_extern_id)
    is_client_extern = bool(cached_fd_id)
    if is_client_extern:
        contact_id = cached_fd_id
        nom_fd = cached_nom or ''
        # Per a client extern sempre fem PVD (preu taller), no PVP.
        mode_preu = 'cost'
        print(f"[albara_individual] usant fd_contact_id cached: {contact_id} per a client extern {client_extern_id}")
        owner_is_admin = bool(_row_get(com, 'owner_is_admin', 0))
    else:
        # Determine FD contact: if admin's own order, use client_nom; if professional's order, use nom_fiscal/nom_empresa
        owner_is_admin = bool(_row_get(com, 'owner_is_admin', 0))
        if owner_is_admin:
            # Admin's own order → the FD contact is the end client
            nom_fd    = (_row_get(com, 'client_nom', '') or '').strip()
            fiscal_id = ''
            telefon   = (_row_get(com, 'client_tel', '') or '').strip()
        else:
            # Professional's order → the FD contact is the professional
            nom_fiscal  = (_row_get(com, 'nom_fiscal', '') or '').strip()
            nom_empresa = (_row_get(com, 'nom_empresa', '') or '').strip()
            usuari_nom  = (_row_get(com, 'usuari_nom', '') or '').strip()
            nom_fd      = nom_fiscal or nom_empresa or usuari_nom
            fiscal_id   = (_row_get(com, 'fiscal_id', '') or '').strip()
            telefon     = (_row_get(com, 'empresa_tel', '') or '').strip()

        if not nom_fd:
            return jsonify({'ok': False, 'error': 'Cal un nom de contacte per crear l\'albarà.'}), 400

        contacte = _fd_cerca_contacte(nif=fiscal_id) if fiscal_id else None
        if not contacte:
            contacte = _fd_crear_contacte(nom_fd, nif=fiscal_id or None, telefon=telefon or None)
        if '_error' in (contacte or {}):
            return jsonify({'ok': False, 'error': f'Error contacte FD {contacte.get("_error")}: {contacte.get("_msg","")}'}), 500

        contact_id = _fd_extract_contact_id(contacte)
        if not contact_id:
            return jsonify({'ok': False, 'error': f'Contacte FD sense ID.'}), 500

    # Build line item
    marc        = (_row_get(com, 'marc_principal', '') or '').strip()
    pre_marc    = (_row_get(com, 'pre_marc', '') or '').strip()
    passpartout = (_row_get(com, 'passpartout', '') or '').strip()
    vidre       = (_row_get(com, 'vidre', '') or '').strip()
    encolat     = (_row_get(com, 'encolat', '') or '').strip()
    impressio   = (_row_get(com, 'impressio', '') or '').strip()
    revers_peu  = str(_row_get(com, 'revers_peu', '') or '').strip().lower() in ('1', 'true', 'yes', 'on')
    quantitat   = int(_row_get(com, 'quantitat', 1) or 1)
    cost_prod   = float(_row_get(com, 'cost_produccio', 0) or 0)
    preu_net    = float(_row_get(com, 'preu_net', 0) or 0)
    client_nom  = (_row_get(com, 'client_nom', '') or '').strip()
    opcio_nom   = (_row_get(com, 'opcio_nom', '') or '').strip()

    parts_opc = []
    if passpartout and passpartout != 'cap': parts_opc.append(passpartout)
    if vidre:                                parts_opc.append(vidre)
    if pre_marc and pre_marc != '-':         parts_opc.append(f'+ {pre_marc}')
    if encolat and encolat != '-':           parts_opc.append(encolat)
    if revers_peu:                           parts_opc.append('Revers amb peu')
    if impressio and impressio != '-':       parts_opc.append(impressio)
    desc_marc = f'Marc {marc}' if marc else 'Emmarcació'
    if parts_opc:
        desc_marc += f' · {", ".join(parts_opc)}'
    if opcio_nom and opcio_nom != 'Opció A':
        desc_marc += f' ({opcio_nom})'
    if client_nom and not owner_is_admin:
        desc_marc = f'[{client_nom}] ' + desc_marc

    # mode_preu: 'pvp' uses preu_net (PVP sense IVA), 'cost' uses cost_produccio
    base_total = preu_net if mode_preu == 'pvp' else cost_prod
    unit_price = round(base_total / quantitat, 2) if quantitat > 0 else round(base_total, 2)
    linies = [{
        'text':      desc_marc,
        'quantity':  float(quantitat),
        'unitPrice': unit_price,
        'tax':       _fd_line_tax(recarrec),
    }]

    notes_parts = []
    num_pres = (_row_get(com, 'num_pressupost', '') or '').strip()
    if num_pres:
        notes_parts.append(f'Pressupost: {num_pres}')
    observacions = (_row_get(com, 'observacions', '') or '').strip()
    if observacions:
        notes_parts.append(f'Obs: {observacions}')
    notes = ' | '.join(notes_parts)

    albara = _fd_crear_albara(contact_id, linies, notes=notes)
    if '_error' in (albara or {}):
        return jsonify({'ok': False, 'error': f'Error albarà FD {albara.get("_error")}: {albara.get("_msg","")}'}), 500

    num_albara = albara.get('number') or albara.get('documentNumber') or albara.get('id', '—')
    execute("UPDATE comandes SET fd_albara=? WHERE id=?", [str(num_albara), comanda_id])

    return jsonify({'ok': True, 'albara': num_albara, 'contact': nom_fd})


@app.route('/api/passpartous')
@login_required
def api_passpartous():
    """Llista de referències de color/estil de passpartú per al selector
    de la calculadora. La taula passpartout té dos tipus de fila barrejades:
      - preus per mida (PAS13x18, PAS20x30…) — sense color, no es mostren
      - referències de color (P001, P003…) — amb color, sí es mostren
    Filtrem per color != '' per mostrar només les segones."""
    rows = query(
        "SELECT referencia, color, textura, descripcio FROM passpartout "
        "WHERE color IS NOT NULL AND color <> '' "
        "ORDER BY referencia"
    ) or []
    return jsonify([{
        'ref': r['referencia'],
        'color': r['color'] or '',
        'textura': r['textura'] or '',
        'descripcio': r['descripcio'] or '',
        'foto': _passpartou_photo_url(r['referencia']),
    } for r in rows])


@app.route('/admin/passpartous', methods=['GET', 'POST'])
@admin_required
def admin_passpartous():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'crear':
            ref = request.form.get('referencia', '').strip().upper()
            color = request.form.get('color', '').strip()
            textura = request.form.get('textura', '').strip()
            descripcio = request.form.get('descripcio', '').strip()
            if ref:
                # UPSERT manual: INSERT OR REPLACE no es comporta igual a PG.
                existing = query('SELECT 1 FROM passpartout WHERE referencia=?', [ref], one=True)
                if existing:
                    execute('UPDATE passpartout SET color=?, textura=?, descripcio=? WHERE referencia=?',
                            [color, textura, descripcio, ref])
                else:
                    execute('INSERT INTO passpartout (referencia, color, textura, descripcio) VALUES (?,?,?,?)',
                            [ref, color, textura, descripcio])
                flash(f'Referència {ref} desada.', 'ok')
        elif action == 'eliminar':
            ref = request.form.get('ref', '').strip()
            execute('DELETE FROM passpartout WHERE referencia=?', [ref])
            flash('Referència eliminada.', 'ok')
        return redirect(url_for('admin_passpartous'))

    # Només mostrem files que són referències de color (P001, P003…), no
    # les files de preus per mida (PAS13x18…) que viuen al mateix mateix table.
    passpartous_raw = query(
        "SELECT referencia, color, textura, descripcio FROM passpartout "
        "WHERE color IS NOT NULL AND color <> '' "
        "ORDER BY referencia"
    ) or []
    passpartous = []
    for r in passpartous_raw:
        d = dict(r) if not isinstance(r, dict) else r
        d['foto'] = _passpartou_photo_url(_row_get(r, 'referencia', ''))
        passpartous.append(d)
    return render_template('admin_passpartous.html', passpartous=passpartous)


# NOTA: /admin/proeco eliminat. ProEco és obsolet (àlies de foam) i ja no té
# UI per crear/editar/eliminar. Les files existents a la taula `proeco` es
# mantenen com a alias de lectura per a pressupostos antics (via calcular_cost_foam).


@app.route('/admin/impressio', methods=['GET', 'POST'])
@admin_required
def admin_impressio():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'crear':
            ref = request.form.get('referencia','').strip().upper()
            desc = request.form.get('descripcio','').strip()
            preu = float(request.form.get('preu', 0))
            # UPSERT manual per a PG (INSERT OR REPLACE no fa upsert allà)
            existing = query('SELECT 1 FROM impressio WHERE referencia=?', [ref], one=True)
            if existing:
                execute('UPDATE impressio SET descripcio=?, preu=? WHERE referencia=?', [desc, preu, ref])
            else:
                execute('INSERT INTO impressio (referencia, descripcio, preu) VALUES (?,?,?)', [ref, desc, preu])
            flash(f'Format {ref} desat.', 'ok')
        elif action == 'eliminar':
            execute('DELETE FROM impressio WHERE referencia=?', [request.form.get('ref')])
            flash('Format eliminat.', 'ok')
        return redirect(url_for('admin_impressio'))

    impressio = query('SELECT referencia, descripcio, preu FROM impressio ORDER BY referencia') or []
    return render_template('admin_impressio.html', impressio=impressio)

@app.route('/admin/foto', methods=['POST'])
@admin_required
def admin_foto():
    ref = request.form.get('referencia', '').strip().lower()
    f = request.files.get('foto')
    if f and ref:
        try:
            foto = _save_moldura_photo(f, ref)
            execute('UPDATE moldures SET foto=? WHERE LOWER(referencia)=?', [foto, ref])
            flash(f'Foto de {ref} pujada.', 'ok')
        except ValueError as e:
            flash(str(e), 'error')
    return redirect(url_for('admin'))

# â"€â"€ PDF generator â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€

def generar_num_pressupost():
    """Genera número tipus RR-2503-001"""
    from datetime import datetime
    # Get initials from empresa_nom in config
    r = query("SELECT valor FROM config WHERE clau='empresa_nom'", one=True)
    nom = r['valor'] if r and r['valor'] else 'XX'
    # Take first letters of each word (max 3)
    inicials = ''.join(w[0].upper() for w in nom.split()[:3] if w)[:3] or 'XX'
    # YYMM
    yymm = datetime.now().strftime('%y%m')
    prefix = f"{inicials}-{yymm}-"
    # Find last number with same prefix
    r2 = query("SELECT num_pressupost FROM comandes WHERE num_pressupost LIKE ? ORDER BY id DESC LIMIT 1",
               [prefix + '%'], one=True)
    if r2 and r2['num_pressupost']:
        try:
            last_n = int(r2['num_pressupost'].split('-')[-1])
            return f"{prefix}{last_n+1:03d}"
        except:
            pass
    return f"{prefix}001"


def crear_pdf_comparativa(comandes):
    from reportlab.lib.pagesizes import landscape
    buf = io.BytesIO()
    PAGE = landscape(A4)
    W_page, H_page = PAGE
    margin = 15*mm
    W = W_page - 2*margin
    n = len(comandes)
    col_lbl = 38*mm
    col_w = (W - col_lbl) / n

    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            rightMargin=margin, leftMargin=margin,
                            topMargin=12*mm, bottomMargin=15*mm)
    DARK  = colors.HexColor("#1C1B18")
    ACC   = colors.HexColor("#1A6B45")
    LIG   = colors.HexColor("#F5F6FA")
    BRD   = colors.HexColor("#E5E2DB")
    GREEN = colors.HexColor("#1A6B45")
    RED   = colors.HexColor("#B84040")
    WHITE = colors.white

    def p(txt, bold=False, size=10, color=DARK, align='LEFT'):
        st = ParagraphStyle('x', fontName='DejaVu-Bold' if bold else 'DejaVu',
                            fontSize=size, textColor=color,
                            alignment={'LEFT':0,'CENTER':1,'RIGHT':2}[align])
        return Paragraph(str(txt), st)

    story = []

    lang = (comandes[0].get('lang') or 'ca').lower()
    t = PDF_T.get(lang, PDF_T['ca'])

    # Header
    c0 = comandes[0]
    _u_comp = query('SELECT nom_empresa FROM usuaris WHERE id=?', [c0.get('user_id',0)], one=True)
    nom_empresa_comp = (_row_get(_u_comp, 'nom_empresa', '') or '') if _u_comp else ''
    if not nom_empresa_comp:
        _r_comp = query("SELECT valor FROM config WHERE clau='empresa_nom'", one=True)
        nom_empresa_comp = (_r_comp['valor'] if _r_comp else '') or 'Reus Revela'
    header = Table([[
        p(f"{t['comparativa']}", bold=True, size=14, color=colors.white),
        p(nom_empresa_comp, size=9, color=colors.HexColor("#B2BEC3"), align='RIGHT')
    ]], colWidths=[W*0.65, W*0.35])
    header.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), DARK),
        ('TOPPADDING',(0,0),(-1,-1), 10), ('BOTTOMPADDING',(0,0),(-1,-1), 10),
        ('LEFTPADDING',(0,0),(-1,-1), 12), ('RIGHTPADDING',(0,0),(-1,-1), 12),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(header)

    client_info = Table([[
        p(f"{t['client']}: {c0['client_nom'] or '—'}", bold=True, size=11),
        p(f"{t['data']}: {c0['data']}", size=10, align='RIGHT')
    ]], colWidths=[W*0.6, W*0.4])
    client_info.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), LIG),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),12),('RIGHTPADDING',(0,0),(-1,-1),12),
        ('BOX',(0,0),(-1,-1),0.5,BRD),
    ]))
    story.append(client_info)
    from reportlab.platypus import Spacer as Sp
    story.append(Sp(1,4*mm))

    # Column headers (with label column)
    hdr = [p('', bold=True, size=10, color=WHITE)] +           [p(c.get('opcio_nom','Opció'), bold=True, size=11, color=WHITE, align='CENTER') for c in comandes]
    hdr_row = Table([hdr], colWidths=[col_lbl]+[col_w]*n)
    hdr_row.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), DARK),
        ('TOPPADDING',(0,0),(-1,-1),9),('BOTTOMPADDING',(0,0),(-1,-1),9),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('BOX',(0,0),(-1,-1),0.5,BRD),
        ('INNERGRID',(0,1),(-1,-1),0.5,colors.HexColor("#3d4d5e")),
    ]))
    story.append(hdr_row)

    # Photo row (frame images for each option)
    import os as _os_comp
    from reportlab.platypus import Image as RLImageComp
    _foto_cells = [p('', size=8)]
    _has_foto_comp = False
    for _c in comandes:
        _cell = p('', size=8)
        if _c.get('marc_principal'):
            _r_f = query('SELECT foto, ref2 FROM moldures WHERE LOWER(referencia)=LOWER(?)',
                         [_c['marc_principal']], one=True)
            _foto_url = _resolve_moldura_photo(_c['marc_principal'], _r_f['foto'] if _r_f else '',
                                               ref2=(_row_get(_r_f,'ref2','') or '') if _r_f else '')
            if _foto_url.startswith('/static/'):
                _rel = _foto_url.lstrip('/')
                _full = _os_comp.path.join(app.root_path, 'static', _rel.replace('static/',''))
                if _os_comp.path.exists(_full):
                    try:
                        _img_c = RLImageComp(_full, width=min(col_w - 8*mm, 35*mm), height=22*mm)
                        _cell = _img_c
                        _has_foto_comp = True
                    except Exception:
                        pass
        _foto_cells.append(_cell)
    if _has_foto_comp:
        foto_row_tbl = Table([_foto_cells], colWidths=[col_lbl]+[col_w]*n)
        foto_row_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), LIG),
            ('BOX',(0,0),(-1,-1),0.5,BRD),
            ('INNERGRID',(0,0),(-1,-1),0.3,BRD),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('ALIGN',(1,0),(-1,0),'CENTER'),
        ]))
        story.append(foto_row_tbl)

    # Build comparison rows
    def val_clean(c, key):
        v = c.get(key,'') or ''
        return '—' if v in ['','-','None'] else str(v)

    fields = [
        (t['marc_principal'], 'marc_principal',   False),
        (t['tipus_peca'],     'piece_type',       False),
        (t['mida_final'],     'final_size',      False),
        (t['mides_foto'],     'photo_size',      False),
        (t['muntatge'],       'muntatge_label',  False),
        (t['vidre_mirall'],   'proteccio_label', False),
        (t['interior'],       'interior_label',  False),
        (t['revers_peu'],     'revers_peu_label', False),
        (t['impressio'],      'impressio_label', False),
        (t['observacions'],   'observacions',    False),
        (t['preu_net_label'], 'preu_net',        True),
        (t['iva'],            'iva',             True),
        (t['total_iva'],      'preu_final',      True),
        (t['entrega'],        'entrega',         True),
        (t['pendent_short'],  'pendent',         True),
    ]

    rows = []
    for lbl, key, is_price in fields:
        lbl_color = DARK if not is_price else colors.HexColor("#1A6B45") if key=='preu_final' else                     RED if key=='pendent' else colors.HexColor("#6B6860")
        row = [p(lbl, bold=is_price, size=9, color=lbl_color)]
        for c in comandes:
            if key == 'marc_principal':
                val = c.get('marc_principal') or t['sense_marc']
            elif key == 'piece_type':
                val = _display_piece_type(c, t)
            elif key == 'final_size':
                val = _final_size_text(c, sep=' x ', with_unit=True)
            elif key == 'photo_size':
                val = _photo_size_text(c, sep=' x ', with_unit=True)
            elif key == 'muntatge_label':
                val = _display_muntatge(c, t)
            elif key == 'proteccio_label':
                val = _display_proteccio(c, t)
            elif key == 'interior_label':
                val = _display_interior(c, t)
            elif key == 'revers_peu_label':
                val = _display_revers_peu(c, t)
            elif key == 'impressio_label':
                val = _display_impressio(c, t)
            elif key == 'iva':
                pn = float(c.get('preu_net',0) or 0)
                val = f"{pn*0.21:.2f} €"
            elif is_price:
                v = float(c.get(key,0) or 0)
                val = f"{v:.2f} €"
            else:
                val = val_clean(c, key)
            bold_cell = is_price and key in ('preu_final','pendent')
            cell_color = GREEN if key=='preu_final' else RED if key=='pendent' else DARK
            row.append(p(val, bold=bold_cell, size=9 if not bold_cell else 11,
                        color=cell_color, align='CENTER'))
        rows.append(row)

    col_widths = [col_lbl] + [col_w]*n
    detail_table = Table(rows, colWidths=col_widths)
    price_start = len([f for f in fields if not f[2]])  # index where prices start
    detail_table.setStyle(TableStyle([
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, WHITE]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),
        ('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('BACKGROUND',(0,0),(0,-1),colors.HexColor("#F2F0EC")),
        ('FONTNAME',(0,0),(0,-1),'DejaVu-Bold'),
        ('LINEABOVE',(0,price_start),(-1,price_start),1.5,ACC),
        ('BACKGROUND',(0,price_start+2),(-1,price_start+2),colors.HexColor("#E8F3EE")),
        ('BACKGROUND',(0,price_start+4),(-1,price_start+4),colors.HexColor("#FAEAEA")),
    ]))
    story.append(detail_table)

    doc.build(story)
    buf.seek(0)
    return buf


def crear_pdf_cataleg_admin(moldures, q='', proveidor=''):
    from reportlab.lib.pagesizes import landscape
    from reportlab.platypus import Image as RLImage

    buf = io.BytesIO()
    page = landscape(A4)
    margin = 10 * mm
    width = page[0] - (margin * 2)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        rightMargin=margin,
        leftMargin=margin,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    dark = colors.HexColor("#1C1B18")
    green = colors.HexColor("#1A6B45")
    border = colors.HexColor("#E5E2DB")
    light = colors.HexColor("#F5F4F1")
    muted = colors.HexColor("#6B6860")

    def p(txt, *, bold=False, size=8.5, color=dark, align=0):
        return Paragraph(
            str(txt),
            ParagraphStyle(
                name=f"cat-{size}-{int(bold)}-{align}",
                fontName='DejaVu-Bold' if bold else 'DejaVu',
                fontSize=size,
                leading=size + 2,
                textColor=color,
                alignment=align,
            )
        )

    story = [
        p("Cataleg de motllures - Admin", bold=True, size=16),
        Spacer(1, 3 * mm),
        p(
            f"Total de motllures: {len(moldures)}"
            + (f" | Cerca: {q}" if q else "")
            + (f" | Proveidor: {proveidor}" if proveidor else ""),
            size=9,
            color=muted,
        ),
        Spacer(1, 5 * mm),
    ]

    headers = [[
        p("Imatge", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Referencia", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Descripcio", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Preu taller", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Gruix", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Cost", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Proveidor", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Ubicacio", bold=True, size=8, color=colors.white, align=TA_CENTER),
        p("Ref2", bold=True, size=8, color=colors.white, align=TA_CENTER),
    ]]

    rows = []
    for moldura in moldures:
        photo_path = _moldura_photo_path(moldura.get('referencia', ''), moldura.get('foto', ''))
        if photo_path:
            photo_cell = RLImage(photo_path, width=16 * mm, height=16 * mm)
        else:
            photo_cell = p("Sense foto", size=7, color=muted, align=TA_CENTER)

        rows.append([
            photo_cell,
            p(moldura.get('referencia', '-') or '-', bold=True, size=8),
            p(moldura.get('descripcio', '-') or '-', size=8),
            p(f"{float(moldura.get('preu_taller', 0) or 0):.2f} EUR/m", size=8, align=TA_RIGHT),
            p(f"{float(moldura.get('gruix', 0) or 0):.1f} cm", size=8, align=TA_RIGHT),
            p(f"{float(moldura.get('cost', 0) or 0):.2f} EUR", size=8, align=TA_RIGHT),
            p(moldura.get('proveidor', '-') or '-', size=8),
            p(moldura.get('ubicacio', '-') or '-', size=8),
            p(moldura.get('ref2', '-') or '-', size=8),
        ])

    if not rows:
        rows.append([p("Sense resultats", size=9, color=muted)] + [''] * 8)

    table = Table(
        headers + rows,
        repeatRows=1,
        colWidths=[18 * mm, 24 * mm, 66 * mm, 22 * mm, 16 * mm, 18 * mm, 30 * mm, 25 * mm, 24 * mm],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), dark),
        ('BOX', (0, 0), (-1, -1), 0.5, border),
        ('INNERGRID', (0, 0), (-1, -1), 0.35, border),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)

    doc.build(story)
    buf.seek(0)
    return buf

# â"€â"€ Translations for PDF â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
PDF_T = {
    'ca': {
        'pressupost': 'PRESSUPOST',
        'client': 'Client',
        'telefon': 'Telèfon',
        'data': 'Data',
        'marc': 'Marc',
        'premarc': 'Pre-Marc',
        'mides': 'Mides',
        'mida_final': 'Mida final',
        'tipus_peca': 'Peça a emmarcar',
        'muntatge': 'Muntatge',
        'proteccio': 'Protecció',
        'interior': 'Interior',
        'revers_peu': 'Revers amb peu',
        'impressio': 'Impressió',
        'observacions': 'Observacions',
        'preu_net': 'Preu net (sense IVA)',
        'iva': 'IVA 21%',
        'total': 'Total amb IVA',
        'entrega': 'Entrega a compte',
        'pendent': 'Pendent de cobrar',
        'num': 'Num. Pressupost',
        'comparativa': 'COMPARATIVA DE PRESSUPOSTOS',
        'marc_principal': 'Marc principal',
        'mides_foto': 'Mides peça (cm)',
        'vidre_mirall': 'Vidre / Mirall',
        'preu_net_label': 'Preu net',
        'total_iva': 'TOTAL amb IVA',
        'pendent_short': 'PENDENT',
        'sense_marc': '(sense marc)',
        'encolat_label': 'Encolat',
        'laminat_label': 'Laminat',
        'protter': 'Protter',
        'vidre_label': 'Vidre',
        'doble_vidre': 'Doble vidre',
        'mirall': 'Mirall',
        'passpartu_label': 'Passpartú',
        'doble_passpartu': 'Doble passpartú',
        'proeco_label': 'ProEco',
        'inclosa': 'Inclosa',
        'piece_photo': 'Fotografia',
        'piece_lamina': 'Làmina',
        'piece_painting_unstretched': 'Pintura sense bastidor',
        'piece_painting_stretched': 'Pintura amb bastidor',
        'piece_puzzle': 'Puzzle',
        'piece_cross_stitch': 'Punt de creu',
        'conservar_vidre': 'Conservar vidre existent',
        'preu_net_pvp': 'Preu net PVP (sense IVA)',
        'descompte_sobre_pvp': 'Descompte {pct}% sobre PVP',
        'total_pvp_iva': 'TOTAL PVP amb IVA',
        'pendent_cobrar': 'PENDENT de cobrar',
    },
    'es': {
        'pressupost': 'PRESUPUESTO',
        'client': 'Cliente',
        'telefon': 'Teléfono',
        'data': 'Fecha',
        'marc': 'Marco',
        'premarc': 'Pre-Marco',
        'mides': 'Medidas',
        'mida_final': 'Medida final',
        'tipus_peca': 'Pieza a enmarcar',
        'muntatge': 'Montaje',
        'proteccio': 'Protección',
        'interior': 'Interior',
        'revers_peu': 'Reverso con pie',
        'impressio': 'Impresión',
        'observacions': 'Observaciones',
        'preu_net': 'Precio neto (sin IVA)',
        'iva': 'IVA 21%',
        'total': 'Total con IVA',
        'entrega': 'Pago a cuenta',
        'pendent': 'Pendiente de cobro',
        'num': 'Num. Presupuesto',
        'comparativa': 'COMPARATIVA DE PRESUPUESTOS',
        'marc_principal': 'Marco principal',
        'mides_foto': 'Medidas pieza (cm)',
        'vidre_mirall': 'Vidrio / Espejo',
        'preu_net_label': 'Precio neto',
        'total_iva': 'TOTAL con IVA',
        'pendent_short': 'PENDIENTE',
        'sense_marc': '(sin marco)',
        'encolat_label': 'Encolado',
        'laminat_label': 'Laminado',
        'protter': 'Protter',
        'vidre_label': 'Vidrio',
        'doble_vidre': 'Doble vidrio',
        'mirall': 'Espejo',
        'passpartu_label': 'Passpartú',
        'doble_passpartu': 'Doble passpartú',
        'proeco_label': 'ProEco',
        'inclosa': 'Incluida',
        'piece_photo': 'Fotografía',
        'piece_lamina': 'Lámina',
        'piece_painting_unstretched': 'Pintura sin bastidor',
        'piece_painting_stretched': 'Pintura con bastidor',
        'piece_puzzle': 'Puzzle',
        'piece_cross_stitch': 'Punto de cruz',
        'conservar_vidre': 'Conservar vidrio existente',
        'preu_net_pvp': 'Precio neto PVP (sin IVA)',
        'descompte_sobre_pvp': 'Descuento {pct}% sobre PVP',
        'total_pvp_iva': 'TOTAL PVP con IVA',
        'pendent_cobrar': 'PENDIENTE de cobro',
    },
    'en': {
        'pressupost': 'QUOTE',
        'client': 'Client',
        'telefon': 'Phone',
        'data': 'Date',
        'marc': 'Frame',
        'premarc': 'Inner frame',
        'mides': 'Dimensions',
        'mida_final': 'Final size',
        'tipus_peca': 'Item to frame',
        'muntatge': 'Mounting',
        'proteccio': 'Protection',
        'interior': 'Interior',
        'revers_peu': 'Backing with stand',
        'impressio': 'Print',
        'observacions': 'Notes',
        'preu_net': 'Net price (excl. VAT)',
        'iva': 'VAT 21%',
        'total': 'Total incl. VAT',
        'entrega': 'Deposit',
        'pendent': 'Outstanding',
        'num': 'Quote No.',
        'comparativa': 'QUOTE COMPARISON',
        'marc_principal': 'Main frame',
        'mides_foto': 'Item size (cm)',
        'vidre_mirall': 'Glass / Mirror',
        'preu_net_label': 'Net price',
        'total_iva': 'TOTAL incl. VAT',
        'pendent_short': 'OUTSTANDING',
        'sense_marc': '(no frame)',
        'encolat_label': 'Mounting',
        'laminat_label': 'Laminate only',
        'protter': 'Protter',
        'vidre_label': 'Glass',
        'doble_vidre': 'Double glass',
        'mirall': 'Mirror',
        'passpartu_label': 'Mat',
        'doble_passpartu': 'Double mat',
        'proeco_label': 'ProEco',
        'inclosa': 'Included',
        'piece_photo': 'Photograph',
        'piece_lamina': 'Poster',
        'piece_painting_unstretched': 'Unstretched painting',
        'piece_painting_stretched': 'Stretched painting',
        'piece_puzzle': 'Puzzle',
        'piece_cross_stitch': 'Cross stitch',
        'conservar_vidre': 'Keep existing glass',
        'preu_net_pvp': 'Net retail price (excl. VAT)',
        'descompte_sobre_pvp': 'Discount {pct}% on retail price',
        'total_pvp_iva': 'TOTAL retail incl. VAT',
        'pendent_cobrar': 'OUTSTANDING amount',
    }
}

def crear_pdf(c, mode=''):
    import os as _os
    from reportlab.platypus import Image as RLImage
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm
    DARK  = colors.HexColor("#1C1B18")
    GREEN = colors.HexColor("#1A6B45")
    AMBER = colors.HexColor("#C8873A")
    RED   = colors.HexColor("#B84040")
    LIG   = colors.HexColor("#F5F6FA")
    BRD   = colors.HexColor("#E5E2DB")
    WHITE = colors.white

    def p(txt, bold=False, size=10, color=DARK, align='LEFT'):
        st = ParagraphStyle('x', fontName='DejaVu-Bold' if bold else 'DejaVu',
                            fontSize=size, textColor=color,
                            alignment={'LEFT':0,'CENTER':1,'RIGHT':2}[align],
                            leading=size*1.4)
        return Paragraph(str(txt) if txt else '—', st)

    def fila(lbl, val, color_val=None):
        return [p(lbl, bold=True, size=9, color=colors.HexColor("#6B6860")),
                p(str(val) if val and val not in ['-',''] else '—', size=10,
                  color=color_val or DARK)]

    story = []

    # ── Capçalera ─────────────────────────────────────────────────────────
    # Get empresa info for this user
    u_data = query('SELECT nom_empresa, empresa_adreca, empresa_tel, brand_color FROM usuaris WHERE id=?', [c.get('user_id',0)], one=True)
    nom_empresa = ''
    if u_data and _row_get(u_data, 'nom_empresa', ''):
        nom_empresa = _row_get(u_data, 'nom_empresa', '')
    green_hex = _normalize_hex_color(_row_get(u_data, 'brand_color', DEFAULT_BRAND_COLOR))
    if not nom_empresa:
        _r = query("SELECT valor FROM config WHERE clau='empresa_nom'", one=True)
        nom_empresa = (_r['valor'] if _r else '') or 'Reus Revela'
    adreca = _row_get(u_data, 'empresa_adreca', '') or ''
    if not adreca:
        r_adr = query("SELECT valor FROM config WHERE clau='empresa_adreca'", one=True)
        adreca = (r_adr['valor'] if r_adr else '') or 'C/ Mare Molas, 26 · Reus'

    # Marca segons el mode: pressupost a PVD (taller) → Reus Revela; PVP → marca
    # actual. Es detecta pel mode o perquè la comanda és d'un client de taller.
    is_pvd_doc = str(mode).strip().lower() in ('pvd', 'cost', 'taller')
    if not is_pvd_doc:
        _cext = _row_get(c, 'client_extern_id')
        if _cext:
            try:
                _trow = query("SELECT tipus FROM clients_externs WHERE id=?", [_cext], one=True)
                if _trow and (_row_get(_trow, 'tipus', '') or '').strip().lower() == 'taller':
                    is_pvd_doc = True
            except Exception:
                pass
    pvd_logo_override = ''
    if is_pvd_doc:
        _rbn = query("SELECT valor FROM config WHERE clau='pvd_brand_nom'", one=True)
        nom_empresa = ((_rbn['valor'] if _rbn else '') or '').strip() or 'Reus Revela'
        _rba = query("SELECT valor FROM config WHERE clau='pvd_brand_adreca'", one=True)
        _rba_val = ((_rba['valor'] if _rba else '') or '').strip()
        if _rba_val:
            adreca = _rba_val
        _rbl = query("SELECT valor FROM config WHERE clau='pvd_brand_logo_b64'", one=True)
        pvd_logo_override = (_rbl['valor'] if _rbl else '') or ''

    GREEN = colors.HexColor(green_hex)
    lang = (c.get('lang') or 'ca').lower()
    t = PDF_T.get(lang, PDF_T['ca'])

    header = Table([[
        p(t['pressupost'], bold=True, size=20, color=WHITE),
        p(nom_empresa + '\n' + adreca, size=8,
          color=colors.HexColor("#9E9B94"), align='RIGHT')
    ]], colWidths=[W*0.55, W*0.45])
    header.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), DARK),
        ('TOPPADDING',(0,0),(-1,-1),14),('BOTTOMPADDING',(0,0),(-1,-1),14),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(header)
    story.append(Spacer(1, 3*mm))

    # â"€â"€ Logo (si existeix) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    try:
        if is_pvd_doc and pvd_logo_override:
            logo_data_url = pvd_logo_override
        else:
            u_logo = query('SELECT logo_b64 FROM usuaris WHERE id=?', [c.get('user_id',0)], one=True)
            logo_data_url = _row_get(u_logo, 'logo_b64', '') or ''
        if logo_data_url and logo_data_url.startswith('data:'):
            import base64 as _b64
            data_url = logo_data_url
            _, b64data = data_url.split(',', 1)
            img_data = _b64.b64decode(b64data)
            from reportlab.platypus import Image as RLImg2
            from PIL import Image as PILImg
            import io as _io2
            # Get real dimensions to preserve aspect ratio
            pil = PILImg.open(_io2.BytesIO(img_data))
            orig_w, orig_h = pil.size
            max_h = 25 * mm
            max_w = 60 * mm
            ratio = min(max_w / orig_w, max_h / orig_h)
            logo_w = orig_w * ratio
            logo_h = orig_h * ratio
            logo_img = RLImg2(io.BytesIO(img_data), width=logo_w, height=logo_h)
            logo_img.hAlign = 'CENTER'
            story.append(Spacer(1, 3*mm))
            story.append(logo_img)
            story.append(Spacer(1, 3*mm))
    except Exception as _e:
        print(f"Logo PDF error: {_e}")

    # â"€â"€ Dades client + data â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    opcio_txt = c.get('opcio_nom','') or ''
    num_pres = c.get('num_pressupost','') or ''
    t1_rows = [
        fila(t['num']+':', num_pres, color_val=colors.HexColor('#1A6B45')) if num_pres else None,
        fila(t['client']+':', c['client_nom'] or '—'),
        fila(t['telefon']+':', c['client_tel'] or '—'),
        fila(t['data']+':', c['data']),
    ]
    t1_rows = [r for r in t1_rows if r is not None]
    if opcio_txt and opcio_txt != 'Opció A':
        t1_rows.append(fila('Opció:', opcio_txt))

    t1 = Table(t1_rows, colWidths=[W*0.32, W*0.68])
    t1.setStyle(TableStyle([
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, WHITE]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),10),('RIGHTPADDING',(0,0),(-1,-1),10),
    ]))
    story.append(t1)
    story.append(Spacer(1, 5*mm))

    # â"€â"€ Foto del marc (si existeix) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
    # Foto del marc + referencia (side-by-side)
    foto_path = None
    marc_ref = c.get('marc_principal') or ''
    marc_descripcio = ''
    if marc_ref:
        r = query('SELECT foto, descripcio, ref2 FROM moldures WHERE LOWER(referencia)=LOWER(?)',
                  [marc_ref], one=True)
        marc_descripcio = (_row_get(r, 'descripcio', '') or '') if r else ''
        foto_url = _resolve_moldura_photo(marc_ref, r['foto'] if r else '',
                                          ref2=(_row_get(r, 'ref2', '') or '') if r else '')
        if foto_url.startswith('/static/'):
            rel = foto_url.lstrip('/')
            full = _os.path.join(app.root_path, 'static', rel.replace('static/',''))
            if _os.path.exists(full):
                foto_path = full

    if foto_path:
        try:
            _ref_lbl = marc_descripcio or marc_ref
            _img_rl = RLImage(foto_path, width=45*mm, height=32*mm)
            _ref_cell_rows = [[p(t['marc_principal'], bold=True, size=8,
                                  color=colors.HexColor("#6B6860"))],
                               [p(_ref_lbl, bold=True, size=11)]]
            if marc_descripcio:
                _ref_cell_rows.append([p(marc_ref, size=8,
                                         color=colors.HexColor("#6B6860"))])
            _ref_cell = Table(_ref_cell_rows, colWidths=[W - 55*mm])
            _ref_cell.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'TOP'),
                ('TOPPADDING',(0,0),(-1,-1),2),('BOTTOMPADDING',(0,0),(-1,-1),2),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ]))
            foto_side = Table([[_img_rl, _ref_cell]], colWidths=[50*mm, W - 50*mm])
            foto_side.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('BACKGROUND',(0,0),(-1,-1), LIG),
                ('BOX',(0,0),(-1,-1),0.5,BRD),
                ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
                ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
            ]))
            story.append(foto_side)
            story.append(Spacer(1, 4*mm))
        except Exception as _e:
            print(f"Foto marc PDF error: {_e}")

    # Detall de la comanda
    final_size = _final_size_text(c, with_unit=True)
    photo_size = _photo_size_text(c, with_unit=True)
    piece_type = _display_piece_type(c, t)
    muntatge_label = _display_muntatge(c, t)
    proteccio_label = _display_proteccio(c, t)
    interior_label = _display_interior(c, t)
    revers_peu_label = _display_revers_peu(c, t)
    impressio_label = _display_impressio(c, t)

    det_rows = []
    if piece_type:
        det_rows.append(fila(t['tipus_peca']+':', piece_type))
    if final_size:
        det_rows.append(fila(t['mida_final']+':', final_size))
    if photo_size:
        det_rows.append(fila(t['mides_foto']+':', photo_size))
    if not foto_path:
        # Show marc as text row only when there's no photo (photo already shown above)
        det_rows.append(fila(t['marc_principal']+':',
                             marc_ref or t['sense_marc']))
    if muntatge_label:
        det_rows.append(fila(t['muntatge']+':', muntatge_label))
    if proteccio_label:
        det_rows.append(fila(t['vidre_mirall']+':', proteccio_label))
    if interior_label:
        det_rows.append(fila(t['interior']+':', interior_label))
    if revers_peu_label:
        det_rows.append(fila(t['revers_peu']+':', revers_peu_label))
    if impressio_label:
        det_rows.append(fila(t['impressio']+':', impressio_label))
    if c.get('observacions'):
        det_rows.append(fila(t['observacions']+':', c['observacions']))

    if det_rows:
        t2 = Table(det_rows, colWidths=[W*0.38, W*0.62])
        t2.setStyle(TableStyle([
            ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, WHITE]),
            ('BOX',(0,0),(-1,-1),0.5,BRD),('INNERGRID',(0,0),(-1,-1),0.3,BRD),
            ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
            ('LEFTPADDING',(0,0),(-1,-1),10),('RIGHTPADDING',(0,0),(-1,-1),10),
        ]))
        story.append(t2)
    story.append(Spacer(1, 5*mm))

    # ── Resum econòmic ────────────────────────────────────────────────────
    desc_pct = float(c.get('descompte') or 0)
    pnet  = float(c.get('preu_net')   or 0)
    pfin  = float(c.get('preu_final') or 0)
    piva  = pnet * 1.21
    pent  = float(c.get('entrega')    or 0)
    ppend = float(c.get('pendent')    or 0)

    t3_data = [
        [p(t['preu_net_pvp'], bold=True, size=9, color=colors.HexColor("#6B6860")),
         p(f'{pnet:.2f} €', size=10, align='RIGHT')],
    ]
    if desc_pct > 0:
        desc_eur = pnet * (desc_pct/100)
        t3_data.append([
            p(t['descompte_sobre_pvp'].format(pct=int(desc_pct)), bold=True, size=9, color=colors.HexColor("#C8873A")),
            p(f'- {desc_eur:.2f} €', size=10, color=colors.HexColor("#C8873A"), align='RIGHT'),
        ])
    t3_data += [
        [p(t['iva'], bold=True, size=9, color=colors.HexColor("#6B6860")),
         p(f'{(pnet*(1-desc_pct/100))*0.21:.2f} €', size=10, align='RIGHT')],
        [p(t['total_pvp_iva'], bold=True, size=11, color=GREEN),
         p(f'{pfin:.2f} €', bold=True, size=14, color=GREEN, align='RIGHT')],
        [p(t['entrega'], bold=True, size=9, color=colors.HexColor("#6B6860")),
         p(f'{pent:.2f} €', size=10, align='RIGHT')],
        [p(t['pendent_cobrar'], bold=True, size=11, color=RED),
         p(f'{ppend:.2f} €', bold=True, size=14, color=RED, align='RIGHT')],
    ]
    t3 = Table(t3_data, colWidths=[W*0.6, W*0.4])
    # Find total row index (index 2 or 3 depending on discount)
    total_idx = 3 if desc_pct > 0 else 2
    pend_idx  = total_idx + 2
    t3.setStyle(TableStyle([
        ('BACKGROUND',(0,total_idx),(-1,total_idx), colors.HexColor("#E8F3EE")),
        ('BACKGROUND',(0,pend_idx), (-1,pend_idx),  colors.HexColor("#FAEAEA")),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, colors.white]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),
        ('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('LINEABOVE',(0,total_idx),(-1,total_idx),1.5,colors.HexColor("#1A6B45")),
        ('LINEABOVE',(0,pend_idx), (-1,pend_idx), 1.5,colors.HexColor("#B84040")),
    ]))
    story.append(t3)

    # ── Secció dual taller (només si mode='dual' i client de taller) ──────
    if mode == 'dual':
        client_extern_id = _row_get(c, 'client_extern_id')
        taller_client = None
        taller_marge = 60.0
        if client_extern_id:
            taller_client = query('SELECT tipus, usuari_id FROM clients_externs WHERE id=?', [client_extern_id], one=True)
        if taller_client and _row_get(taller_client, 'tipus') == 'taller':
            usuari_id = _row_get(taller_client, 'usuari_id')
            if usuari_id:
                u_taller = query('SELECT marge FROM usuaris WHERE id=?', [usuari_id], one=True)
                if u_taller and _row_get(u_taller, 'marge') is not None:
                    taller_marge = float(_row_get(u_taller, 'marge'))
            cost_prod = float(c.get('cost_produccio') or 0)
            qty = int(c.get('quantitat') or 1)
            pvd_unit = cost_prod / qty if qty > 0 else cost_prod
            pvp_suggerit_unit = round(pvd_unit * (1 + taller_marge / 100), 2)
            pvp_suggerit_total = round(pvp_suggerit_unit * qty, 2)
            pvp_suggerit_iva = round(pvp_suggerit_total * 1.21, 2)

            story.append(Spacer(1, 5*mm))
            dual_header = Table([[
                p('Preus per al teu client final', bold=True, size=11, color=WHITE),
            ]], colWidths=[W])
            dual_header.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,-1), AMBER),
                ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
                ('LEFTPADDING',(0,0),(-1,-1),10),('RIGHTPADDING',(0,0),(-1,-1),10),
            ]))
            story.append(dual_header)

            dual_data = [
                [p('Preu taller (el que pagues)', bold=True, size=9, color=colors.HexColor("#6B6860")),
                 p(f'{cost_prod:.2f} €', size=10, align='RIGHT')],
                [p(f'PVP suggerit (marge {taller_marge:.0f}%)', bold=True, size=9, color=colors.HexColor("#6B6860")),
                 p(f'{pvp_suggerit_total:.2f} €', size=10, align='RIGHT')],
                [p('PVP suggerit amb IVA 21%', bold=True, size=11, color=AMBER),
                 p(f'{pvp_suggerit_iva:.2f} €', bold=True, size=14, color=AMBER, align='RIGHT')],
            ]
            if qty > 1:
                dual_data.insert(1, [
                    p(f'PVP suggerit per unitat', bold=True, size=9, color=colors.HexColor("#6B6860")),
                    p(f'{pvp_suggerit_unit:.2f} € × {qty}', size=10, align='RIGHT'),
                ])

            t_dual = Table(dual_data, colWidths=[W*0.6, W*0.4])
            pvp_idx = len(dual_data) - 1
            t_dual.setStyle(TableStyle([
                ('BACKGROUND',(0,pvp_idx),(-1,pvp_idx), colors.HexColor("#FDF3E8")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, colors.white]),
                ('BOX',(0,0),(-1,-1),0.5,BRD),
                ('INNERGRID',(0,0),(-1,-1),0.3,BRD),
                ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
                ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
                ('LINEABOVE',(0,pvp_idx),(-1,pvp_idx),1.5,AMBER),
            ]))
            story.append(t_dual)

    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=W, thickness=0.5, color=BRD))
    story.append(Spacer(1, 2*mm))
    story.append(p('Objectiu Emmarcació · C/ Mare Molas, 26 · Reus', size=8,
                   color=colors.HexColor("#636E72"), align='CENTER'))

    doc.build(story)
    buf.seek(0)
    return buf


def crear_pdf_marcs(items, client, mode='pvp', num_pressupost='', observacions='', user_id=0):
    """PDF d'un pressupost amb DIVERSOS marcs (cistella multi-marc) per a un
    mateix client. `items` = [{text, quantity, preu_net, cost_produccio}].
    `mode` = 'pvp' (preu de venda) o 'pvd' (preu taller/cost)."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=12*mm, bottomMargin=15*mm)
    W = A4[0] - 30*mm
    DARK  = colors.HexColor("#1C1B18")
    LIG   = colors.HexColor("#F5F6FA")
    BRD   = colors.HexColor("#E5E2DB")
    WHITE = colors.white

    def p(txt, bold=False, size=10, color=DARK, align='LEFT'):
        st = ParagraphStyle('x', fontName='DejaVu-Bold' if bold else 'DejaVu',
                            fontSize=size, textColor=color,
                            alignment={'LEFT':0,'CENTER':1,'RIGHT':2}[align],
                            leading=size*1.4)
        return Paragraph(str(txt) if txt not in (None, '') else '—', st)

    from reportlab.platypus import Image as RLImage

    def _marc_thumb(ref):
        """Retorna una miniatura (RLImage) de la foto de mostra d'una motllura,
        o None si no en té una de local. Mateixa resolució que el PDF d'un marc."""
        if not ref:
            return None
        try:
            r = query('SELECT foto, ref2 FROM moldures WHERE LOWER(referencia)=LOWER(?)',
                      [ref], one=True)
            if not r:
                return None
            foto_url = _resolve_moldura_photo(ref, _row_get(r, 'foto', '') or '',
                                              ref2=_row_get(r, 'ref2', '') or '')
            if not foto_url.startswith('/static/'):
                return None
            rel = foto_url.lstrip('/')
            full = _os.path.join(app.root_path, 'static', rel.replace('static/', ''))
            if _os.path.exists(full):
                return RLImage(full, width=22*mm, height=16*mm)
        except Exception as _e:
            print(f"[pdf-marcs] thumb error {ref}: {_e}")
        return None

    def _concept_cell(text, ref):
        """Cel·la de concepte: si la motllura té foto de mostra, la posa a
        l'esquerra del text; si no, només el text."""
        thumb = _marc_thumb(ref)
        if thumb is None:
            return p(text, size=9)
        inner = Table([[thumb, p(text, size=9)]], colWidths=[24*mm, W*0.52 - 24*mm])
        inner.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(0,0),4),
            ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0),
        ]))
        return inner

    story = []

    # ── Capçalera + dades empresa ─────────────────────────────────────────
    u_data = query('SELECT nom_empresa, empresa_adreca, brand_color FROM usuaris WHERE id=?',
                   [user_id or 0], one=True)
    nom_empresa = (_row_get(u_data, 'nom_empresa', '') or '') if u_data else ''
    green_hex = _normalize_hex_color(_row_get(u_data, 'brand_color', DEFAULT_BRAND_COLOR) if u_data else DEFAULT_BRAND_COLOR)
    if not nom_empresa:
        _r = query("SELECT valor FROM config WHERE clau='empresa_nom'", one=True)
        nom_empresa = (_r['valor'] if _r else '') or 'Reus Revela'
    adreca = (_row_get(u_data, 'empresa_adreca', '') or '') if u_data else ''
    if not adreca:
        r_adr = query("SELECT valor FROM config WHERE clau='empresa_adreca'", one=True)
        adreca = (r_adr['valor'] if r_adr else '') or 'C/ Mare Molas, 26 · Reus'
    # Marca segons el mode: pressupost a PVD (preu taller) → Reus Revela;
    # a PVP (preu client) → marca actual (Objectiu Fotògrafs). El logo i, si
    # s'ha configurat, l'adreça també canvien.
    is_pvd_doc = str(mode).strip().lower() in ('pvd', 'cost', 'taller')
    pvd_logo_override = ''
    if is_pvd_doc:
        _rbn = query("SELECT valor FROM config WHERE clau='pvd_brand_nom'", one=True)
        nom_empresa = ((_rbn['valor'] if _rbn else '') or '').strip() or 'Reus Revela'
        _rba = query("SELECT valor FROM config WHERE clau='pvd_brand_adreca'", one=True)
        _rba_val = ((_rba['valor'] if _rba else '') or '').strip()
        if _rba_val:
            adreca = _rba_val
        _rbl = query("SELECT valor FROM config WHERE clau='pvd_brand_logo_b64'", one=True)
        pvd_logo_override = (_rbl['valor'] if _rbl else '') or ''
    GREEN = colors.HexColor(green_hex)

    header = Table([[
        p('Pressupost', bold=True, size=20, color=WHITE),
        p(nom_empresa + '\n' + adreca, size=8,
          color=colors.HexColor("#9E9B94"), align='RIGHT')
    ]], colWidths=[W*0.55, W*0.45])
    header.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,-1), DARK),
        ('TOPPADDING',(0,0),(-1,-1),14),('BOTTOMPADDING',(0,0),(-1,-1),14),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(header)
    story.append(Spacer(1, 3*mm))

    # ── Logo (si existeix) ────────────────────────────────────────────────
    try:
        if is_pvd_doc and pvd_logo_override:
            logo_data_url = pvd_logo_override
        else:
            u_logo = query('SELECT logo_b64 FROM usuaris WHERE id=?', [user_id or 0], one=True)
            logo_data_url = _row_get(u_logo, 'logo_b64', '') or ''
        if logo_data_url and logo_data_url.startswith('data:'):
            import base64 as _b64
            from reportlab.platypus import Image as RLImg2
            from PIL import Image as PILImg
            import io as _io2
            _, b64data = logo_data_url.split(',', 1)
            img_data = _b64.b64decode(b64data)
            pil = PILImg.open(_io2.BytesIO(img_data))
            orig_w, orig_h = pil.size
            ratio = min(60*mm / orig_w, 25*mm / orig_h)
            logo_img = RLImg2(io.BytesIO(img_data), width=orig_w*ratio, height=orig_h*ratio)
            logo_img.hAlign = 'CENTER'
            story.append(Spacer(1, 3*mm)); story.append(logo_img); story.append(Spacer(1, 3*mm))
    except Exception as _e:
        print(f"Logo PDF marcs error: {_e}")

    # ── Dades client ──────────────────────────────────────────────────────
    def fila(lbl, val, color_val=None):
        return [p(lbl, bold=True, size=9, color=colors.HexColor("#6B6860")),
                p(str(val) if val not in (None, '', '-') else '—', size=10, color=color_val or DARK)]
    cli_rows = []
    if (num_pressupost or '').strip():
        cli_rows.append(fila('Núm.:', num_pressupost, color_val=GREEN))
    cli_rows.append(fila('Client:', (client or {}).get('nom') or '—'))
    if (client or {}).get('tel'):
        cli_rows.append(fila('Telèfon:', client.get('tel')))
    if (client or {}).get('nif'):
        cli_rows.append(fila('NIF/DNI:', client.get('nif')))
    cli_rows.append(fila('Data:', datetime.now().strftime('%d/%m/%Y')))
    t1 = Table(cli_rows, colWidths=[W*0.25, W*0.75])
    t1.setStyle(TableStyle([
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[LIG, WHITE]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),10),('RIGHTPADDING',(0,0),(-1,-1),10),
    ]))
    story.append(t1)
    story.append(Spacer(1, 5*mm))

    # ── Taula de marcs ────────────────────────────────────────────────────
    es_pvd = (mode == 'pvd' or mode == 'cost')
    head = [p('Concepte', bold=True, size=9, color=WHITE),
            p('Unitats', bold=True, size=9, color=WHITE, align='CENTER'),
            p('Preu/u', bold=True, size=9, color=WHITE, align='RIGHT'),
            p('Import', bold=True, size=9, color=WHITE, align='RIGHT')]
    rows = [head]
    subtotal = 0.0
    for it in (items or []):
        if not isinstance(it, dict):
            continue
        text = (str(it.get('text') or 'Emmarcació')).strip()
        ref = (str(it.get('ref') or '')).strip()
        try:
            qty = float(it.get('quantity') or 1) or 1
        except Exception:
            qty = 1
        try:
            base = float(it.get('cost_produccio') or 0) if es_pvd else float(it.get('preu_net') or 0)
        except Exception:
            base = 0.0
        unit = round(base / qty, 2) if qty > 0 else round(base, 2)
        import_linia = round(unit * qty, 2)
        subtotal += import_linia
        qty_txt = str(int(qty)) if float(qty).is_integer() else f'{qty:g}'
        rows.append([
            _concept_cell(text, ref),
            p(qty_txt, size=9, align='CENTER'),
            p(f'{unit:.2f} €', size=9, align='RIGHT'),
            p(f'{import_linia:.2f} €', size=9, align='RIGHT'),
        ])
    tmarcs = Table(rows, colWidths=[W*0.52, W*0.14, W*0.16, W*0.18])
    tmarcs.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0), DARK),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[WHITE, LIG]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
    ]))
    story.append(tmarcs)
    story.append(Spacer(1, 5*mm))

    # ── Totals ────────────────────────────────────────────────────────────
    iva = round(subtotal * 0.21, 2)
    total = round(subtotal + iva, 2)
    tot_rows = [
        [p('Subtotal (sense IVA)', bold=True, size=9, color=colors.HexColor("#6B6860")),
         p(f'{subtotal:.2f} €', size=10, align='RIGHT')],
        [p('IVA 21%', bold=True, size=9, color=colors.HexColor("#6B6860")),
         p(f'{iva:.2f} €', size=10, align='RIGHT')],
        [p('TOTAL amb IVA', bold=True, size=11, color=GREEN),
         p(f'{total:.2f} €', bold=True, size=14, color=GREEN, align='RIGHT')],
    ]
    ttot = Table(tot_rows, colWidths=[W*0.62, W*0.38])
    ttot.setStyle(TableStyle([
        ('BACKGROUND',(0,2),(-1,2), colors.HexColor("#E8F3EE")),
        ('ROWBACKGROUNDS',(0,0),(-1,1),[LIG, WHITE]),
        ('BOX',(0,0),(-1,-1),0.5,BRD),('INNERGRID',(0,0),(-1,-1),0.3,BRD),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('LINEABOVE',(0,2),(-1,2),1.5,GREEN),
    ]))
    story.append(ttot)

    if (observacions or '').strip():
        story.append(Spacer(1, 5*mm))
        story.append(p('Observacions', bold=True, size=9, color=colors.HexColor("#6B6860")))
        story.append(Spacer(1, 1*mm))
        story.append(p(observacions, size=9))

    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width=W, thickness=0.5, color=BRD))
    story.append(Spacer(1, 2*mm))
    story.append(p(f'{nom_empresa} · {adreca}', size=8,
                   color=colors.HexColor("#636E72"), align='CENTER'))

    doc.build(story)
    buf.seek(0)
    return buf


@app.route('/api/pdf-marcs', methods=['POST'])
@login_required
def api_pdf_marcs():
    """Genera el PDF d'un pressupost multi-marc (cistella) per a un client.
    Disponible per a tots els usuaris (al seu PVP); el mode 'cost'/PVD només
    el pot fer servir l'admin, perquè és preu intern de taller."""
    d = request.get_json(force=True) or {}
    items = d.get('items')
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'El pressupost no té cap marc.'}), 400
    client = {
        'nom': (d.get('client_nom') or '').strip(),
        'tel': (d.get('client_tel') or '').strip(),
        'nif': (d.get('client_nif') or '').strip(),
    }
    mode = (d.get('mode_preu') or 'pvp').strip().lower()
    # Els usuaris no-admin sempre generen el PDF a PVP (mai a cost de taller).
    if mode != 'pvp' and not session.get('is_admin'):
        mode = 'pvp'
    try:
        pdf = crear_pdf_marcs(items, client, mode=mode,
                              num_pressupost=(d.get('num_pressupost') or '').strip(),
                              observacions=(d.get('observacions') or '').strip(),
                              user_id=session.get('user_id', 0))
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Error generant PDF: {e}'}), 500
    nom_fitxer = (client['nom'] or 'pressupost').replace(' ', '_')[:40]
    return send_file(pdf, mimetype='application/pdf', as_attachment=True,
                     download_name=f"pressupost_marcs_{nom_fitxer}.pdf")


@app.route('/ajustos')
@login_required
def ajustos():
    # Si 'email' encara no s'ha migrat, aplica-ho de forma idempotent i reintenta.
    try:
        u = query(
            'SELECT marge, marge_impressio, nom_empresa, nom_fiscal, fiscal_id, empresa_adreca, empresa_tel, email, margins_json, brand_color, brand_color_secondary, brand_color_menu FROM usuaris WHERE id=?',
            [session['user_id']],
            one=True,
        )
    except Exception as e:
        if 'email' in str(e).lower() and ('does not exist' in str(e).lower() or 'no such column' in str(e).lower()):
            try:
                execute("ALTER TABLE usuaris ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''")
            except Exception:
                pass
            u = query(
                'SELECT marge, marge_impressio, nom_empresa, nom_fiscal, fiscal_id, empresa_adreca, empresa_tel, email, margins_json, brand_color, brand_color_secondary, brand_color_menu FROM usuaris WHERE id=?',
                [session['user_id']],
                one=True,
            )
        else:
            raise
    marge_actual = float(u['marge']) if u and u['marge'] is not None else 60
    marge_imp = float(u['marge_impressio']) if u and u['marge_impressio'] is not None else 0
    if float(marge_actual).is_integer():
        marge_actual = int(marge_actual)
    if float(marge_imp).is_integer():
        marge_imp = int(marge_imp)
    margins = _load_user_commercial_margins(u)
    margin_entries = [
        {'key': 'general', 'label': 'General', 'description': 'Marge base per a productes generals i acabats que no tenen marge propi.', 'value': _format_margin_for_view(margins['general'])},
        {'key': 'frames', 'label': 'Marcs', 'description': 'Marge principal de la calculadora de marcs.', 'value': _format_margin_for_view(margins['frames'])},
        {'key': 'canvas', 'label': 'Llenços', 'description': 'Marge dels llenços de la calculadora (mides de catàleg i a mida).', 'value': _format_margin_for_view(margins['canvas'])},
        {'key': 'prints', 'label': 'Impressió fotogràfica', 'description': 'S\'aplica a còpia fotogràfica i serveix també de base per a acabats d\'impressió.', 'value': _format_margin_for_view(margins['prints'])},
        {'key': 'foam', 'label': 'Foam', 'description': 'Permet separar el marge de foam del de la impressió si ho necessites.', 'value': _format_margin_for_view(margins['foam'])},
        {'key': 'laminate_foam', 'label': 'Laminat + foam', 'description': 'Per si voleu treballar aquesta combinació amb un marge propi.', 'value': _format_margin_for_view(margins['laminate_foam'])},
        {'key': 'fine_art', 'label': 'Fine art', 'description': 'Marge específic per a papers fine art i treballs més cuidats.', 'value': _format_margin_for_view(margins['fine_art'])},
        {'key': 'albums', 'label': 'Àlbums', 'description': 'Marge dels àlbums i libretos (individual i packs de boda/comunió).', 'value': _format_margin_for_view(margins['albums'])},
        {'key': 'digital', 'label': 'Digitalització', 'description': 'Marge de la digitalització (cintes, DVD, Súper 8). Es cobra sense IVA.', 'value': _format_margin_for_view(margins['digital'])},
        {'key': 'orles', 'label': 'Orles', 'description': 'Marge de les orles escolars (impressió per trams + muntatge).', 'value': _format_margin_for_view(margins['orles'])},
        {'key': 'regals', 'label': 'Regals', 'description': 'Marge dels regals personalitzats (tasses i imants).', 'value': _format_margin_for_view(margins['regals'])},
        {'key': 'offset', 'label': 'Offset / Papereria', 'description': 'Marge de la impressió digital en offset (targetes, invitacions, calendaris, díptics…).', 'value': _format_margin_for_view(margins['offset'])},
    ]
    nom_emp = u['nom_empresa'] if u and u['nom_empresa'] else ''
    brand_color = _normalize_hex_color(_row_get(u, 'brand_color', DEFAULT_BRAND_COLOR))
    brand_color_secondary = _normalize_hex_color(
        _row_get(u, 'brand_color_secondary', DEFAULT_BRAND_SECONDARY_COLOR),
        DEFAULT_BRAND_SECONDARY_COLOR,
    )
    brand_color_menu = _normalize_hex_color(
        _row_get(u, 'brand_color_menu', brand_color),
        brand_color,
    )
    cfg_rows = query("SELECT clau, valor FROM config WHERE clau LIKE 'empresa_%'")
    cfg = {r['clau']: r['valor'] for r in (cfg_rows or [])}
    if not nom_emp:
        nom_emp = cfg.get('empresa_nom', '')
    user_adreca  = _row_get(u, 'empresa_adreca', '') or ''
    user_tel     = _row_get(u, 'empresa_tel', '') or ''
    user_email   = _row_get(u, 'email', '') or ''
    nom_fiscal   = _row_get(u, 'nom_fiscal', '') or ''
    fiscal_id    = _row_get(u, 'fiscal_id', '') or ''

    # Trams d'impressió de l'usuari + defaults globals (per als placeholders)
    user_full = query(
        'SELECT imp_tram1, imp_tram2, imp_tram3, imp_tram4, imp_tram5, imp_tram6 '
        'FROM usuaris WHERE id=?', [session['user_id']], one=True,
    )
    tram_areas = [
        float(get_config_value('imp_tram1_area', '900')),
        float(get_config_value('imp_tram2_area', '2000')),
        float(get_config_value('imp_tram3_area', '4200')),
        float(get_config_value('imp_tram4_area', '6000')),
        float(get_config_value('imp_tram5_area', '14400')),
        None,
    ]
    tram_labels = ['fins 30×30', 'fins 40×50', 'fins 60×70', 'fins 75×100', 'fins 80×180', '> 80×180']
    imp_trams = []
    for i in range(1, 7):
        v = _row_get(user_full, f'imp_tram{i}') if user_full else None
        d = float(get_config_value(f'imp_tram{i}_marge_default', '0'))
        imp_trams.append({
            'idx': i,
            'label': tram_labels[i - 1],
            'area_max': tram_areas[i - 1],
            'value': float(v) if v is not None else None,
            'default': d,
        })

    return render_template('ajustos.html', marge_actual=marge_actual, marge_imp=marge_imp,
                           margin_entries=[
                               dict(entry, description='S\'aplica a la fotografia impresa. Foam, laminat + foam i ProEco treballen amb el marge general.') if entry['key'] == 'prints' else entry
                               for entry in margin_entries if entry['key'] not in ('foam', 'laminate_foam')
                           ],
                           nom_empresa=nom_emp,
                           nom_fiscal=nom_fiscal,
                           fiscal_id=fiscal_id,
                           brand_color=brand_color,
                           brand_color_secondary=brand_color_secondary,
                           brand_color_menu=brand_color_menu,
                           empresa_adreca=user_adreca if user_adreca else cfg.get('empresa_adreca',''),
                           empresa_tel=user_tel if user_tel else cfg.get('empresa_tel',''),
                           user_email=user_email,
                           imp_trams=imp_trams)


@app.route('/ajustos/impressio-trams', methods=['POST'])
@login_required
def ajustos_impressio_trams():
    """Desa els 6 trams de marge d'impressió de l'usuari logat. Si el camp
    arriba buit, queda NULL (s'aplicarà el default global del tram)."""
    valors = []
    for i in range(1, 7):
        raw = (request.form.get(f'imp_tram{i}') or '').strip()
        try:
            valors.append(float(raw) if raw else None)
        except ValueError:
            valors.append(None)
    execute(
        'UPDATE usuaris SET imp_tram1=?, imp_tram2=?, imp_tram3=?, '
        'imp_tram4=?, imp_tram5=?, imp_tram6=? WHERE id=?',
        valors + [session['user_id']],
    )
    flash('Trams d\'impressió actualitzats.', 'ok')
    return redirect(url_for('ajustos'))


@app.route('/ajustos/impressio-trams/reset', methods=['POST'])
@login_required
def ajustos_impressio_trams_reset():
    """Torna els 6 trams a NULL — l'app aplicarà el default global."""
    execute(
        'UPDATE usuaris SET imp_tram1=NULL, imp_tram2=NULL, imp_tram3=NULL, '
        'imp_tram4=NULL, imp_tram5=NULL, imp_tram6=NULL WHERE id=?',
        [session['user_id']],
    )
    flash('Trams d\'impressió restaurats als valors per defecte.', 'ok')
    return redirect(url_for('ajustos'))


import re as _re

def _parse_dims(ref):
    m = _re.search(r'(\d+)[xX](\d+)', ref)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = _re.search(r'[A-Z]+(\d+)$', ref)
    if m:
        d = m.group(1); mid = len(d)//2
        return int(d[:mid]), int(d[mid:])
    return None, None

def _fmt_measure(value):
    try:
        num = float(value or 0)
    except (TypeError, ValueError):
        return ''
    if num.is_integer():
        return str(int(num))
    return f"{num:.1f}".rstrip('0').rstrip('.')

def _format_size_text(w, h, sep=' × ', with_unit=False):
    if w in (None, '') or h in (None, ''):
        return ''
    try:
        wf = float(w)
        hf = float(h)
    except (TypeError, ValueError):
        return ''
    if wf <= 0 or hf <= 0:
        return ''
    text = f"{_fmt_measure(wf)}{sep}{_fmt_measure(hf)}"
    return text + (' cm' if with_unit else '')

def _photo_size_text(c, sep=' × ', with_unit=False):
    return _format_size_text(c.get('amplada'), c.get('alcada'), sep=sep, with_unit=with_unit)

def _final_size_text(c, sep=' × ', with_unit=False):
    text = _format_size_text(c.get('final_amplada'), c.get('final_alcada'), sep=sep, with_unit=with_unit)
    if text:
        return text
    ref = (c.get('passpartout') or '').strip()
    if ref and ref not in ('-', 'CONSERVAR'):
        w, h = _parse_dims(ref)
        text = _format_size_text(w, h, sep=sep, with_unit=with_unit)
        if text:
            return text
    return _photo_size_text(c, sep=sep, with_unit=with_unit)

def _display_piece_type(c, t):
    value = (c.get('tipus_peca') or '').strip().lower()
    if not value:
        return ''
    labels = {
        'fotografia': t['piece_photo'],
        'lamina': t['piece_lamina'],
        'pintura_sense_bastidor': t['piece_painting_unstretched'],
        'pintura_amb_bastidor': t['piece_painting_stretched'],
        'puzzle': t['piece_puzzle'],
        'punt_de_creu': t['piece_cross_stitch'],
    }
    return labels.get(value, value.replace('_', ' ').title())


def _display_piece_detail(c, lang='ca'):
    lang = (lang or 'ca').lower()
    piece_type = (c.get('tipus_peca') or '').strip().lower()
    detail = (c.get('tipus_peca_detall') or '').strip().lower()
    labels = {
        'ca': {
            ('fotografia', 'client'): 'La porta el client',
            ('fotografia', 'laboratori'): 'La imprimeix el laboratori',
            ('puzzle', 'sobre_base'): 'Ja va sobre una base',
            ('puzzle', 'sense_base'): 'Cal afegir suport abans d’emmarcar-lo',
            ('pintura_sense_bastidor', ''): 'Es resol amb encolat i es pot protegir amb vidre',
            ('pintura_amb_bastidor', ''): 'Pot portar vidre si es vol protegir la peça',
        },
        'es': {
            ('fotografia', 'client'): 'La trae el cliente',
            ('fotografia', 'laboratori'): 'La imprime el laboratorio',
            ('puzzle', 'sobre_base'): 'Ya va sobre una base',
            ('puzzle', 'sense_base'): 'Hay que añadir soporte antes de enmarcarlo',
            ('pintura_sense_bastidor', ''): 'Se trabaja con encolado y puede protegerse con vidrio',
            ('pintura_amb_bastidor', ''): 'Puede llevar vidrio si se quiere proteger la pieza',
        },
        'en': {
            ('fotografia', 'client'): 'Provided by the client',
            ('fotografia', 'laboratori'): 'Printed by the lab',
            ('puzzle', 'sobre_base'): 'Already mounted on a backing board',
            ('puzzle', 'sense_base'): 'A backing support must be added before framing',
            ('pintura_sense_bastidor', ''): 'Mounted with adhesive and optionally protected with glass',
            ('pintura_amb_bastidor', ''): 'Glass can be added if the piece needs protection',
        },
    }
    current = labels.get(lang, labels['ca'])
    return current.get((piece_type, detail)) or current.get((piece_type, '')) or ''

def _display_muntatge(c, t):
    ref = (c.get('encolat') or '').strip().upper()
    if not ref or ref == '-':
        return ''
    # Puzzle: el muntatge sempre porta base de foam. El preset "foam + laminat"
    # es desa amb ref de protter (foam+laminat), però l'etiquetem com a foam+laminat.
    if (c.get('tipus_peca') or '') == 'puzzle' and ref.startswith('PRO'):
        return f"{t['encolat_label']} + {t['laminat_label']}"
    if ref.startswith('LAM'):
        return t['laminat_label']
    if ref.startswith('PRO'):
        return t['protter']
    return t['encolat_label']

def _display_proteccio(c, t):
    ref = (c.get('vidre') or '').strip().upper()
    if not ref or ref == '-':
        return ''
    if ref == 'CONSERVAR':
        return t['conservar_vidre']
    if ref.startswith('MIR'):
        return t['mirall']
    if ref.startswith('DV-'):
        return t['doble_vidre']
    return t['vidre_label']

def _display_interior(c, t):
    ref = (c.get('passpartout') or '').strip().upper()
    if not ref or ref == '-':
        return ''
    if ref.startswith('DOBPAS'):
        base = t['doble_passpartu']
    elif ref.startswith('PROECO'):
        base = t['proeco_label']
    else:
        base = t['passpartu_label']
    color_ref = (c.get('passpartu_ref') or '').strip()
    if color_ref:
        return f'{base} ({color_ref})'
    return base

def _display_revers_peu(c, t):
    raw = c.get('revers_peu')
    if isinstance(raw, bool):
        enabled = raw
    else:
        enabled = str(raw or '').strip().lower() in ('1', 'true', 'yes', 'on')
    if not enabled:
        return ''
    price = _safe_float(c.get('revers_peu_preu'), 0.0)
    return f'{price:.2f} €'

def _display_impressio(c, t):
    ref = (c.get('impressio') or '').strip()
    if not ref or ref == '-':
        return ''
    return t['inclosa']

def _find_closest(rows, w, h, prefix=None):
    """Among all sizes that contain (w, h), return the cheapest one.
    Ties broken by smallest perimeter overshoot to avoid unnecessary waste.
    This guarantees a larger frame never costs more than a smaller one due to
    non-monotonic price tables."""
    candidates = []
    for row in rows:
        ref = row['referencia']
        if prefix and not ref.upper().startswith(prefix.upper()):
            continue
        rw, rh = _parse_dims(ref)
        if rw is None: continue
        best_ov = None
        for fw, fh in [(rw,rh),(rh,rw)]:
            if fw >= w and fh >= h:
                ov = (fw-w)+(fh-h)
                if best_ov is None or ov < best_ov:
                    best_ov = ov
        if best_ov is not None:
            candidates.append((float(row.get('preu', 0) or 0), best_ov, dict(row)))
    if not candidates:
        return None
    # Sort by price first, then by overshoot to minimise waste among equally-priced options
    candidates.sort(key=lambda x: (x[0], x[1]))
    return candidates[0][2]

def _find_closest_area(rows, w, h, prefix=None, exclude_multi=None):
    """Closest by surface area — works for all products"""
    target = w * h
    best, best_diff = None, float('inf')
    for row in rows:
        ref = row['referencia']
        if prefix and not ref.upper().startswith(prefix.upper()): continue
        if exclude_multi and any(ref.upper().startswith(e.upper()) for e in exclude_multi): continue
        rw, rh = _parse_dims(ref)
        if rw is None: continue
        diff = abs(rw * rh - target)
        if diff < best_diff:
            best_diff = diff
            best = dict(row)
    return best

def _find_min_contain(rows, w, h, prefix=None):
    """Among all formats that physically contain (w, h), return the cheapest.
    Ties broken by smallest area to minimise waste.
    Fallback: largest available if nothing fits."""
    candidates = []
    for row in rows:
        ref = row['referencia']
        if prefix and not ref.upper().startswith(prefix.upper()): continue
        rw, rh = _parse_dims(ref)
        if rw is None: continue
        fits = (rw >= w and rh >= h) or (rh >= w and rw >= h)
        if fits:
            candidates.append((float(row.get('preu', 0) or 0), rw * rh, row))
    if candidates:
        candidates.sort(key=lambda x: (x[0], x[1]))
        return candidates[0][2]
    # Fallback: if nothing contains it, take the largest available
    return max(rows, key=lambda r: (_parse_dims(r['referencia'])[0] or 0) * (_parse_dims(r['referencia'])[1] or 0), default=None)

# Papers amb anques de gran format pròpies (refs prefixades a la taula
# `impressio`). El Lustre —i el Silk, que en comparteix preu— usa els refs
# SENSE prefix. La clau interna del paper cotó segueix sent 'baryta' (de cara
# al client es diu Hahnemühle Photo Rag Baryta).
IMP_PAPER_PREFIX = {'baryta': 'BARYTA', 'poster_mate': 'MATE'}


def _imp_closest(fw, fh, paper='lustre'):
    """Tarifa d'impressió fotogràfica amb lògica híbrida + threshold:
      1. Min-contain sobre la taula 'impressio'.
      2. Si fila trobada i àrea_taula / àrea_sol ≤ encolat_ratio_max
         (default 1.40) → usar preu de taula directament.
      3. Si massa gran o no hi ha → calcular per fórmula:
            cost_real = àrea_sol · cost_cm2(paper)
            factor    = preu(ref_calibration) / (àrea_calib · cost_cm2)
            preu      = round(cost_real · factor, 2)
         on ref_calibration és la fila de taula amb àrea més propera per
         sota a la sol·licitada (fallback: la fila més petita disponible).

    Retorna {ref, preu, origen, area} o None si no hi ha cap fila ni
    càlcul possible. /api/closest enriqueix amb tram, marge, pvp."""
    all_rows = [dict(r) for r in query('SELECT * FROM impressio')] or []
    if not all_rows:
        return None
    # Filtratge per paper: cada paper de gran format (baryta/poster_mate) té les
    # seves anques prefixades; el Lustre/Silk usa les no-prefixades (que inclouen
    # les anques petites històriques, intactes). Fallback al joc del Lustre si un
    # paper encara no s'ha sembrat.
    _known_pfx = tuple(v + '-' for v in IMP_PAPER_PREFIX.values())
    _pfx = IMP_PAPER_PREFIX.get(paper)
    if _pfx:
        rows = [r for r in all_rows if (r.get('referencia') or '').upper().startswith(_pfx + '-')]
        if not rows:
            rows = [r for r in all_rows if not (r.get('referencia') or '').upper().startswith(_known_pfx)]
    else:
        rows = [r for r in all_rows if not (r.get('referencia') or '').upper().startswith(_known_pfx)]
    if not rows:
        rows = all_rows

    area_sol = max(1.0, float(fw) * float(fh))
    ratio_max = float(get_config_value('encolat_ratio_max', '1.40'))

    # 0) Sistema de trams per paper (override del càlcul híbrid)
    #    Si imp_{paper}_trams_actius == '1' s'aplica cost·multiplicador segons àrea.
    if get_config_value(f'imp_{paper}_trams_actius', '0') == '1':
        cost_cm2_t = float(get_config_value(f'imp_{paper}_cost_cm2',
                                            get_config_value('imp_lustre_cost_cm2', '0.000703')))
        cost_real_t = area_sol * cost_cm2_t
        mult_aplicat = None
        for i in range(1, 7):
            max_str = get_config_value(f'imp_{paper}_t{i}_max', None)
            mult_str = get_config_value(f'imp_{paper}_t{i}_mult', None)
            if mult_str is None:
                continue
            try:
                mult = float(mult_str)
            except (TypeError, ValueError):
                continue
            if max_str is None or str(max_str).strip() == '':
                mult_aplicat = mult
                break
            try:
                max_area = float(max_str)
            except (TypeError, ValueError):
                continue
            if area_sol <= max_area:
                mult_aplicat = mult
                break
        if mult_aplicat is None:
            mult_aplicat = float(get_config_value(f'imp_{paper}_t6_mult', '3.1'))
        preu_tram = round(cost_real_t * mult_aplicat, 2)
        return {
            'ref': f'imp-{paper}-{int(fw)}x{int(fh)}',
            'preu': preu_tram,
            'origen': 'tram',
            'area': round(area_sol, 2),
        }

    # 1) Min-contain (només per a la ref de referència al resultat)
    fila = _find_min_contain(rows, fw, fh)

    # 2) Impressió s'imprimeix de BOBINA (43, 60, 111 cm) — el cost és
    #    proporcional a l'àrea real, no a mides de full fix. Per tant
    #    SEMPRE usem fórmula per calcular el preu. Les refs del catàleg
    #    serveixen com a punts de calibració, no com a preus finals.
    #    (Altres productes com vidre o passpartú sí usen el threshold
    #    perquè es tallen de planchas fixes.)

    # 3) Fórmula amb DOBLE calibració (Option B): per a una mida
    # sol·licitada agafem la ref de catàleg més propera per SOTA i la
    # més propera per SOBRE en àrea. Calculem el preu a partir de
    # cadascuna escalant linealment per cm² i tornem el MAX de les
    # dues. Per què: la calibració amb ref menor sola subestima el
    # preu de mides intermèdies (el €/cm² creix amb àrea per malbarat
    # de paper i feina). La ref major dóna un sostre raonable. El max
    # garanteix que el preu MAI baixa per sota de la pendent del
    # catàleg en cap dels dos extrems.
    cost_cm2 = float(get_config_value(f'imp_{paper}_cost_cm2',
                                      get_config_value('imp_lustre_cost_cm2', '0.000703')))
    cost_real = area_sol * cost_cm2

    calib_low = None    # max àrea entre les ≤ area_sol
    calib_high = None   # min àrea entre les ≥ area_sol
    smallest = None
    for r in rows:
        try:
            preu_r = float(r.get('preu') or 0)
        except (TypeError, ValueError):
            continue
        if preu_r <= 0:
            continue
        rw, rh = _parse_dims(r.get('referencia') or '')
        if not rw or not rh:
            continue
        a = rw * rh
        if smallest is None or a < smallest[1]:
            smallest = (r, a, preu_r)
        if a <= area_sol and (calib_low is None or a > calib_low[1]):
            calib_low = (r, a, preu_r)
        if a >= area_sol and (calib_high is None or a < calib_high[1]):
            calib_high = (r, a, preu_r)

    if calib_low is None and calib_high is None:
        calib_low = smallest
    if calib_low is None and calib_high is None:
        return None  # no hi ha cap fila utilitzable per calibrar

    def _preu_from_calib(calib):
        if calib is None:
            return None
        _r, ca, cp = calib
        f = cp / (ca * cost_cm2) if ca > 0 and cost_cm2 > 0 else 1.0
        return round(cost_real * f, 2)

    candidates = [p for p in (_preu_from_calib(calib_low), _preu_from_calib(calib_high))
                  if p is not None]
    preu_formula = max(candidates) if candidates else 0.0

    # Blindatge de l'extrapolació cap avall: si la mida sol·licitada és més petita
    # que TOTES les anques d'aquest paper (cas dels papers només de gran format,
    # com Hahnemühle i Pòster Mate), no la venem més barata que la seva anca més
    # petita. Evita preus irrisoris si algú tria aquest paper per a una còpia
    # petita. El Lustre té anques petites → calib_low no és None → no l'afecta.
    if calib_low is None and calib_high is not None:
        try:
            preu_floor = round(float(calib_high[2]), 2)
            if preu_formula < preu_floor:
                preu_formula = preu_floor
        except (TypeError, ValueError):
            pass

    # Si existeix una ref exacta al catàleg per aquesta mida, la usem
    # com a etiqueta (millor per a facturació); si no, generem imp-WxH.
    if fila:
        rw, rh = _parse_dims(fila['referencia'])
        if rw and rh and rw * rh == area_sol:
            ref_label = fila['referencia']
        else:
            ref_label = f'IMP{int(fw)}x{int(fh)}'
    else:
        ref_label = f'IMP{int(fw)}x{int(fh)}'

    return {
        'ref': ref_label,
        'preu': preu_formula,
        'origen': 'formula',
        'area': round(area_sol, 2),
    }


def _laminate_only_closest(fw, fh):
    rows = []
    for ref, price in LAMINATE_ONLY_PRICES.items():
        rows.append({'referencia': ref, 'preu': float(price)})
    r = _find_min_contain(rows, fw, fh)
    if r:
        return {'ref': r['referencia'], 'preu': r.get('preu', 0)}
    return None

@app.route('/api/closest')
@login_required
def api_closest():
    w = float(request.args.get('w', 0))
    h = float(request.args.get('h', 0))
    foto_w = float(request.args.get('foto_w', w))
    foto_h = float(request.args.get('foto_h', h))
    tipus_laminat = request.args.get('laminat', 'semibrillo')  # 'semibrillo' | 'mate'
    # Paper d'impressió fotogràfica: 'lustre' (default) o 'baryta' (premium).
    # Si el client envia un altre valor, ho ignorem i caem a 'lustre'.
    paper = (request.args.get('paper') or 'lustre').strip().lower()
    if paper not in ('lustre', 'silk', 'baryta', 'poster_mate'):
        paper = 'lustre'
    if w <= 0 or h <= 0:
        return jsonify({})

    def _build_result(r, preu_col='preu'):
        """Build a result dict with preu_cost only for admins."""
        res = {'ref': r['referencia'], 'preu': r.get(preu_col, 0)}
        if session.get('is_admin') and r.get('preu_cost') is not None:
            res['preu_cost'] = float(r['preu_cost'])
        return res

    def closest(table, prefix=None, preu_col='preu', exclude_prefix=None, exclude_multi=None, w_override=None, h_override=None):
        rows = [dict(r) for r in query(f'SELECT * FROM {table}')]
        if exclude_multi:
            rows = [r for r in rows if not any(r['referencia'].upper().startswith(e.upper()) for e in exclude_multi)]
        elif exclude_prefix:
            rows = [r for r in rows if not r['referencia'].upper().startswith(exclude_prefix.upper())]
        uw = w_override if w_override is not None else w
        uh = h_override if h_override is not None else h
        r = _find_closest(rows, uw, uh, prefix)
        return _build_result(r, preu_col) if r else None

    def ca(table, fw, fh, prefix=None, exclude_multi=None, preu_col='preu'):
        """Closest by surface area for all products"""
        rows = [dict(r) for r in query(f'SELECT * FROM {table}')]
        r = _find_closest_area(rows, fw, fh, prefix=prefix, exclude_multi=exclude_multi)
        return _build_result(r, preu_col) if r else None

    # Tots els productes físics (vidre, passpartú, encolat, protter) usen
    # min-contain: la mida ha de cobrir físicament el marc. Això garanteix
    # que una peça més gran no pot sortir mai més barata que una de més petita.
    # Impressió: min-contain també (el paper ha de cobrir la foto).
    def cc(table, cw, ch, prefix=None, exclude_multi=None, preu_col='preu'):
        """Closest by min overshoot (must contain the size)"""
        rows = [dict(r) for r in query(f'SELECT * FROM {table}')]
        if exclude_multi:
            rows = [r for r in rows if not any(r['referencia'].upper().startswith(e.upper()) for e in exclude_multi)]
        r = _find_closest(rows, cw, ch, prefix)
        return _build_result(r, preu_col) if r else None

    def _pvd_result(res, categoria):
        """Add pvd field and update preu from preu_cost if available."""
        if res and res.get('preu_cost') is not None:
            pvd = calcular_pvd(res['preu_cost'], categoria)
            if pvd is not None:
                res['pvd'] = pvd
                res['preu'] = pvd  # backward compat: JS reads 'preu'
        return res

    def _fn_result(r):
        """Format foam/laminat/passpartu result for closest API."""
        res = {'ref': r['ref'], 'preu': r['pvd'], 'pvd': r['pvd'], 'origen': r['origen']}
        if session.get('is_admin'):
            res['preu_cost'] = r['cost']
        return res

    # Foam (encolat simple). ProEco: àlies de compatibilitat històrica.
    foam = _fn_result(calcular_cost_foam(w, h))
    # Laminat sol (sense foam): retornem ambdós acabats
    laminat_semi = _fn_result(calcular_cost_laminat(w, h, tipus='semibrillo'))
    laminat_mate = _fn_result(calcular_cost_laminat(w, h, tipus='mate'))
    # Protter = foam + laminat: retornem ambdós acabats
    protter_semi = _fn_result(calcular_cost_protter(w, h, tipus='semibrillo'))
    protter_mate = _fn_result(calcular_cost_protter(w, h, tipus='mate'))
    protter_actual = protter_mate if tipus_laminat == 'mate' else protter_semi
    laminat_actual = laminat_mate if tipus_laminat == 'mate' else laminat_semi

    result = {
        'encolat':      foam,
        'proeco':       foam,  # àlies històric, mateix cost
        'protter':      protter_actual,
        'protter_semi': protter_semi,
        'protter_mate': protter_mate,
        'laminat':      laminat_actual,
        'laminat_semi': laminat_semi,
        'laminat_mate': laminat_mate,
        'vidre':        _fn_result(calcular_cost_vidre(w, h)),
        'doble_vidre':  _fn_result(calcular_cost_doble_vidre(w, h)),
        'mirall':       _fn_result(calcular_cost_mirall(w, h)),
        'passpartu':    _fn_result(calcular_cost_passpartu(w, h, tipus='simple')),
        'doble_pas':    _fn_result(calcular_cost_passpartu(w, h, tipus='doble')),
        'impressio':    _imp_closest(foto_w, foto_h, paper=paper),
        'paper':        paper,
    }

    # Enriquir la impressió amb el marge de tram aplicat (PVD→PVP).
    # 'preu' segueix sent el PVD (per compat amb consumidors antics);
    # afegim 'pvp' (preu·(1+marge/100)), 'marge_aplicat' i 'tram'.
    imp_res = result.get('impressio')
    if imp_res:
        usuari_actual = query('SELECT * FROM usuaris WHERE id=?', [session.get('user_id')], one=True)
        tram_info = get_marge_impressio_tram(foto_w * foto_h, usuari_actual)
        preu_pvd = float(imp_res.get('preu') or 0)
        marge_aplicat = float(tram_info['marge'])
        imp_res['marge_aplicat'] = marge_aplicat
        imp_res['tram'] = tram_info['tram']
        imp_res['area'] = round(foto_w * foto_h, 2)
        imp_res['pvp'] = round(preu_pvd * (1 + marge_aplicat / 100), 4)
        if session.get('is_admin'):
            imp_res['preu_cost'] = preu_pvd

    return jsonify(result)

# â"€â"€ Routes: Email (mailto) â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
@app.route('/mailto-data', methods=['POST'])
@login_required
def mailto_data():
    d = request.json or {}
    cid = d.get('comanda_id')
    nom_comerc = d.get('nom_comerc', session.get('nom',''))
    c = _get_comanda_for_session(cid)
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    lang = (c.get('lang') or 'ca').lower()
    MAILTO_T = {
        'ca': {
            'greet': 'Bon dia,',
            'intro': "Us faig arribar el pressupost d'emmarcació per al client {client}.",
            'detail': 'DETALL DE LA COMANDA:',
            'mides': 'Mides peça',
            'mida_final': 'Mida final',
            'tipus_peca': 'Peça a emmarcar',
            'vidre': 'Vidre',
            'passpartout': 'Passpartout/ProEco',
            'encolat': 'Encolat',
            'revers_peu': 'Revers amb peu',
            'impressio': 'Impressió',
            'inclosa': 'Inclosa',
            'encolat_label': 'Encolat',
            'laminat_label': 'Laminat',
            'protter': 'Protter',
            'vidre_label': 'Vidre',
            'doble_vidre': 'Doble vidre',
            'mirall': 'Mirall',
            'passpartu_label': 'Passpartú',
            'doble_passpartu': 'Doble passpartú',
            'proeco_label': 'ProEco',
            'piece_photo': 'Fotografia',
            'piece_lamina': 'Làmina',
            'piece_painting_unstretched': 'Pintura sense bastidor',
            'piece_painting_stretched': 'Pintura amb bastidor',
            'piece_puzzle': 'Puzzle',
            'piece_cross_stitch': 'Punt de creu',
            'preu': 'PREU FINAL: {price:.2f} EUR (IVA inclòs)',
            'pendent': 'Pendent de cobrar: {pend:.2f} EUR',
            'obs': 'Observacions: {obs}',
            'attach': 'Trobareu el pressupost en PDF adjunt.',
            'bye': 'Atentament,',
            'subject': 'Pressupost emmarcació - {client}'
        },
        'es': {
            'greet': 'Buenos días,',
            'intro': 'Os envío el presupuesto de enmarcación para el cliente {client}.',
            'detail': 'DETALLE DEL PEDIDO:',
            'mides': 'Medidas pieza',
            'mida_final': 'Medida final',
            'tipus_peca': 'Pieza a enmarcar',
            'vidre': 'Vidrio',
            'passpartout': 'Passpartout/ProEco',
            'encolat': 'Montaje',
            'revers_peu': 'Reverso con pie',
            'impressio': 'Impresión',
            'inclosa': 'Incluida',
            'encolat_label': 'Encolado',
            'laminat_label': 'Laminado',
            'protter': 'Protter',
            'vidre_label': 'Vidrio',
            'doble_vidre': 'Doble vidrio',
            'mirall': 'Espejo',
            'passpartu_label': 'Passpartú',
            'doble_passpartu': 'Doble passpartú',
            'proeco_label': 'ProEco',
            'piece_photo': 'Fotografía',
            'piece_lamina': 'Lámina',
            'piece_painting_unstretched': 'Pintura sin bastidor',
            'piece_painting_stretched': 'Pintura con bastidor',
            'piece_puzzle': 'Puzzle',
            'piece_cross_stitch': 'Punto de cruz',
            'preu': 'PRECIO FINAL: {price:.2f} EUR (IVA incluido)',
            'pendent': 'Pendiente de cobro: {pend:.2f} EUR',
            'obs': 'Observaciones: {obs}',
            'attach': 'Encontraréis el presupuesto en PDF adjunto.',
            'bye': 'Atentamente,',
            'subject': 'Presupuesto enmarcación - {client}'
        },
        'en': {
            'greet': 'Good morning,',
            'intro': 'Please find the framing quote for client {client}.',
            'detail': 'QUOTE DETAILS:',
            'mides': 'Item size',
            'mida_final': 'Final size',
            'tipus_peca': 'Item to frame',
            'vidre': 'Glass',
            'passpartout': 'Passpartout/ProEco',
            'encolat': 'Mounting',
            'revers_peu': 'Backing with stand',
            'impressio': 'Print',
            'inclosa': 'Included',
            'encolat_label': 'Mounting',
            'laminat_label': 'Laminate only',
            'protter': 'Protter',
            'vidre_label': 'Glass',
            'doble_vidre': 'Double glass',
            'mirall': 'Mirror',
            'passpartu_label': 'Mat',
            'doble_passpartu': 'Double mat',
            'proeco_label': 'ProEco',
            'piece_photo': 'Photograph',
            'piece_lamina': 'Poster',
            'piece_painting_unstretched': 'Unstretched painting',
            'piece_painting_stretched': 'Stretched painting',
            'piece_puzzle': 'Puzzle',
            'piece_cross_stitch': 'Cross stitch',
            'preu': 'FINAL PRICE: {price:.2f} EUR (VAT included)',
            'pendent': 'Outstanding: {pend:.2f} EUR',
            'obs': 'Notes: {obs}',
            'attach': 'The quote PDF is attached.',
            'bye': 'Kind regards,',
            'subject': 'Framing quote - {client}'
        }
    }
    tt = MAILTO_T.get(lang, MAILTO_T['ca'])
    final_size = _final_size_text(c, sep=' x ', with_unit=True)
    photo_size = _photo_size_text(c, sep=' x ', with_unit=True)
    piece_type = _display_piece_type(c, tt)
    piece_detail = _display_piece_detail(c, lang)
    if piece_type and piece_detail:
        piece_type = f"{piece_type} · {piece_detail}"
    proteccio_label = _display_proteccio(c, tt)
    interior_label = _display_interior(c, tt)
    revers_peu_label = _display_revers_peu(c, tt)
    muntatge_label = _display_muntatge(c, tt)
    impressio_label = _display_impressio(c, tt)
    lines = [
        tt['greet'],
        "",
        tt['intro'].format(client=c['client_nom']),
        "",
        tt['detail'],
    ]
    if piece_type:
        lines.append(f"  {tt['tipus_peca']}: {piece_type}")
    if final_size:
        lines.append(f"  {tt['mida_final']}: {final_size}")
    if photo_size:
        lines.append(f"  {tt['mides']}: {photo_size}")
    if proteccio_label:
        lines.append(f"  {tt['vidre']}: {proteccio_label}")
    if interior_label:
        lines.append(f"  {tt['passpartout']}: {interior_label}")
    if revers_peu_label:
        lines.append(f"  {tt['revers_peu']}: {revers_peu_label}")
    if muntatge_label:
        lines.append(f"  {tt['encolat']}: {muntatge_label}")
    if impressio_label:
        lines.append(f"  {tt['impressio']}: {tt['inclosa']}")
    lines += [
        "",
        tt['preu'].format(price=c['preu_final']),
    ]
    if c['pendent'] and float(c['pendent']) > 0.01:
        lines.append(tt['pendent'].format(pend=c['pendent']))
    if c['observacions']:
        lines.append(tt['obs'].format(obs=c['observacions']))
    lines += [
        "",
        tt['attach'],
        "",
        tt['bye'],
        f"{nom_comerc}",
    ]
    body = "%0D%0A".join(lines)
    subject = tt['subject'].format(client=c['client_nom'])
    mailto = f"mailto:reusrevela@gmail.com?subject={subject}&body={body}"
    return jsonify({'ok': True, 'mailto': mailto})

@app.route('/enviar-email', methods=['POST'])
@login_required
def enviar_email():
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    d = request.json or {}
    cid = d.get('comanda_id')
    nom_comerc = d.get('nom_comerc', session.get('nom',''))
    nota = d.get('nota','')
    c = _get_comanda_for_session(cid)
    if not c:
        return jsonify({'ok': False, 'error': 'No autoritzat'}), 403
    cfg = {r['clau']: r['valor'] for r in query('SELECT * FROM config')}
    gmail_user = cfg.get('gmail_user','')
    gmail_pass = cfg.get('gmail_pass','')
    if not gmail_user or not gmail_pass:
        return jsonify({'ok': False, 'error': "Configura el Gmail a l'admin primer."})
    dest = 'reusrevela@gmail.com'
    obs = c['observacions'] or ''
    pdf_lang = PDF_T.get((c.get('lang') or 'ca').lower(), PDF_T['ca'])
    final_size = _final_size_text(c, sep=' x ', with_unit=True) or '-'
    photo_size = _photo_size_text(c, sep=' x ', with_unit=True) or '-'
    piece_type = _display_piece_type(c, pdf_lang)
    piece_detail = _display_piece_detail(c, (c.get('lang') or 'ca').lower())
    if piece_type and piece_detail:
        piece_type = f"{piece_type} · {piece_detail}"
    proteccio_label = _display_proteccio(c, pdf_lang) or '-'
    interior_label = _display_interior(c, pdf_lang) or '-'
    revers_peu_label = _display_revers_peu(c, pdf_lang) or '-'
    muntatge_label = _display_muntatge(c, pdf_lang) or '-'
    impressio_label = _display_impressio(c, pdf_lang) or '-'
    nota_html = f"<p style='font-family:sans-serif;color:#C8873A'><b>Nota:</b> {nota}</p>" if nota else ""
    html = f"""
    <h2 style='color:#1A6B45;font-family:sans-serif;border-bottom:2px solid #1A6B45;padding-bottom:8px'>
        Nou pressupost d'emmarcacio</h2>
    <p style='font-family:sans-serif;font-size:15px'><b>Comers:</b> {nom_comerc}</p>
    <p style='font-family:sans-serif;font-size:15px'><b>Client:</b> {c['client_nom']} &nbsp;·&nbsp; {c['client_tel'] or '-'}</p>
    <p style='font-family:sans-serif;font-size:15px'><b>Data:</b> {c['data']}</p>
    <hr style='border:1px solid #E5E2DB;margin:12px 0'>
    {"<p style='font-family:sans-serif;font-size:14px'><b>" + pdf_lang['tipus_peca'] + ":</b> " + piece_type + "</p>" if piece_type else ""}
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['mida_final']}:</b> {final_size}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['mides_foto']}:</b> {photo_size}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['muntatge']}:</b> {muntatge_label}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['vidre_mirall']}:</b> {proteccio_label}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['interior']}:</b> {interior_label}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['revers_peu']}:</b> {revers_peu_label}</p>
    <p style='font-family:sans-serif;font-size:14px'><b>{pdf_lang['impressio']}:</b> {impressio_label}</p>
    {"<p style='font-family:sans-serif;font-size:14px'><b>Obs:</b> " + obs + "</p>" if obs else ""}
    <hr style='border:1px solid #E5E2DB;margin:12px 0'>
    <p style='font-family:sans-serif;font-size:16px'><b>Preu final:</b>
        <span style='color:#1A6B45;font-size:22px;font-weight:bold'>{c['preu_final']:.2f} EUR</span></p>
    <p style='font-family:sans-serif;font-size:14px'><b>Pendent:</b> {c['pendent']:.2f} EUR</p>
    {nota_html}
    <p style='font-family:sans-serif;font-size:12px;color:#9E9B94;margin-top:24px'>
        Enviat des de l'aplicacio Objectiu Emmarcacio</p>
    """
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"Pressupost #{cid} - {nom_comerc} - {c['client_nom']}"
        msg['From'] = gmail_user
        msg['To'] = dest
        msg.attach(MIMEText(html, 'html'))
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=8) as s:
            s.login(gmail_user, gmail_pass)
            s.sendmail(gmail_user, dest, msg.as_string())
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

# â"€â"€ Init DB â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€â"€
def init_db():
    with app.app_context():
        db = get_db()
        if USE_PG:
            # Use a SEPARATE connection with autocommit for DDL.
            # Timeouts evitten que un bloqueig a PG penji l'arrencada del worker.
            import psycopg2 as _pg2
            ddl_conn = _pg2.connect(
                DATABASE_URL,
                connect_timeout=10,
                options='-c statement_timeout=30000',
            )
            ddl_conn.autocommit = True
            ddl_cur = ddl_conn.cursor()
            ddl = [
                """CREATE TABLE IF NOT EXISTS usuaris (
                    id SERIAL PRIMARY KEY, username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL, nom TEXT NOT NULL,
                    is_admin INTEGER DEFAULT 0, marge REAL DEFAULT 60,
                    marge_impressio REAL DEFAULT 100, nom_empresa TEXT DEFAULT '',
                    brand_color TEXT DEFAULT '#1A6B45',
                    brand_color_secondary TEXT DEFAULT '#C8873A',
                    brand_color_menu TEXT DEFAULT '#1A6B45',
                    margins_json TEXT DEFAULT '',
                    access_status TEXT DEFAULT 'active',
                    profile_type TEXT DEFAULT 'professional',
                    web_url TEXT DEFAULT '',
                    instagram TEXT DEFAULT '',
                    fiscal_id TEXT DEFAULT '',
                    notes_validacio TEXT DEFAULT '',
                    email TEXT DEFAULT ''
                )""",
                """CREATE TABLE IF NOT EXISTS impressio (
                    referencia TEXT PRIMARY KEY, descripcio TEXT, preu REAL)""",
                """CREATE TABLE IF NOT EXISTS moldures (
                    referencia TEXT PRIMARY KEY, preu_taller REAL, gruix REAL,
                    cost REAL, proveidor TEXT, ref2 TEXT, ubicacio TEXT,
                    descripcio TEXT, foto TEXT)""",
                """CREATE TABLE IF NOT EXISTS vidres (
                    referencia TEXT PRIMARY KEY, preu REAL)""",
                """CREATE TABLE IF NOT EXISTS passpartout (
                    referencia TEXT PRIMARY KEY, preu REAL)""",
                """CREATE TABLE IF NOT EXISTS encolat_pro (
                    referencia TEXT PRIMARY KEY, preu REAL)""",
                """CREATE TABLE IF NOT EXISTS proeco (
                    referencia TEXT PRIMARY KEY, preu REAL)""",
                """CREATE TABLE IF NOT EXISTS comandes (
                    id SERIAL PRIMARY KEY, user_id INTEGER, data TEXT,
                    client_nom TEXT, client_tel TEXT, pre_marc TEXT,
                    marc_principal TEXT, amplada REAL, alcada REAL, copia REAL,
                    encolat TEXT, vidre TEXT, passpartout TEXT, impressio TEXT,
                    revers_peu INTEGER DEFAULT 0,
                    revers_peu_preu REAL DEFAULT 0,
                    tipus_peca TEXT DEFAULT '',
                    tipus_peca_detall TEXT DEFAULT '',
                    final_amplada REAL DEFAULT 0, final_alcada REAL DEFAULT 0,
                    marge REAL, descompte REAL, quantitat REAL,
                    preu_net REAL, preu_final REAL, entrega REAL, pendent REAL,
                    observacions TEXT, sessio_id TEXT,
                    opcio_nom TEXT DEFAULT 'Opció A')""",
                """CREATE TABLE IF NOT EXISTS config (
                    clau TEXT PRIMARY KEY, valor TEXT)""",
            ]
            for s in ddl:
                try:
                    ddl_cur.execute(s)
                    print("DDL OK:", s[:40])
                except Exception as e:
                    print("DDL err:", e)
            for tbl, col, typ in [
                ('comandes','impressio','TEXT'),
                ('comandes','revers_peu','INTEGER DEFAULT 0'),
                ('comandes','revers_peu_preu','REAL DEFAULT 0'),
                ('comandes','tipus_peca',"TEXT DEFAULT ''"),
                ('comandes','tipus_peca_detall',"TEXT DEFAULT ''"),
                ('comandes','final_amplada','REAL DEFAULT 0'),
                ('comandes','final_alcada','REAL DEFAULT 0'),
                ('comandes','sessio_id','TEXT'),
                ('comandes','opcio_nom','TEXT'),
                ('usuaris','nom_empresa',"TEXT DEFAULT ''"),
                ('usuaris','empresa_adreca',"TEXT DEFAULT ''"),
                ('usuaris','empresa_tel',"TEXT DEFAULT ''"),
                ('usuaris','brand_color',"TEXT DEFAULT '#1A6B45'"),
                ('usuaris','brand_color_secondary',"TEXT DEFAULT '#C8873A'"),
                ('usuaris','brand_color_menu',"TEXT DEFAULT '#1A6B45'"),
                ('usuaris','margins_json',"TEXT DEFAULT ''"),
                ('usuaris','setup_done','INTEGER DEFAULT 0'),
                ('usuaris','logo_b64','TEXT'),
                ('usuaris','marge_impressio_setup','INTEGER DEFAULT 0'),
                ('usuaris','access_status',"TEXT DEFAULT 'active'"),
                ('usuaris','profile_type',"TEXT DEFAULT 'professional'"),
                ('usuaris','web_url',"TEXT DEFAULT ''"),
                ('usuaris','instagram',"TEXT DEFAULT ''"),
                ('usuaris','fiscal_id',"TEXT DEFAULT ''"),
                ('usuaris','notes_validacio',"TEXT DEFAULT ''"),
                ('comandes','num_pressupost','TEXT'),
                ('comandes','pagat','INTEGER DEFAULT 0'),
                ('comandes','entregat','INTEGER DEFAULT 0'),
                ('comandes','lang','TEXT DEFAULT \'ca\''),
                ('comandes','foto_comanda','TEXT'),
                ('comandes','foto_ts','REAL'),
                ('comandes','passpartu_ref',"TEXT DEFAULT ''"),
                ('comandes','cost_produccio','REAL DEFAULT 0'),
                ('comandes','fd_albara',"TEXT DEFAULT ''"),
                ('usuaris','nom_fiscal',"TEXT DEFAULT ''"),
                # --- v2 price model: cost columns ---
                ('moldures','preu_cost','REAL'),
                ('moldures','merma_pct','REAL DEFAULT 10.0'),
                ('moldures','minim_cm','REAL DEFAULT 100.0'),
                ('moldures','preu_cost_ant','REAL'),
                ('moldures','data_cost','TEXT'),
                ('moldures','usuari_cost_id','INTEGER'),
                ('moldures','notes_cost','TEXT'),
                ('moldures','cost_verificat','INTEGER DEFAULT 0'),
                # --- Control d'stock de marcs (cm lineals) ---
                ('moldures','stock_cm','REAL'),
                ('moldures','stock_min_cm','REAL'),
                ('vidres','preu_cost','REAL'),
                ('vidres','preu_cost_ant','REAL'),
                ('vidres','data_cost','TEXT'),
                ('vidres','usuari_cost_id','INTEGER'),
                ('vidres','notes_cost','TEXT'),
                ('vidres','cost_verificat','INTEGER DEFAULT 0'),
                ('encolat_pro','preu_cost','REAL'),
                ('encolat_pro','preu_cost_ant','REAL'),
                ('encolat_pro','data_cost','TEXT'),
                ('encolat_pro','usuari_cost_id','INTEGER'),
                ('encolat_pro','notes_cost','TEXT'),
                ('encolat_pro','cost_verificat','INTEGER DEFAULT 0'),
                ('passpartout','preu_cost','REAL'),
                ('passpartout','preu_cost_ant','REAL'),
                ('passpartout','data_cost','TEXT'),
                ('passpartout','usuari_cost_id','INTEGER'),
                ('passpartout','notes_cost','TEXT'),
                ('passpartout','cost_verificat','INTEGER DEFAULT 0'),
                # --- v2 price model: order snapshots ---
                ('comandes','cost_unitari','REAL'),
                ('comandes','pvd_unitari','REAL'),
                ('comandes','marge_admin_snap','REAL'),
                ('comandes','marge_pro_snap','REAL'),
                # --- v2 price model: user margin aliases ---
                ('usuaris','marge_pro_pct','REAL DEFAULT 60'),
                ('usuaris','marge_impressio_pro_pct','REAL DEFAULT 100'),
                # Trams de marge per a marcs segons àrea (cm²)
                ('usuaris','mr_tram1_limit','INTEGER DEFAULT 2000'),
                ('usuaris','mr_tram2_limit','INTEGER DEFAULT 6000'),
                ('usuaris','mr_tram1_pct','REAL'),
                ('usuaris','mr_tram2_pct','REAL'),
                ('usuaris','mr_tram3_pct','REAL'),
                ('usuaris','mr_trams_vist','INTEGER DEFAULT 0'),
                # Paper Hahnemühle Photo Rag Baryta — activació per usuari
                ('usuaris','baryta_actiu','INTEGER DEFAULT 0'),
                # Novetats (what's new) ja vistes per l'usuari (ids per comes)
                ('usuaris','novetats_vistes','TEXT DEFAULT \'\''),
            ]:
                try:
                    ddl_cur.execute(f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS {col} {typ}")
                except Exception as e:
                    print("ALTER skip:", e)
            # --- v2: historial_preus_cost audit table ---
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS historial_preus_cost (
                    id SERIAL PRIMARY KEY,
                    taula TEXT NOT NULL,
                    referencia TEXT NOT NULL,
                    preu_cost_antic REAL,
                    preu_cost_nou REAL,
                    usuari_id INTEGER,
                    data TEXT,
                    notes TEXT
                )""")
            except Exception as e:
                print("historial_preus_cost skip:", e)
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS moviments_stock_marc (
                    id SERIAL PRIMARY KEY,
                    referencia TEXT NOT NULL,
                    data TEXT,
                    tipus TEXT NOT NULL,
                    cm REAL,
                    motiu TEXT,
                    albara_num TEXT,
                    usuari_id INTEGER,
                    stock_resultant REAL
                )""")
                ddl_cur.execute("CREATE INDEX IF NOT EXISTS idx_mov_stock_ref ON moviments_stock_marc(referencia, data)")
            except Exception as e:
                print("moviments_stock_marc skip:", e)
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS feedback (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER,
                    tipus TEXT NOT NULL DEFAULT 'millora',
                    missatge TEXT NOT NULL,
                    pagina TEXT,
                    data TEXT,
                    llegit INTEGER DEFAULT 0
                )""")
            except Exception as e:
                print("feedback table skip:", e)
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS lab_sends (
                    id SERIAL PRIMARY KEY,
                    comanda_id INTEGER NOT NULL,
                    canal VARCHAR(20) NOT NULL,
                    destinacio TEXT,
                    filename TEXT,
                    mida_kb INTEGER,
                    ok INTEGER DEFAULT 0,
                    error TEXT,
                    link TEXT,
                    sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    user_id INTEGER
                )""")
            except Exception as e:
                print("lab_sends table skip:", e)
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS audit_log (
                    id SERIAL PRIMARY KEY,
                    actor_user_id INTEGER NOT NULL,
                    actor_username VARCHAR(120),
                    target_user_id INTEGER,
                    target_username VARCHAR(120),
                    action VARCHAR(60) NOT NULL,
                    details TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )""")
            except Exception as e:
                print("audit_log table skip:", e)
            try:
                ddl_cur.execute("""CREATE TABLE IF NOT EXISTS clients_externs (
                    id SERIAL PRIMARY KEY,
                    nom VARCHAR(255) NOT NULL,
                    nif VARCHAR(30),
                    fd_contact_id VARCHAR(100) NOT NULL UNIQUE,
                    actiu BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )""")
                ddl_cur.execute("ALTER TABLE comandes ADD COLUMN IF NOT EXISTS client_extern_id INTEGER")
            except Exception as e:
                print("clients_externs table skip:", e)
            ddl_cur.close()
            ddl_conn.close()
            print("DDL done, checking admin...")
            # Now use normal connection for DML
            cur = db.cursor()
            cur.execute("INSERT INTO config (clau,valor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                       ['marge_defecte','60'])
            cur.execute("INSERT INTO config (clau,valor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                       ['empresa_nom','Reus Revela'])
            cur.execute("INSERT INTO config (clau,valor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                       ['empresa_adreca',''])
            cur.execute("INSERT INTO config (clau,valor) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                       ['empresa_tel',''])
            # --- v2 price model: admin margin config ---
            for k, v in [('marge_admin_moldures_pct','60'), ('marge_admin_vidres_pct','60'),
                         ('marge_admin_passpartu_pct','60'), ('marge_admin_encolat_pct','60'),
                         ('marge_pvp_suggerit_pct','40'),
                         ('marge_pro_actiu','1'),
                         ('cost_hora_taller','25.00'),
                         ('passpartu_temps_simple','9'),
                         ('passpartu_temps_doble','16'),
                         ('passpartu_temps_finestra','3.5'),
                         ('passpartu_cost_cm2','0.000620'),
                         ('passpartu_minim_material','0.50'),
                         ('foam_cost_cm2','0.001143'),
                         ('foam_temps_base_min','9'),
                         ('foam_temps_var_cm2','0.0015'),
                         ('laminat_semibrillo_cost_cm2','0.000504'),
                         ('laminat_mate_cost_cm2','0.000685'),
                         ('laminat_temps_base_min','12'),
                         ('laminat_temps_var_cm2','0.0012'),
                         ('vidre_cost_cm2','0.002880'),
                         ('vidre_temps_base_min','3'),
                         ('vidre_temps_lineal_m','0.5'),
                         ('vidre_dv_muntatge_min','5'),
                         ('vidre_dv_muntatge_eur','1.30'),
                         ('mirall_cost_cm2','0.003153'),
                         ('mirall_multiplo_dm2','6'),
                         ('vidre_tolerancia_cm','2'),
                         ('passpartu_tolerancia_cm','2'),
                         ('encolat_tolerancia_cm','2'),
                         ('encolat_ratio_max','1.40'),
                         ('imp_lustre_cost_cm2','0.000703'),
                         ('imp_silk_cost_cm2','0.000756'),
                         ('imp_matte_cost_cm2','0.000447'),
                         ('imp_baryta_cost_cm2','0.005351'),
                         ('imp_baryta_trams_actius','1'),
                         ('imp_baryta_t1_max','300'),
                         ('imp_baryta_t1_mult','9.5'),
                         ('imp_baryta_t2_max','900'),
                         ('imp_baryta_t2_mult','6.5'),
                         ('imp_baryta_t3_max','2000'),
                         ('imp_baryta_t3_mult','4.3'),
                         ('imp_baryta_t4_max','4000'),
                         ('imp_baryta_t4_mult','3.5'),
                         ('imp_baryta_t5_max','8000'),
                         ('imp_baryta_t5_mult','3.2'),
                         ('imp_baryta_t6_mult','3.1'),
                         ('combo_desc_marc_imp_protter','6'),
                         ('combo_desc_marc_imp_foam','5'),
                         ('combo_desc_marc_imp','3'),
                         ('combo_desc_marc_suport','3'),
                         ('combo_desc_minim_pvp','80'),
                         # Trams àrea per al marge d'impressions (cm²)
                         ('imp_tram1_area','900'),
                         ('imp_tram2_area','2000'),
                         ('imp_tram3_area','4200'),
                         ('imp_tram4_area','6000'),
                         ('imp_tram5_area','14400'),
                         # Trams admin (Objectiu Fotògrafs) — referència
                         ('imp_tram1_marge_admin','140'),
                         ('imp_tram2_marge_admin','130'),
                         ('imp_tram3_marge_admin','120'),
                         ('imp_tram4_marge_admin','100'),
                         ('imp_tram5_marge_admin','90'),
                         ('imp_tram6_marge_admin','80'),
                         # Trams per defecte distribuïdor nou
                         ('imp_tram1_marge_default','80'),
                         ('imp_tram2_marge_default','75'),
                         ('imp_tram3_marge_default','70'),
                         ('imp_tram4_marge_default','60'),
                         ('imp_tram5_marge_default','50'),
                         ('imp_tram6_marge_default','45')]:
                cur.execute("INSERT INTO config (clau,valor) VALUES (%s,%s) ON CONFLICT DO NOTHING", [k, v])
            db.commit()
            _seed_admin_if_configured(db)
            # NOTA: tota operació pesada (_seed_proeco_preus, _seed_intermol_moldures,
            # _fix_ref2_errors, _run_v2_price_backfill) s'invoca ara només des de
            # /admin/run-migrations. init_db() es queda amb CREATE/ALTER + config seeds
            # + admin bootstrap perquè l'arrencada del worker no penji a PG.
            db.commit()
        else:
            db.executescript('''
                CREATE TABLE IF NOT EXISTS usuaris (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password TEXT NOT NULL,
                    nom TEXT NOT NULL,
                    is_admin INTEGER DEFAULT 0,
                    marge REAL DEFAULT 60,
                    marge_impressio REAL DEFAULT 100,
                    nom_empresa TEXT DEFAULT '',
                    brand_color TEXT DEFAULT '#1A6B45',
                    brand_color_secondary TEXT DEFAULT '#C8873A',
                    brand_color_menu TEXT DEFAULT '#1A6B45',
                    margins_json TEXT DEFAULT '',
                    access_status TEXT DEFAULT 'active',
                    profile_type TEXT DEFAULT 'professional',
                    web_url TEXT DEFAULT '',
                    instagram TEXT DEFAULT '',
                    fiscal_id TEXT DEFAULT '',
                    notes_validacio TEXT DEFAULT '',
                    email TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS impressio (
                    referencia TEXT PRIMARY KEY, descripcio TEXT, preu REAL
                );
                CREATE TABLE IF NOT EXISTS moldures (
                    referencia TEXT PRIMARY KEY,
                    preu_taller REAL, gruix REAL, cost REAL,
                    proveidor TEXT, ref2 TEXT, ubicacio TEXT,
                    descripcio TEXT, foto TEXT
                );
                CREATE TABLE IF NOT EXISTS vidres (
                    referencia TEXT PRIMARY KEY, preu REAL
                );
                CREATE TABLE IF NOT EXISTS passpartout (
                    referencia TEXT PRIMARY KEY, preu REAL
                );
                CREATE TABLE IF NOT EXISTS encolat_pro (
                    referencia TEXT PRIMARY KEY, preu REAL
                );
                CREATE TABLE IF NOT EXISTS proeco (
                    referencia TEXT PRIMARY KEY, preu REAL
                );
                CREATE TABLE IF NOT EXISTS comandes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER, data TEXT,
                    client_nom TEXT, client_tel TEXT,
                    pre_marc TEXT, marc_principal TEXT,
                    amplada REAL, alcada REAL, copia REAL,
                    encolat TEXT, vidre TEXT, passpartout TEXT, impressio TEXT,
                    revers_peu INTEGER DEFAULT 0,
                    revers_peu_preu REAL DEFAULT 0,
                    tipus_peca TEXT DEFAULT '',
                    final_amplada REAL DEFAULT 0, final_alcada REAL DEFAULT 0,
                    marge REAL, descompte REAL, quantitat REAL,
                    preu_net REAL, preu_final REAL,
                    entrega REAL, pendent REAL, observacions TEXT,
                    sessio_id TEXT, opcio_nom TEXT DEFAULT 'Opció A'
                );
                CREATE TABLE IF NOT EXISTS config (
                    clau TEXT PRIMARY KEY, valor TEXT
                );
                CREATE TABLE IF NOT EXISTS moviments_stock_marc (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    referencia TEXT NOT NULL,
                    data TEXT,
                    tipus TEXT NOT NULL,
                    cm REAL,
                    motiu TEXT,
                    albara_num TEXT,
                    usuari_id INTEGER,
                    stock_resultant REAL
                );
                CREATE INDEX IF NOT EXISTS idx_mov_stock_ref ON moviments_stock_marc(referencia, data);
            ''')
            db.commit()
            for sql in [
                "ALTER TABLE usuaris ADD COLUMN setup_done INTEGER DEFAULT 0",
                "ALTER TABLE usuaris ADD COLUMN logo_b64 TEXT",
                "ALTER TABLE usuaris ADD COLUMN brand_color TEXT DEFAULT '#1A6B45'",
                "ALTER TABLE usuaris ADD COLUMN brand_color_secondary TEXT DEFAULT '#C8873A'",
                "ALTER TABLE usuaris ADD COLUMN brand_color_menu TEXT DEFAULT '#1A6B45'",
                "ALTER TABLE usuaris ADD COLUMN margins_json TEXT DEFAULT ''",
                "ALTER TABLE usuaris ADD COLUMN marge_impressio_setup INTEGER DEFAULT 0",
                "ALTER TABLE usuaris ADD COLUMN access_status TEXT DEFAULT 'active'",
                "ALTER TABLE usuaris ADD COLUMN profile_type TEXT DEFAULT 'professional'",
                "ALTER TABLE usuaris ADD COLUMN web_url TEXT DEFAULT ''",
                "ALTER TABLE usuaris ADD COLUMN instagram TEXT DEFAULT ''",
                "ALTER TABLE usuaris ADD COLUMN fiscal_id TEXT DEFAULT ''",
                "ALTER TABLE usuaris ADD COLUMN nom_fiscal TEXT DEFAULT ''",
                "ALTER TABLE usuaris ADD COLUMN notes_validacio TEXT DEFAULT ''",
                "ALTER TABLE comandes ADD COLUMN num_pressupost TEXT",
                "ALTER TABLE comandes ADD COLUMN revers_peu INTEGER DEFAULT 0",
                "ALTER TABLE comandes ADD COLUMN revers_peu_preu REAL DEFAULT 0",
                "ALTER TABLE comandes ADD COLUMN pagat INTEGER DEFAULT 0",
                "ALTER TABLE comandes ADD COLUMN entregat INTEGER DEFAULT 0",
                "ALTER TABLE comandes ADD COLUMN lang TEXT DEFAULT 'ca'",
                "ALTER TABLE comandes ADD COLUMN foto_comanda TEXT",
                "ALTER TABLE comandes ADD COLUMN foto_ts REAL",
                "ALTER TABLE comandes ADD COLUMN passpartu_ref TEXT DEFAULT ''",
                "ALTER TABLE comandes ADD COLUMN cost_produccio REAL DEFAULT 0",
                "ALTER TABLE comandes ADD COLUMN fd_albara TEXT DEFAULT ''",
                "ALTER TABLE comandes ADD COLUMN estat TEXT DEFAULT ''",
                "UPDATE comandes SET estat = CASE WHEN entregat=1 THEN 'entregat' WHEN observacions LIKE '%[ACCEPTAT]%' THEN 'produccio' ELSE 'nou' END WHERE estat IS NULL OR estat=''",
                "ALTER TABLE passpartout ADD COLUMN color TEXT DEFAULT ''",
                "ALTER TABLE passpartout ADD COLUMN textura TEXT DEFAULT ''",
                "ALTER TABLE passpartout ADD COLUMN descripcio TEXT DEFAULT ''",
                # --- v2 price model: cost columns ---
                "ALTER TABLE moldures ADD COLUMN preu_cost REAL",
                "ALTER TABLE moldures ADD COLUMN merma_pct REAL DEFAULT 10.0",
                "ALTER TABLE moldures ADD COLUMN minim_cm REAL DEFAULT 100.0",
                "ALTER TABLE moldures ADD COLUMN preu_cost_ant REAL",
                "ALTER TABLE moldures ADD COLUMN data_cost TEXT",
                "ALTER TABLE moldures ADD COLUMN usuari_cost_id INTEGER",
                "ALTER TABLE moldures ADD COLUMN notes_cost TEXT",
                "ALTER TABLE moldures ADD COLUMN cost_verificat INTEGER DEFAULT 0",
                # --- Control d'stock de marcs (cm lineals) ---
                "ALTER TABLE moldures ADD COLUMN stock_cm REAL",
                "ALTER TABLE moldures ADD COLUMN stock_min_cm REAL",
                "ALTER TABLE vidres ADD COLUMN preu_cost REAL",
                "ALTER TABLE vidres ADD COLUMN preu_cost_ant REAL",
                "ALTER TABLE vidres ADD COLUMN data_cost TEXT",
                "ALTER TABLE vidres ADD COLUMN usuari_cost_id INTEGER",
                "ALTER TABLE vidres ADD COLUMN notes_cost TEXT",
                "ALTER TABLE vidres ADD COLUMN cost_verificat INTEGER DEFAULT 0",
                "ALTER TABLE encolat_pro ADD COLUMN preu_cost REAL",
                "ALTER TABLE encolat_pro ADD COLUMN preu_cost_ant REAL",
                "ALTER TABLE encolat_pro ADD COLUMN data_cost TEXT",
                "ALTER TABLE encolat_pro ADD COLUMN usuari_cost_id INTEGER",
                "ALTER TABLE encolat_pro ADD COLUMN notes_cost TEXT",
                "ALTER TABLE encolat_pro ADD COLUMN cost_verificat INTEGER DEFAULT 0",
                "ALTER TABLE passpartout ADD COLUMN preu_cost REAL",
                "ALTER TABLE passpartout ADD COLUMN preu_cost_ant REAL",
                "ALTER TABLE passpartout ADD COLUMN data_cost TEXT",
                "ALTER TABLE passpartout ADD COLUMN usuari_cost_id INTEGER",
                "ALTER TABLE passpartout ADD COLUMN notes_cost TEXT",
                "ALTER TABLE passpartout ADD COLUMN cost_verificat INTEGER DEFAULT 0",
                # --- v2 price model: order snapshots ---
                "ALTER TABLE comandes ADD COLUMN cost_unitari REAL",
                "ALTER TABLE comandes ADD COLUMN pvd_unitari REAL",
                "ALTER TABLE comandes ADD COLUMN marge_admin_snap REAL",
                "ALTER TABLE comandes ADD COLUMN marge_pro_snap REAL",
                # --- v2 price model: user margin aliases ---
                "ALTER TABLE usuaris ADD COLUMN marge_pro_pct REAL DEFAULT 60",
                "ALTER TABLE usuaris ADD COLUMN marge_impressio_pro_pct REAL DEFAULT 100",
                # Trams de marge per a marcs (segons àrea cm²)
                "ALTER TABLE usuaris ADD COLUMN mr_tram1_limit INTEGER DEFAULT 2000",
                "ALTER TABLE usuaris ADD COLUMN mr_tram2_limit INTEGER DEFAULT 6000",
                "ALTER TABLE usuaris ADD COLUMN mr_tram1_pct REAL",
                "ALTER TABLE usuaris ADD COLUMN mr_tram2_pct REAL",
                "ALTER TABLE usuaris ADD COLUMN mr_tram3_pct REAL",
                "ALTER TABLE usuaris ADD COLUMN mr_trams_vist INTEGER DEFAULT 0",
                # Paper Hahnemühle Photo Rag Baryta — activació per usuari
                "ALTER TABLE usuaris ADD COLUMN baryta_actiu INTEGER DEFAULT 0",
                # Novetats (what's new) ja vistes per l'usuari (ids per comes)
                "ALTER TABLE usuaris ADD COLUMN novetats_vistes TEXT DEFAULT ''",
            ]:
                try:
                    db.execute(sql)
                except Exception:
                    pass
            db.commit()
            # --- v2: historial_preus_cost audit table ---
            db.execute("""CREATE TABLE IF NOT EXISTS historial_preus_cost (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                taula TEXT NOT NULL,
                referencia TEXT NOT NULL,
                preu_cost_antic REAL,
                preu_cost_nou REAL,
                usuari_id INTEGER,
                data TEXT,
                notes TEXT
            )""")
            db.execute("""CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                tipus TEXT NOT NULL DEFAULT 'millora',
                missatge TEXT NOT NULL,
                pagina TEXT,
                data TEXT,
                llegit INTEGER DEFAULT 0
            )""")
            db.execute("""CREATE TABLE IF NOT EXISTS lab_sends (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                comanda_id INTEGER NOT NULL,
                canal TEXT NOT NULL,
                destinacio TEXT,
                filename TEXT,
                mida_kb INTEGER,
                ok INTEGER DEFAULT 0,
                error TEXT,
                link TEXT,
                sent_at TEXT,
                user_id INTEGER
            )""")
            db.execute("""CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                actor_user_id INTEGER NOT NULL,
                actor_username TEXT,
                target_user_id INTEGER,
                target_username TEXT,
                action TEXT NOT NULL,
                details TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )""")
            db.execute("""CREATE TABLE IF NOT EXISTS clients_externs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT NOT NULL,
                nif TEXT,
                fd_contact_id TEXT NOT NULL UNIQUE,
                actiu INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )""")
            # client_extern_id a comandes — un try/except per si la taula
            # encara no té la columna (cas update progressiu).
            try:
                db.execute("ALTER TABLE comandes ADD COLUMN client_extern_id INTEGER")
            except Exception:
                pass
            db.commit()
            db.execute("INSERT OR IGNORE INTO config (clau,valor) VALUES ('marge_defecte','60')")
            db.execute("INSERT OR IGNORE INTO config (clau,valor) VALUES ('empresa_nom','Reus Revela')")
            db.execute("INSERT OR IGNORE INTO config (clau,valor) VALUES ('empresa_adreca','')")
            db.execute("INSERT OR IGNORE INTO config (clau,valor) VALUES ('empresa_tel','')")
            # --- v2 price model: admin margin config ---
            for k, v in [('marge_admin_moldures_pct','60'), ('marge_admin_vidres_pct','60'),
                         ('marge_admin_passpartu_pct','60'), ('marge_admin_encolat_pct','60'),
                         ('marge_pvp_suggerit_pct','40'),
                         ('marge_pro_actiu','1'),
                         ('cost_hora_taller','25.00'),
                         ('passpartu_temps_simple','9'),
                         ('passpartu_temps_doble','16'),
                         ('passpartu_temps_finestra','3.5'),
                         ('passpartu_cost_cm2','0.000620'),
                         ('passpartu_minim_material','0.50'),
                         ('foam_cost_cm2','0.001143'),
                         ('foam_temps_base_min','9'),
                         ('foam_temps_var_cm2','0.0015'),
                         ('laminat_semibrillo_cost_cm2','0.000504'),
                         ('laminat_mate_cost_cm2','0.000685'),
                         ('laminat_temps_base_min','12'),
                         ('laminat_temps_var_cm2','0.0012'),
                         ('vidre_cost_cm2','0.002880'),
                         ('vidre_temps_base_min','3'),
                         ('vidre_temps_lineal_m','0.5'),
                         ('vidre_dv_muntatge_min','5'),
                         ('vidre_dv_muntatge_eur','1.30'),
                         ('mirall_cost_cm2','0.003153'),
                         ('mirall_multiplo_dm2','6'),
                         ('vidre_tolerancia_cm','2'),
                         ('passpartu_tolerancia_cm','2'),
                         ('encolat_tolerancia_cm','2'),
                         ('encolat_ratio_max','1.40'),
                         ('imp_lustre_cost_cm2','0.000703'),
                         ('imp_silk_cost_cm2','0.000756'),
                         ('imp_matte_cost_cm2','0.000447'),
                         ('imp_baryta_cost_cm2','0.005351'),
                         ('imp_baryta_trams_actius','1'),
                         ('imp_baryta_t1_max','300'),
                         ('imp_baryta_t1_mult','9.5'),
                         ('imp_baryta_t2_max','900'),
                         ('imp_baryta_t2_mult','6.5'),
                         ('imp_baryta_t3_max','2000'),
                         ('imp_baryta_t3_mult','4.3'),
                         ('imp_baryta_t4_max','4000'),
                         ('imp_baryta_t4_mult','3.5'),
                         ('imp_baryta_t5_max','8000'),
                         ('imp_baryta_t5_mult','3.2'),
                         ('imp_baryta_t6_mult','3.1'),
                         ('combo_desc_marc_imp_protter','6'),
                         ('combo_desc_marc_imp_foam','5'),
                         ('combo_desc_marc_imp','3'),
                         ('combo_desc_marc_suport','3'),
                         ('combo_desc_minim_pvp','80'),
                         ('imp_tram1_area','900'),
                         ('imp_tram2_area','2000'),
                         ('imp_tram3_area','4200'),
                         ('imp_tram4_area','6000'),
                         ('imp_tram5_area','14400'),
                         ('imp_tram1_marge_admin','140'),
                         ('imp_tram2_marge_admin','130'),
                         ('imp_tram3_marge_admin','120'),
                         ('imp_tram4_marge_admin','100'),
                         ('imp_tram5_marge_admin','90'),
                         ('imp_tram6_marge_admin','80'),
                         ('imp_tram1_marge_default','80'),
                         ('imp_tram2_marge_default','75'),
                         ('imp_tram3_marge_default','70'),
                         ('imp_tram4_marge_default','60'),
                         ('imp_tram5_marge_default','50'),
                         ('imp_tram6_marge_default','45')]:
                db.execute("INSERT OR IGNORE INTO config (clau,valor) VALUES (?,?)", [k, v])
            db.commit()
            _seed_admin_if_configured(db)
            # Veure nota a la branca PG: tota operació pesada s'executa ara només
            # via /admin/run-migrations.
            db.commit()


# ── Trams de marge per a marcs ───────────────────────────────────────────
MR_TRAM_LIMITS_DEFAULT = {
    'tram1_limit': 2000,   # ≤ 2.000 cm² (mides petites)
    'tram2_limit': 6000,   # 2.000–6.000 cm² (mides mitjanes)
}
MR_TRAM_DEFAULTS_RECOMANATS = {
    # Fallback absolut quan l'usuari no té marge propi configurat (=NULL o 0).
    # Per a usuaris amb marge configurat, usar get_mr_recomendats(marge_actual).
    'tram1_pct': 70.0,
    'tram2_pct': 60.0,
    'tram3_pct': 50.0,
}


def get_mr_recomendats(marge_actual):
    """Recomanats personalitzats per a un usuari segons el seu marge_pro_pct actual.
    - T1: manté el marge actual (mides petites no canvien).
    - T2: marge_actual × 0.80, mínim 30.
    - T3: marge_actual × 0.65, mínim 20.
    Fallback si NULL/0: defaults absoluts 70/60/50."""
    try:
        m = float(marge_actual or 0)
    except (TypeError, ValueError):
        m = 0
    if m <= 0:
        return {'t1': 70, 't2': 60, 't3': 50}
    return {
        't1': round(m),
        't2': max(round(m * 0.80), 30),
        't3': max(round(m * 0.65), 20),
    }


def get_mr_tram_pct(area_cm2, user):
    """Retorna el percentatge del tram aplicable segons l'àrea del marc.
    Compatible amb dict o sqlite Row. Si trams no configurats, fallback a marge_pro_pct."""
    def _g(key, default=None):
        try:
            v = user[key] if user is not None else None
        except (KeyError, IndexError, TypeError):
            v = None
        if v is None:
            try:
                v = user.get(key) if hasattr(user, 'get') else None
            except Exception:
                v = None
        return v if v is not None else default

    fallback = _g('marge_pro_pct') or _g('marge') or 60.0
    t1_lim = _g('mr_tram1_limit', MR_TRAM_LIMITS_DEFAULT['tram1_limit'])
    t2_lim = _g('mr_tram2_limit', MR_TRAM_LIMITS_DEFAULT['tram2_limit'])
    t1_pct = _g('mr_tram1_pct', fallback)
    t2_pct = _g('mr_tram2_pct', fallback)
    t3_pct = _g('mr_tram3_pct', fallback)
    try:
        a = float(area_cm2 or 0)
    except (TypeError, ValueError):
        a = 0.0
    if a <= float(t1_lim or 0):
        return float(t1_pct if t1_pct is not None else fallback)
    if a <= float(t2_lim or 0):
        return float(t2_pct if t2_pct is not None else fallback)
    return float(t3_pct if t3_pct is not None else fallback)


def _run_mr_trams_backfill(db):
    """Inicialitza els 3 trams de marcs amb el marge_pro_pct actual (= cap canvi de preu el dia 1).
    Es marca com a done amb config['migration_mr_trams_done']=1. Idempotent."""
    check = query("SELECT valor FROM config WHERE clau='migration_mr_trams_done'", one=True)
    if check:
        return
    print("Running mr_trams backfill...")
    # Per cada usuari, inicialitzar els tres trams al seu marge_pro_pct (o marge legacy o 60).
    execute("""UPDATE usuaris SET
        mr_tram1_pct = COALESCE(mr_tram1_pct, marge_pro_pct, marge, 60),
        mr_tram2_pct = COALESCE(mr_tram2_pct, marge_pro_pct, marge, 60),
        mr_tram3_pct = COALESCE(mr_tram3_pct, marge_pro_pct, marge, 60),
        mr_tram1_limit = COALESCE(mr_tram1_limit, 2000),
        mr_tram2_limit = COALESCE(mr_tram2_limit, 6000)
        WHERE mr_tram1_pct IS NULL OR mr_tram2_pct IS NULL OR mr_tram3_pct IS NULL""")
    if USE_PG:
        execute("INSERT INTO config (clau, valor) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                ['migration_mr_trams_done', '1'])
    else:
        execute("INSERT OR IGNORE INTO config (clau, valor) VALUES (?, ?)",
                ['migration_mr_trams_done', '1'])
    db.commit()
    print("mr_trams backfill complete.")


# Tarifa d'impressió de GRAN FORMAT (≥40 cm costat curt, impresa a la Pro 4000).
# PVD (cost/taller), sense IVA. Són les anques de calibració; el motor interpola
# per àrea entre elles. NO inclou mides petites (van per una altra impressora amb
# estructura de cost diferent i es queden com estan al motor).
#   paper_key: (prefix_ref, prefix_descripció, [(w, h, preu_pvd), ...])
IMP_TARIFA_GRANFORMAT = {
    'lustre': ('', 'Lustre', [
        (40, 50, 10.29), (50, 70, 15.88), (60, 80, 21.75), (60, 90, 24.00),
        (70, 100, 32.00), (80, 120, 42.40), (100, 150, 66.21),
    ]),
    'baryta': ('BARYTA-', 'Hahnemühle', [
        (40, 50, 18.29), (50, 70, 30.29), (60, 80, 42.75), (60, 90, 47.63),
        (70, 100, 64.67), (80, 120, 87.20), (100, 150, 138.62),
    ]),
    'poster_mate': ('MATE-', 'Pòster Mate', [
        (40, 50, 8.00), (50, 70, 11.76), (60, 80, 15.75), (60, 90, 17.25),
        (70, 100, 22.67), (80, 120, 29.60), (100, 150, 45.52),
    ]),
}


def _seed_impressio_tarifa_granformat(db, use_pg=False):
    """Sembra/actualitza les anques de gran format dels 3 papers a la taula
    `impressio` (upsert per referència). NO toca les mides petites històriques.
    També desactiva els trams del baryta perquè el paper cotó passi a calibrar
    per anques com el Lustre. Idempotent via flag de versió."""
    if get_config_value('imp_tarifa_gf_v1', '0') == '1':
        return
    print('Seeding impressió gran format (Lustre / Hahnemühle / Pòster Mate)...')
    for paper, (prefix, desc_pref, files) in IMP_TARIFA_GRANFORMAT.items():
        for (w, h, preu) in files:
            ref = f'{prefix}{w}X{h}'
            desc = f'{desc_pref} {w}x{h}'
            existing = query('SELECT 1 FROM impressio WHERE referencia=?', [ref], one=True)
            if existing:
                execute('UPDATE impressio SET descripcio=?, preu=? WHERE referencia=?', [desc, preu, ref])
            else:
                execute('INSERT INTO impressio (referencia, descripcio, preu) VALUES (?,?,?)', [ref, desc, preu])
    # El baryta (Hahnemühle) passa a calibració per anques: desactivem el seu
    # sistema de trams perquè _imp_closest usi les noves anques BARYTA-*.
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('imp_baryta_trams_actius', '0')")
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('imp_poster_mate_cost_cm2', '0.000447')")
    execute("INSERT OR REPLACE INTO config (clau, valor) VALUES ('imp_tarifa_gf_v1', '1')")
    db.commit()
    print('Impressió gran format seeded.')


def _run_v2_price_backfill(db):
    """Populate preu_cost from existing prices and copy marge to new columns. Idempotent."""
    check = query("SELECT valor FROM config WHERE clau='migration_v2_done'", one=True)
    if check:
        return
    print("Running v2 price model backfill...")
    # Read current default margin to reverse-calculate cost
    cfg = query("SELECT valor FROM config WHERE clau='marge_defecte'", one=True)
    divisor = 1 + float(cfg['valor'] if cfg else 60) / 100  # e.g. 1.60

    # PG needs ::numeric cast for ROUND with precision; SQLite handles ROUND(real, int) natively
    if USE_PG:
        rnd = lambda col: f'ROUND(({col} / %s)::numeric, 4)'
    else:
        rnd = lambda col: f'ROUND({col} / ?, 4)'
    # Moldures: preu_cost = preu_taller / divisor
    execute(f"UPDATE moldures SET preu_cost = {rnd('preu_taller')} WHERE preu_cost IS NULL AND preu_taller IS NOT NULL", [divisor])
    # Vidres
    execute(f"UPDATE vidres SET preu_cost = {rnd('preu')} WHERE preu_cost IS NULL AND preu IS NOT NULL", [divisor])
    # Encolat
    execute(f"UPDATE encolat_pro SET preu_cost = {rnd('preu')} WHERE preu_cost IS NULL AND preu IS NOT NULL", [divisor])
    # Passpartout
    execute(f"UPDATE passpartout SET preu_cost = {rnd('preu')} WHERE preu_cost IS NULL AND preu IS NOT NULL", [divisor])
    # Copy existing margins to new alias columns (always overwrite defaults on first migration)
    execute("UPDATE usuaris SET marge_pro_pct = marge WHERE marge IS NOT NULL")
    execute("UPDATE usuaris SET marge_impressio_pro_pct = marge_impressio WHERE marge_impressio IS NOT NULL")
    # Mark migration as done
    if USE_PG:
        execute("INSERT INTO config (clau, valor) VALUES (%s, %s) ON CONFLICT DO NOTHING", ['migration_v2_done', '1'])
    else:
        execute("INSERT OR IGNORE INTO config (clau, valor) VALUES (?, ?)", ['migration_v2_done', '1'])
    db.commit()
    print("v2 price backfill complete.")


# init_db() ja NO s'executa automàticament a cap request. S'invoca explícitament
# des de /admin/run-migrations perquè els workers arranquin sempre sense tocar BD
# (evita penjades d'arrencada a PG). Cal visitar /admin/run-migrations com a admin
# després de cada deploy amb canvis d'schema.

# ============================================================================
#  MAILING ALS CLIENTS DEL TALLER  (campanyes + avisos d'horari/vacances)
#  - Llista de marqueting propia (mailing_contacts), separada de clients_externs
#  - Importacio per enganxat/CSV + sync des de clients_externs
#  - Enviament per Resend en lots (guiat des del client) amb prova previa
#  - Baixa RGPD tokenitzada (/baixa/<token>) + llista de supressio
# ============================================================================

_MAILING_EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')


def _mailing_base_url():
    return (os.environ.get('CALC_BASE_URL', '').strip().rstrip('/')
            or 'https://calculadora.reusrevela.cat')


def _mailing_now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _mailing_valid_email(e):
    e = (e or '').strip().lower()
    return e if _MAILING_EMAIL_RE.match(e) else ''


def _ensure_mailing_schema():
    """Crea les taules de mailing si no existeixen (idempotent i dialecte-aware).
    Es crida a l'inici de cada ruta de mailing perque el panell funcioni sense
    dependre d'una migracio manual."""
    if USE_PG:
        stmts = [
            """CREATE TABLE IF NOT EXISTS mailing_contacts (
                id SERIAL PRIMARY KEY,
                nom VARCHAR(255),
                email VARCHAR(255) NOT NULL UNIQUE,
                idioma VARCHAR(5) DEFAULT 'ca',
                origen VARCHAR(20) DEFAULT 'manual',
                subscrit BOOLEAN DEFAULT TRUE,
                token VARCHAR(64) NOT NULL UNIQUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                unsubscribed_at TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS mailing_campaigns (
                id SERIAL PRIMARY KEY,
                uid VARCHAR(32) NOT NULL UNIQUE,
                assumpte VARCHAR(300) NOT NULL,
                cos_html TEXT NOT NULL,
                tipus VARCHAR(20) DEFAULT 'campanya',
                idioma_filtre VARCHAR(5) DEFAULT 'tot',
                estat VARCHAR(20) DEFAULT 'sending',
                total INTEGER DEFAULT 0,
                enviats INTEGER DEFAULT 0,
                errors INTEGER DEFAULT 0,
                creat_per VARCHAR(120),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                sent_at TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS mailing_sends (
                id SERIAL PRIMARY KEY,
                campaign_id INTEGER NOT NULL,
                contact_id INTEGER,
                email VARCHAR(255) NOT NULL,
                nom VARCHAR(255),
                token VARCHAR(64),
                estat VARCHAR(20) DEFAULT 'pending',
                error TEXT,
                sent_at TIMESTAMP
            )""",
            "CREATE INDEX IF NOT EXISTS idx_mailing_sends_campaign ON mailing_sends(campaign_id, estat)",
        ]
    else:
        stmts = [
            """CREATE TABLE IF NOT EXISTS mailing_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nom TEXT,
                email TEXT NOT NULL UNIQUE,
                idioma TEXT DEFAULT 'ca',
                origen TEXT DEFAULT 'manual',
                subscrit INTEGER DEFAULT 1,
                token TEXT NOT NULL UNIQUE,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                unsubscribed_at TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS mailing_campaigns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uid TEXT NOT NULL UNIQUE,
                assumpte TEXT NOT NULL,
                cos_html TEXT NOT NULL,
                tipus TEXT DEFAULT 'campanya',
                idioma_filtre TEXT DEFAULT 'tot',
                estat TEXT DEFAULT 'sending',
                total INTEGER DEFAULT 0,
                enviats INTEGER DEFAULT 0,
                errors INTEGER DEFAULT 0,
                creat_per TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                sent_at TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS mailing_sends (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                campaign_id INTEGER NOT NULL,
                contact_id INTEGER,
                email TEXT NOT NULL,
                nom TEXT,
                token TEXT,
                estat TEXT DEFAULT 'pending',
                error TEXT,
                sent_at TEXT
            )""",
            "CREATE INDEX IF NOT EXISTS idx_mailing_sends_campaign ON mailing_sends(campaign_id, estat)",
        ]
    for s in stmts:
        try:
            execute(s)
        except Exception as e:
            try:
                get_db().rollback()
            except Exception:
                pass
            print(f"[mailing_schema] skip: {str(e)[:120]}")
    # Columnes afegides despres de la v1 (campanyes bilingues ca/es). A PG
    # son IF NOT EXISTS; a SQLite l'ALTER duplicat falla i s'ignora en silenci.
    if USE_PG:
        alters = [
            "ALTER TABLE mailing_campaigns ADD COLUMN IF NOT EXISTS assumpte_es VARCHAR(300)",
            "ALTER TABLE mailing_campaigns ADD COLUMN IF NOT EXISTS cos_html_es TEXT",
            "ALTER TABLE mailing_sends ADD COLUMN IF NOT EXISTS idioma VARCHAR(5)",
        ]
    else:
        alters = [
            "ALTER TABLE mailing_campaigns ADD COLUMN assumpte_es TEXT",
            "ALTER TABLE mailing_campaigns ADD COLUMN cos_html_es TEXT",
            "ALTER TABLE mailing_sends ADD COLUMN idioma TEXT",
        ]
    for s in alters:
        try:
            execute(s)
        except Exception:
            try:
                get_db().rollback()
            except Exception:
                pass


def _mailing_upsert_contact(nom, email, idioma='ca', origen='manual'):
    """Insereix o actualitza un contacte. MAI reactiva una baixa (no toca
    subscrit). Retorna 'added' | 'updated' | 'skipped'."""
    email = _mailing_valid_email(email)
    if not email:
        return 'skipped'
    existing = query('SELECT id, nom FROM mailing_contacts WHERE email=?', [email], one=True)
    if existing:
        if (nom or '').strip() and not (_row_get(existing, 'nom', '') or '').strip():
            execute('UPDATE mailing_contacts SET nom=? WHERE id=?', [nom.strip(), _row_get(existing, 'id')])
            return 'updated'
        return 'skipped'
    token = secrets.token_urlsafe(24)
    execute('INSERT INTO mailing_contacts (nom, email, idioma, origen, token) VALUES (?,?,?,?,?)',
            [(nom or '').strip(), email, (idioma or 'ca'), origen, token])
    return 'added'


def _mailing_parse_lines(text):
    """Parseja text enganxat: una linia per contacte. Accepta 'email',
    'nom,email', 'nom;email', 'nom <email>' o 'email,nom'. Retorna [(nom,email)]."""
    out = []
    for raw in (text or '').replace('\r', '').split('\n'):
        line = raw.strip().strip(',;').strip()
        if not line:
            continue
        parts = [p.strip() for p in re.split(r'[,;\t]', line) if p.strip()]
        email = ''
        name_parts = []
        for p in parts:
            m = re.search(r'<([^<>@\s]+@[^<>@\s]+)>', p)
            if m:
                email = m.group(1)
                rest = p.replace(m.group(0), '').strip()
                if rest:
                    name_parts.append(rest)
            elif '@' in p and not email:
                email = p
            else:
                name_parts.append(p)
        if email:
            out.append((' '.join(name_parts).strip(), email))
    return out


def _mailing_text_to_html(text):
    """Converteix el text pla que escriu l'admin en HTML segur: escapa,
    paragrafs per linia en blanc, salts simples en <br>."""
    import html as _html
    safe = _html.escape((text or '').strip())
    paras = [p.strip().replace('\n', '<br>') for p in re.split(r'\n\s*\n', safe) if p.strip()]
    return ''.join(f'<p style="margin:0 0 14px">{p}</p>' for p in paras)


def _mailing_render_html(cos_html, contact):
    """Embolcalla el cos amb la plantilla de marca + peu RGPD amb enllac de baixa.
    La marca de cara al client (mailing_marca, p. ex. "Reus Revela") pot ser
    diferent del nom comercial general (empresa_nom) i del nom fiscal que firma
    el text legal (mailing_responsable, p. ex. "OBJECTIU S.C.P.")."""
    empresa = (get_config_value('empresa_nom', 'Reus Revela') or 'Reus Revela').strip()
    marca = (get_config_value('mailing_marca', '') or '').strip() or empresa
    responsable = (get_config_value('mailing_responsable', '') or '').strip() or marca
    adreca = (get_config_value('empresa_adreca', '') or '').strip()
    tel = (get_config_value('empresa_tel', '') or '').strip()
    nom = (_row_get(contact, 'nom', '') or '').strip()
    token = _row_get(contact, 'token', '') or ''
    # {nom} amb contacte sense nom: s'elimina el marcador i l'espai previ,
    # perque "Hola {nom}," quedi "Hola," i no "Hola hola,".
    if nom:
        body = (cos_html or '').replace('{nom}', nom)
    else:
        body = re.sub(r'[ \t]*\{nom\}', '', cos_html or '')
    baixa_url = f"{_mailing_base_url()}/baixa/{token}"
    peu = ' &middot; '.join([x for x in [marca, adreca, tel] if x])
    return (
        '<!DOCTYPE html><html lang="ca"><body style="margin:0;background:#f6f3ee;'
        'padding:24px 12px;font-family:\'Helvetica Neue\',Arial,sans-serif">'
        '<div style="max-width:580px;margin:0 auto;background:#fff;border:1px solid #DDD3C4;'
        'overflow:hidden">'
        # Capcalera segons manual de marca: logo simbol + nom en serif sobre
        # fons blanc amb separador Linia — sense bandes de color.
        '<div style="padding:18px 24px;background:#FFFFFF;border-bottom:1px solid #DDD3C4">'
        '<img src="https://reusrevela.cat/static/img/logo-reusrevela.png" alt="" '
        'width="44" height="44" style="vertical-align:middle;border:0">'
        f'<span style="font-family:Georgia,\'Times New Roman\',serif;font-size:21px;'
        f'color:#12100C;margin-left:12px;vertical-align:middle">{marca}</span></div>'
        f'<div style="padding:24px;color:#1d1b18;font-size:15px;line-height:1.7">{body}</div>'
        '<div style="padding:18px 24px;border-top:1px solid #eee;color:#8d877d;'
        'font-size:12px;line-height:1.6">'
        f'{peu}<br>Reps aquest correu perqu&egrave; ets client del taller. '
        f'Si no en vols rebre m&eacute;s, pots <a href="{baixa_url}" '
        'style="color:#A67843">donar-te de baixa aqu&iacute;</a>.'
        '<br><br>'
        f'<span style="color:#a9a299">Responsable del tractament: {responsable}. '
        'Tractem les teves dades nom&eacute;s per enviar-te comunicacions del taller, '
        'sobre la base del nostre inter&eacute;s leg&iacute;tim com a client. Pots exercir els teus '
        'drets d\'acc&eacute;s, rectificaci&oacute;, supressi&oacute; i oposici&oacute; responent a aquest '
        'correu o donant-te de baixa. M&eacute;s informaci&oacute; a la '
        '<a href="https://reusrevela.cat/politica-de-privacitat" '
        'style="color:#8d877d;text-decoration:underline">Pol&iacute;tica de Privacitat</a>.</span>'
        '</div></div></body></html>'
    )


def _mailing_send_one(to_addr, subject, html, baixa_url):
    """Envia un correu via Resend amb capcalera List-Unsubscribe (deliverability).
    Retorna (ok, error)."""
    api_key = (get_config_value('resend_api_key', '') or '').strip()
    if not api_key:
        return False, 'resend_api_key no configurat'
    from_addr = (get_config_value('resend_from', '') or '').strip() or 'onboarding@resend.dev'
    payload_dict = {
        'from': from_addr,
        'to': [to_addr],
        'subject': subject,
        'html': html,
        'headers': {
            'List-Unsubscribe': f'<{baixa_url}>',
            'List-Unsubscribe-Post': 'List-Unsubscribe=One-Click',
        },
    }
    # Reply-to opcional: les respostes dels clients van on de debò es llegeix
    # el correu (p. ex. reusrevela@gmail.com), no a la adreca remitent sense bustia.
    reply_to = (get_config_value('resend_reply_to', '') or '').strip()
    if reply_to:
        payload_dict['reply_to'] = reply_to
    payload = json.dumps(payload_dict).encode('utf-8')
    req = urllib_request.Request(
        'https://api.resend.com/emails', data=payload,
        headers={
            'Authorization': 'Bearer ' + api_key,
            'Content-Type': 'application/json',
            'User-Agent': 'Calculadora-Marcs/1.0 (+https://calculadora.reusrevela.cat)',
            'Accept': 'application/json',
        }, method='POST')
    try:
        with urllib_request.urlopen(req, timeout=12) as resp:
            resp.read()
        return True, ''
    except urllib_error.HTTPError as e:
        body = ''
        try:
            body = e.read().decode('utf-8', errors='replace')[:200]
        except Exception:
            pass
        return False, f'HTTP {e.code}: {body}'
    except Exception as e:
        return False, str(e)[:200]


@app.route('/admin/mailing')
@admin_required
def admin_mailing():
    _ensure_mailing_schema()
    total = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts', one=True), 'n', 0)
    subs = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts WHERE subscrit', one=True), 'n', 0)
    camps = query('SELECT id, assumpte, assumpte_es, tipus, estat, total, enviats, errors, created_at, sent_at '
                  'FROM mailing_campaigns ORDER BY id DESC LIMIT 20') or []
    return render_template(
        'admin_mailing.html',
        total=total, subscrits=subs, baixes=(total - subs), campaigns=camps,
        resend_ok=_resend_is_configured(),
        resend_from=(get_config_value('resend_from', '') or ''),
    )


@app.route('/admin/mailing/sync-clients', methods=['POST'])
@admin_required
def admin_mailing_sync_clients():
    _ensure_mailing_schema()
    rows = query("SELECT nom, email FROM clients_externs "
                 "WHERE actiu AND email IS NOT NULL AND email <> ''") or []
    added = updated = 0
    for r in rows:
        res = _mailing_upsert_contact(_row_get(r, 'nom', ''), _row_get(r, 'email', ''), origen='clients')
        if res == 'added':
            added += 1
        elif res == 'updated':
            updated += 1
    total = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts', one=True), 'n', 0)
    return jsonify(ok=True, added=added, updated=updated, total=total)


@app.route('/admin/mailing/import', methods=['POST'])
@admin_required
def admin_mailing_import():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    idioma = (data.get('idioma') or 'ca').strip().lower()
    if idioma not in ('ca', 'es'):
        idioma = 'ca'
    parsed = _mailing_parse_lines(data.get('text', ''))
    added = skipped = invalid = 0
    seen = set()
    for nom, email in parsed:
        ve = _mailing_valid_email(email)
        if not ve:
            invalid += 1
            continue
        if ve in seen:
            skipped += 1
            continue
        seen.add(ve)
        res = _mailing_upsert_contact(nom, ve, idioma=idioma, origen='import')
        if res == 'added':
            added += 1
        else:
            skipped += 1
    total = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts', one=True), 'n', 0)
    return jsonify(ok=True, added=added, skipped=skipped, invalid=invalid, total=total)


@app.route('/admin/mailing/contacts')
@admin_required
def admin_mailing_contacts():
    _ensure_mailing_schema()
    q = (request.args.get('q') or '').strip().lower()
    if q:
        like = f'%{q}%'
        rows = query("SELECT id, nom, email, idioma, origen, subscrit FROM mailing_contacts "
                     "WHERE LOWER(email) LIKE ? OR LOWER(COALESCE(nom,'')) LIKE ? "
                     "ORDER BY id DESC LIMIT 500", [like, like]) or []
    else:
        rows = query("SELECT id, nom, email, idioma, origen, subscrit FROM mailing_contacts "
                     "ORDER BY id DESC LIMIT 500") or []
    total = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts', one=True), 'n', 0)
    items = [{
        'id': _row_get(r, 'id'),
        'nom': _row_get(r, 'nom', '') or '',
        'email': _row_get(r, 'email', ''),
        'idioma': _row_get(r, 'idioma', 'ca'),
        'origen': _row_get(r, 'origen', ''),
        'subscrit': bool(_row_get(r, 'subscrit')),
    } for r in rows]
    return jsonify(ok=True, contacts=items, shown=len(items), total=total)


@app.route('/admin/mailing/contacts/delete', methods=['POST'])
@admin_required
def admin_mailing_contacts_delete():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    cid = data.get('id')
    if not cid:
        return jsonify(ok=False, error='Falta id'), 400
    execute('DELETE FROM mailing_contacts WHERE id=?', [cid])
    total = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts', one=True), 'n', 0)
    subs = _row_get(query('SELECT COUNT(*) AS n FROM mailing_contacts WHERE subscrit', one=True), 'n', 0)
    return jsonify(ok=True, total=total, subscrits=subs, baixes=(total - subs))


@app.route('/admin/mailing/preview', methods=['POST'])
@admin_required
def admin_mailing_preview():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    cos = _mailing_text_to_html(data.get('cos', ''))
    fake = {'nom': 'Nom del client', 'token': 'PREVISUALITZACIO'}
    return jsonify(ok=True, html=_mailing_render_html(cos, fake))


@app.route('/admin/mailing/contacts/set-lang', methods=['POST'])
@admin_required
def admin_mailing_contacts_set_lang():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    cid = data.get('id')
    idioma = (data.get('idioma') or '').strip().lower()
    if not cid or idioma not in ('ca', 'es'):
        return jsonify(ok=False, error='Falta id o idioma no vàlid'), 400
    execute('UPDATE mailing_contacts SET idioma=? WHERE id=?', [idioma, cid])
    return jsonify(ok=True, idioma=idioma)


@app.route('/admin/mailing/test', methods=['POST'])
@admin_required
def admin_mailing_test():
    """Envia la prova al correu indicat. Si la campanya té versió en castellà,
    n'envia dues: [PROVA CA] i [PROVA ES], per revisar-les totes dues d'un clic."""
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    to = _mailing_valid_email(data.get('to', ''))
    if not to:
        return jsonify(ok=False, error='Adreca de prova no valida'), 400
    fake = {'nom': 'prova', 'token': 'PROVA-TOKEN'}
    baixa = f"{_mailing_base_url()}/baixa/PROVA-TOKEN"
    versions = [('CA', (data.get('assumpte') or '').strip() or '(sense assumpte)',
                 data.get('cos', ''))]
    if (data.get('assumpte_es') or '').strip() or (data.get('cos_es') or '').strip():
        versions.append(('ES', (data.get('assumpte_es') or '').strip() or '(sense assumpte)',
                         data.get('cos_es', '')))
    sent = 0
    err = ''
    for tag, assumpte, cos_raw in versions:
        html = _mailing_render_html(_mailing_text_to_html(cos_raw), fake)
        ok, e = _mailing_send_one(to, f'[PROVA {tag}] ' + assumpte, html, baixa)
        if ok:
            sent += 1
        else:
            err = e
    return jsonify(ok=(sent == len(versions)), sent=sent, total=len(versions), error=err)


@app.route('/admin/mailing/create', methods=['POST'])
@admin_required
def admin_mailing_create():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    assumpte = (data.get('assumpte') or '').strip()
    cos_raw = (data.get('cos') or '').strip()
    # Versio castellana opcional: si s'omple, els contactes amb idioma 'es'
    # reben aquesta; la resta, la catalana. Si es deixa buida, tothom rep
    # la catalana (comportament d'abans).
    assumpte_es = (data.get('assumpte_es') or '').strip()
    cos_es_raw = (data.get('cos_es') or '').strip()
    tipus = (data.get('tipus') or 'campanya').strip()
    idioma_filtre = (data.get('idioma') or 'tot').strip().lower()
    if idioma_filtre not in ('tot', 'ca', 'es'):
        idioma_filtre = 'tot'
    if not assumpte or not cos_raw:
        return jsonify(ok=False, error='Cal assumpte i cos en català (és la versió base)'), 400
    if bool(assumpte_es) != bool(cos_es_raw):
        return jsonify(ok=False, error='La versió en castellà necessita assumpte i cos (o cap dels dos)'), 400
    cos_html = _mailing_text_to_html(cos_raw)
    cos_html_es = _mailing_text_to_html(cos_es_raw) if cos_es_raw else ''
    if idioma_filtre == 'tot':
        contacts = query('SELECT id, nom, email, token, idioma FROM mailing_contacts WHERE subscrit') or []
    else:
        contacts = query('SELECT id, nom, email, token, idioma FROM mailing_contacts '
                         'WHERE subscrit AND idioma=?', [idioma_filtre]) or []
    if not contacts:
        return jsonify(ok=False, error='No hi ha destinataris subscrits per a aquest filtre'), 400
    uid = secrets.token_hex(8)
    execute('INSERT INTO mailing_campaigns (uid, assumpte, cos_html, assumpte_es, cos_html_es, tipus, idioma_filtre, estat, total, creat_per) '
            'VALUES (?,?,?,?,?,?,?,?,?,?)',
            [uid, assumpte, cos_html, assumpte_es, cos_html_es, tipus, idioma_filtre, 'sending', len(contacts),
             (session.get('nom') or 'admin')])
    camp = query('SELECT id FROM mailing_campaigns WHERE uid=?', [uid], one=True)
    cid = _row_get(camp, 'id')
    for c in contacts:
        execute('INSERT INTO mailing_sends (campaign_id, contact_id, email, nom, token, idioma, estat) '
                'VALUES (?,?,?,?,?,?,?)',
                [cid, _row_get(c, 'id'), _row_get(c, 'email'), _row_get(c, 'nom', ''),
                 _row_get(c, 'token', ''), (_row_get(c, 'idioma', 'ca') or 'ca'), 'pending'])
    return jsonify(ok=True, campaign_id=cid, total=len(contacts))


@app.route('/admin/mailing/send-chunk', methods=['POST'])
@admin_required
def admin_mailing_send_chunk():
    _ensure_mailing_schema()
    data = request.get_json(silent=True) or {}
    cid = data.get('campaign_id')
    camp = query('SELECT id, assumpte, cos_html, assumpte_es, cos_html_es FROM mailing_campaigns WHERE id=?', [cid], one=True)
    if not camp:
        return jsonify(ok=False, error='Campanya no trobada'), 404
    assumpte = _row_get(camp, 'assumpte')
    cos_html = _row_get(camp, 'cos_html')
    assumpte_es = (_row_get(camp, 'assumpte_es', '') or '').strip()
    cos_html_es = (_row_get(camp, 'cos_html_es', '') or '').strip()
    bilingue = bool(assumpte_es and cos_html_es)
    pend = query('SELECT id, email, nom, token, idioma FROM mailing_sends '
                 'WHERE campaign_id=? AND estat=? ORDER BY id LIMIT 20', [cid, 'pending']) or []
    for s in pend:
        contact = {'nom': _row_get(s, 'nom', ''), 'token': _row_get(s, 'token', '')}
        es = bilingue and (_row_get(s, 'idioma', 'ca') or 'ca') == 'es'
        html = _mailing_render_html(cos_html_es if es else cos_html, contact)
        baixa = f"{_mailing_base_url()}/baixa/{_row_get(s, 'token', '')}"
        ok, err = _mailing_send_one(_row_get(s, 'email'), assumpte_es if es else assumpte, html, baixa)
        if ok:
            execute('UPDATE mailing_sends SET estat=?, sent_at=? WHERE id=?',
                    ['sent', _mailing_now(), _row_get(s, 'id')])
        else:
            execute('UPDATE mailing_sends SET estat=?, error=? WHERE id=?',
                    ['failed', err, _row_get(s, 'id')])
        time.sleep(0.08)
    enviats = _row_get(query('SELECT COUNT(*) AS n FROM mailing_sends WHERE campaign_id=? AND estat=?',
                             [cid, 'sent'], one=True), 'n', 0)
    errs = _row_get(query('SELECT COUNT(*) AS n FROM mailing_sends WHERE campaign_id=? AND estat=?',
                          [cid, 'failed'], one=True), 'n', 0)
    remaining = _row_get(query('SELECT COUNT(*) AS n FROM mailing_sends WHERE campaign_id=? AND estat=?',
                               [cid, 'pending'], one=True), 'n', 0)
    done = (remaining == 0)
    if done:
        execute('UPDATE mailing_campaigns SET estat=?, enviats=?, errors=?, sent_at=? WHERE id=?',
                ['sent', enviats, errs, _mailing_now(), cid])
    else:
        execute('UPDATE mailing_campaigns SET enviats=?, errors=? WHERE id=?', [enviats, errs, cid])
    return jsonify(ok=True, enviats=enviats, errors=errs, remaining=remaining, done=done)


@app.route('/admin/mailing/campaign/<int:cid>')
@admin_required
def admin_mailing_campaign(cid):
    _ensure_mailing_schema()
    camp = query('SELECT * FROM mailing_campaigns WHERE id=?', [cid], one=True)
    if not camp:
        return jsonify(ok=False), 404
    fails = query('SELECT email, error FROM mailing_sends WHERE campaign_id=? AND estat=? '
                  'ORDER BY id LIMIT 100', [cid, 'failed']) or []
    return jsonify(
        ok=True,
        campaign={k: _row_get(camp, k) for k in
                  ['id', 'assumpte', 'tipus', 'estat', 'total', 'enviats', 'errors']},
        failures=[{'email': _row_get(f, 'email'), 'error': _row_get(f, 'error')} for f in fails],
    )


@app.route('/baixa/<token>', methods=['GET', 'POST'])
def mailing_unsubscribe(token):
    _ensure_mailing_schema()
    c = query('SELECT id, email, subscrit FROM mailing_contacts WHERE token=?', [token], one=True)
    if not c:
        if request.method == 'POST':
            return ('', 200)
        return render_template('baixa.html', estat='no_trobat'), 404
    if _row_get(c, 'subscrit'):
        execute('UPDATE mailing_contacts SET subscrit=?, unsubscribed_at=? WHERE id=?',
                [(False if USE_PG else 0), _mailing_now(), _row_get(c, 'id')])
    if request.method == 'POST':
        return ('', 200)
    return render_template('baixa.html', estat='ok', email=_row_get(c, 'email'))


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
